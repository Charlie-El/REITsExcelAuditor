from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from reit_excel_auditor.annual_update import (
    normalize_code,
    normalize_fund_name,
    normalize_project,
    parse_number,
    read_future_cashflow_rows,
    read_standard_rows,
)

AMOUNT_TOLERANCE = 0.01


def amount(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = str(value)
    if not any(ch.isdigit() for ch in text):
        return None
    return round(float(parse_number(value)), 6)


def same_amount(left: Any, right: Any) -> bool:
    left_amount = amount(left)
    right_amount = amount(right)
    if left_amount is None or right_amount is None:
        return left_amount == right_amount
    return abs(left_amount - right_amount) <= AMOUNT_TOLERANCE


def row_key(row: dict[str, Any]) -> tuple[str, str, int | None]:
    year = row.get("年份")
    return (
        normalize_code(row.get("REITs代码")),
        normalize_project(row.get("项目名称")),
        year if isinstance(year, int) else None,
    )


def code_year_key(row: dict[str, Any]) -> tuple[str, int | None]:
    year = row.get("年份")
    return (normalize_code(row.get("REITs代码")), year if isinstance(year, int) else None)


def build_name_code_lookup(standard_rows: list[dict[str, Any]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for row in standard_rows:
        code = normalize_code(row.get("REITs代码"))
        name = normalize_fund_name(row.get("REITs名称"))
        if code and name:
            lookup.setdefault(name, code)
    return lookup


def fill_missing_code_from_standard_names(rows: list[dict[str, Any]], standard_rows: list[dict[str, Any]]) -> int:
    name_lookup = build_name_code_lookup(standard_rows)
    fixed = 0
    for row in rows:
        if normalize_code(row.get("REITs代码")):
            row["REITs代码"] = normalize_code(row.get("REITs代码"))
            continue
        name = normalize_fund_name(row.get("REITs名称"))
        if not name:
            continue
        found = name_lookup.get(name)
        if not found:
            for candidate_name, candidate_code in name_lookup.items():
                if len(name) >= 8 and (name in candidate_name or candidate_name in name):
                    found = candidate_code
                    break
        if found:
            row["REITs代码"] = found
            fixed += 1
    return fixed


def index_exact(rows: list[dict[str, Any]]) -> dict[tuple[str, str, int | None], list[dict[str, Any]]]:
    indexed: dict[tuple[str, str, int | None], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = row_key(row)
        if key[0] and key[2] is not None and amount(row.get("预测现金流金额（万元）")) is not None:
            indexed[key].append(row)
    return indexed


def aggregate_by_code_year(rows: list[dict[str, Any]]) -> dict[tuple[str, int | None], float]:
    grouped: dict[tuple[str, int | None], float] = defaultdict(float)
    for row in rows:
        key = code_year_key(row)
        value = amount(row.get("预测现金流金额（万元）"))
        if key[0] and key[1] is not None and value is not None:
            grouped[key] += value
    return {key: round(value, 6) for key, value in grouped.items()}


def first_value(rows: list[dict[str, Any]], field: str) -> Any:
    for row in rows:
        value = row.get(field)
        if value not in (None, ""):
            return value
    return ""


def compare_rows(standard_rows: list[dict[str, Any]], ocr_rows: list[dict[str, Any]]) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]]]:
    standard_exact = index_exact(standard_rows)
    standard_aggregate = aggregate_by_code_year(standard_rows)
    ocr_aggregate = aggregate_by_code_year(ocr_rows)
    standard_years_by_code: dict[str, set[int]] = defaultdict(set)
    for row in standard_rows:
        code, year = code_year_key(row)
        if code and isinstance(year, int):
            standard_years_by_code[code].add(year)
    ocr_codes = {normalize_code(row.get("REITs代码")) for row in ocr_rows if normalize_code(row.get("REITs代码"))}
    ocr_years_by_code: dict[str, set[int]] = defaultdict(set)
    for row in ocr_rows:
        code, year = code_year_key(row)
        if code and isinstance(year, int):
            ocr_years_by_code[code].add(year)

    row_report: list[list[Any]] = []
    issue_report: list[list[Any]] = []
    for row in ocr_rows:
        value = amount(row.get("预测现金流金额（万元）"))
        code, project, year = row_key(row)
        if not code or year is None or value is None:
            status = "无法比较"
            note = "OCR 行缺少有效代码、年份或现金流金额"
            exact_sum = ""
            standard_sum = ""
            ocr_sum = ""
            diff = ""
        else:
            exact_rows = standard_exact.get((code, project, year), [])
            exact_sum_value = round(sum(amount(item.get("预测现金流金额（万元）")) or 0 for item in exact_rows), 6) if exact_rows else None
            standard_sum_value = standard_aggregate.get((code, year))
            ocr_sum_value = ocr_aggregate.get((code, year))
            exact_sum = exact_sum_value if exact_sum_value is not None else ""
            standard_sum = standard_sum_value if standard_sum_value is not None else ""
            ocr_sum = ocr_sum_value if ocr_sum_value is not None else ""
            if exact_rows and same_amount(exact_sum_value, value):
                status = "完全一致"
                note = ""
                diff = 0
            elif exact_rows:
                status = "金额不同"
                diff = round(value - float(exact_sum_value or 0), 6)
                note = "代码、项目、年份能匹配，但金额不同"
            elif standard_sum_value is None:
                if year not in standard_years_by_code.get(code, set()):
                    status = "标准表无对应年份"
                    note = "OCR 提取了标准表该基金未提供的年份，常见于报告披露周期长于标准表"
                else:
                    status = "标准表无对应代码年度"
                    note = "通常是 OCR 代码/基金名识别错误，或标准表没有该基金年度"
                diff = ""
            elif same_amount(standard_sum_value, ocr_sum_value):
                status = "项目名或颗粒度不一致但合计一致"
                diff = 0
                note = "金额按代码+年份汇总一致，主要差异在项目名称或项目拆分"
            else:
                status = "项目名或颗粒度不一致且合计不同"
                diff = round(float(ocr_sum_value or 0) - float(standard_sum_value or 0), 6)
                note = "需要检查 OCR 是否漏行、多行、年份错位或项目拆分错误"

        report_row = [
            status,
            code,
            row.get("REITs名称"),
            row.get("项目名称"),
            year,
            value,
            exact_sum,
            standard_sum,
            ocr_sum,
            diff,
            row.get("来源文件"),
            row.get("来源页码"),
            note,
            difference_reason(status),
            suggested_action(status),
        ]
        row_report.append(report_row)
        if status != "完全一致":
            issue_report.append(report_row)

    aggregate_report: list[list[Any]] = []
    aggregate_keys = set(ocr_aggregate)
    for code in ocr_codes:
        for year in ocr_years_by_code.get(code, set()):
            aggregate_keys.add((code, year))
    for key in sorted(aggregate_keys, key=lambda item: (item[0], item[1] or 0)):
        code, year = key
        standard_sum = standard_aggregate.get(key)
        ocr_sum = ocr_aggregate.get(key)
        if standard_sum is None:
            if year not in standard_years_by_code.get(code, set()):
                status = "标准表无对应年份"
            else:
                status = "标准表缺失"
            diff = ""
        elif ocr_sum is None:
            status = "OCR缺失"
            diff = ""
        elif same_amount(standard_sum, ocr_sum):
            status = "合计一致"
            diff = 0
        else:
            status = "合计不同"
            diff = round(float(ocr_sum or 0) - float(standard_sum or 0), 6)
        aggregate_report.append([status, code, year, standard_sum or "", ocr_sum or "", diff])

    return row_report, aggregate_report, issue_report


def difference_reason(status: str) -> str:
    reasons = {
        "完全一致": "代码、项目、年份和金额均能对齐。",
        "无法比较": "OCR/AI 标准化结果缺少关键字段，通常是截图不完整、OCR 识别失败或 AI 未能判断表头。",
        "金额不同": "同一代码、项目和年份存在金额差异，常见原因是 OCR 数字识别错位、AI 取错指标行、标准表人工调整或小数精度口径不同。",
        "标准表无对应年份": "OCR 抽到了标准表没有保留的年份，通常是报告披露周期更长，不一定是错误。",
        "标准表无对应代码年度": "该代码年度在标准表里不存在，常见原因是代码识别错误、基金名称未匹配，或标准表没有该基金。",
        "项目名或颗粒度不一致但合计一致": "截图披露的项目名或拆分颗粒度与标准表不同，但同一代码和年份的金额合计一致。",
        "项目名或颗粒度不一致且合计不同": "项目颗粒度不同且金额合计也不一致，通常需要检查是否漏取、多取或年份错位。",
    }
    return reasons.get(status, "未归类差异。")


def suggested_action(status: str) -> str:
    actions = {
        "完全一致": "无需处理。",
        "无法比较": "回看 OCR 原始识别文本和截图，必要时人工补充代码、项目、年份或金额。",
        "金额不同": "优先核对截图中对应年度的现金流行；若差异很小，可按四舍五入口径处理。",
        "标准表无对应年份": "明年正式使用时按基金到期日和报告披露周期判断是否保留，不要简单删除。",
        "标准表无对应代码年度": "检查 REITs 代码是否被 OCR 识别错；可用补全表按基金名称纠正。",
        "项目名或颗粒度不一致但合计一致": "确认正式表需要项目整体还是子项目；若只需要整体，可接受或合并。",
        "项目名或颗粒度不一致且合计不同": "回看截图，检查是否有合计行和明细行同时被取、年份列是否错位、是否遗漏某个子项目。",
    }
    return actions.get(status, "人工复核。")


def project_coverage_rows(standard_rows: list[dict[str, Any]], ocr_rows: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    codes = sorted({normalize_code(row.get("REITs代码")) for row in ocr_rows if normalize_code(row.get("REITs代码"))})
    for code in codes:
        standard_for_code = [row for row in standard_rows if normalize_code(row.get("REITs代码")) == code]
        ocr_for_code = [row for row in ocr_rows if normalize_code(row.get("REITs代码")) == code]
        standard_projects = sorted({row.get("项目名称") or "项目整体" for row in standard_for_code})
        ocr_projects = sorted({row.get("项目名称") or "项目整体" for row in ocr_for_code})
        rows.append(
            [
                code,
                first_value(ocr_for_code, "REITs名称") or first_value(standard_for_code, "REITs名称"),
                len(ocr_for_code),
                min([row.get("年份") for row in ocr_for_code if isinstance(row.get("年份"), int)] or [""]),
                max([row.get("年份") for row in ocr_for_code if isinstance(row.get("年份"), int)] or [""]),
                "；".join(ocr_projects),
                "；".join(standard_projects),
                "一致" if {normalize_project(item) for item in ocr_projects} == {normalize_project(item) for item in standard_projects} else "项目名称/颗粒度不同",
            ]
        )
    return rows


def append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def style_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for column in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 80)
            worksheet.column_dimensions[column[0].column_letter].width = width


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare only OCR/AI extracted cash-flow content against the standard future cash-flow workbook.")
    parser.add_argument("--ocr-standard", required=True, help="OCR/AI generated 年度更新_标准化导入表.xlsx")
    parser.add_argument("--standard-future", required=True, help="Standard 202604reits未来现金流.xlsx")
    parser.add_argument("--report", required=True, help="Output comparison report path")
    args = parser.parse_args()

    ocr_path = Path(args.ocr_standard)
    standard_path = Path(args.standard_future)
    report_path = Path(args.report)
    standard_rows = read_future_cashflow_rows(standard_path)
    ocr_rows = [
        row
        for row in read_standard_rows(ocr_path)
        if amount(row.get("预测现金流金额（万元）")) is not None
    ]
    fixed_codes = fill_missing_code_from_standard_names(ocr_rows, standard_rows)

    row_report, aggregate_report, issue_report = compare_rows(standard_rows, ocr_rows)
    coverage_report = project_coverage_rows(standard_rows, ocr_rows)

    status_counts: dict[str, int] = defaultdict(int)
    for row in row_report:
        status_counts[str(row[0])] += 1
    aggregate_counts: dict[str, int] = defaultdict(int)
    for row in aggregate_report:
        aggregate_counts[str(row[0])] += 1

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["项目", "数值", "说明"])
    summary_rows = [
        ["OCR现金流行数", len(ocr_rows), "仅统计有预测现金流金额的 OCR/AI 标准化行"],
        ["OCR基金数量", len({normalize_code(row.get("REITs代码")) for row in ocr_rows if normalize_code(row.get("REITs代码"))}), ""],
        ["用标准表基金名修正代码行数", fixed_codes, "用于处理 OCR 把代码识别成无效字符的情况"],
        ["逐行完全一致", status_counts.get("完全一致", 0), "代码+项目+年份+金额一致"],
        ["逐行需复核", len(issue_report), "包括金额不同、项目颗粒度不同、标准表缺失等"],
        ["代码年度合计一致", aggregate_counts.get("合计一致", 0), "忽略项目拆分后，代码+年份合计金额一致"],
        ["代码年度合计不同", aggregate_counts.get("合计不同", 0), "通常需要回看 OCR 文本或截图"],
        ["标准表无对应年份", aggregate_counts.get("标准表无对应年份", 0), "OCR 抽到了标准表该基金未提供的年份"],
        ["标准表缺失代码年度", aggregate_counts.get("标准表缺失", 0), "多为代码/基金名识别问题"],
    ]
    for row in summary_rows:
        summary.append(row)

    append_sheet(
        workbook,
        "逐行核对",
        ["状态", "REITs代码", "REITs名称", "OCR项目名称", "年份", "OCR金额", "标准同项目金额", "标准代码年度合计", "OCR代码年度合计", "差额", "来源文件", "来源页码", "说明", "差异原因", "建议处理"],
        row_report,
    )
    append_sheet(
        workbook,
        "代码年度汇总",
        ["状态", "REITs代码", "年份", "标准代码年度合计", "OCR代码年度合计", "差额"],
        aggregate_report,
    )
    append_sheet(
        workbook,
        "项目覆盖",
        ["REITs代码", "REITs名称", "OCR行数", "OCR最早年份", "OCR最晚年份", "OCR项目名称", "标准表项目名称", "判断"],
        coverage_report,
    )
    append_sheet(
        workbook,
        "需复核",
        ["状态", "REITs代码", "REITs名称", "OCR项目名称", "年份", "OCR金额", "标准同项目金额", "标准代码年度合计", "OCR代码年度合计", "差额", "来源文件", "来源页码", "说明", "差异原因", "建议处理"],
        issue_report,
    )

    style_workbook(workbook)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)
    workbook.close()

    print(f"report={report_path}")
    print(f"ocr_rows={len(ocr_rows)}")
    print(f"issue_rows={len(issue_report)}")
    print(f"fixed_codes={fixed_codes}")


if __name__ == "__main__":
    main()
