from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
import sys
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from reit_excel_auditor.annual_update import (
    AI_CALL_OUTPUT_NAME,
    STANDARD_INPUT_OUTPUT_NAME,
    classify_content_difference,
    compare_workbooks_for_report,
    explain_content_difference,
    normalize_code,
    normalize_project,
    projects_match,
    read_future_cashflow_rows,
    read_standard_rows,
)


COMPARE_FIELDS = [
    "预测现金流金额（万元）",
    "营业收入（万元）",
    "基础资产评估价值（万元）",
    "折现率",
    "评估基准日",
    "公告日期",
    "到期日",
    "基金净资产（万元）",
    "折旧及摊销（万元）",
    "固定管理费率(%)",
    "托管费率(%)",
]


def row_key(row: dict[str, Any]) -> tuple[str, str, Any]:
    return (normalize_code(row.get("REITs代码")), normalize_project(row.get("项目名称")), row.get("年份"))


def amount(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(value), 6)
    except Exception:
        return None


def index_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str, Any], list[dict[str, Any]]]:
    indexed: dict[tuple[str, str, Any], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        indexed[row_key(row)].append(row)
    return indexed


def compare_row_sets(
    label: str,
    standard_rows: list[dict[str, Any]],
    output_rows: list[dict[str, Any]],
    *,
    only_output_codes: bool = True,
    output_start_year: int | None = None,
) -> list[list[Any]]:
    standard_index = index_rows(standard_rows)
    output_index = index_rows(output_rows)
    output_codes = {key[0] for key in output_index if key[0]}
    standard_keys = set(standard_index)
    if only_output_codes:
        standard_keys = {key for key in standard_keys if key[0] in output_codes}

    diffs: list[list[Any]] = []
    alias_pairs = match_project_alias_keys(standard_index, output_index)
    paired_standard_keys = set(alias_pairs.values())
    paired_output_keys = set(alias_pairs)

    for output_key, standard_key in sorted(alias_pairs.items(), key=lambda item: (item[0][0], item[0][1], str(item[0][2]))):
        standard_group = standard_index.get(standard_key, [])
        output_group = output_index.get(output_key, [])
        if not standard_group or not output_group:
            continue
        standard_value = standard_group[0].get("预测现金流金额（万元）")
        output_value = output_group[0].get("预测现金流金额（万元）")
        level = "需复核" if amount(standard_value) == amount(output_value) else "风险差异"
        diffs.append(
            [
                label,
                level,
                "项目名称/颗粒度不同",
                output_key[0],
                f"{standard_key[1]} -> {output_key[1]}",
                output_key[2],
                standard_value,
                output_value,
                standard_group[0].get("来源文件"),
                output_group[0].get("来源文件"),
                "同一代码和年份能按项目名模糊匹配；金额一致时通常是简称/全称或项目拆分口径差异。",
                "项目名与标准审核表不完全一致，程序已按模糊规则识别为同一项目或近似项目。",
                "确认正式表应使用标准审核表项目名；如辅助表更准确，可人工调整项目名映射规则。",
            ]
        )

    keys = sorted(standard_keys | set(output_index), key=lambda item: (item[0], item[1], str(item[2])))
    for key in keys:
        if key in paired_standard_keys or key in paired_output_keys:
            continue
        standard_group = standard_index.get(key, [])
        output_group = output_index.get(key, [])
        if not standard_group:
            for output in output_group:
                diffs.append(
                    [
                        label,
                        "需复核",
                        "输出多出",
                        key[0],
                        key[1],
                        key[2],
                        "",
                        output.get("预测现金流金额（万元）"),
                        "",
                        output.get("来源文件"),
                        "标准表无该代码/项目/年份",
                        "输出包含标准审核表没有的项目或年份，可能是新项目、新年度，或项目名称未匹配。",
                        "如果是新增项目可以保留；如果只是简称/全称差异，建议补充项目别名后重跑。",
                    ]
                )
            continue

        if not output_group:
            for standard in standard_group:
                missing_before_start = is_year_before_start(key[2], output_start_year)
                diffs.append(
                    [
                        label,
                        "正常更新" if missing_before_start else "风险差异",
                        "输出缺失",
                        key[0],
                        key[1],
                        key[2],
                        standard.get("预测现金流金额（万元）"),
                        "",
                        standard.get("来源文件"),
                        "",
                        "输出未生成该代码/项目/年份",
                        "该年份低于本次正式底稿输出起始年度，按规则不输出。" if missing_before_start else "标准审核表有对应记录但输出缺失，可能是标准导入表漏项、资产性质识别错误或项目名未匹配。",
                        "通常无需处理；如今年仍需追溯展示，请调低输出起始年度后重跑。" if missing_before_start else "优先检查标准化导入表是否包含该代码/项目/年份，以及底层资产性质是否正确。",
                    ]
                )
            continue

        if len(standard_group) != len(output_group):
            diffs.append(
                [
                    label,
                    "需复核",
                    "重复数量不同",
                    key[0],
                    key[1],
                    key[2],
                    len(standard_group),
                    len(output_group),
                    "",
                    "",
                    "同一代码/项目/年份的行数不同",
                    "同一键下标准表和输出表重复行数量不同，通常来自项目拆分或重复 OCR/导入记录。",
                    "检查标准化导入表是否有重复行，必要时删除重复后重跑。",
                ]
            )

        for index in range(max(len(standard_group), len(output_group))):
            standard = standard_group[min(index, len(standard_group) - 1)]
            output = output_group[min(index, len(output_group) - 1)]
            for field in COMPARE_FIELDS:
                standard_value = standard.get(field)
                output_value = output.get(field)
                if amount(standard_value) is not None or amount(output_value) is not None:
                    same = amount(standard_value) == amount(output_value)
                else:
                    same = str(standard_value or "") == str(output_value or "")
                if same:
                    continue
                detail = ""
                if field == "预测现金流金额（万元）" and amount(standard_value) is not None and amount(output_value) is not None:
                    detail = f"差额={round(float(output_value) - float(standard_value), 6)}"
                level = classify_content_difference(field, standard_value, output_value)
                reason, action = explain_content_difference(field, standard_value, output_value)
                diffs.append(
                    [
                        label,
                        level,
                        f"{field}不同",
                        key[0],
                        key[1],
                        key[2],
                        standard_value,
                        output_value,
                        standard.get("来源文件"),
                        output.get("来源文件"),
                        detail,
                        reason,
                        action,
                    ]
                )
    return diffs


def match_project_alias_keys(
    standard_index: dict[tuple[str, str, Any], list[dict[str, Any]]],
    output_index: dict[tuple[str, str, Any], list[dict[str, Any]]],
) -> dict[tuple[str, str, Any], tuple[str, str, Any]]:
    pairs: dict[tuple[str, str, Any], tuple[str, str, Any]] = {}
    used_standard: set[tuple[str, str, Any]] = set()
    missing_standard = set(standard_index) - set(output_index)
    extra_output = set(output_index) - set(standard_index)
    for output_key in sorted(extra_output, key=lambda item: (item[0], item[1], str(item[2]))):
        code, output_project, year = output_key
        candidates = [
            standard_key
            for standard_key in missing_standard
            if standard_key not in used_standard
            and standard_key[0] == code
            and standard_key[2] == year
            and projects_match(standard_key[1], output_project)
        ]
        if not candidates:
            continue
        best = max(candidates, key=lambda item: project_similarity(item[1], output_project))
        pairs[output_key] = best
        used_standard.add(best)
    return pairs


def project_similarity(left: str, right: str) -> float:
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.96
    # Avoid importing difflib here; projects_match already handles the fuzzy
    # threshold, so length proximity is enough to pick a stable candidate.
    return 1 - abs(len(left) - len(right)) / max(len(left), len(right), 1)


def is_year_before_start(year: Any, output_start_year: int | None) -> bool:
    if output_start_year is None:
        return False
    try:
        return int(year) < output_start_year
    except Exception:
        return False


def read_ai_summary(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()

    success = [row for row in rows if row.get("状态") == "成功"]
    failed = [row for row in rows if row.get("状态") == "失败"]
    skipped = [row for row in rows if row.get("状态") == "跳过"]
    return {
        "total_batches": len(rows),
        "success_batches": len(success),
        "failed_batches": len(failed),
        "skipped_batches": len(skipped),
        "rows_from_ai": sum(row.get("输出行数") or 0 for row in success),
        "elapsed_seconds": round(sum(float(row.get("耗时秒") or 0) for row in rows), 2),
        "input_tokens": sum(row.get("输入Token") or 0 for row in rows),
        "output_tokens": sum(row.get("输出Token") or 0 for row in rows),
        "total_tokens": sum(row.get("总Token") or 0 for row in rows),
        "max_elapsed": max([float(row.get("耗时秒") or 0) for row in rows] or [0]),
    }


def first_match(output_dir: Path, patterns: list[str]) -> Path | None:
    for pattern in patterns:
        matches = sorted(output_dir.glob(pattern))
        if matches:
            return matches[0]
    return None


def write_rows(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]], limit: int = 5000) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows[:limit]:
        worksheet.append(row)
    if len(rows) > limit:
        worksheet.append(["已截断", f"仅写入前{limit}条，共{len(rows)}条"])


def level_counts(rows: list[list[Any]]) -> dict[str, int]:
    counts = {"正常更新": 0, "需复核": 0, "风险差异": 0}
    for row in rows:
        if len(row) < 2:
            continue
        level = str(row[1] or "")
        counts[level] = counts.get(level, 0) + 1
    return counts


def style_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for column in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            worksheet.column_dimensions[column[0].column_letter].width = width


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare annual OCR/AI outputs against standard workbooks.")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--standard-future", required=True)
    parser.add_argument("--standard-property", required=True)
    parser.add_argument("--standard-concession", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--output-start-year", type=int, default=2027)
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    standard_future = Path(args.standard_future)
    standard_property = Path(args.standard_property)
    standard_concession = Path(args.standard_concession)
    report_path = Path(args.report)

    output_standard = first_match(output_dir, [STANDARD_INPUT_OUTPUT_NAME, "标准化导入表.xlsx"])
    output_future = first_match(output_dir, ["年度更新_未来现金流汇总表.xlsx", "*未来现金流*自动更新.xlsx"])
    output_property = first_match(output_dir, ["*产权*自动更新.xlsx"])
    output_concession = first_match(output_dir, ["*特许经营权*自动更新.xlsx"])
    ai_calls = first_match(output_dir, [AI_CALL_OUTPUT_NAME, "AI调用记录.xlsx"]) or (output_dir / AI_CALL_OUTPUT_NAME)

    standard_future_rows = read_future_cashflow_rows(standard_future)
    output_standard_rows = read_standard_rows(output_standard) if output_standard else []
    output_future_rows = read_future_cashflow_rows(output_future) if output_future else []
    standard_property_rows = read_standard_rows(standard_property)
    output_property_rows = read_standard_rows(output_property) if output_property else []
    standard_concession_rows = read_standard_rows(standard_concession)
    output_concession_rows = read_standard_rows(output_concession) if output_concession else []

    standard_diffs = compare_row_sets("标准导入表 vs 未来现金流标准表", standard_future_rows, output_standard_rows)
    future_diffs = compare_row_sets("未来现金流宽表", standard_future_rows, output_future_rows)
    property_diffs = compare_row_sets(
        "产权正式表语义行",
        standard_property_rows,
        output_property_rows,
        output_start_year=args.output_start_year,
    )
    concession_diffs = compare_row_sets(
        "特许经营权正式表语义行",
        standard_concession_rows,
        output_concession_rows,
        output_start_year=args.output_start_year,
    )

    workbook_rows: list[list[Any]] = []
    for label, standard_path, output_path in [
        ("产权正式表工作簿", standard_property, output_property),
        ("特许经营权正式表工作簿", standard_concession, output_concession),
        ("未来现金流宽表工作簿", standard_future, output_future),
    ]:
        if not output_path or not output_path.exists():
            workbook_rows.append([label, str(standard_path), str(output_path or ""), "未生成输出", "", "", ""])
            continue
        try:
            structure, content = compare_workbooks_for_report(label, standard_path, output_path)
        except Exception as exc:
            workbook_rows.append([label, str(standard_path), str(output_path), "对比失败", "", "", str(exc)])
            continue
        workbook_rows.append([label, str(standard_path), str(output_path), "已对比", len(structure), len(content), "详见后续样例，最多各200条"])
        for item in structure[:200]:
            workbook_rows.append([label, "", "", "结构/格式差异样例", "", "", " | ".join(map(str, item))])
        for item in content[:200]:
            workbook_rows.append([label, "", "", "内容差异样例", "", "", " | ".join(map(str, item))])

    ai_summary = read_ai_summary(ai_calls)
    output_codes = {normalize_code(row.get("REITs代码")) for row in output_standard_rows if normalize_code(row.get("REITs代码"))}
    same_code_standard_rows = [row for row in standard_future_rows if normalize_code(row.get("REITs代码")) in output_codes]

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    rows = [
        ["项目", "数值", "说明"],
        ["AI标准化输出行数", len(output_standard_rows), "本次 AI 生成并归一化后的标准行"],
        ["输出涉及REITs代码数", len(output_codes), ""],
        ["标准表同代码范围行数", len(same_code_standard_rows), "function 根目录未来现金流标准表中同代码范围"],
        ["标准导入表行级差异数", len(standard_diffs), "仅比较输出涉及代码范围"],
        ["未来现金流宽表行级差异数", len(future_diffs), ""],
        ["产权正式表语义差异数", len(property_diffs), ""],
        ["特许经营权正式表语义差异数", len(concession_diffs), "未生成时会有缺失差异"],
    ]
    for label, diffs in [
        ("标准导入表", standard_diffs),
        ("未来现金流宽表", future_diffs),
        ("产权正式表", property_diffs),
        ("特许经营权正式表", concession_diffs),
    ]:
        counts = level_counts(diffs)
        rows.append([f"{label}-正常更新", counts.get("正常更新", 0), "通常来自今年截图、标准导入表或辅助表写入，可抽样复核。"])
        rows.append([f"{label}-需复核", counts.get("需复核", 0), "通常是项目颗粒度、公式、重复记录或日期口径，需要人工确认。"])
        rows.append([f"{label}-风险差异", counts.get("风险差异", 0), "可能影响最终底稿，建议优先检查。"])
    for key, value in ai_summary.items():
        rows.append([f"AI-{key}", value, f"来自 {AI_CALL_OUTPUT_NAME}"])
    for row in rows:
        summary.append(row)

    headers = [
        "对比对象",
        "判断级别",
        "差异类型",
        "REITs代码",
        "项目名称",
        "年份",
        "标准值",
        "输出值",
        "标准来源",
        "输出来源",
        "说明",
        "差异原因",
        "建议处理",
    ]
    write_rows(workbook, "StandardDiffs", headers, standard_diffs)
    write_rows(workbook, "FutureDiffs", headers, future_diffs)
    write_rows(workbook, "PropertyDiffs", headers, property_diffs)
    write_rows(workbook, "ConcessionDiffs", headers, concession_diffs)
    write_rows(workbook, "WorkbookSamples", ["表类型", "标准表", "输出表", "状态/差异类型", "结构差异数", "内容差异数", "说明"], workbook_rows)
    style_workbook(workbook)
    workbook.save(report_path)
    workbook.close()

    print(f"report={report_path}")
    print(f"output_standard_rows={len(output_standard_rows)}")
    print(f"output_future_rows={len(output_future_rows)}")
    print(f"standard_future_rows={len(standard_future_rows)}")
    print(f"standard_diffs={len(standard_diffs)}")
    print(f"future_diffs={len(future_diffs)}")
    print(f"property_diffs={len(property_diffs)}")
    print(f"concession_diffs={len(concession_diffs)}")
    print(f"ai_summary={ai_summary}")


if __name__ == "__main__":
    main()
