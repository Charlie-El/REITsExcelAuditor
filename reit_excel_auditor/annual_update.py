from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from functools import lru_cache
import hashlib
import importlib.util
import json
import os
from pathlib import Path
import re
import shutil
import tempfile
import time
import urllib.error
import urllib.request
from zipfile import ZipFile
from typing import Any, Callable, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection


DEFAULT_DASHSCOPE_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DEFAULT_AI_MODEL = "qwen-flash"
DEFAULT_VISION_OCR_MODEL = "qwen-vl-ocr-latest"
DEFAULT_AI_BATCH_CHAR_LIMIT = 3000
DEFAULT_AI_REQUEST_TIMEOUT_SECONDS = 60
DEFAULT_AI_TOTAL_TIMEOUT_SECONDS = 300
DEFAULT_REPORT_YEAR = 2026
DEFAULT_OUTPUT_START_YEAR = 2027
MAX_EXCEL_TEXT_LENGTH = 32767
PROJECT_ROOT = Path(__file__).resolve().parents[1]
ANNUAL_TEMPLATE_DIR = PROJECT_ROOT / "standard_templates" / "annual_update"
ANNUAL_TEMPLATE_FILENAMES = {
    "property": "产权年报提取模板.xlsx",
    "concession": "特许经营权年报提取模板.xlsx",
    "future_cashflow": "未来现金流模板.xlsx",
}
SUMMARY_OUTPUT_NAME = "年度更新_结果汇总与复核清单.xlsx"
STANDARD_INPUT_OUTPUT_NAME = "年度更新_标准化导入表.xlsx"
OCR_OUTPUT_NAME = "年度更新_OCR原始识别结果.xlsx"
PLAN_OUTPUT_NAME = "年度更新_更新计划预览.xlsx"
REVIEW_OUTPUT_NAME = "年度更新_人工复核清单.xlsx"
COMPARISON_OUTPUT_NAME = "年度更新_输出对比检查.xlsx"
AI_CALL_OUTPUT_NAME = "年度更新_AI调用记录.xlsx"
ANNUAL_REPORT_EXTRACT_OUTPUT_NAME = "年度更新_年报基金净资产与折旧摊销提取表.xlsx"
FUTURE_CASHFLOW_OUTPUT_NAME = "年度更新_未来现金流汇总表.xlsx"
PROPERTY_DETAIL_FONT_NAME = "Times New Roman"
CONCESSION_DETAIL_FONT_NAME = "等线"
FUTURE_CASHFLOW_FONT_NAME = "微软雅黑"
CONCESSION_INTEREST_VAT_RATE = 0.0326
DETAIL_TERMINAL_FILL = PatternFill(fill_type=None)
DETAIL_FALLBACK_BORDER_SIDE = Side(style="thin", color="000000")
DETAIL_MIN_FORMATTED_ROWS = 120
YELLOW_FILL_RGB_SUFFIXES = {"FFFF00", "FFF2CC", "FFE699", "FFFF99", "FFFFCC"}
LEGACY_ANNUAL_OUTPUT_NAMES = {
    "更新结果汇总.xlsx",
    "标准化导入表.xlsx",
    "OCR原始识别结果.xlsx",
    "更新计划预览.xlsx",
    "人工复核清单.xlsx",
    "输出对比检查.xlsx",
    "AI调用记录.xlsx",
    "基金净资产及折旧摊销提取表_程序自动提取.xlsx",
    "202604reits未来现金流_自动更新.xlsx",
}

_RAPIDOCR_ENGINE: Any | None = None
_PADDLE_OCR_ENGINE: Any | None = None

OCR_KEYWORDS = (
    "预测现金流",
    "现金流",
    "折现率",
    "评估价值",
    "评估基准日",
    "残值",
    "折旧",
    "摊销",
    "基金净资产",
    "所有者权益",
)

STANDARD_FIELDS = [
    "REITs代码",
    "REITs名称",
    "基础设施项目类型",
    "底层资产性质",
    "股债",
    "上市日期",
    "上市年份",
    "到期日",
    "公告日期",
    "报告期",
    "项目名称",
    "是否整体项目",
    "年份",
    "预测现金流金额（万元）",
    "增长率预测起始年度",
    "预测现金流增长率",
    "经营期末",
    "现金流折现日期",
    "折现率",
    "评估基准日",
    "基础资产评估价值（万元）",
    "残值年度",
    "残值基础数据预测",
    "考虑残值现金流",
    "营业收入（万元）",
    "EBITDA（万元）",
    "运营资本披露值（万元）",
    "资本支出披露值（万元）",
    "借款本金（万元）",
    "借款利息（万元）",
    "基金净资产（万元）",
    "折旧及摊销（万元）",
    "固定管理费率(%)",
    "托管费率(%)",
    "调整浮动管理费",
    "来源文件",
    "来源页码",
    "备注",
]

FIELD_ALIASES = {
    "REITs代码": ["代码", "基金代码", "证券代码", "REITs代码", "REITs Code", "reits_code"],
    "REITs名称": ["名称", "基金名称", "基金简称", "证券简称", "REITs名称", "REITs Name", "reits_name"],
    "基础设施项目类型": ["基础资产类型", "基础设施项目类型", "项目类型", "资产类型"],
    "底层资产性质": ["资产性质", "底层资产性质", "权属性质"],
    "股债": ["股债", "股/债", "权益属性"],
    "上市日期": ["基金上市日", "上市日期", "发行日期"],
    "上市年份": ["上市年份"],
    "到期日": ["基金到期日", "到期日", "预期到期日", "经营期限"],
    "公告日期": ["公告日期", "披露日期", "信息披露日期"],
    "报告期": ["报告期", "数据来源", "报告来源"],
    "项目名称": ["项目名称", "资产项目名称", "基础设施项目名称", "基础资产名称"],
    "是否整体项目": ["是否整体项目", "项目层级", "整体项目"],
    "年份": ["年份", "年度", "预测年份", "现金流年份"],
    "预测现金流金额（万元）": [
        "预测现金流金额",
        "预测现金流金额（万元）",
        "基础资产预测现金流",
        "基础资产预测现金流（万元）",
        "现金流金额",
        "现金流",
        "运营净收益",
        "运营净现金流",
        "预测净现金流",
        "项目净现金流",
        "自由现金流",
        "全周期运营净收益",
        "评估对象全周期运营净收益",
        "验证-评估报告披露预测现金流金额（万元）",
    ],
    "增长率预测起始年度": ["增长率预测起始年度", "自年份", "增长起始年度", "预测起始年度"],
    "预测现金流增长率": ["预测现金流增长率", "未来增长率", "增长率", "现金流增长率"],
    "经营期末": ["经营期末", "期末回收", "期末回收金额", "终期回收", "残余价值", "期末余值", "期末补偿款", "移交补偿", "经营期满回收"],
    "现金流折现日期": ["现金流折现日期", "折现日期"],
    "折现率": ["折现率", "贴现率"],
    "评估基准日": ["评估基准日", "估值基准日期", "估值基准日"],
    "基础资产评估价值（万元）": [
        "基础资产评估价值（万元）",
        "评估报告评估价值（万元）",
        "评估价值（万元）",
        "基础资产评估价值",
        "评估价值",
    ],
    "残值年度": ["残值年度"],
    "残值基础数据预测": ["残值基础数据预测", "残值基础数据"],
    "考虑残值现金流": ["考虑残值现金流", "残值现金流"],
    "营业收入（万元）": ["营业收入（万元）", "营业收入"],
    "EBITDA（万元）": ["EBITDA（万元）", "EBITDA", "息税折旧摊销前利润"],
    "运营资本披露值（万元）": ["运营资本披露值（万元）", "运营资本披露值", "营运资金"],
    "资本支出披露值（万元）": ["资本支出披露值（万元）", "资本支出披露值", "资本支出"],
    "借款本金（万元）": ["借款本金（万元）", "借款本金"],
    "借款利息（万元）": ["借款利息（万元）", "借款利息"],
    "基金净资产（万元）": ["基金净资产（万元）", "基金净资产", "所有者权益合计", "净资产"],
    "折旧及摊销（万元）": ["折旧及摊销（万元）", "折旧及摊销", "本期折旧和摊销"],
    "固定管理费率(%)": ["固定管理费率(%)", "固定管理费率", "固定管理费"],
    "托管费率(%)": ["托管费率(%)", "托管费率", "托管费"],
    "调整浮动管理费": ["调整浮动管理费", "浮动管理费"],
    "来源文件": ["来源文件", "文件名"],
    "来源页码": ["来源页码", "页码"],
    "备注": ["备注", "说明", "识别置信说明"],
}

NUMERIC_FIELDS = {
    "预测现金流金额（万元）",
    "预测现金流增长率",
    "经营期末",
    "折现率",
    "基础资产评估价值（万元）",
    "残值基础数据预测",
    "考虑残值现金流",
    "营业收入（万元）",
    "EBITDA（万元）",
    "运营资本披露值（万元）",
    "资本支出披露值（万元）",
    "借款本金（万元）",
    "借款利息（万元）",
    "基金净资产（万元）",
    "折旧及摊销（万元）",
    "固定管理费率(%)",
    "托管费率(%)",
    "调整浮动管理费",
}
FEE_RATE_FIELDS = {"固定管理费率(%)", "托管费率(%)"}

INTEGER_FIELDS = {"年份", "增长率预测起始年度", "残值年度", "上市年份"}
DATE_FIELDS = {"上市日期", "到期日", "公告日期", "现金流折现日期", "评估基准日"}
RESIDUAL_PARAMETER_KEY = "__residual_parameters__"
RESIDUAL_PARAMETER_FIELDS = {"残值年度", "残值基础数据预测", "考虑残值现金流", "折现率"}

CONCESSION_HINTS = (
    "特许经营权",
    "收费权",
    "高速",
    "高速公路",
    "公路",
    "水务",
    "污水",
    "水利",
    "原水",
    "清洁能源",
    "新能源",
    "供热",
    "生物质",
)

PROPERTY_HINTS = (
    "产权",
    "产业园",
    "软件园",
    "仓储",
    "物流",
    "租赁住房",
    "保障性租赁住房",
    "商业",
    "购物中心",
    "奥特莱斯",
    "消费基础设施",
    "数据中心",
    "农贸市场",
)

METRIC_PROJECT_TERMS = (
    "运营净收益",
    "预测现金流",
    "净现金流",
    "现金流金额",
    "基础资产预测现金流",
)


class AnnualUpdateError(Exception):
    """Raised when the annual cash-flow update cannot continue."""


def find_internal_annual_template_path(kind: str) -> Path | None:
    filename = ANNUAL_TEMPLATE_FILENAMES.get(kind)
    if not filename:
        return None
    path = ANNUAL_TEMPLATE_DIR / filename
    return path if path.exists() else None


def is_internal_annual_template_path(path: Path | None) -> bool:
    if not path:
        return False
    try:
        resolved = path.resolve()
    except OSError:
        return False
    return any(
        template_path and template_path.resolve() == resolved
        for template_path in (
            find_internal_annual_template_path("property"),
            find_internal_annual_template_path("concession"),
            find_internal_annual_template_path("future_cashflow"),
        )
    )


@dataclass
class AnnualUpdateOptions:
    workspace_path: Path | str
    output_dir: Path | str | None = None
    standard_input_path: Path | str | None = None
    ocr_source_path: Path | str | None = None
    annual_report_source_path: Path | str | None = None
    ocr_engine: str = "auto"
    use_ai: bool = False
    api_key: str | None = None
    api_key_env: str = "DASHSCOPE_API_KEY"
    base_url: str = DEFAULT_DASHSCOPE_BASE_URL
    model: str = DEFAULT_AI_MODEL
    ocr_api_key: str | None = None
    ocr_api_key_env: str = "DASHSCOPE_API_KEY"
    ocr_base_url: str = DEFAULT_DASHSCOPE_BASE_URL
    ocr_model: str = DEFAULT_VISION_OCR_MODEL
    report_year: int = DEFAULT_REPORT_YEAR
    output_start_year: int = DEFAULT_OUTPUT_START_YEAR
    max_ocr_pages_per_file: int = 3
    max_ai_chars: int = DEFAULT_AI_BATCH_CHAR_LIMIT
    ai_items_per_batch: int = 1
    ai_request_timeout_seconds: int = DEFAULT_AI_REQUEST_TIMEOUT_SECONDS
    ai_total_timeout_seconds: int = DEFAULT_AI_TOTAL_TIMEOUT_SECONDS
    ai_stop_after_failures: int = 2
    excel_open_check: bool = True
    compact_outputs: bool = True
    allow_existing_context_fill: bool = True
    progress: Callable[[str], None] | None = None


@dataclass
class OcrItem:
    source_file: Path
    page: int | None
    method: str
    text: str
    used_for_ai: bool = False
    warning: str = ""


@dataclass
class AnnualUpdateResult:
    output_dir: Path
    ocr_file: Path
    standard_file: Path
    plan_file: Path
    review_file: Path
    summary_file: Path
    comparison_file: Path
    property_file: Path | None = None
    concession_file: Path | None = None
    future_cashflow_file: Path | None = None
    ai_call_file: Path | None = None
    annual_report_extract_file: Path | None = None
    standard_row_count: int = 0
    ocr_item_count: int = 0
    updated_cell_count: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class DiscoveredFiles:
    workspace: Path
    property_workbook: Path | None = None
    concession_workbook: Path | None = None
    future_cashflow_workbook: Path | None = None
    standard_input_workbook: Path | None = None
    property_format_reference: Path | None = None
    concession_format_reference: Path | None = None
    future_cashflow_format_reference: Path | None = None
    fee_workbook: Path | None = None
    valuation_workbook: Path | None = None
    residual_workbook: Path | None = None
    announcement_workbook: Path | None = None
    project_alias_workbook: Path | None = None
    net_asset_reference_workbook: Path | None = None
    excel_files: list[Path] = field(default_factory=list)
    annual_report_pdf_files: list[Path] = field(default_factory=list)
    annual_report_source_path: Path | None = None
    pdf_files: list[Path] = field(default_factory=list)
    image_files: list[Path] = field(default_factory=list)
    docx_files: list[Path] = field(default_factory=list)
    ocr_source_path: Path | None = None


def run_annual_update(options: AnnualUpdateOptions) -> AnnualUpdateResult:
    workspace = Path(options.workspace_path).expanduser().resolve()
    if not workspace.exists():
        raise AnnualUpdateError(f"工作文件夹不存在：{workspace}")

    output_dir = Path(options.output_dir).expanduser().resolve() if options.output_dir else default_output_dir(workspace)
    output_dir.mkdir(parents=True, exist_ok=True)
    progress = options.progress or (lambda _message: None)

    progress("正在识别工作文件夹中的表格和年报文件...")
    ocr_source_path = Path(options.ocr_source_path).expanduser().resolve() if options.ocr_source_path else None
    annual_report_source_path = (
        Path(options.annual_report_source_path).expanduser().resolve() if options.annual_report_source_path else None
    )
    discovered = discover_annual_files(
        workspace,
        ocr_source_path,
        annual_report_source_path,
        skip_paths=[output_dir],
    )
    internal_property_template = find_internal_annual_template_path("property")
    internal_concession_template = find_internal_annual_template_path("concession")
    internal_future_template = find_internal_annual_template_path("future_cashflow")
    property_source_workbook = discovered.property_workbook or internal_property_template
    concession_source_workbook = discovered.concession_workbook or internal_concession_template
    property_format_reference = internal_property_template or discovered.property_format_reference
    concession_format_reference = internal_concession_template or discovered.concession_format_reference
    warnings: list[str] = []
    if not discovered.property_workbook:
        if internal_property_template:
            warnings.append(f"未识别到产权年报提取表，将回退到内置模板 {internal_property_template.name} 生成产权输出。")
        else:
            warnings.append("未识别到产权年报提取表，产权输出将跳过。")
    if not discovered.concession_workbook:
        if internal_concession_template:
            warnings.append(
                f"未识别到特许经营权年报提取表，将回退到内置模板 {internal_concession_template.name} 生成特许经营权输出。"
            )
        else:
            warnings.append("未识别到特许经营权年报提取表，特许经营权输出将跳过。")
    if discovered.future_cashflow_workbook:
        warnings.append("识别到未来现金流宽表；为避免误读旧表或参考表，本流程不会把它作为默认输入，只会作为输出格式参考。")
    for label, reference_path in [
        ("产权格式参考表", discovered.property_format_reference),
        ("特许经营权格式参考表", discovered.concession_format_reference),
        ("未来现金流格式参考表", discovered.future_cashflow_format_reference),
    ]:
        if reference_path:
            if label == "未来现金流格式参考表":
                warnings.append(f"识别到{label}：{reference_path.name}；不会把它作为默认输入，只会借用其格式和布局。")
            elif label == "产权格式参考表" and discovered.property_workbook and not is_internal_annual_template_path(property_source_workbook):
                warnings.append(f"识别到{label}：{reference_path.name}；本次产权正式输出会借用其表头、列宽、颜色和数据行样式。")
            elif label == "特许经营权格式参考表" and discovered.concession_workbook and not is_internal_annual_template_path(concession_source_workbook):
                warnings.append(f"识别到{label}：{reference_path.name}；本次特许经营权正式输出会借用其表头、列宽、颜色和数据行样式。")
            else:
                warnings.append(f"识别到{label}：{reference_path.name}，将优先借用其格式和布局。")
    if not discovered.fee_workbook:
        warnings.append("未识别到管理费率表；固定管理费率、托管费率等字段会保持原值或留空。")
    if not discovered.valuation_workbook:
        warnings.append("未识别到评估价值/资产性质表；评估价值、资产性质等字段会保持原值或留空。")
    if not discovered.announcement_workbook:
        warnings.append("未识别到公告日期表；公告日期会优先使用标准导入表、OCR/AI 标准化结果或源模板既有信息。")
    if discovered.project_alias_workbook:
        warnings.append(f"识别到项目别名映射表：{discovered.project_alias_workbook.name}，将按映射表修正标准化项目名称。")
    if not discovered.annual_report_pdf_files:
        warnings.append("未识别到公募年报 PDF；基金净资产、折旧及摊销不会从年报自动提取。")

    summary_file = output_dir / SUMMARY_OUTPUT_NAME
    ocr_file = summary_file if options.compact_outputs else output_dir / OCR_OUTPUT_NAME

    progress("正在进行本地 PDF/图片文本提取与 OCR...")
    ocr_items, ocr_warnings = extract_ocr_items(discovered, options)
    warnings.extend(ocr_warnings)
    if not options.compact_outputs:
        write_ocr_workbook(ocr_file, ocr_items)

    progress("正在读取标准化现金流输入...")
    standard_rows: list[dict[str, Any]] = []
    standard_input_path = Path(options.standard_input_path) if options.standard_input_path else discovered.standard_input_workbook
    if standard_input_path:
        standard_rows.extend(read_standard_rows(standard_input_path))
        if not options.standard_input_path:
            warnings.append(f"识别到标准导入/统一补充表：{standard_input_path.name}，已作为本轮主输入。")

    local_ocr_rows: list[dict[str, Any]] = []
    if ocr_items:
        local_ocr_rows, local_ocr_warnings = standardize_ocr_locally(ocr_items)
        warnings.extend(local_ocr_warnings)
        if local_ocr_rows:
            standard_rows.extend(local_ocr_rows)
            warnings.append(
                f"已用本地规则从 OCR 文本中整理出 {len(local_ocr_rows)} 行现金流记录；"
                "AI 可作为可选增强，但不是必需步骤。"
            )

    ai_records: list[AiCallRecord] = []
    if options.use_ai:
        progress("正在调用 AI 将 OCR 文本整理为标准导入表...")
        ai_rows, ai_warnings, ai_records = standardize_ocr_with_ai(ocr_items, options, progress)
        warnings.extend(ai_warnings)
        ai_rows_to_write, skipped_ai_rows = filter_ai_rows_for_unparsed_ocr_sources(ai_rows, local_ocr_rows)
        if skipped_ai_rows:
            warnings.append(
                f"AI 返回的 {skipped_ai_rows} 行来自本地规则已成功解析的 OCR 来源，"
                "为避免重复或模型误识别，已仅保留在 AI 调用记录中供复核，未写入正式标准导入表。"
            )
        standard_rows.extend(ai_rows_to_write)
        if not options.compact_outputs:
            write_ocr_workbook(ocr_file, ocr_items)
    elif ocr_items and not local_ocr_rows:
        warnings.append("已生成 OCR 原始识别结果；未启用 AI 标准化，因此 OCR 文本不会直接写入目标表。")

    standard_rows = normalize_standard_rows(standard_rows)
    stash_residual_parameters_from_rows(standard_rows)
    if not standard_rows:
        warnings.append("未读取到可用于填表的标准化现金流数据；程序仍会输出过程表，目标字段将保持原值或留空。")
    review_items: list[dict[str, Any]] = []
    project_aliases = read_project_alias_rows(discovered.project_alias_workbook)
    apply_project_aliases_to_rows(standard_rows, project_aliases, review_items)
    fee_rows = read_lookup_rows(discovered.fee_workbook) if discovered.fee_workbook else {}
    valuation_rows = read_lookup_rows(discovered.valuation_workbook) if discovered.valuation_workbook else {}
    residual_rows = read_lookup_rows(discovered.residual_workbook) if discovered.residual_workbook else {}
    announcement_rows = read_lookup_rows(discovered.announcement_workbook) if discovered.announcement_workbook else {}
    enrich_rows_from_lookups(standard_rows, fee_rows, valuation_rows, {}, residual_rows, announcement_rows)

    annual_report_extract_file = None
    annual_report_lookup_rows: dict[str, dict[str, Any]] = {}
    annual_report_financial_rows: list[dict[str, Any]] = []
    reference_rows = read_annual_report_financial_reference(discovered.net_asset_reference_workbook)
    if discovered.annual_report_pdf_files and (standard_rows or reference_rows):
        progress("正在从公募年报 PDF 提取基金净资产和折旧摊销...")
        annual_report_financial_rows, annual_warnings = extract_annual_report_financial_rows(
            discovered.annual_report_pdf_files,
            standard_rows,
            reference_rows,
        )
        warnings.extend(annual_warnings)
        annual_report_lookup_rows = build_annual_report_financial_lookup(annual_report_financial_rows)
        if not options.compact_outputs:
            annual_report_extract_file = output_dir / ANNUAL_REPORT_EXTRACT_OUTPUT_NAME
            write_annual_report_financial_workbook(
                annual_report_extract_file,
                annual_report_financial_rows,
                reference_rows,
            )
    elif discovered.net_asset_reference_workbook:
        warnings.append("识别到基金净资产及折旧摊销参考表，但未读取到可用于匹配的标准行；参考表不会直接参与填表。")

    enrich_rows_from_lookups(standard_rows, {}, {}, annual_report_lookup_rows, residual_rows)
    standard_rows = collapse_generic_estimate_object_rows(standard_rows)
    apply_annual_period_defaults_to_rows(
        standard_rows,
        [discovered.property_workbook, discovered.concession_workbook],
        options.output_start_year,
        review_items,
        options.allow_existing_context_fill,
    )

    # The future-cashflow workbook is an output audit aid. Snapshot it before
    # detail-template canonicalization so it reflects the disclosed/imported
    # project wording and never includes formula-projected rows.
    future_rows = filter_disclosed_cashflow_rows(standard_rows)
    canonicalize_project_names_from_existing_workbooks(
        standard_rows,
        [discovered.property_workbook, discovered.concession_workbook],
    )
    apply_existing_detail_context_to_rows(
        standard_rows,
        [discovered.property_workbook, discovered.concession_workbook],
        review_items,
        options.allow_existing_context_fill,
    )
    fill_group_static_context(standard_rows)
    expand_standard_rows_with_year_skeletons(
        standard_rows,
        [discovered.property_workbook, discovered.concession_workbook],
        review_items,
    )

    standard_file = summary_file if options.compact_outputs else output_dir / STANDARD_INPUT_OUTPUT_NAME
    if not options.compact_outputs:
        write_standard_rows_workbook(standard_file, standard_rows)

    updated_cell_count = 0

    progress("正在输出未来现金流宽表...")
    future_template = (
        discovered.future_cashflow_format_reference
        or discovered.future_cashflow_workbook
        or find_future_reference_workbook(workspace)
        or internal_future_template
    )
    future_file = update_future_cashflow_workbook(future_template, future_rows, output_dir, review_items)

    property_file = None
    detail_rows = filter_rows_by_start_year(standard_rows, options.output_start_year)
    property_rows = filter_rows_by_asset_nature(detail_rows, "产权")
    if property_source_workbook and property_rows:
        progress("正在更新产权年报提取表...")
        property_file, count = update_detail_workbook(
            property_source_workbook,
            property_rows,
            "property",
            output_dir,
            review_items,
            property_format_reference,
        )
        updated_cell_count += count
    elif property_source_workbook:
        warnings.append("识别到产权年报提取表，但本次标准化数据没有产权记录；未生成产权正式输出。")

    concession_file = None
    concession_rows = filter_rows_by_asset_nature(detail_rows, "特许经营权")
    if concession_source_workbook and concession_rows:
        progress("正在更新特许经营权年报提取表...")
        concession_file, count = update_detail_workbook(
            concession_source_workbook,
            concession_rows,
            "concession",
            output_dir,
            review_items,
            concession_format_reference,
        )
        updated_cell_count += count
    elif concession_source_workbook:
        warnings.append("识别到特许经营权年报提取表，但本次标准化数据没有特许经营权记录；未生成特许经营权正式输出。")

    progress("正在输出校验和汇总表...")
    plan_file = summary_file if options.compact_outputs else output_dir / PLAN_OUTPUT_NAME
    review_file = summary_file if options.compact_outputs else output_dir / REVIEW_OUTPUT_NAME
    comparison_file = summary_file if options.compact_outputs else output_dir / COMPARISON_OUTPUT_NAME
    ai_call_file = None if options.compact_outputs else (output_dir / AI_CALL_OUTPUT_NAME if ai_records else None)
    comparison_pairs = [
        ("产权", property_source_workbook, property_file, property_format_reference),
        ("特许经营权", concession_source_workbook, concession_file, concession_format_reference),
        ("未来现金流", future_template, future_file, future_template),
    ]
    if not options.compact_outputs:
        write_plan_workbook(plan_file, standard_rows)
        if ai_call_file:
            write_ai_call_workbook(ai_call_file, ai_records)
        write_comparison_workbook(comparison_file, comparison_pairs)
    if options.excel_open_check:
        progress("正在用本机 Excel 校验输出文件可打开性...")
        warnings.extend(
            repair_workbooks_with_excel_if_needed(
                [
                    property_file,
                    concession_file,
                    future_file,
                    standard_file,
                    ocr_file,
                    plan_file,
                    comparison_file,
                    annual_report_extract_file,
                    ai_call_file,
                ]
            )
        )
    if options.compact_outputs:
        write_process_workbook(
            summary_file,
            discovered,
            standard_rows,
            ocr_items,
            ai_records,
            review_items,
            warnings,
            [property_file, concession_file, future_file],
            comparison_pairs,
            annual_report_financial_rows,
            reference_rows,
        )
    else:
        write_review_workbook(review_file, review_items, warnings)
        write_summary_workbook(
            summary_file,
            discovered,
            standard_rows,
            ocr_items,
            warnings,
            [property_file, concession_file, future_file, comparison_file, annual_report_extract_file],
        )

    return AnnualUpdateResult(
        output_dir=output_dir,
        ocr_file=ocr_file,
        standard_file=standard_file,
        plan_file=plan_file,
        review_file=review_file,
        summary_file=summary_file,
        comparison_file=comparison_file,
        property_file=property_file,
        concession_file=concession_file,
        future_cashflow_file=future_file,
        ai_call_file=ai_call_file,
        annual_report_extract_file=annual_report_extract_file,
        standard_row_count=len(standard_rows),
        ocr_item_count=len(ocr_items),
        updated_cell_count=updated_cell_count,
        warnings=warnings,
    )


def default_output_dir(workspace: Path) -> Path:
    base = workspace if workspace.is_dir() else workspace.parent
    return base / "年度更新_输出结果"


HELPER_WORKBOOK_DIR_HINTS = (
    "辅助",
    "helper",
    "补充",
    "补全",
    "导入",
    "补录",
    "填写",
)

ANNUAL_REPORT_DIR_HINTS = (
    "公募reits年报",
    "公募年报",
    "年报pdf",
    "年报",
    "annualreport",
    "annual_report",
)

OCR_SOURCE_DIR_HINTS = (
    "现金流",
    "ocr",
    "截图",
    "截圖",
    "摘页",
    "评估报告",
    "手工",
    "人工",
    "screen",
    "capture",
)

CHECKED_REFERENCE_DIR_HINTS = (
    "对比今年参考样表",
    "参考样表",
    "标准审核表",
    "已核表",
    "已审核表",
    "格式参考",
    "reference",
)


def path_matches_any_hint(path: Path | str, hints: Iterable[str]) -> bool:
    normalized = normalize_text(str(path)).lower()
    return any(normalize_text(hint).lower() in normalized for hint in hints)


def is_helper_workbook_dir(path: Path | str) -> bool:
    return path_matches_any_hint(path, HELPER_WORKBOOK_DIR_HINTS)


def is_annual_report_dir(path: Path | str) -> bool:
    normalized = normalize_text(str(path)).lower()
    if "评估报告" in normalized:
        return False
    return path_matches_any_hint(path, ANNUAL_REPORT_DIR_HINTS)


def is_likely_ocr_source_dir(path: Path | str) -> bool:
    return path_matches_any_hint(path, OCR_SOURCE_DIR_HINTS)


def is_checked_reference_named_dir(path: Path | str) -> bool:
    return path_matches_any_hint(path, CHECKED_REFERENCE_DIR_HINTS)


def progress_message(options: AnnualUpdateOptions, message: str) -> None:
    if options.progress:
        options.progress(message)


def discover_annual_files(
    workspace: Path,
    ocr_source_path: Path | None = None,
    annual_report_source_path: Path | None = None,
    skip_paths: Iterable[Path] | None = None,
) -> DiscoveredFiles:
    root = workspace if workspace.is_dir() else workspace.parent
    files = [workspace] if workspace.is_file() else [path for path in root.rglob("*") if path.is_file()]
    discovered = DiscoveredFiles(workspace=root)
    skip_roots = [path.resolve() for path in (skip_paths or []) if Path(path).exists()]
    format_references = discover_checked_format_references(root)
    discovered.property_format_reference = format_references.get("property")
    discovered.concession_format_reference = format_references.get("concession")
    discovered.future_cashflow_format_reference = format_references.get("future_cashflow")
    format_reference_paths = {path.resolve() for path in format_references.values() if path.exists()}
    annual_report_files: list[Path] = []
    workbook_candidates: dict[str, list[tuple[int, Path]]] = {}
    if annual_report_source_path:
        if not annual_report_source_path.exists():
            raise AnnualUpdateError(f"公募年报文件夹不存在：{annual_report_source_path}")
        discovered.annual_report_source_path = annual_report_source_path
        annual_report_files = (
            [annual_report_source_path]
            if annual_report_source_path.is_file()
            else [path for path in annual_report_source_path.rglob("*") if path.is_file()]
        )

    for path in files:
        if should_skip_path(path, skip_roots):
            continue
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"} and not path.name.startswith("~$"):
            if path.resolve() in format_reference_paths or is_likely_checked_reference_path(path):
                continue
            if is_generated_annual_output_file(path):
                continue
            discovered.excel_files.append(path)
            kind = detect_annual_workbook_kind(path)
            if kind:
                workbook_candidates.setdefault(kind, []).append((annual_workbook_priority(path, kind), path))
        elif suffix == ".pdf" and is_annual_report_pdf(path):
            discovered.annual_report_pdf_files.append(path)

    for kind, candidates in workbook_candidates.items():
        _score, selected = max(candidates, key=lambda item: (item[0], len(str(item[1]))))
        if kind == "property":
            discovered.property_workbook = selected
        elif kind == "concession":
            discovered.concession_workbook = selected
        elif kind == "future_cashflow":
            discovered.future_cashflow_workbook = selected
        elif kind == "standard_input":
            discovered.standard_input_workbook = selected
        elif kind == "fee":
            discovered.fee_workbook = selected
        elif kind == "valuation":
            discovered.valuation_workbook = selected
        elif kind == "residual":
            discovered.residual_workbook = selected
        elif kind == "announcement":
            discovered.announcement_workbook = selected
        elif kind == "project_alias":
            discovered.project_alias_workbook = selected
        elif kind == "net_asset_reference":
            discovered.net_asset_reference_workbook = selected
    seen_annual_reports = {path.resolve() for path in discovered.annual_report_pdf_files}
    for path in annual_report_files:
        if should_skip_path(path, skip_roots):
            continue
        if path.suffix.lower() != ".pdf":
            continue
        resolved = path.resolve()
        if resolved in seen_annual_reports:
            continue
        # When the user explicitly selects a report folder, every PDF in it is
        # treated as a candidate annual report. This avoids forcing a rigid file
        # naming convention on public REITs report downloads.
        discovered.annual_report_pdf_files.append(path)
        seen_annual_reports.add(resolved)
    selected_ocr_source = resolve_ocr_source(root, ocr_source_path)
    if selected_ocr_source:
        discovered.ocr_source_path = selected_ocr_source
        ocr_files = [selected_ocr_source] if selected_ocr_source.is_file() else [path for path in selected_ocr_source.rglob("*") if path.is_file()]
    else:
        ocr_files = files

    for path in ocr_files:
        if should_skip_path(path, skip_roots):
            continue
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            if not selected_ocr_source and not is_likely_cashflow_ocr_pdf(path):
                continue
            discovered.pdf_files.append(path)
        elif suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}:
            discovered.image_files.append(path)
        elif suffix == ".docx":
            discovered.docx_files.append(path)

    return discovered


def annual_workbook_priority(path: Path, kind: str) -> int:
    normalized_path = normalize_text(str(path))
    normalized_parent = normalize_text(str(path.parent))
    normalized_name = normalize_text(path.stem)
    score = 0
    if "validation_baseline" in normalized_path or "baseline" in normalized_path:
        score -= 200
    if "模板" in normalized_name or "template" in normalized_name:
        score -= 120
    if is_helper_workbook_dir(path.parent):
        score += 120
    if "去年表" in normalized_parent:
        score += 90
    if "今年标准表" in normalized_parent:
        score += 50
    if kind in {"property", "concession"} and ("过程稿" in normalized_name or "irr" in normalized_name):
        score += 40
    if kind in {"fee", "valuation", "residual", "announcement", "project_alias", "net_asset_reference"}:
        if "辅助表" in normalized_name or "补充" in normalized_name or "导入" in normalized_name:
            score += 30
        if "提取" in normalized_name or "资产性质" in normalized_name:
            score += 20
    if kind == "announcement" and is_annual_report_dir(path.parent):
        score -= 20
    return score


def discover_checked_format_references(workspace: Path) -> dict[str, Path]:
    """Find checked annual workbooks that should be used only as style references."""
    references: dict[str, Path] = {}
    for directory in checked_reference_directories(workspace):
        try:
            candidates = sorted(
                path
                for path in directory.glob("*.xlsx")
                if path.is_file() and not path.name.startswith("~$")
            )
        except OSError:
            continue
        scored: list[tuple[int, str, str, Path]] = []
        for path in candidates:
            kind = detect_checked_reference_kind(path)
            if not kind or kind in references:
                continue
            scored.append((checked_reference_score(path, kind), kind, path.name, path))
        for _score, kind, _name, path in sorted(scored, reverse=True):
            references.setdefault(kind, path)
        if {"property", "concession", "future_cashflow"}.issubset(references):
            break
    return references


def checked_reference_directories(workspace: Path) -> list[Path]:
    root = workspace if workspace.is_dir() else workspace.parent
    directories: list[Path] = []

    for base in [root, *root.parents]:
        if should_skip_path(base):
            continue
        directories.append(base)
        direct_reference = base / "对比今年参考样表"
        if direct_reference.exists() and direct_reference.is_dir():
            directories.append(direct_reference)
        try:
            for child in base.iterdir():
                if child.is_dir() and is_checked_reference_dir(child):
                    directories.append(child)
        except OSError:
            pass
        # Searching beyond the repository folder can accidentally pick up
        # unrelated desktop workbooks, so stop once the project root is reached.
        if (base / "reit_excel_auditor").exists() and (base / "pyproject.toml").exists():
            break

    unique: list[Path] = []
    seen: set[Path] = set()
    for directory in directories:
        try:
            resolved = directory.resolve()
        except OSError:
            continue
        if resolved in seen:
            continue
        seen.add(resolved)
        unique.append(directory)
    return unique


def is_checked_reference_dir(path: Path) -> bool:
    name = normalize_text(path.name).lower()
    if "辅助" in name or "helper" in name or "测试" in name or "输出" in name:
        return False
    return is_checked_reference_named_dir(path)


def is_likely_checked_reference_path(path: Path) -> bool:
    stem = normalize_text(path.stem).lower()
    if any(token in stem for token in ("辅助", "helper", "模拟", "去年", "自动更新", "输出", "结果汇总", "模板")):
        return False
    return any(token in stem for token in ("已核", "已审核", "参考样表", "标准审核")) or stem == "202604reits未来现金流"


def detect_checked_reference_kind(path: Path) -> str | None:
    stem = normalize_text(path.stem)
    if not is_likely_checked_reference_path(path):
        return None
    if "未来现金流" in stem:
        return "future_cashflow"
    if "特许" in stem:
        return "concession"
    if "产权" in stem:
        return "property"

    kind = detect_annual_workbook_kind(path)
    if kind in {"property", "concession", "future_cashflow"}:
        return kind
    return None


def checked_reference_score(path: Path, kind: str) -> int:
    stem = normalize_text(path.stem)
    score = 0
    if "已核" in stem or "已审核" in stem:
        score += 100
    if "参考" in stem or "标准" in stem:
        score += 60
    if "年报提取" in stem:
        score += 30
    if kind == "future_cashflow" and "202604" in stem:
        score += 30
    if is_checked_reference_named_dir(path.parent):
        score += 20
    return score


def resolve_ocr_source(workspace: Path, explicit_source: Path | None) -> Path | None:
    if explicit_source:
        if not explicit_source.exists():
            raise AnnualUpdateError(f"OCR 来源路径不存在：{explicit_source}")
        return explicit_source

    candidates: list[tuple[int, Path]] = []
    for folder in [workspace, *[path for path in workspace.rglob("*") if path.is_dir()]]:
        files = [path for path in folder.iterdir() if path.is_file()]
        useful_count = sum(1 for path in files if path.suffix.lower() in {".pdf", ".docx", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"})
        if useful_count == 0:
            continue
        normalized_path = normalize_text(str(folder))
        score = 0
        if is_likely_ocr_source_dir(folder):
            score += 60
        if "人工" in normalized_path or "示例" in normalized_path:
            score += 30
        if useful_count <= 40:
            score += 15
        if is_annual_report_dir(folder):
            score -= 50
        if useful_count > 60:
            score -= 20
        candidates.append((score, folder))

    candidates.sort(key=lambda item: (item[0], -len(str(item[1]))), reverse=True)
    if candidates and candidates[0][0] >= 60:
        return candidates[0][1]
    return None


def should_skip_path(path: Path, skip_roots: Iterable[Path] | None = None) -> bool:
    resolved_path = path.resolve()
    for root in skip_roots or []:
        try:
            resolved_path.relative_to(root)
            return True
        except ValueError:
            pass
    skip_dirs = {".git", "build", "dist", "__pycache__", ".pytest_cache"}
    generated_dir_names = {"输出结果"}
    generated_dir_prefixes = (
        "年度更新_输出结果",
        "年报现金流自动更新输出",
        "最终验证输出",
        "辅助表验证输出",
        "全流程验收",
        "口径优化验证",
        "aux_validation",
        "exe_aux_validation",
        "aux_step_debug",
        "aux_validation_debug",
        "测试输出",
    )
    for part in path.parts:
        if part in skip_dirs:
            return True
        if part in generated_dir_names:
            return True
        if any(part.startswith(prefix) for prefix in generated_dir_prefixes):
            return True
    return False


def is_generated_annual_output_file(path: Path) -> bool:
    if "_自动更新" in path.stem:
        return True
    return path.name in {
        SUMMARY_OUTPUT_NAME,
        STANDARD_INPUT_OUTPUT_NAME,
        OCR_OUTPUT_NAME,
        PLAN_OUTPUT_NAME,
        REVIEW_OUTPUT_NAME,
        COMPARISON_OUTPUT_NAME,
        AI_CALL_OUTPUT_NAME,
        ANNUAL_REPORT_EXTRACT_OUTPUT_NAME,
        FUTURE_CASHFLOW_OUTPUT_NAME,
        *LEGACY_ANNUAL_OUTPUT_NAMES,
    }


def detect_annual_workbook_kind(path: Path) -> str | None:
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return None
    try:
        for worksheet in workbook.worksheets[:3]:
            header_row, headers = find_header_row(worksheet)
            normalized = {normalize_text(header) for header in headers if header not in (None, "")}
            year_headers = [header for header in headers if isinstance(header, int)]
            if {"代码", "项目名称", "年份", "预测现金流金额"}.issubset(normalized):
                return "property"
            if "验证评估报告披露预测现金流金额万元" in normalized and {"代码", "项目名称", "年份"}.issubset(normalized):
                return "concession"
            has_standard_identity = "reits代码" in normalized or "REITs代码" in normalized
            has_standard_cashflow = "预测现金流金额万元" in normalized or "预测现金流金额" in normalized
            if has_standard_identity and {"项目名称", "年份"}.issubset(normalized) and has_standard_cashflow:
                return "standard_input"
            if {"名称", "代码", "项目名称"}.issubset(normalized) and len(year_headers) >= 5:
                return "future_cashflow"
            if {"固定管理费率", "托管费"}.issubset(normalized) or {"固定管理费率", "托管费率"}.issubset(normalized):
                return "fee"
            has_valuation_value = any(
                header in normalized
                for header in {
                    "评估价值元",
                    "基础资产评估价值元",
                    "评估报告评估价值元",
                    "评估价值万元",
                    "基础资产评估价值万元",
                    "评估报告评估价值万元",
                    "基础资产评估价值",
                    "评估价值",
                }
            )
            has_asset_nature = any(header in normalized for header in {"资产性质", "底层资产性质", "权属性质"})
            if has_valuation_value and has_asset_nature:
                return "valuation"
            if (
                {"残值年度", "残值基础数据预测"}.intersection(normalized)
                and ("代码" in normalized or "REITs代码" in normalized or "reits代码" in normalized)
                and not {"年份", "预测现金流金额"}.issubset(normalized)
            ):
                return "residual"
            has_alias_source = any(header in normalized for header in {"原项目名称", "来源项目名称", "输入项目名称", "OCR项目名称", "ocr项目名称", "项目别名", "别名"})
            has_alias_target = any(header in normalized for header in {"标准项目名称", "目标项目名称", "正式项目名称", "输出项目名称"})
            if has_alias_source and has_alias_target:
                return "project_alias"
            if (
                ("基金净资产万元" in normalized or "基金净资产元" in normalized)
                and ("折旧及摊销万元" in normalized or "折旧及摊销元" in normalized)
            ):
                return "net_asset_reference"
            has_announcement_date = any(header in normalized for header in {"公告日期", "披露日期", "信息披露日期"})
            has_fund_identity = any(
                header in normalized
                for header in {"代码", "REITs代码", "reits代码", "基金代码", "证券代码", "名称", "REITs名称", "reits名称", "基金名称", "相关基金"}
            )
            if {"标题", "相关基金"}.issubset(normalized) or (
                has_announcement_date
                and has_fund_identity
                and not year_headers
                and not {"年份", "预测现金流金额"}.issubset(normalized)
            ):
                return "announcement"
            if header_row == 1 and {"名称", "代码", "项目名称"}.issubset(normalized):
                return "future_cashflow"
    finally:
        workbook.close()
    return None


def is_annual_report_pdf(path: Path) -> bool:
    normalized_name = normalize_text(path.name)
    if is_annual_report_dir(path.parent):
        return True
    if "年度报告" in normalized_name and "评估报告" not in normalized_name:
        return True
    return False


def is_likely_cashflow_ocr_pdf(path: Path) -> bool:
    normalized_path = normalize_text(str(path.parent)) + normalize_text(path.name)
    if is_annual_report_dir(path.parent) or is_annual_report_dir(path):
        return False
    if is_annual_report_pdf(path):
        return False
    return any(keyword in normalized_path for keyword in ("现金流", "人工", "截图", "ocr", "示例", "摘页"))


def find_header_row(worksheet: Any, max_scan_rows: int = 12) -> tuple[int, list[Any]]:
    cache = getattr(worksheet, "_reit_header_row_cache", None)
    if (
        cache
        and cache.get("max_scan_rows") == max_scan_rows
        and cache.get("max_column") == worksheet.max_column
    ):
        return cache["header_row"], list(cache["headers"])

    best_row = 1
    best_headers: list[Any] = []
    best_score = -1
    max_col = min(max(worksheet.max_column, 1), 120)
    for row_idx in range(1, min(worksheet.max_row, max_scan_rows) + 1):
        headers = [worksheet.cell(row_idx, col_idx).value for col_idx in range(1, max_col + 1)]
        normalized = [normalize_text(header) for header in headers if header not in (None, "")]
        normalized_set = set(normalized)
        year_count = len([header for header in headers if isinstance(header, int) and 1900 <= header <= 2200])
        score = len(normalized)
        score += 25 if "代码" in normalized_set else 0
        score += 20 if "项目名称" in normalized_set else 0
        score += 12 if "名称" in normalized_set else 0
        score += 12 if "年份" in normalized_set else 0
        score += min(year_count, 20)
        if {"名称", "代码", "项目名称"}.issubset(normalized_set) and year_count >= 3:
            score += 50
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = headers
    try:
        worksheet._reit_header_row_cache = {
            "max_scan_rows": max_scan_rows,
            "max_column": worksheet.max_column,
            "header_row": best_row,
            "headers": tuple(best_headers),
        }
    except Exception:
        pass
    return best_row, best_headers


def extract_ocr_items(discovered: DiscoveredFiles, options: AnnualUpdateOptions) -> tuple[list[OcrItem], list[str]]:
    items: list[OcrItem] = []
    warnings: list[str] = []
    if options.max_ocr_pages_per_file < 0:
        warnings.append("已按设置跳过 PDF/图片 OCR，仅使用已有标准导入表；未来现金流宽表会作为本轮输出重新生成。")
        return items, warnings
    resolved_ocr_engine, engine_warning = resolve_ocr_engine(options.ocr_engine)
    if engine_warning:
        warnings.append(engine_warning)
    for docx_path in discovered.docx_files:
        method = "docx-text"
        try:
            text = extract_docx_text(docx_path)
        except Exception as exc:
            warnings.append(f"DOCX 文本提取失败：{docx_path.name}，原因：{exc}")
            text = ""
        if not text.strip() and resolved_ocr_engine:
            try:
                text = extract_docx_embedded_image_text(docx_path, resolved_ocr_engine, options)
                if text.strip():
                    method = "docx-image-ocr"
            except Exception as exc:
                warnings.append(f"DOCX 图片 OCR 失败：{docx_path.name}，原因：{exc}")
                text = ""
        if text.strip():
            items.append(OcrItem(source_file=docx_path, page=None, method=method, text=trim_excel_text(compact_text(text))))

    for pdf_path in discovered.pdf_files:
        try:
            items.extend(extract_pdf_text_and_ocr(pdf_path, options, warnings, resolved_ocr_engine))
        except Exception as exc:
            warnings.append(f"PDF 处理失败：{pdf_path.name}，原因：{exc}")

    for image_path in discovered.image_files:
        if not resolved_ocr_engine:
            continue
        try:
            text = ocr_image_file(image_path, resolved_ocr_engine, options)
        except Exception as exc:
            warnings.append(f"图片 OCR 失败：{image_path.name}，原因：{exc}")
            text = ""
        if text.strip():
            items.append(OcrItem(source_file=image_path, page=None, method="image-ocr", text=text.strip()))
    return items, warnings


OCR_NUMBER_RE = re.compile(r"-?(?:\d{1,3}(?:[,，]\d{3})+|\d+)(?:\.\d+)?")


def standardize_ocr_locally(ocr_items: list[OcrItem]) -> tuple[list[dict[str, Any]], list[str]]:
    """Best-effort local extraction for common cash-flow screenshots.

    The AI step is useful for messy OCR, but the workflow should still produce
    reviewable rows without sending anything out of the user's machine.
    """
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    last_project_by_file: dict[Path, str] = {}
    for item in ocr_items:
        item_rows = extract_cashflow_rows_from_ocr_item(item)
        if item_rows:
            continuation_project = last_project_by_file.get(item.source_file)
            if continuation_project:
                for row in item_rows:
                    if normalize_project(row.get("项目名称")) == "项目整体":
                        row["项目名称"] = continuation_project
                        row["备注"] = append_warning(str(row.get("备注") or ""), f"跨页延续上一页项目：{continuation_project}")
            for row in item_rows:
                project = row.get("项目名称")
                if project not in (None, "") and normalize_project(project) != "项目整体":
                    last_project_by_file[item.source_file] = str(project)
            rows.extend(item_rows)
        elif item.text.strip():
            warnings.append(f"本地规则未能解析现金流：{item.source_file.name} 第 {item.page or ''} 页；可启用 AI 或人工整理标准导入表。")
    return rows, warnings


def extract_cashflow_rows_from_ocr_item(item: OcrItem) -> list[dict[str, Any]]:
    lines = normalize_ocr_lines(item.text)
    if not lines:
        return []
    fund_name = clean_report_fund_name_from_filename(item.source_file)
    report_year = extract_report_year_from_filename(item.source_file.name)
    parsed_rows = parse_cashflow_table_segments(lines)
    if not parsed_rows:
        return []

    growth_start, growth_rate = extract_growth_instruction(item.text)
    terminal_recovery = extract_terminal_recovery(item.text)
    discount_rate = extract_discount_rate(item.text)
    valuation_amount = extract_valuation_amount_ten_thousand(item.text)
    max_disclosed_year = max((row["年份"] for row in parsed_rows if isinstance(row.get("年份"), int)), default=None)
    if growth_start and max_disclosed_year and growth_start <= max_disclosed_year:
        growth_start = max_disclosed_year + 1

    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, int, float]] = set()
    for parsed in parsed_rows:
        year = parsed.get("年份")
        amount = parsed.get("预测现金流金额（万元）")
        if not isinstance(year, int) or amount in (None, ""):
            continue
        key = (normalize_project(parsed.get("项目名称")), year, round(float(amount), 6))
        if key in seen:
            continue
        seen.add(key)
        row = {
            "REITs名称": fund_name,
            "项目名称": parsed.get("项目名称") or "项目整体",
            "年份": year,
            "预测现金流金额（万元）": amount,
            "来源文件": item.source_file.name,
            "来源页码": item.page,
            "备注": "本地OCR规则提取；请结合过程表复核。",
        }
        if report_year:
            row["报告期"] = f"{report_year}年评估报告"
        if growth_start and growth_rate not in (None, ""):
            row["增长率预测起始年度"] = growth_start
            row["预测现金流增长率"] = growth_rate
        if terminal_recovery is not None:
            row["经营期末"] = terminal_recovery
        if discount_rate is not None:
            row["折现率"] = discount_rate
        if valuation_amount is not None:
            row["基础资产评估价值（万元）"] = valuation_amount
        rows.append(row)
    return rows


def normalize_ocr_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw in text.replace("\r", "\n").splitlines():
        line = raw.strip()
        if not line:
            continue
        lines.append(line)
    return lines


def parse_cashflow_table_segments(lines: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    index = 0
    current_project = "项目整体"
    while index < len(lines):
        project_label = project_label_from_line(lines[index])
        if project_label:
            current_project = project_label
        years = years_from_line(lines[index])
        if not years:
            index += 1
            continue

        year_start = index
        cursor = index + 1
        while cursor < len(lines):
            more_years = years_from_line(lines[cursor])
            normalized = normalize_text(lines[cursor])
            if more_years:
                years.extend(more_years)
                cursor += 1
                continue
            if is_unit_or_table_scaffold_line(normalized):
                cursor += 1
                continue
            break

        if len(years) < 2:
            index += 1
            continue

        amount_start = find_cashflow_amount_start(lines, year_start, cursor)
        if amount_start is None:
            index += 1
            continue

        unit_multiplier = cashflow_unit_multiplier(lines, year_start, amount_start)
        values, end_index = collect_cashflow_amounts(lines, amount_start, len(years), unit_multiplier)
        if len(values) >= min(2, len(years)):
            project_name = project_name_near(lines, year_start)
            if project_name == "项目整体":
                project_name = current_project
            for year, value in zip(years, values):
                rows.append(
                    {
                        "项目名称": project_name,
                        "年份": year,
                        "预测现金流金额（万元）": round(value, 6),
                    }
                )
            index = max(end_index, cursor)
            continue
        index += 1
    return rows


def years_from_line(line: str) -> list[int]:
    normalized = normalize_text(line)
    if any(token in normalized for token in ("价值时点", "预测期自", "其余期间", "增长率", "收益期届满", "报告出具日")):
        return []
    years: list[int] = []
    for match in re.finditer(r"(20\d{2})\s*年(?:\s*\d{1,2}\s*月\s*\d{1,2}\s*日\s*止?)?", line):
        year = int(match.group(1))
        if 2020 <= year <= 2200:
            years.append(year)
    return years


def is_unit_or_table_scaffold_line(normalized: str) -> bool:
    return normalized in {"科目", "单位万元", "单位元"} or "单位万元" in normalized or "单位元" in normalized


def find_cashflow_amount_start(lines: list[str], year_start: int, cursor: int) -> int | None:
    title_has_cashflow = any(
        is_cashflow_metric_label(lines[row_idx], title_context=True)
        for row_idx in range(max(0, year_start - 6), year_start)
    )
    next_year_row = None
    for row_idx in range(cursor, min(len(lines), cursor + 250)):
        if years_from_line(lines[row_idx]):
            next_year_row = row_idx
            break
        if is_cashflow_metric_label(lines[row_idx]):
            return row_idx + 1
    if title_has_cashflow:
        return cursor
    return None if next_year_row is not None else None


def is_cashflow_metric_label(line: str, title_context: bool = False) -> bool:
    normalized = normalize_text(line)
    if not normalized:
        return False
    if any(stop in normalized for stop in ("说明", "提示声明", "上述年度", "不代表", "不可直接作为")):
        return False
    metric_terms = (
        "运营净收益",
        "运营净现金流",
        "预测现金流",
        "基础资产预测现金流",
        "自由现金流",
        "税前净现金流",
        "现金流量",
    )
    if not any(term in normalized for term in metric_terms):
        return False
    if title_context:
        return True
    return len(normalized) <= 30 or normalized in metric_terms


def cashflow_unit_multiplier(lines: list[str], start: int, end: int) -> float:
    window = "\n".join(lines[max(0, start - 3) : min(len(lines), end + 3)])
    normalized = normalize_text(window)
    if "单位元" in normalized and "单位万元" not in normalized:
        return 0.0001
    return 1.0


def collect_cashflow_amounts(lines: list[str], start: int, expected_count: int, unit_multiplier: float) -> tuple[list[float], int]:
    values: list[float] = []
    end_index = start
    for row_idx in range(start, min(len(lines), start + expected_count + 30)):
        line = lines[row_idx].strip()
        normalized = normalize_text(line)
        if values and (years_from_line(line) or any(term in normalized for term in ("说明", "提示声明", "上述年度"))):
            end_index = row_idx
            break
        if is_non_amount_line(normalized):
            continue
        for raw_number in ocr_numbers_from_line(line):
            if looks_like_year(raw_number, line):
                continue
            value = parse_ocr_amount(raw_number)
            if value is None:
                continue
            values.append(value * unit_multiplier)
            if len(values) >= expected_count:
                return values, row_idx + 1
        end_index = row_idx + 1
    return values, end_index


def ocr_numbers_from_line(line: str) -> list[str]:
    numbers: list[tuple[int, int, str]] = []
    for match in re.finditer(r"\d{1,3}(?:\.\d{3})+\.\d{1,4}", line):
        numbers.append((match.start(), match.end(), match.group()))
    covered = [(start, end) for start, end, _value in numbers]
    for match in OCR_NUMBER_RE.finditer(line):
        if any(start <= match.start() < end for start, end in covered):
            continue
        numbers.append((match.start(), match.end(), match.group()))
    return [value for _start, _end, value in sorted(numbers, key=lambda item: item[0])]


def is_non_amount_line(normalized: str) -> bool:
    if not normalized:
        return True
    return any(
        token in normalized
        for token in (
            "运营收入",
            "成本费用",
            "税金及附加",
            "资本性支出",
            "附加资本性支出",
            "单位万元",
            "单位元",
            "科目",
        )
    )


def looks_like_year(raw_number: str, line: str) -> bool:
    try:
        value = int(raw_number.replace(",", "").replace("，", ""))
    except ValueError:
        return False
    return 1900 <= value <= 2200 and "年" in line


def parse_ocr_amount(raw_number: str) -> float | None:
    text = raw_number.strip().replace("，", ",")
    if not text:
        return None
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+\.\d{1,4}", text):
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(Decimal(text.replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def project_name_near(lines: list[str], row_idx: int) -> str:
    for probe in range(row_idx, max(-1, row_idx - 12), -1):
        label = project_label_from_line(lines[probe])
        if label:
            return label
    return "项目整体"


def project_label_from_line(line: str) -> str | None:
    match = re.search(r"^(估价对象[一二三四五六七八九十\d]+|评估对象[一二三四五六七八九十\d]+|项目[一二三四五六七八九十\d]+)\s*[：:]", line.strip())
    return match.group(1) if match else None


def extract_growth_instruction(text: str) -> tuple[int | None, float | None]:
    patterns = [
        r"自\s*(20\d{2})\s*年\s*起[^。\n]{0,60}?增长率\s*([0-9]+(?:\.[0-9]+)?)\s*%",
        r"(20\d{2})\s*年\s*起[^。\n]{0,60}?按照年增长率\s*([0-9]+(?:\.[0-9]+)?)\s*%",
        r"增长率\s*([0-9]+(?:\.[0-9]+)?)\s*%[^。\n]{0,20}?自\s*(20\d{2})\s*年",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        if pattern.startswith("增长率"):
            rate_text, year_text = match.groups()
        else:
            year_text, rate_text = match.groups()
        return int(year_text), parse_number(rate_text + "%")
    return None, None


def extract_terminal_recovery(text: str) -> float | None:
    lines = normalize_ocr_lines(text)
    labels = ("期末回收", "期末余值", "残余价值", "经营期满回收", "经营期末", "移交补偿")
    for index, line in enumerate(lines):
        normalized = normalize_text(line)
        if not any(label in normalized for label in labels):
            continue
        window = " ".join(lines[index : min(len(lines), index + 4)])
        values = [
            value
            for value in (parse_ocr_amount(raw) for raw in ocr_numbers_from_line(window))
            if value is not None and abs(value) > 0
        ]
        if values:
            return values[0]

    has_terminal_column = any(normalize_text(line) == "期末" for line in lines)
    if not has_terminal_column:
        return None

    for index, line in enumerate(lines):
        normalized_line = normalize_text(line)
        if not is_cashflow_metric_label(line):
            continue
        if any(token in normalized_line for token in ("预测表", "附表", "评估基准日", "金额单位", "单位人民币")):
            continue
        if not (
            normalized_line.startswith(("一", "二", "三", "四", "五", "六", "七", "八", "九", "十"))
            or normalized_line.startswith(("税前净现金流", "现金流量", "预测现金流", "运营净现金流"))
            or "、税前净现金流" in normalized_line
            or "、现金流量" in normalized_line
        ):
            continue
        values: list[float] = []
        for candidate in lines[index + 1 : min(len(lines), index + 25)]:
            normalized = normalize_text(candidate)
            if not normalized:
                continue
            raw_numbers = ocr_numbers_from_line(candidate)
            parsed_numbers = [
                value
                for value in (parse_ocr_amount(raw) for raw in raw_numbers)
                if value is not None and abs(value) > 0
            ]
            if parsed_numbers:
                values.extend(parsed_numbers)
                continue
            if values and any(token in normalized for token in ("折现期", "折现率", "现值", "市场价值", "资产组市场价值", "评估结果")):
                break
            if values and normalized.endswith("现金流"):
                break
        if len(values) >= 2:
            return values[-1]
    return None


def extract_discount_rate(text: str) -> float | None:
    lines = normalize_ocr_lines(text)
    for index, line in enumerate(lines):
        normalized = normalize_text(line)
        if "折现率" not in normalized and "贴现率" not in normalized:
            continue
        window = " ".join(lines[index : min(len(lines), index + 3)])
        for raw in ocr_numbers_from_line(window):
            if looks_like_year(raw, window):
                continue
            value = parse_number(raw + "%") if "%" in window or "％" in window else parse_number(raw)
            if value > 1:
                value /= 100
            if 0 < value < 1:
                return value
    return None


def extract_valuation_amount_ten_thousand(text: str) -> float | None:
    lines = normalize_ocr_lines(text)
    labels = ("评估价值", "评估报告评估价值", "基础资产评估价值", "估值")
    for index, line in enumerate(lines):
        normalized = normalize_text(line)
        if not any(label in normalized for label in labels):
            continue
        window = " ".join(lines[index : min(len(lines), index + 4)])
        normalized_window = normalize_text(window)
        multiplier = 0.0001 if "元" in normalized_window and "万元" not in normalized_window else 1.0
        values = []
        for raw in ocr_numbers_from_line(window):
            if looks_like_year(raw, window):
                continue
            value = parse_ocr_amount(raw)
            if value is not None and abs(value) >= 1000:
                values.append(value * multiplier)
        if values:
            return values[-1]
    return None


def extract_report_year_from_filename(name: str) -> int | None:
    matches = re.findall(r"20\d{2}", name)
    if not matches:
        return None
    return int(matches[-1])


def resolve_ocr_engine(engine: str) -> tuple[str | None, str | None]:
    normalized = engine.lower().strip()
    if normalized == "pdf_text":
        return None, None

    rapidocr_available = importlib.util.find_spec("rapidocr_onnxruntime") is not None
    paddle_available = importlib.util.find_spec("paddleocr") is not None
    tesseract_available = (
        importlib.util.find_spec("pytesseract") is not None
        and importlib.util.find_spec("PIL") is not None
        and shutil.which("tesseract") is not None
    )

    if normalized in {"rapidocr", "rapid"}:
        if rapidocr_available:
            return "rapidocr", None
        return None, "已选择 RapidOCR，但当前环境未安装 rapidocr-onnxruntime；图片型 PDF 将不会 OCR。"
    if normalized in {"paddleocr", "paddle"}:
        if paddle_available:
            return "paddleocr", None
        return None, "已选择 PaddleOCR，但当前环境未安装 paddleocr/paddlepaddle；图片型 PDF 将不会 OCR。"
    if normalized in {"pytesseract", "tesseract"}:
        if tesseract_available:
            return "pytesseract", None
        return None, "已选择 Tesseract，但当前环境未安装 pytesseract/Pillow 或找不到 tesseract 命令；图片型 PDF 将不会 OCR。"
    if normalized in {"vision_api", "cloud_ocr", "api_ocr"}:
        return "vision_api", "已选择云端视觉 OCR：会把需要 OCR 的截图或 PDF 渲染页发送到所选 OCR API；本地 OCR 不会上传图片。"
    if normalized == "auto":
        if rapidocr_available:
            return "rapidocr", None
        if paddle_available:
            return "paddleocr", None
        if tesseract_available:
            return "pytesseract", None
        return None, "未检测到可用本地 OCR 引擎；如需识别图片型 PDF，请安装 RapidOCR、PaddleOCR，或安装 Tesseract + pytesseract。"
    return None, f"未知 OCR 引擎：{engine}；图片型 PDF 将不会 OCR。"


def extract_docx_text(path: Path) -> str:
    with ZipFile(path) as archive:
        names = [name for name in archive.namelist() if name.startswith("word/") and name.endswith(".xml")]
        chunks: list[str] = []
        for name in names:
            if name.startswith("word/_rels/"):
                continue
            xml = archive.read(name).decode("utf-8", errors="ignore")
            chunks.extend(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, flags=re.S))
    text = "\n".join(unescape_xml_text(chunk) for chunk in chunks)
    return compact_text(text)


def extract_docx_embedded_image_text(path: Path, engine: str, options: AnnualUpdateOptions | None = None) -> str:
    texts: list[str] = []
    with ZipFile(path) as archive, tempfile.TemporaryDirectory(prefix="reit_docx_ocr_") as temp_dir:
        for name in archive.namelist():
            if not name.startswith("word/media/"):
                continue
            suffix = Path(name).suffix or ".bin"
            temp_path = Path(temp_dir) / Path(name).name
            temp_path.write_bytes(archive.read(name))
            if suffix.lower() not in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}:
                continue
            text = ocr_image_file(temp_path, engine, options)
            if text.strip():
                texts.append(text.strip())
    return compact_text("\n".join(texts))


def unescape_xml_text(text: str) -> str:
    return (
        text.replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&amp;", "&")
        .replace("&quot;", '"')
        .replace("&apos;", "'")
    )


def extract_pdf_text_and_ocr(
    pdf_path: Path,
    options: AnnualUpdateOptions,
    warnings: list[str],
    resolved_ocr_engine: str | None,
) -> list[OcrItem]:
    try:
        import fitz  # type: ignore
    except ImportError as exc:
        raise AnnualUpdateError("缺少 PyMuPDF，无法读取 PDF。请安装 pymupdf。") from exc

    items: list[OcrItem] = []
    document = fitz.open(pdf_path)
    try:
        page_count = document.page_count
        if options.max_ocr_pages_per_file > 0:
            page_count = min(page_count, options.max_ocr_pages_per_file)
        for page_idx in range(page_count):
            page = document.load_page(page_idx)
            text = page.get_text("text") or ""
            method = "pdf-text"
            warning = ""

            if should_keep_ocr_text(text):
                items.append(OcrItem(source_file=pdf_path, page=page_idx + 1, method=method, text=compact_text(text), warning=warning))
                continue

            should_ocr_page = bool(resolved_ocr_engine) and (
                options.max_ocr_pages_per_file == 0 or page_idx < options.max_ocr_pages_per_file
            )
            if should_ocr_page and options.ocr_engine != "pdf_text":
                try:
                    image_path = render_pdf_page_to_temp_image(page)
                    try:
                        ocr_text = ocr_image_file(image_path, resolved_ocr_engine, options)
                    finally:
                        safe_unlink(image_path)
                    if ocr_text.strip():
                        items.append(OcrItem(source_file=pdf_path, page=page_idx + 1, method="pdf-page-ocr", text=compact_text(ocr_text)))
                except Exception as exc:
                    warnings.append(f"PDF 第 {page_idx + 1} 页 OCR 跳过：{pdf_path.name}，原因：{exc}")
            elif text.strip():
                items.append(OcrItem(source_file=pdf_path, page=page_idx + 1, method=method, text=compact_text(text), warning="未命中关键词"))
    finally:
        document.close()
    return items


def should_keep_ocr_text(text: str) -> bool:
    stripped = text.strip()
    if len(stripped) < 30:
        return False
    return any(keyword in stripped for keyword in OCR_KEYWORDS)


def compact_text(text: str) -> str:
    return re.sub(r"[ \t]+", " ", text).strip()


def render_pdf_page_to_temp_image(page: Any) -> Path:
    matrix = page.parent.Matrix(2, 2) if hasattr(page.parent, "Matrix") else None
    if matrix is None:
        import fitz  # type: ignore

        matrix = fitz.Matrix(2, 2)
    pixmap = page.get_pixmap(matrix=matrix, alpha=False)
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    temp.close()
    pixmap.save(temp.name)
    return Path(temp.name)


def safe_unlink(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


def ocr_image_file(image_path: Path, engine: str, options: AnnualUpdateOptions | None = None) -> str:
    normalized_engine = engine.lower().strip()
    if normalized_engine in {"vision_api", "cloud_ocr", "api_ocr"}:
        if options is None:
            options = AnnualUpdateOptions(workspace_path=Path("."))
        return ocr_image_with_vision_api(image_path, options)
    if normalized_engine in {"auto", "rapidocr", "rapid"}:
        try:
            return ocr_image_with_rapidocr(image_path)
        except ImportError:
            if normalized_engine != "auto":
                raise
        except Exception:
            if normalized_engine != "auto":
                raise
    if normalized_engine in {"auto", "paddleocr", "paddle"}:
        try:
            return ocr_image_with_paddle(image_path)
        except ImportError:
            if normalized_engine != "auto":
                raise
        except Exception:
            if normalized_engine != "auto":
                raise
    if normalized_engine in {"auto", "pytesseract", "tesseract"}:
        try:
            return ocr_image_with_tesseract(image_path)
        except ImportError:
            if normalized_engine != "auto":
                raise
        except Exception:
            if normalized_engine != "auto":
                raise
    if normalized_engine == "pdf_text":
        return ""
    raise AnnualUpdateError("未找到可用的本地 OCR 引擎。可安装 rapidocr-onnxruntime、paddleocr，或安装 Tesseract 后使用 pytesseract。")


def ocr_image_with_rapidocr(image_path: Path) -> str:
    global _RAPIDOCR_ENGINE
    try:
        from rapidocr_onnxruntime import RapidOCR  # type: ignore
    except ImportError as exc:
        raise ImportError("未安装 rapidocr-onnxruntime。") from exc
    if _RAPIDOCR_ENGINE is None:
        _RAPIDOCR_ENGINE = RapidOCR()
    result, _elapsed = _RAPIDOCR_ENGINE(str(image_path))
    lines: list[str] = []
    for item in result or []:
        if not item or len(item) < 2:
            continue
        text = item[1]
        if text not in (None, ""):
            lines.append(str(text))
    return "\n".join(lines)


def ocr_image_with_paddle(image_path: Path) -> str:
    global _PADDLE_OCR_ENGINE
    try:
        from paddleocr import PaddleOCR  # type: ignore
    except ImportError as exc:
        raise ImportError("未安装 paddleocr。") from exc
    if _PADDLE_OCR_ENGINE is None:
        _PADDLE_OCR_ENGINE = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)
    result = _PADDLE_OCR_ENGINE.ocr(str(image_path), cls=True)
    lines: list[str] = []
    for page_result in result or []:
        for item in page_result or []:
            if not item or len(item) < 2:
                continue
            text_info = item[1]
            if isinstance(text_info, (list, tuple)) and text_info:
                lines.append(str(text_info[0]))
    return "\n".join(lines)


def ocr_image_with_tesseract(image_path: Path) -> str:
    try:
        from PIL import Image  # type: ignore
        import pytesseract  # type: ignore
    except ImportError as exc:
        raise ImportError("未安装 pytesseract 或 Pillow。") from exc
    return pytesseract.image_to_string(Image.open(image_path), lang="chi_sim+eng")


def ocr_image_with_vision_api(image_path: Path, options: AnnualUpdateOptions) -> str:
    api_key = options.ocr_api_key or os.getenv(options.ocr_api_key_env, "")
    base_url = options.ocr_base_url or DEFAULT_DASHSCOPE_BASE_URL
    if not api_key and not is_local_ai_endpoint(base_url):
        env_hint = f"环境变量 {options.ocr_api_key_env}" if options.ocr_api_key_env else "OCR API Key"
        raise AnnualUpdateError(f"云端视觉 OCR 需要 API Key，请填写临时 Key 或设置 {env_hint}。")
    data_url = image_file_to_data_url(image_path)
    response_text, _usage = call_openai_compatible_chat(
        api_key=api_key,
        base_url=base_url,
        model=options.ocr_model or DEFAULT_VISION_OCR_MODEL,
        timeout_seconds=options.ai_request_timeout_seconds,
        response_format={},
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "请对这张 REITs 现金流/评估报告截图做 OCR。"
                            "只输出图片中能看见的原始文字和数字，按阅读顺序换行；"
                            "不要解释，不要总结，不要补充图片中没有的信息。"
                        ),
                    },
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }
        ],
    )
    return response_text.strip()


def image_file_to_data_url(image_path: Path) -> str:
    suffix = image_path.suffix.lower()
    mime = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".bmp": "image/bmp",
        ".tif": "image/tiff",
        ".tiff": "image/tiff",
    }.get(suffix, "image/png")
    import base64

    encoded = base64.b64encode(image_path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def standardize_ocr_with_ai(
    ocr_items: list[OcrItem],
    options: AnnualUpdateOptions,
    progress: Callable[[str], None] | None = None,
) -> tuple[list[dict[str, Any]], list[str], list["AiCallRecord"]]:
    batches = build_ai_payload_batches(ocr_items, options.max_ai_chars, options.ai_items_per_batch)
    if not batches:
        return [], [], []

    api_key = options.api_key or os.getenv(options.api_key_env, "")
    if not api_key and not is_local_ai_endpoint(options.base_url):
        env_hint = f"设置环境变量 {options.api_key_env}，或" if options.api_key_env else ""
        return [], [f"已启用 AI 标准化，但未找到 API Key。请{env_hint}在软件中临时输入；本次仅输出 OCR 原始结果。"], []

    all_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    records: list[AiCallRecord] = []
    started_at = time.monotonic()
    consecutive_failures = 0
    for batch_index, batch in enumerate(batches, 1):
        remaining_batches = batches[batch_index - 1 :]
        if ai_time_budget_exceeded(started_at, options):
            warning = (
                f"AI 标准化达到总时长上限 {options.ai_total_timeout_seconds} 秒，"
                f"已停止剩余 {len(remaining_batches)} 批；请查看 OCR 原始结果并可稍后重试。"
            )
            mark_ai_batches_skipped(remaining_batches, warning)
            warnings.append(warning)
            records.extend(build_skipped_ai_records(remaining_batches, batch_index, len(batches), warning))
            break

        request_timeout = effective_ai_request_timeout(started_at, options)
        if request_timeout <= 0:
            warning = f"AI 标准化剩余时间不足，已停止剩余 {len(remaining_batches)} 批。"
            mark_ai_batches_skipped(remaining_batches, warning)
            warnings.append(warning)
            records.extend(build_skipped_ai_records(remaining_batches, batch_index, len(batches), warning))
            break

        if progress:
            elapsed = int(time.monotonic() - started_at)
            progress(
                f"正在调用 AI 标准化 OCR 文本：第 {batch_index}/{len(batches)} 批，"
                f"已用 {elapsed} 秒，单批上限 {request_timeout} 秒..."
            )
        prompt = build_standardization_prompt(batch.payload_text)
        call_started_at = time.monotonic()
        record = base_ai_call_record(batch, batch_index, len(batches), len(prompt))
        try:
            response_result = call_openai_compatible_chat(
                api_key=api_key,
                base_url=options.base_url,
                model=options.model,
                timeout_seconds=request_timeout,
                messages=[
                    {"role": "system", "content": "你是严谨的 REITs 年报现金流表格结构化助手，只输出 JSON。"},
                    {"role": "user", "content": prompt},
                ],
            )
            if isinstance(response_result, tuple):
                response_text, usage = response_result
            else:
                response_text, usage = str(response_result), {}
            rows = parse_ai_standard_rows(response_text)
        except Exception as exc:
            consecutive_failures += 1
            record.status = "失败"
            record.elapsed_seconds = round(time.monotonic() - call_started_at, 2)
            record.error = str(exc)
            records.append(record)
            for item in batch.items:
                item.warning = append_warning(item.warning, f"AI第{batch_index}批失败：{exc}")
            warnings.append(f"AI 第 {batch_index}/{len(batches)} 批标准化失败，已跳过该批：{exc}")
            if consecutive_failures >= max(options.ai_stop_after_failures, 1):
                rest_batches = batches[batch_index:]
                if rest_batches:
                    warning = (
                        f"AI 已连续失败 {consecutive_failures} 批，程序已主动停止剩余 {len(rest_batches)} 批，"
                        "避免长时间卡住；请检查网络、模型服务或改用人工标准导入表。"
                    )
                    mark_ai_batches_skipped(rest_batches, warning)
                    warnings.append(warning)
                    records.extend(build_skipped_ai_records(rest_batches, batch_index + 1, len(batches), warning))
                break
            continue

        consecutive_failures = 0
        source_hash = hashlib.sha256(batch.payload_text.encode("utf-8")).hexdigest()[:12]
        for item in batch.items:
            item.used_for_ai = True
        normalized_rows = []
        for row in rows:
            normalized_row = normalize_one_standard_row(row)
            if batch.items and not normalized_row.get("REITs名称"):
                normalized_row["REITs名称"] = guess_reits_name_from_source_file(batch.items[0].source_file)
            if batch.items and not normalized_row.get("来源页码"):
                normalized_row["来源页码"] = batch.items[0].page
            if batch.items:
                normalized_row.setdefault("来源文件", batch.items[0].source_file.name)
            else:
                normalized_row.setdefault("来源文件", f"OCR文本-{source_hash}")
            normalized_row.setdefault("备注", f"AI标准化结果，第{batch_index}批，需人工复核")
            normalized_rows.append(normalized_row)
        all_rows.extend(normalized_rows)
        record.status = "成功"
        record.row_count = len(normalized_rows)
        record.elapsed_seconds = round(time.monotonic() - call_started_at, 2)
        record.input_tokens = parse_usage_token_count(usage, "prompt_tokens", "input_tokens")
        record.output_tokens = parse_usage_token_count(usage, "completion_tokens", "output_tokens")
        record.total_tokens = parse_usage_token_count(usage, "total_tokens")
        records.append(record)
        if progress:
            progress(f"AI 第 {batch_index}/{len(batches)} 批完成：新增 {len(normalized_rows)} 行标准化记录。")

    if not all_rows and warnings:
        warnings.append("所有 AI 标准化批次均失败；程序仍会输出 OCR 原始识别结果和复核清单。")
    return all_rows, warnings, records


@dataclass
class AiPayloadBatch:
    payload_text: str
    items: list[OcrItem]


@dataclass
class AiCallRecord:
    batch_index: int
    total_batches: int
    source_files: str
    source_pages: str
    prompt_chars: int
    status: str
    row_count: int = 0
    elapsed_seconds: float = 0
    input_tokens: int | None = None
    output_tokens: int | None = None
    total_tokens: int | None = None
    error: str = ""


def build_ai_payload_batches(ocr_items: list[OcrItem], max_chars: int, max_items_per_batch: int = 1) -> list[AiPayloadBatch]:
    budget = max(max_chars, 1000)
    item_limit = max(max_items_per_batch, 1)
    batches: list[AiPayloadBatch] = []
    current_chunks: list[str] = []
    current_items: list[OcrItem] = []
    current_chars = 0

    for item in ocr_items:
        text = item.text.strip()
        if not text:
            continue
        chunk = format_ocr_item_for_ai(item, text)
        if current_chunks and (current_chars + len(chunk) > budget or len(current_items) >= item_limit):
            batches.append(AiPayloadBatch(payload_text="\n\n---\n\n".join(current_chunks), items=current_items))
            current_chunks = []
            current_items = []
            current_chars = 0
        current_chunks.append(chunk)
        current_items.append(item)
        current_chars += len(chunk)

    if current_chunks:
        batches.append(AiPayloadBatch(payload_text="\n\n---\n\n".join(current_chunks), items=current_items))
    return batches


def build_ai_payload_text(ocr_items: list[OcrItem], max_chars: int) -> str:
    chunks: list[str] = []
    remaining = max_chars
    for item in ocr_items:
        if remaining <= 0:
            break
        text = item.text[:remaining]
        item.used_for_ai = bool(text)
        chunks.append(f"【文件】{item.source_file.name}\n【页码】{item.page or ''}\n{text}")
        remaining -= len(text)
    return "\n\n---\n\n".join(chunks)


def format_ocr_item_for_ai(item: OcrItem, text: str) -> str:
    return f"【文件】{item.source_file.name}\n【页码】{item.page or ''}\n{text}"


def guess_reits_name_from_source_file(path: Path) -> str:
    stem = path.stem
    match = re.search(r"(.+?封闭式基础设施证券投资基金)", stem)
    if match:
        return match.group(1)
    stem = re.sub(r"\d{4}.*$", "", stem)
    stem = re.sub(r"(年度报告|年报|评估报告|资产评估报告|房地产估价报告|市场价值|不动产项目)", "", stem)
    return stem.strip(" -_【】[]（）()") or path.stem


def append_warning(existing: str, warning: str) -> str:
    return f"{existing}；{warning}" if existing else warning


def base_ai_call_record(batch: "AiPayloadBatch", batch_index: int, total_batches: int, prompt_chars: int) -> "AiCallRecord":
    return AiCallRecord(
        batch_index=batch_index,
        total_batches=total_batches,
        source_files="；".join(sorted({item.source_file.name for item in batch.items})),
        source_pages="；".join(str(item.page or "") for item in batch.items),
        prompt_chars=prompt_chars,
        status="待处理",
    )


def build_skipped_ai_records(
    batches: Iterable["AiPayloadBatch"],
    start_index: int,
    total_batches: int,
    reason: str,
) -> list["AiCallRecord"]:
    records: list[AiCallRecord] = []
    for offset, batch in enumerate(batches):
        record = base_ai_call_record(batch, start_index + offset, total_batches, 0)
        record.status = "跳过"
        record.error = reason
        records.append(record)
    return records


def parse_usage_token_count(usage: dict[str, Any], *keys: str) -> int | None:
    for key in keys:
        value = usage.get(key)
        if isinstance(value, int):
            return value
        if isinstance(value, float) and value.is_integer():
            return int(value)
    return None


def ai_time_budget_exceeded(started_at: float, options: AnnualUpdateOptions) -> bool:
    return options.ai_total_timeout_seconds > 0 and time.monotonic() - started_at >= options.ai_total_timeout_seconds


def effective_ai_request_timeout(started_at: float, options: AnnualUpdateOptions) -> int:
    request_timeout = max(options.ai_request_timeout_seconds, 5)
    if options.ai_total_timeout_seconds <= 0:
        return request_timeout
    remaining = int(options.ai_total_timeout_seconds - (time.monotonic() - started_at))
    if remaining < 10:
        return 0
    return max(5, min(request_timeout, remaining))


def mark_ai_batches_skipped(batches: Iterable[AiPayloadBatch], warning: str) -> None:
    for batch in batches:
        for item in batch.items:
            item.warning = append_warning(item.warning, warning)


def build_standardization_prompt(ocr_text: str) -> str:
    fields = "、".join(STANDARD_FIELDS)
    return (
        "请把下面 OCR 识别出的 REITs 年报/评估报告现金流相关文本整理为标准 JSON 对象。\n"
        "要求：\n"
        "1. 一行代表一个 REITs + 一个项目 + 一个年份的现金流记录。\n"
        "2. 如果同一张表横向列出多个年份，必须把每个年份拆成独立行；不要只输出第一年，不要合并年份。\n"
        "3. 预测现金流金额优先使用“运营净收益/运营净现金流/预测现金流/基础资产预测现金流/自由现金流”对应的年度金额。\n"
        "4. 只能使用文本中明确出现或能从同一表头延续判断的信息，缺失字段留空字符串，不要猜。\n"
        "5. 重点提取“运营净收益、预测现金流、评估价值、折现率、评估基准日、残值、基金净资产、折旧及摊销”等字段；无关叙述不要输出。\n"
        "6. 金额字段统一按万元输出；如果原文是元，需要除以 10000。\n"
        "7. 百分比字段输出小数，例如 3% 输出 0.03。\n"
        "8. REITs 名称可优先从【文件】名中提取；REITs 代码若文本没有明确出现则留空。\n"
        "9. 项目名称必须是真实项目、路段、资产包或“项目整体”；不要把“运营净收益、现金流、全周期运营净收益”等指标名称当作项目名称。\n"
        "10. 如果同一材料有多个项目、路段、水厂、估价对象，请分别输出各自项目；只有明确是合计或整体时才写“项目整体”。\n"
        "11. 若文本、文件名或项目类型明显属于高速公路、公路、水务、污水、水利、清洁能源、新能源、供热、生物质等收费权/特许经营权资产，底层资产性质填“特许经营权”；商业、产业园、仓储物流、租赁住房、数据中心、消费基础设施等通常填“产权”；不确定则留空。\n"
        "12. 特许经营权的期末回收、期末余值、移交补偿、经营期满回收金额填入“经营期末”，不要并入普通年度现金流。\n"
        "13. 为节省成本，每行只输出有值字段，不要输出空字符串字段，不要把字段清单完整照抄到每一行。\n"
        "14. 只输出 JSON，不要 Markdown，不要解释；格式必须是 {\"rows\": [...]}。\n"
        "示例：若 OCR 中有“2027年 2028年”和“运营净收益 100 120”，应输出两行："
        "{\"rows\":[{\"年份\":2027,\"预测现金流金额（万元）\":100},{\"年份\":2028,\"预测现金流金额（万元）\":120}]}。\n"
        f"字段清单：{fields}\n\n"
        f"OCR文本：\n{ocr_text}"
    )


def is_local_ai_endpoint(base_url: str) -> bool:
    normalized = base_url.lower()
    return "localhost" in normalized or "127.0.0.1" in normalized


def call_openai_compatible_chat(
    api_key: str,
    base_url: str,
    model: str,
    messages: list[dict[str, Any]],
    timeout_seconds: int = DEFAULT_AI_REQUEST_TIMEOUT_SECONDS,
    response_format: dict[str, Any] | None = None,
) -> tuple[str, dict[str, Any]]:
    url = base_url.rstrip("/") + "/chat/completions"
    body_data: dict[str, Any] = {
        "model": model,
        "messages": messages,
        "temperature": 0,
    }
    if response_format is None:
        body_data["response_format"] = {"type": "json_object"}
    elif response_format:
        body_data["response_format"] = response_format
    if "dashscope" in base_url.lower():
        body_data["enable_thinking"] = False
    body = json.dumps(
        body_data,
        ensure_ascii=False,
    ).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    request = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=max(timeout_seconds, 30)) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise AnnualUpdateError(f"AI 请求失败：HTTP {exc.code} {detail[:500]}") from exc
    except urllib.error.URLError as exc:
        raise AnnualUpdateError(f"AI 请求失败：{exc}") from exc
    except TimeoutError as exc:
        raise AnnualUpdateError(f"AI 请求超时（{timeout_seconds} 秒）。可减少 OCR 文件数量、换更快模型，或先人工整理标准导入表。") from exc
    return str(data["choices"][0]["message"]["content"]), dict(data.get("usage") or {})


def parse_ai_standard_rows(response_text: str) -> list[dict[str, Any]]:
    text = response_text.strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"(\[.*\]|\{.*\})", text, flags=re.S)
        if not match:
            raise AnnualUpdateError("AI 返回内容不是可解析的 JSON。")
        data = json.loads(match.group(1))

    if isinstance(data, dict):
        if isinstance(data.get("rows"), list):
            data = data["rows"]
        elif isinstance(data.get("data"), list):
            data = data["data"]
        else:
            data = [data]
    if not isinstance(data, list):
        raise AnnualUpdateError("AI 返回 JSON 不是数组结构。")
    return [dict(item) for item in data if isinstance(item, dict)]


def read_standard_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            header_row, headers = find_header_row(worksheet)
            if contains_year_headers(headers):
                rows.extend(read_future_rows_from_sheet(worksheet, header_row))
            else:
                rows.extend(read_long_rows_from_sheet(worksheet, header_row, headers))
    finally:
        workbook.close()
    return rows


def read_future_cashflow_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook.active
        return read_future_rows_from_sheet(worksheet, 1)
    finally:
        workbook.close()


def contains_year_headers(headers: Iterable[Any]) -> bool:
    header_list = list(headers)
    year_count = sum(1 for header in header_list if isinstance(header, int) and 1900 <= header <= 2200)
    normalized = {normalize_text(header) for header in header_list if header not in (None, "")}
    if {"名称", "代码", "项目名称"}.issubset(normalized) and year_count >= 2:
        return True
    return year_count >= 3


def read_future_rows_from_sheet(worksheet: Any, header_row: int) -> list[dict[str, Any]]:
    headers = [worksheet.cell(header_row, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    year_columns = {header: col_idx for col_idx, header in enumerate(headers, 1) if isinstance(header, int)}
    fixed_map = {
        "名称": "REITs名称",
        "代码": "REITs代码",
        "公告日期": "公告日期",
        "基础设施项目类型": "基础设施项目类型",
        "底层资产性质": "底层资产性质",
        "股债": "股债",
        "上市日期": "上市日期",
        "上市年份": "上市年份",
        "到期日": "到期日",
        "报告期": "报告期",
        "项目名称": "项目名称",
        "自年份": "增长率预测起始年度",
        "未来增长率": "预测现金流增长率",
        "经营期末": "经营期末",
    }
    header_to_col = {str(header): col_idx for col_idx, header in enumerate(headers, 1) if header not in (None, "")}
    rows: list[dict[str, Any]] = []
    empty_streak = 0
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        base: dict[str, Any] = {}
        for source_name, target_name in fixed_map.items():
            col_idx = header_to_col.get(source_name)
            if col_idx:
                base[target_name] = worksheet.cell(row_idx, col_idx).value
        has_year_value = any(worksheet.cell(row_idx, col_idx).value not in (None, "") for col_idx in year_columns.values())
        if not any(base.get(key) for key in ("REITs代码", "REITs名称", "项目名称")) and not has_year_value:
            empty_streak += 1
            if empty_streak >= 100:
                break
            continue
        empty_streak = 0
        for year, col_idx in year_columns.items():
            value = worksheet.cell(row_idx, col_idx).value
            if value in (None, ""):
                continue
            item = dict(base)
            item["年份"] = year
            item["预测现金流金额（万元）"] = value
            item.setdefault("来源文件", Path(getattr(worksheet.parent, "path", "") or "").name)
            rows.append(item)
    return rows


def read_long_rows_from_sheet(worksheet: Any, header_row: int, headers: list[Any]) -> list[dict[str, Any]]:
    column_map = build_standard_column_map(headers)
    rows: list[dict[str, Any]] = []
    empty_streak = 0
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row: dict[str, Any] = {}
        for col_idx, field_name in column_map.items():
            value = worksheet.cell(row_idx, col_idx).value
            if value not in (None, ""):
                row[field_name] = value
        if any(row.get(key) for key in ("REITs代码", "REITs名称", "项目名称", "预测现金流金额（万元）")):
            empty_streak = 0
            rows.append(row)
        else:
            empty_streak += 1
            if empty_streak >= 100:
                break
    return rows


def build_standard_column_map(headers: list[Any]) -> dict[int, str]:
    normalized_aliases: dict[str, str] = {}
    for field_name, aliases in FIELD_ALIASES.items():
        normalized_aliases[normalize_text(field_name)] = field_name
        for alias in aliases:
            normalized_aliases[normalize_text(alias)] = field_name

    mapping: dict[int, str] = {}
    for col_idx, header in enumerate(headers, 1):
        normalized = normalize_text(header)
        if not normalized:
            continue
        if normalized in {"评估价值元", "基础资产评估价值元", "评估报告评估价值元"}:
            mapping[col_idx] = "基础资产评估价值（元）"
            continue
        field_name = normalized_aliases.get(normalized)
        if field_name:
            mapping[col_idx] = field_name
    return mapping


def filter_ai_rows_for_unparsed_ocr_sources(
    ai_rows: list[dict[str, Any]],
    local_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int]:
    """Use AI as a safe fallback when deterministic OCR parsing did not produce rows."""
    if not ai_rows:
        return [], 0
    parsed_sources = {
        str(row.get("来源文件") or "").strip()
        for row in local_rows
        if str(row.get("来源文件") or "").strip()
    }
    if not parsed_sources:
        return ai_rows, 0

    filtered_rows: list[dict[str, Any]] = []
    skipped_count = 0
    for row in ai_rows:
        source_file = str(row.get("来源文件") or "").strip()
        if source_file and source_file in parsed_sources:
            skipped_count += 1
            continue
        filtered_rows.append(row)
    return filtered_rows, skipped_count


def normalize_standard_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_rows: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for raw in rows:
        row = normalize_one_standard_row(raw)
        key = (
            normalize_code(row.get("REITs代码")),
            normalize_project(row.get("项目名称")),
            row.get("年份"),
            row.get("预测现金流金额（万元）"),
            row.get("来源文件"),
        )
        if key in seen:
            continue
        seen.add(key)
        normalized_rows.append(row)
    return normalized_rows


def stash_residual_parameters_from_rows(rows: list[dict[str, Any]]) -> None:
    """Keep residual parameters out of disclosed year rows and use them for the residual block."""
    for row in rows:
        parameters = row.get(RESIDUAL_PARAMETER_KEY)
        if not isinstance(parameters, dict):
            parameters = {}
            row[RESIDUAL_PARAMETER_KEY] = parameters
        for field in ("残值年度", "残值基础数据预测", "考虑残值现金流"):
            value = row.pop(field, None)
            if value not in (None, ""):
                parameters.setdefault(field, value)
        if not parameters:
            row.pop(RESIDUAL_PARAMETER_KEY, None)


def normalize_one_standard_row(raw: dict[str, Any]) -> dict[str, Any]:
    row: dict[str, Any] = {field: raw.get(field) for field in STANDARD_FIELDS if raw.get(field) not in (None, "")}
    if row.get("基础资产评估价值（万元）") in (None, ""):
        valuation_yuan = first_raw_value_by_normalized_header(
            raw,
            {
                "评估价值元",
                "基础资产评估价值元",
                "评估报告评估价值元",
            },
        )
        if valuation_yuan not in (None, ""):
            row["基础资产评估价值（万元）"] = parse_number(valuation_yuan) / 10000
    for key, value in raw.items():
        if key in row or value in (None, ""):
            continue
        field = alias_to_standard_field(str(key))
        if field and field not in row:
            row[field] = value

    row["REITs代码"] = normalize_code(row.get("REITs代码"))
    if "项目名称" not in row or not row.get("项目名称"):
        row["项目名称"] = "项目整体"
    else:
        row["项目名称"] = clean_project_name(row.get("项目名称"))

    for field in NUMERIC_FIELDS:
        if field in row:
            if field in FEE_RATE_FIELDS:
                row[field] = parse_fee_rate_value(row[field])
            else:
                row[field] = parse_number(row[field])
    for field in INTEGER_FIELDS:
        if field in row:
            row[field] = parse_int(row[field])
    for field in DATE_FIELDS:
        if field in row:
            row[field] = parse_date_like(row[field])

    if row.get("预测现金流增长率") and abs(float(row["预测现金流增长率"])) > 1:
        row["预测现金流增长率"] = float(row["预测现金流增长率"]) / 100
    if row.get("折现率") and abs(float(row["折现率"])) > 1:
        row["折现率"] = float(row["折现率"]) / 100
    return row


def first_raw_value_by_normalized_header(raw: dict[str, Any], normalized_headers: set[str]) -> Any:
    for key, value in raw.items():
        if normalize_text(key) in normalized_headers and value not in (None, ""):
            return value
    return None


def alias_to_standard_field(name: str) -> str | None:
    normalized = normalize_text(name)
    for field, aliases in FIELD_ALIASES.items():
        if normalize_text(field) == normalized:
            return field
        if any(normalize_text(alias) == normalized for alias in aliases):
            return field
    return None


def read_lookup_rows(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path:
        return {}
    workbook = load_workbook(path, data_only=True, read_only=True)
    records: dict[str, dict[str, Any]] = {}
    try:
        for worksheet in workbook.worksheets:
            header_row, headers = find_header_row(worksheet)
            column_map = build_standard_column_map(headers)
            if not column_map:
                continue
            empty_streak = 0
            max_scan_row = min(worksheet.max_row, header_row + 5000)
            for offset, values in enumerate(
                worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=max_scan_row,
                    max_col=min(worksheet.max_column, 120),
                    values_only=True,
                ),
                start=header_row + 1,
            ):
                row: dict[str, Any] = {}
                for col_idx, field_name in column_map.items():
                    value = values[col_idx - 1] if col_idx - 1 < len(values) else None
                    if value not in (None, ""):
                        row[field_name] = value
                normalized_row = normalize_one_standard_row(row) if row else {}
                code = normalize_code(normalized_row.get("REITs代码"))
                name_key = normalize_fund_name(normalized_row.get("REITs名称"))
                project_key = normalize_project(normalized_row.get("项目名称"))
                if code:
                    empty_streak = 0
                    records.setdefault(code, {}).update(normalized_row)
                    records.setdefault(f"code:{code}", {}).update(normalized_row)
                    if project_key:
                        records.setdefault(f"code_project:{code}:{project_key}", {}).update(normalized_row)
                    if name_key:
                        records.setdefault(f"name:{name_key}", {}).update(normalized_row)
                        if project_key:
                            records.setdefault(f"name_project:{name_key}:{project_key}", {}).update(normalized_row)
                elif name_key:
                    empty_streak = 0
                    records.setdefault(f"name:{name_key}", {}).update(normalized_row)
                    if project_key:
                        records.setdefault(f"name_project:{name_key}:{project_key}", {}).update(normalized_row)
                elif row:
                    empty_streak = 0
                else:
                    empty_streak += 1
                    if empty_streak >= 100:
                        break
            if worksheet.max_row > max_scan_row:
                records.setdefault(
                    "__warnings__",
                    {},
                )[worksheet.title] = f"辅助表 {worksheet.title} 有 {worksheet.max_row} 行记录范围，已仅读取前 {max_scan_row} 行以避免扫描格式化空白区域。"
    finally:
        workbook.close()
    return records


def extract_annual_report_financial_rows(
    pdf_files: list[Path],
    standard_rows: list[dict[str, Any]],
    reference_rows: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not pdf_files:
        return [], warnings
    try:
        import fitz  # type: ignore
    except ImportError:
        return [], ["缺少 PyMuPDF，无法从公募年报 PDF 提取基金净资产和折旧摊销。"]

    reference_seed_rows = list((reference_rows or {}).values())
    fund_refs = build_fund_references([*standard_rows, *reference_seed_rows])
    rows: list[dict[str, Any]] = []
    for pdf_path in sorted(pdf_files, key=lambda item: item.name):
        match = match_pdf_to_fund_reference(pdf_path, fund_refs)
        row = {
            "代码": match.get("REITs代码", ""),
            "基金名称": match.get("REITs名称") or clean_report_fund_name_from_filename(pdf_path),
            "匹配PDF": pdf_path.name,
            "匹配方式": match.get("匹配方式", "未匹配"),
            "基金净资产(元)": None,
            "基金净资产(万元)": None,
            "折旧及摊销(元)": None,
            "折旧及摊销(万元)": None,
            "数据来源": "PDF程序自动提取",
            "备注": "",
        }
        try:
            document = fitz.open(pdf_path)
        except Exception as exc:
            row["备注"] = f"PDF打开失败：{exc}"
            rows.append(row)
            continue
        try:
            net_asset: float | None = None
            depreciation: float | None = None
            net_page = ""
            dep_page = ""
            for page_index in range(document.page_count):
                text = document.load_page(page_index).get_text("text")
                if net_asset is None:
                    net_asset = extract_amount_by_labels(
                        text,
                        [
                            "期末不动产基金净资产",
                            "期末基金净资产",
                            "基金净资产",
                            "fund net assets",
                            "net assets",
                        ],
                        min_abs_value=50_000_000,
                    )
                    if net_asset is not None:
                        net_page = str(page_index + 1)
                if depreciation is None:
                    depreciation = extract_amount_by_labels(
                        text,
                        [
                            "本期折旧和摊销",
                            "本期折旧及摊销",
                            "折旧和摊销",
                            "折旧及摊销",
                            "depreciation and amortization",
                            "depreciation amortization",
                        ],
                        min_abs_value=100_000,
                    )
                    if depreciation is not None:
                        dep_page = str(page_index + 1)
                if net_asset is not None and depreciation is not None:
                    break
        finally:
            document.close()

        notes: list[str] = []
        if net_asset is not None:
            row["基金净资产(元)"] = net_asset
            row["基金净资产(万元)"] = net_asset / 10000
            notes.append(f"净资产页码 {net_page}")
        else:
            notes.append("未提取到基金净资产")
        if depreciation is not None:
            row["折旧及摊销(元)"] = depreciation
            row["折旧及摊销(万元)"] = depreciation / 10000
            notes.append(f"折旧摊销页码 {dep_page}")
        else:
            notes.append("未提取到折旧及摊销")
        row["备注"] = "；".join(notes)
        rows.append(row)

    matched_count = sum(1 for row in rows if row.get("代码"))
    if matched_count < len(rows):
        warnings.append(f"公募年报提取：{len(rows) - matched_count} 个 PDF 未能按文件名匹配到 REITs 代码。")
    return rows, warnings


def build_fund_references(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    references: dict[str, dict[str, str]] = {}
    for row in rows:
        code = normalize_code(row.get("REITs代码") or row.get("代码"))
        name = str(row.get("REITs名称") or row.get("基金名称") or row.get("名称") or "").strip()
        if not code or not name:
            continue
        references.setdefault(code, {"REITs代码": code, "REITs名称": name})
    return sorted(references.values(), key=lambda item: len(normalize_fund_name(item["REITs名称"])), reverse=True)


def match_pdf_to_fund_reference(pdf_path: Path, references: list[dict[str, str]]) -> dict[str, str]:
    file_key = normalize_fund_name(pdf_path.stem)
    for reference in references:
        code = reference["REITs代码"]
        if code and normalize_text(code) in normalize_text(pdf_path.stem):
            return {**reference, "匹配方式": "代码匹配"}
    for reference in references:
        name_key = normalize_fund_name(reference["REITs名称"])
        if name_key and (name_key in file_key or file_key in name_key):
            return {**reference, "匹配方式": "文件名包含匹配"}
    scored = [
        (SequenceMatcher(None, file_key, normalize_fund_name(reference["REITs名称"])).ratio(), reference)
        for reference in references
        if normalize_fund_name(reference["REITs名称"])
    ]
    scored.sort(key=lambda item: item[0], reverse=True)
    if scored:
        best_score, best_reference = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0
        if best_score >= 0.84 and best_score - second_score >= 0.03:
            return {**best_reference, "匹配方式": f"文件名相似匹配({best_score:.2f})"}
    return {}


def clean_report_fund_name_from_filename(path: Path) -> str:
    text = re.sub(r"\.(pdf|docx?|png|jpe?g|tiff?|bmp)$", "", path.name, flags=re.I)
    text = re.sub(r"20\d{2}年(度)?", "", text)
    text = re.sub(r"(年度报告|年报|评估报告|资产评估报告|不动产项目评估报告|市场价值|持有的全部不动产项目的)", "", text)
    text = re.sub(r"年度?$", "", text)
    return text.strip(" _-—")


def extract_amount_by_labels(text: str, labels: list[str], min_abs_value: float = 10000) -> float | None:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    normalized_labels = [normalize_text(label).lower() for label in labels]
    for index, line in enumerate(lines):
        normalized_line = normalize_text(line).lower()
        # PDF text often splits one Chinese table header into several adjacent lines.
        header_fragments: list[str] = []
        for fragment in lines[index : index + 5]:
            if re.search(r"\d", fragment):
                break
            header_fragments.append(fragment)
        normalized_window_label = normalize_text("".join(header_fragments)).lower()
        if not any(label and (label in normalized_line or label in normalized_window_label) for label in normalized_labels):
            continue
        if "比例" in normalized_line or "比率" in normalized_line or "%" in line:
            continue
        window = "\n".join(lines[index : index + 12])
        for raw_value in re.findall(r"-?\d[\d,]*(?:\.\d+)?", window):
            value = parse_number(raw_value)
            if value is None:
                continue
            # Skip years and small footnote numbers near the label.
            if abs(float(value)) < min_abs_value:
                continue
            return float(value)
    return None


def build_annual_report_financial_lookup(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for row in rows:
        code = normalize_code(row.get("代码"))
        if not code:
            continue
        normalized = {
            "REITs代码": code,
            "REITs名称": row.get("基金名称"),
            "基金净资产（万元）": row.get("基金净资产(万元)"),
            "折旧及摊销（万元）": row.get("折旧及摊销(万元)"),
            "来源文件": row.get("匹配PDF"),
        }
        cleaned = {key: value for key, value in normalized.items() if value not in (None, "")}
        lookup.setdefault(code, {}).update(cleaned)
        lookup.setdefault(f"code:{code}", {}).update(cleaned)
        name_key = normalize_fund_name(row.get("基金名称"))
        if name_key:
            lookup.setdefault(f"name:{name_key}", {}).update(cleaned)
    return lookup


def read_annual_report_financial_reference(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not Path(path).exists():
        return {}
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: dict[str, dict[str, Any]] = {}
    try:
        worksheet = workbook.active
        header_row, headers = find_header_row(worksheet)
        col_map = build_header_col_map(headers)
        code_col = col_map.get("代码")
        if not code_col:
            return {}
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            code = normalize_code(worksheet.cell(row_idx, code_col).value)
            if not code:
                continue
            rows[code] = {
                "代码": code,
                "基金名称": get_cell_by_normalized_header(worksheet, row_idx, col_map, "基金名称"),
                "匹配PDF": get_cell_by_normalized_header(worksheet, row_idx, col_map, "匹配PDF"),
                "基金净资产(万元)": get_cell_by_normalized_header(worksheet, row_idx, col_map, "基金净资产万元"),
                "折旧及摊销(万元)": get_cell_by_normalized_header(worksheet, row_idx, col_map, "折旧及摊销万元"),
            }
    finally:
        workbook.close()
    return rows


def get_cell_by_normalized_header(worksheet: Any, row_idx: int, col_map: dict[str, int], header: str) -> Any:
    col_idx = col_map.get(normalize_text(header))
    return worksheet.cell(row_idx, col_idx).value if col_idx else None


def write_annual_report_financial_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    reference_rows: dict[str, dict[str, Any]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "程序提取结果"
    headers = [
        "代码",
        "基金名称",
        "匹配PDF",
        "匹配方式",
        "基金净资产(元)",
        "基金净资产(万元)",
        "折旧及摊销(元)",
        "折旧及摊销(万元)",
        "数据来源",
        "备注",
    ]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    comparison = workbook.create_sheet("与参考表对比")
    comparison.append(["代码", "基金名称", "字段", "程序值", "参考值", "差额", "判断", "程序PDF", "参考PDF"])
    append_annual_report_financial_comparison_rows(comparison, rows, reference_rows)
    for worksheet in workbook.worksheets:
        style_simple_table(worksheet)
        worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(path)
    workbook.close()


def append_annual_report_financial_comparison_rows(
    worksheet: Any,
    rows: list[dict[str, Any]],
    reference_rows: dict[str, dict[str, Any]],
) -> None:
    rows_by_code = {normalize_code(row.get("代码")): row for row in rows if normalize_code(row.get("代码"))}
    for code in sorted(set(rows_by_code) | set(reference_rows)):
        extracted = rows_by_code.get(code, {})
        reference = reference_rows.get(code, {})
        for field in ("基金净资产(万元)", "折旧及摊销(万元)"):
            extracted_value = parse_number(extracted.get(field))
            reference_value = parse_number(reference.get(field))
            if extracted_value is None and reference_value is None:
                judgment = "两边均空"
                diff = ""
            elif extracted_value is None or reference_value is None:
                judgment = "缺失"
                diff = ""
            else:
                diff = round(float(extracted_value) - float(reference_value), 6)
                judgment = "一致" if abs(diff) <= 0.01 else "不一致"
            comparison.append(
                [
                    code,
                    extracted.get("基金名称") or reference.get("基金名称"),
                    field,
                    extracted_value,
                    reference_value,
                    diff,
                    judgment,
                    extracted.get("匹配PDF"),
                    reference.get("匹配PDF"),
                ]
            )


def enrich_rows_from_lookups(
    rows: list[dict[str, Any]],
    fee_rows: dict[str, dict[str, Any]],
    valuation_rows: dict[str, dict[str, Any]],
    annual_report_rows: dict[str, dict[str, Any]] | None = None,
    residual_rows: dict[str, dict[str, Any]] | None = None,
    announcement_rows: dict[str, dict[str, Any]] | None = None,
) -> None:
    for row in rows:
        if row.get("REITs代码") and not is_valid_reits_code(row.get("REITs代码")):
            row["REITs代码"] = ""
        for lookup_rows, lookup_kind in (
            (fee_rows, "fee"),
            (valuation_rows, "valuation"),
            (annual_report_rows or {}, "annual_report"),
            (announcement_rows or {}, "announcement"),
        ):
            lookup = find_lookup_for_standard_row(row, lookup_rows)
            apply_lookup_to_standard_row(row, lookup, lookup_kind)
        residual_lookup = find_project_lookup_for_standard_row(row, residual_rows or {})
        apply_residual_parameters_to_standard_row(row, residual_lookup)
        if row.get("基金净资产（万元）") is not None and parse_number(row.get("基金净资产（万元）")) <= 0:
            row["基金净资产（万元）"] = None
        infer_missing_asset_nature(row)


def apply_residual_parameters_to_standard_row(row: dict[str, Any], lookup: dict[str, Any]) -> None:
    if not lookup:
        return
    parameters = {
        field: lookup.get(field)
        for field in RESIDUAL_PARAMETER_FIELDS
        if lookup.get(field) not in (None, "")
    }
    if not parameters:
        return
    existing = row.get(RESIDUAL_PARAMETER_KEY)
    if isinstance(existing, dict):
        existing.update(parameters)
    else:
        row[RESIDUAL_PARAMETER_KEY] = parameters
    if row.get("折现率") in (None, "") and parameters.get("折现率") not in (None, ""):
        row["折现率"] = parameters["折现率"]


def read_project_alias_rows(path: Path | None) -> list[dict[str, Any]]:
    if not path or not Path(path).exists():
        return []
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            header_row, headers = find_header_row(worksheet)
            normalized_to_index = {normalize_text(header): index for index, header in enumerate(headers) if header not in (None, "")}
            source_col = first_existing_col(
                normalized_to_index,
                ["原项目名称", "来源项目名称", "输入项目名称", "OCR项目名称", "ocr项目名称", "项目别名", "别名"],
            )
            target_col = first_existing_col(normalized_to_index, ["标准项目名称", "目标项目名称", "正式项目名称", "输出项目名称"])
            if source_col is None or target_col is None:
                continue
            code_col = first_existing_col(normalized_to_index, ["REITs代码", "reits代码", "代码", "基金代码", "证券代码"])
            name_col = first_existing_col(normalized_to_index, ["REITs名称", "reits名称", "名称", "基金名称"])
            for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                source = values[source_col] if source_col < len(values) else None
                target = values[target_col] if target_col < len(values) else None
                if source in (None, "") or target in (None, ""):
                    continue
                rows.append(
                    {
                        "REITs代码": normalize_code(values[code_col]) if code_col is not None and code_col < len(values) else "",
                        "REITs名称": values[name_col] if name_col is not None and name_col < len(values) else "",
                        "原项目名称": str(source).strip(),
                        "标准项目名称": str(target).strip(),
                    }
                )
    finally:
        workbook.close()
    return rows


def first_existing_col(normalized_to_index: dict[str, int], candidates: list[str]) -> int | None:
    for candidate in candidates:
        col_idx = normalized_to_index.get(normalize_text(candidate))
        if col_idx is not None:
            return col_idx
    return None


def apply_project_aliases_to_rows(
    rows: list[dict[str, Any]],
    aliases: list[dict[str, Any]],
    review_items: list[dict[str, Any]],
) -> None:
    if not rows or not aliases:
        return
    by_code: dict[tuple[str, str], str] = {}
    by_name: dict[tuple[str, str], str] = {}
    global_aliases: dict[str, str] = {}
    for item in aliases:
        source = normalize_project(item.get("原项目名称"))
        target = str(item.get("标准项目名称") or "").strip()
        if not source or not target:
            continue
        code = normalize_code(item.get("REITs代码"))
        name = normalize_fund_name(item.get("REITs名称"))
        if code:
            by_code[(code, source)] = target
        elif name:
            by_name[(name, source)] = target
        else:
            global_aliases[source] = target

    changed: set[tuple[str, str, str]] = set()
    for row in rows:
        source_project = normalize_project(row.get("项目名称"))
        if not source_project:
            continue
        code = normalize_code(row.get("REITs代码"))
        name = normalize_fund_name(row.get("REITs名称"))
        target = by_code.get((code, source_project)) or by_name.get((name, source_project)) or global_aliases.get(source_project)
        if not target or normalize_project(target) == source_project:
            continue
        original = str(row.get("项目名称") or "")
        row["项目名称"] = target
        row["备注"] = append_warning(str(row.get("备注") or ""), f"项目名称按别名映射修正：{original} -> {target}")
        marker = (code, source_project, normalize_project(target))
        if marker not in changed:
            review_items.append(
                {
                    "类型": "项目别名映射",
                    "对象": f"{code or name or '未识别基金'} / {original}",
                    "说明": f"已按项目别名映射表将项目名称修正为：{target}",
                }
            )
            changed.add(marker)


def apply_lookup_to_standard_row(row: dict[str, Any], lookup: dict[str, Any], lookup_kind: str) -> None:
    if row.get("REITs代码") in (None, "") and lookup.get("REITs代码") not in (None, ""):
        row["REITs代码"] = lookup["REITs代码"]
    if lookup.get("REITs名称") not in (None, ""):
        row["REITs名称"] = lookup["REITs名称"]
    row_project = normalize_project(row.get("项目名称"))
    lookup_project = normalize_project(lookup.get("项目名称")) if lookup.get("项目名称") not in (None, "") else ""
    for field in STANDARD_FIELDS:
        if row.get(field) not in (None, "") or lookup.get(field) in (None, ""):
            continue
        if (
            lookup_kind == "valuation"
            and field == "基础资产评估价值（万元）"
            and row_project != normalize_project("项目整体")
            and (not lookup_project or lookup_project == normalize_project("项目整体"))
        ):
            # A code-level valuation is a fund/overall-project value. Do not
            # write it into each asset detail row, otherwise every project gets
            # the same total valuation.
            continue
        row[field] = lookup[field]


def collapse_generic_estimate_object_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    passthrough: list[dict[str, Any]] = []
    for row in rows:
        code = normalize_code(row.get("REITs代码"))
        source = str(row.get("来源文件") or "")
        if code and source:
            grouped.setdefault((code, source), []).append(row)
        else:
            passthrough.append(row)

    collapsed: list[dict[str, Any]] = list(passthrough)
    for group_rows in grouped.values():
        project_keys = {normalize_project(row.get("项目名称")) for row in group_rows if row.get("项目名称") not in (None, "")}
        non_whole_projects = {project for project in project_keys if project != "项目整体"}
        if not non_whole_projects or not all(is_generic_estimate_project(project) for project in non_whole_projects):
            collapsed.extend(group_rows)
            continue

        by_year: dict[int, dict[str, Any]] = {}
        no_year_rows: list[dict[str, Any]] = []
        for row in group_rows:
            year = row.get("年份")
            value = row.get("预测现金流金额（万元）")
            if not isinstance(year, int) or value in (None, ""):
                copy_row = dict(row)
                copy_row["项目名称"] = "项目整体"
                no_year_rows.append(copy_row)
                continue
            target = by_year.setdefault(year, {**row, "项目名称": "项目整体", "预测现金流金额（万元）": 0.0})
            target["预测现金流金额（万元）"] = parse_number(target.get("预测现金流金额（万元）")) + parse_number(value)
            target["备注"] = append_warning(str(target.get("备注") or ""), "估价对象一/二等通用标签已按年度汇总为项目整体。")
        collapsed.extend(no_year_rows)
        collapsed.extend(by_year[year] for year in sorted(by_year))
    return collapsed


def is_generic_estimate_project(project: Any) -> bool:
    normalized = normalize_project(project)
    return bool(re.fullmatch(r"(估价对象|评估对象|项目)[一二三四五六七八九十\d]+", normalized))


def canonicalize_project_names_from_existing_workbooks(rows: list[dict[str, Any]], paths: Iterable[Path | None]) -> None:
    high_priority_candidates: dict[str, set[str]] = {}
    low_priority_candidates: dict[str, set[str]] = {}
    for index, path in enumerate(paths):
        if not path or not Path(path).exists():
            continue
        candidates = high_priority_candidates if index < 2 else low_priority_candidates
        try:
            source_rows = read_standard_rows(Path(path))
        except Exception:
            continue
        for source_row in source_rows:
            code = normalize_code(source_row.get("REITs代码"))
            project = source_row.get("项目名称")
            if code and project not in (None, ""):
                candidates.setdefault(code, set()).add(str(project))

    for row in rows:
        code = normalize_code(row.get("REITs代码"))
        project = row.get("项目名称")
        if not code or project in (None, ""):
            continue
        high_candidates = high_priority_candidates.get(code, set())
        all_candidates = set(high_candidates) | low_priority_candidates.get(code, set())
        if normalize_project(project) == "项目整体":
            if any(normalize_project(candidate) == "项目整体" for candidate in all_candidates):
                continue
            single = single_non_whole_project_candidate(high_candidates) or single_non_whole_project_candidate(all_candidates)
            if single:
                row["备注"] = append_warning(str(row.get("备注") or ""), f"项目名称按原表唯一项目修正：{project} -> {single}")
                row["项目名称"] = single
            continue
        matched = best_project_candidate(str(project), high_candidates) or best_project_candidate(str(project), all_candidates)
        if matched and normalize_project(matched) != normalize_project(project):
            row["备注"] = append_warning(str(row.get("备注") or ""), f"项目名称按原表模糊修正：{project} -> {matched}")
            row["项目名称"] = matched


def single_non_whole_project_candidate(candidates: Iterable[str]) -> str | None:
    real_projects = sorted({str(candidate) for candidate in candidates if normalize_project(candidate) != "项目整体"})
    return real_projects[0] if len(real_projects) == 1 else None


def best_project_candidate(project: str, candidates: Iterable[str]) -> str | None:
    project_key = normalize_project(project)
    if not project_key or project_key == "项目整体":
        return None
    best_name = None
    best_score = 0.0
    for candidate in candidates:
        candidate_key = normalize_project(candidate)
        if not candidate_key or candidate_key == "项目整体":
            continue
        if candidate_key == project_key:
            return candidate
        if project_key in candidate_key or candidate_key in project_key:
            score = 0.96
        else:
            score = SequenceMatcher(None, project_key, candidate_key).ratio()
        if score > best_score:
            best_name = candidate
            best_score = score
    return best_name if best_score >= 0.82 else None


STANDARD_PRESERVED_FIELDS = {
    "上市日期": "基金上市日",
    "到期日": "基金到期日",
}

STANDARD_FILL_FROM_EXISTING_WHEN_MISSING_FIELDS = {
    "折现率": "折现率",
}

STANDARD_CONTEXT_ONLY_FIELDS = {
    "公告日期": "公告日期",
}

AUTO_EXPANDED_ROW_KEY = "__auto_expanded_year_row__"


def apply_existing_detail_context_to_rows(
    rows: list[dict[str, Any]],
    paths: Iterable[Path | None],
    review_items: list[dict[str, Any]],
    allow_fill_missing: bool = True,
) -> None:
    contexts: dict[str, list[dict[str, Any]]] = {}
    for path in paths:
        for context in read_existing_detail_contexts(path):
            contexts.setdefault(context["code"], []).append(context)

    reported: set[tuple[str, str, str]] = set()
    for group_key, group_rows in group_standard_rows(rows).items():
        code, project_key = group_key
        context = find_existing_detail_context(code, project_key, contexts.get(code, []))
        if not context:
            continue
        original_project = str(first_value(group_rows, "项目名称") or "")
        if normalize_project(context["project"]) != normalize_project(original_project):
            for row in group_rows:
                row["备注"] = append_warning(
                    str(row.get("备注") or ""),
                    f"项目名称按标准审核表口径修正：{original_project} -> {context['project']}",
                )
                row["项目名称"] = context["project"]
        for field_name, source_header in STANDARD_PRESERVED_FIELDS.items():
            standard_value = context["values"].get(field_name)
            if standard_value in (None, ""):
                continue
            changed = False
            for row in group_rows:
                current = row.get(field_name)
                if not annual_values_equivalent(current, standard_value):
                    row[field_name] = standard_value
                    row["备注"] = append_warning(
                        str(row.get("备注") or ""),
                        f"{source_header}按标准审核表保留；辅助表/导入表口径为 {display_value(current)}，标准审核表口径为 {display_value(standard_value)}",
                    )
                    changed = True
            if changed and (code, normalize_project(context["project"]), field_name) not in reported:
                review_items.append(
                    {
                        "类型": "标准表口径覆盖",
                        "对象": f"{code} / {context['project']} / {source_header}",
                        "说明": f"辅助表或标准导入表与标准审核表不一致，已按标准审核表保留 {display_value(standard_value)}；新项目无标准表历史记录时才使用辅助表/导入表值。",
                    }
                )
                reported.add((code, normalize_project(context["project"]), field_name))
        if not allow_fill_missing:
            continue
        for field_name, source_header in STANDARD_FILL_FROM_EXISTING_WHEN_MISSING_FIELDS.items():
            standard_value = context["values"].get(field_name)
            if standard_value in (None, ""):
                continue
            changed = False
            for row in group_rows:
                if row.get(field_name) in (None, ""):
                    row[field_name] = standard_value
                    row["备注"] = append_warning(
                        str(row.get("备注") or ""),
                        f"{source_header}本轮未提供，暂按源明细表同项目历史口径补齐：{display_value(standard_value)}",
                    )
                    changed = True
            if changed and (code, normalize_project(context["project"]), field_name) not in reported:
                review_items.append(
                    {
                        "类型": "源表口径补齐",
                        "对象": f"{code} / {context['project']} / {source_header}",
                        "说明": f"本轮导入表或辅助表未提供 {source_header}，已使用源明细表同项目已有值 {display_value(standard_value)}；如今年报告披露新口径，请在辅助表中补充后重跑。",
                    }
                )
                reported.add((code, normalize_project(context["project"]), field_name))


def expand_standard_rows_with_year_skeletons(
    rows: list[dict[str, Any]],
    paths: Iterable[Path | None],
    review_items: list[dict[str, Any]],
) -> None:
    contexts: dict[str, list[dict[str, Any]]] = {}
    for path in paths:
        for context in read_existing_detail_year_contexts(path):
            contexts.setdefault(context["code"], []).append(context)

    appended_rows: list[dict[str, Any]] = []
    reported: set[tuple[str, str]] = set()
    for group_key, group_rows in group_standard_rows(rows).items():
        code, project_key = group_key
        context = find_existing_detail_year_context(code, project_key, contexts.get(code, []))
        existing_years = sorted({row.get("年份") for row in group_rows if isinstance(row.get("年份"), int)})
        if not existing_years:
            continue

        target_years = []
        if context:
            target_years = [year for year in context["years"] if year >= existing_years[0]]
        else:
            terminal_year = date_year(first_value(group_rows, "到期日"))
            if terminal_year and terminal_year > existing_years[-1]:
                target_years = list(range(existing_years[0], terminal_year + 1))
        if not target_years:
            continue

        target_years = sorted({year for year in target_years if isinstance(year, int)})
        growth_start = first_value(group_rows, "增长率预测起始年度")
        growth_rate = first_value(group_rows, "预测现金流增长率")
        base_row = next(
            (row for row in group_rows if row.get("年份") == existing_years[-1]),
            group_rows[-1],
        )
        rows_by_year = {row.get("年份"): row for row in group_rows if isinstance(row.get("年份"), int)}
        added_count = 0
        computed_count = 0

        for year in target_years:
            if year in rows_by_year:
                continue
            new_row = dict(base_row)
            new_row["年份"] = year
            new_row[AUTO_EXPANDED_ROW_KEY] = True
            previous_row = rows_by_year.get(year - 1)
            previous_cashflow = previous_row.get("预测现金流金额（万元）") if previous_row else None
            if (
                isinstance(growth_start, int)
                and growth_rate not in (None, "")
                and year >= growth_start
                and previous_cashflow not in (None, "")
            ):
                new_row["预测现金流金额（万元）"] = round(
                    parse_number(previous_cashflow) * (1 + parse_number(growth_rate)),
                    6,
                )
                new_row["备注"] = append_warning(
                    str(new_row.get("备注") or ""),
                    f"按增长率自动补算 {year} 年现金流，用于补齐至存续期末。",
                )
                computed_count += 1
            else:
                new_row["预测现金流金额（万元）"] = None
                new_row["备注"] = append_warning(
                    str(new_row.get("备注") or ""),
                    f"按标准审核表/到期日补齐 {year} 年行；因缺少增长率或前序现金流，金额暂留空。",
                )
            rows_by_year[year] = new_row
            appended_rows.append(new_row)
            added_count += 1

        if added_count and group_key not in reported:
            review_items.append(
                {
                    "类型": "年份骨架补齐",
                    "对象": f"{code} / {first_value(group_rows, '项目名称') or '项目整体'}",
                    "说明": (
                        f"已按标准审核表或到期日补齐 {added_count} 个缺失年度行"
                        + (f"，其中 {computed_count} 行按增长率自动补算。" if computed_count else "，其余金额留空待人工复核。")
                    ),
                }
            )
            reported.add(group_key)

    if appended_rows:
        rows.extend(appended_rows)


def read_existing_detail_year_contexts(path: Path | None) -> list[dict[str, Any]]:
    if not path or not Path(path).exists():
        return []
    try:
        source_rows = read_standard_rows(Path(path))
    except Exception:
        return []

    grouped: dict[tuple[str, str], set[int]] = {}
    labels: dict[tuple[str, str], str] = {}
    for row in source_rows:
        code = normalize_code(row.get("REITs代码"))
        project = str(row.get("项目名称") or "项目整体")
        project_key = normalize_project(project)
        year = row.get("年份")
        if not code or not project_key or not isinstance(year, int):
            continue
        grouped.setdefault((code, project_key), set()).add(year)
        labels.setdefault((code, project_key), project)

    return [
        {
            "code": code,
            "project": labels[(code, project_key)],
            "years": sorted(years),
        }
        for (code, project_key), years in grouped.items()
        if years
    ]


def find_existing_detail_year_context(code: str, project_key: str, contexts: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not contexts:
        return None
    exact = [context for context in contexts if normalize_project(context["project"]) == project_key]
    if exact:
        return exact[0]
    matches = [
        context
        for context in contexts
        if projects_match(normalize_project(context["project"]), project_key)
    ]
    if matches:
        return max(matches, key=lambda context: project_match_score(normalize_project(context["project"]), project_key))
    return None


def read_existing_detail_contexts(path: Path | None) -> list[dict[str, Any]]:
    if not path or not Path(path).exists():
        return []
    workbook = load_workbook(Path(path), data_only=False)
    try:
        worksheet = workbook.active
        header_row, headers = find_header_row(worksheet)
        col_map = build_header_col_map(headers)
        code_col = col_map.get("代码")
        project_col = col_map.get("项目名称")
        if not code_col or not project_col:
            return []
        grouped: dict[tuple[str, str], dict[str, Any]] = {}
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            code = normalize_code(worksheet.cell(row_idx, code_col).value)
            project = worksheet.cell(row_idx, project_col).value
            project_key = normalize_project(project)
            if not code or not project_key:
                continue
            item = grouped.setdefault(
                (code, project_key),
                {
                    "code": code,
                    "project": str(project or "项目整体"),
                    "raw_values": {
                        field: []
                        for field in {
                            *STANDARD_PRESERVED_FIELDS,
                            *STANDARD_FILL_FROM_EXISTING_WHEN_MISSING_FIELDS,
                            *STANDARD_CONTEXT_ONLY_FIELDS,
                        }
                    },
                },
            )
            for field_name, header in {
                **STANDARD_PRESERVED_FIELDS,
                **STANDARD_FILL_FROM_EXISTING_WHEN_MISSING_FIELDS,
                **STANDARD_CONTEXT_ONLY_FIELDS,
            }.items():
                col_idx = col_map.get(normalize_text(header))
                if not col_idx:
                    continue
                value = worksheet.cell(row_idx, col_idx).value
                if value not in (None, ""):
                    item["raw_values"][field_name].append(value)
        contexts: list[dict[str, Any]] = []
        for item in grouped.values():
            values = {
                field_name: choose_existing_standard_value(field_name, raw_values)
                for field_name, raw_values in item["raw_values"].items()
            }
            item["values"] = values
            del item["raw_values"]
            contexts.append(item)
        return contexts
    finally:
        workbook.close()


def choose_existing_standard_value(field_name: str, values: list[Any]) -> Any:
    cleaned = [value for value in values if value not in (None, "")]
    if not cleaned:
        return None
    if field_name == "到期日":
        return max(cleaned, key=date_sort_key)
    return most_common_value(cleaned)


def most_common_value(values: list[Any]) -> Any:
    counts: dict[str, tuple[int, Any]] = {}
    for value in values:
        key = value_identity(value)
        count, original = counts.get(key, (0, value))
        counts[key] = (count + 1, original)
    return max(counts.values(), key=lambda item: item[0])[1]


def apply_annual_period_defaults_to_rows(
    rows: list[dict[str, Any]],
    paths: Iterable[Path | None],
    output_start_year: int,
    review_items: list[dict[str, Any]],
    allow_existing_context_fill: bool = True,
) -> None:
    """Update year-sensitive fields so detail outputs do not carry last year's period."""
    if not rows or not isinstance(output_start_year, int):
        return

    report_year = output_start_year - 1
    report_period = f"{report_year}年评估报告"
    valuation_date = date(report_year, 12, 31)
    contexts: dict[str, list[dict[str, Any]]] = {}
    for path in paths:
        for context in read_existing_detail_contexts(path):
            contexts.setdefault(context["code"], []).append(context)

    reported: set[tuple[str, str]] = set()
    for (code, project_key), group_rows in group_standard_rows(rows).items():
        context = find_existing_detail_context(code, project_key, contexts.get(code, []))
        context_announcement = context["values"].get("公告日期") if context and allow_existing_context_fill else None
        changed = False
        for row in group_rows:
            original_report = row.get("报告期")
            original_valuation = row.get("评估基准日")
            original_announcement = row.get("公告日期")

            row["报告期"] = report_period
            row["评估基准日"] = valuation_date
            row["公告日期"] = annualized_announcement_date(
                original_announcement,
                context_announcement,
                output_start_year,
            )
            if (
                not annual_values_equivalent(original_report, report_period)
                or not annual_values_equivalent(original_valuation, valuation_date)
                or not annual_values_equivalent(original_announcement, row["公告日期"])
            ):
                changed = True

        if changed and (code, project_key) not in reported:
            review_items.append(
                {
                    "类型": "年度字段更新",
                    "对象": f"{code} / {first_value(group_rows, '项目名称') or '项目整体'}",
                    "说明": (
                        f"已按输出起始年份 {output_start_year} 更新公告日期、报告期和评估/估值基准日；"
                        f"报告期={report_period}，评估/估值基准日={valuation_date.isoformat()}。"
                        "公告日期优先使用本轮导入/辅助表；如缺失，则按源模板同项目公告日期的月日平移到本轮输出年份。"
                    ),
                }
            )
            reported.add((code, project_key))


def annualized_announcement_date(current_value: Any, context_value: Any, output_year: int) -> date:
    current_date = parse_date_like(current_value)
    if isinstance(current_date, datetime):
        current_date = current_date.date()
    if isinstance(current_date, date):
        return current_date if current_date.year == output_year else replace_date_year(current_date, output_year)

    context_date = parse_date_like(context_value)
    if isinstance(context_date, datetime):
        context_date = context_date.date()
    if isinstance(context_date, date):
        return replace_date_year(context_date, output_year)

    return date(output_year, 3, 31)


def replace_date_year(value: date, target_year: int) -> date:
    try:
        return value.replace(year=target_year)
    except ValueError:
        return date(target_year, 2, 28)


def value_identity(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def date_sort_key(value: Any) -> tuple[int, str]:
    parsed = parse_date_like(value)
    if isinstance(parsed, datetime):
        return (parsed.toordinal(), "")
    if isinstance(parsed, date):
        return (parsed.toordinal(), "")
    return (0, str(value))


def find_existing_detail_context(code: str, project_key: str, contexts: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not contexts:
        return None
    exact = [context for context in contexts if normalize_project(context["project"]) == project_key]
    if exact:
        return exact[0]
    matches = [
        context
        for context in contexts
        if projects_match(normalize_project(context["project"]), project_key)
    ]
    if matches:
        return max(matches, key=lambda context: project_match_score(normalize_project(context["project"]), project_key))
    if project_key == "项目整体":
        real_projects = [context for context in contexts if normalize_project(context["project"]) != "项目整体"]
        if len(real_projects) == 1:
            return real_projects[0]
    return None


def project_match_score(left: str, right: str) -> float:
    if left == right:
        return 1.0
    if not left or not right:
        return 0.0
    if left in right or right in left:
        return 0.96
    return SequenceMatcher(None, left, right).ratio()


def annual_values_equivalent(left: Any, right: Any) -> bool:
    if left in (None, "") and right in (None, ""):
        return True
    left_date = parse_date_like(left)
    right_date = parse_date_like(right)
    if isinstance(left_date, (date, datetime)) and isinstance(right_date, (date, datetime)):
        return date_sort_key(left_date) == date_sort_key(right_date)
    if isinstance(left, (int, float)) or isinstance(right, (int, float)):
        try:
            return abs(float(left) - float(right)) <= 0.000001
        except Exception:
            pass
    return str(left or "") == str(right or "")


def display_value(value: Any) -> str:
    if value in (None, ""):
        return "空"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def find_lookup_for_standard_row(row: dict[str, Any], lookup_rows: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if not lookup_rows:
        return {}

    code = normalize_code(row.get("REITs代码"))
    if code:
        found = lookup_rows.get(code) or lookup_rows.get(f"code:{code}")
        if found:
            return found

    name_key = normalize_fund_name(row.get("REITs名称"))
    if name_key:
        found = lookup_rows.get(f"name:{name_key}")
        if found:
            return found
        for key, candidate in lookup_rows.items():
            if not key.startswith("name:"):
                continue
            candidate_key = key.removeprefix("name:")
            if len(name_key) >= 8 and (name_key in candidate_key or candidate_key in name_key):
                return candidate
    return {}


def find_project_lookup_for_standard_row(row: dict[str, Any], lookup_rows: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if not lookup_rows:
        return {}

    code = normalize_code(row.get("REITs代码"))
    project_key = normalize_project(row.get("项目名称") or "项目整体")
    if code and project_key:
        found = lookup_rows.get(f"code_project:{code}:{project_key}")
        if found:
            return found

    name_key = normalize_fund_name(row.get("REITs名称"))
    if name_key and project_key:
        found = lookup_rows.get(f"name_project:{name_key}:{project_key}")
        if found:
            return found

    if project_key == normalize_project("项目整体"):
        return find_lookup_for_standard_row(row, lookup_rows)
    return {}


def filter_rows_by_asset_nature(rows: list[dict[str, Any]], asset_nature: str) -> list[dict[str, Any]]:
    target = normalize_text(asset_nature)
    return [row for row in rows if target in normalize_text(row.get("底层资产性质"))]


def filter_rows_by_start_year(rows: list[dict[str, Any]], start_year: int) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    grouped = group_standard_rows(rows)
    for group_rows in grouped.values():
        keep_group = [
            row
            for row in group_rows
            if not isinstance(row.get("年份"), int) or row.get("年份") >= start_year
        ]
        if keep_group:
            filtered.extend(keep_group)
            continue

        # Preserve non-year parameters, such as growth start/rate, even when the
        # disclosed rows are all before the configured output start year.
        parameter_row = next(
            (
                row
                for row in group_rows
                if row.get("增长率预测起始年度") or row.get("预测现金流增长率") or row.get("经营期末")
            ),
            None,
        )
        if parameter_row:
            copy_row = dict(parameter_row)
            copy_row["年份"] = None
            copy_row["预测现金流金额（万元）"] = None
            filtered.append(copy_row)
    return filtered


GROUP_STATIC_CONTEXT_FIELDS = (
    "REITs名称",
    "基础设施项目类型",
    "底层资产性质",
    "股债",
    "上市日期",
    "上市年份",
    "到期日",
    "公告日期",
    "报告期",
    "是否整体项目",
    "增长率预测起始年度",
    "预测现金流增长率",
    "经营期末",
    "折现率",
    "评估基准日",
    "基础资产评估价值（万元）",
    "固定管理费率(%)",
    "托管费率(%)",
    "调整浮动管理费",
)


def fill_group_static_context(rows: list[dict[str, Any]]) -> None:
    """Propagate stable group metadata to OCR-only rows before writing details."""
    for group_rows in group_standard_rows(rows).values():
        context = {field: first_value(group_rows, field) for field in GROUP_STATIC_CONTEXT_FIELDS}
        for row in group_rows:
            for field, value in context.items():
                if row.get(field) in (None, "") and value not in (None, ""):
                    row[field] = value


def filter_disclosed_cashflow_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Rows written to the future-cashflow wide table.

    The wide table is an audit aid showing only disclosed/imported years. Rows
    created later from growth-rate formulas stay in the detail workbook only.
    """
    return [
        dict(row)
        for row in rows
        if isinstance(row.get("年份"), int) and row.get("预测现金流金额（万元）") not in (None, "")
    ]


def update_future_cashflow_workbook(
    source_path: Path | None,
    rows: list[dict[str, Any]],
    output_dir: Path,
    review_items: list[dict[str, Any]],
) -> Path:
    if source_path and source_path.exists():
        workbook = load_workbook(source_path)
        worksheet = workbook.active
        output_path = output_dir / FUTURE_CASHFLOW_OUTPUT_NAME
        prepare_future_cashflow_template(worksheet)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        output_path = output_dir / FUTURE_CASHFLOW_OUTPUT_NAME
        build_future_cashflow_header(worksheet)

    header_map = ensure_future_year_columns(worksheet)
    grouped = group_standard_rows(rows)
    for group_key, group_rows in grouped.items():
        row_idx = find_or_append_future_row(worksheet, header_map, group_key, group_rows)
        fill_future_metadata(worksheet, row_idx, group_rows)
        for item in group_rows:
            year = item.get("年份")
            value = item.get("预测现金流金额（万元）")
            if isinstance(year, int) and value not in (None, ""):
                col_idx = header_map.get(year)
                if col_idx:
                    worksheet.cell(row_idx, col_idx).value = value
        growth_start = first_value(group_rows, "增长率预测起始年度")
        growth_rate = first_value(group_rows, "预测现金流增长率")
        terminal = first_value(group_rows, "经营期末")
        if growth_start:
            worksheet.cell(row_idx, 12).value = growth_start
        if growth_rate not in (None, ""):
            worksheet.cell(row_idx, 13).value = growth_rate
        if terminal not in (None, ""):
            worksheet.cell(row_idx, 14).value = terminal

    if source_path and source_path.exists():
        apply_future_reference_format(worksheet, source_path)
    normalize_future_cashflow_output_styles(worksheet)
    finalize_worksheet_view(worksheet)
    workbook.save(output_path)
    workbook.close()
    review_items.append({"类型": "输出", "对象": output_path.name, "说明": "未来现金流宽表已输出。"})
    return output_path


def find_future_reference_workbook(workspace: Path) -> Path | None:
    reference_dirs: list[Path] = []
    root = workspace if workspace.is_dir() else workspace.parent
    for base in [root, *root.parents]:
        if not base.exists() or not base.is_dir():
            continue
        reference_dirs.append(base)
        direct = base / "对比今年参考样表"
        if direct.exists() and direct.is_dir():
            reference_dirs.append(direct)
        try:
            for child in base.iterdir():
                if child.is_dir() and is_checked_reference_dir(child):
                    reference_dirs.append(child)
        except OSError:
            continue
        if (base / "reit_excel_auditor").exists() and (base / "pyproject.toml").exists():
            break

    seen: set[Path] = set()
    for directory in reference_dirs:
        if directory in seen:
            continue
        seen.add(directory)
        try:
            candidates = sorted(
                path
                for path in directory.glob("*.xlsx")
                if not path.name.startswith("~$")
                and "未来现金流" in path.stem
                and is_likely_checked_reference_path(path)
                and "自动" not in path.stem
            )
        except OSError:
            continue
        if candidates:
            return candidates[0]
    return None


def apply_future_reference_format(worksheet: Any, reference_path: Path) -> None:
    try:
        reference_workbook = load_workbook(reference_path, keep_links=False)
    except Exception:
        return
    try:
        reference_worksheet = reference_workbook.active
        max_col = min(worksheet.max_column, reference_worksheet.max_column)
        worksheet.freeze_panes = reference_worksheet.freeze_panes
        worksheet.auto_filter.ref = reference_worksheet.auto_filter.ref
        worksheet.sheet_format.defaultRowHeight = reference_worksheet.sheet_format.defaultRowHeight
        worksheet.sheet_format.defaultColWidth = reference_worksheet.sheet_format.defaultColWidth
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            source_dimension = reference_worksheet.column_dimensions[letter]
            target_dimension = worksheet.column_dimensions[letter]
            target_dimension.width = source_dimension.width
            target_dimension.hidden = source_dimension.hidden
        for row_idx in range(1, worksheet.max_row + 1):
            source_row_idx = row_idx if row_idx <= reference_worksheet.max_row else 2
            worksheet.row_dimensions[row_idx].height = reference_worksheet.row_dimensions[source_row_idx].height
            for col_idx in range(1, max_col + 1):
                copy_cell_format(reference_worksheet.cell(source_row_idx, col_idx), worksheet.cell(row_idx, col_idx))
    finally:
        reference_workbook.close()


def normalize_future_cashflow_output_styles(worksheet: Any) -> None:
    """Apply the final future-cashflow font policy after reference-format copying."""
    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_idx, col_idx)
            if type(cell).__name__ == "MergedCell":
                continue
            cell.font = clone_font_with_name(cell.font, FUTURE_CASHFLOW_FONT_NAME)


def normalize_detail_autofilter_range(worksheet: Any, header_row: int) -> None:
    """Keep filters aligned to the generated table size instead of old template rows."""
    if not worksheet.auto_filter.ref:
        return
    last_col = get_column_letter(worksheet.max_column)
    worksheet.auto_filter.ref = f"A{header_row}:{last_col}{worksheet.max_row}"


def prepare_future_cashflow_template(worksheet: Any) -> None:
    """Keep the future-cashflow style template but remove old data values."""
    header_row, _headers = find_header_row(worksheet)
    if header_row != 1:
        return
    if worksheet.max_row >= 2:
        if worksheet.max_row > 2:
            worksheet.delete_rows(3, worksheet.max_row - 2)
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(2, col_idx).value = None
    else:
        worksheet.append([None for _ in range(worksheet.max_column)])
        for col_idx in range(1, worksheet.max_column + 1):
            target = worksheet.cell(2, col_idx)
            target.font = Font(name="Microsoft YaHei", size=11)
            target.fill = PatternFill(fill_type=None)
            target.alignment = Alignment(vertical="center")


def build_future_cashflow_header(worksheet: Any) -> None:
    headers = [
        "名称",
        "代码",
        "公告日期",
        "基础设施项目类型",
        "底层资产性质",
        "股债",
        "上市日期",
        "上市年份",
        "到期日",
        "报告期",
        "项目名称",
        "自年份",
        "未来增长率",
        "经营期末",
    ] + list(range(2020, 2101))
    worksheet.append(headers)
    worksheet.freeze_panes = "A2"
    for col_idx in range(1, len(headers) + 1):
        worksheet.cell(1, col_idx).font = Font(name="Microsoft YaHei", bold=True)
        worksheet.cell(1, col_idx).alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 13
    worksheet.column_dimensions["A"].width = 42
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["J"].width = 22
    worksheet.column_dimensions["K"].width = 22


def ensure_future_year_columns(worksheet: Any) -> dict[int, int]:
    header_values = [worksheet.cell(1, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    year_map = {header: col_idx for col_idx, header in enumerate(header_values, 1) if isinstance(header, int)}
    if year_map:
        return year_map
    for year in range(2020, 2101):
        col_idx = worksheet.max_column + 1
        worksheet.cell(1, col_idx).value = year
        year_map[year] = col_idx
    return year_map


def find_or_append_future_row(
    worksheet: Any,
    year_map: dict[int, int],
    group_key: tuple[str, str],
    group_rows: list[dict[str, Any]],
) -> int:
    code, project_key = group_key
    first_blank_row: int | None = None
    for row_idx in range(2, worksheet.max_row + 1):
        row_code = normalize_code(worksheet.cell(row_idx, 2).value)
        row_project = normalize_project(worksheet.cell(row_idx, 11).value)
        if row_code == code and projects_match(row_project, project_key):
            return row_idx
        if first_blank_row is None and is_blank_future_data_row(worksheet, row_idx, year_map):
            first_blank_row = row_idx
    if first_blank_row:
        return first_blank_row

    row_idx = worksheet.max_row + 1
    style_row = best_future_style_row(worksheet, group_rows)
    copy_row_style_and_formulas(worksheet, style_row, row_idx, copy_static_values=False)
    return row_idx


def is_blank_future_data_row(worksheet: Any, row_idx: int, year_map: dict[int, int]) -> bool:
    check_columns = set(range(1, 15)) | set(year_map.values())
    return all(worksheet.cell(row_idx, col_idx).value in (None, "") for col_idx in check_columns)


def best_future_style_row(worksheet: Any, group_rows: list[dict[str, Any]]) -> int:
    asset_nature = normalize_text(first_value(group_rows, "底层资产性质"))
    code = normalize_code(first_value(group_rows, "REITs代码"))
    for row_idx in range(2, worksheet.max_row + 1):
        if normalize_code(worksheet.cell(row_idx, 2).value) == code:
            return row_idx
    for row_idx in range(2, worksheet.max_row + 1):
        if asset_nature and asset_nature in normalize_text(worksheet.cell(row_idx, 5).value):
            return row_idx
    return 2 if worksheet.max_row >= 2 else 1


def fill_future_metadata(worksheet: Any, row_idx: int, group_rows: list[dict[str, Any]]) -> None:
    metadata = {
        1: first_value(group_rows, "REITs名称"),
        2: first_value(group_rows, "REITs代码"),
        3: first_value(group_rows, "公告日期"),
        4: first_value(group_rows, "基础设施项目类型"),
        5: first_value(group_rows, "底层资产性质"),
        6: first_value(group_rows, "股债"),
        7: to_yyyymmdd(first_value(group_rows, "上市日期")),
        8: first_value(group_rows, "上市年份") or date_year(first_value(group_rows, "上市日期")),
        9: to_yyyymmdd(first_value(group_rows, "到期日")),
        10: first_value(group_rows, "报告期"),
        11: first_value(group_rows, "项目名称") or "项目整体",
    }
    for col_idx, value in metadata.items():
        if value not in (None, ""):
            worksheet.cell(row_idx, col_idx).value = value


def update_detail_workbook(
    source_path: Path,
    rows: list[dict[str, Any]],
    kind: str,
    output_dir: Path,
    review_items: list[dict[str, Any]],
    format_reference_path: Path | None = None,
) -> tuple[Path, int]:
    workbook = load_workbook(source_path, keep_links=False)
    strip_unstable_template_artifacts(workbook)
    worksheet = workbook.active
    keep_only_active_worksheet(workbook, worksheet)
    preserve_source_visuals = not is_internal_annual_template_path(source_path)
    if kind == "property":
        migrate_property_detail_layout_if_needed(worksheet)
    elif kind == "concession":
        ensure_min_column_count(worksheet, 46)
    output_path = output_dir / f"{source_path.stem}_自动更新{source_path.suffix}"
    header_row, headers = find_header_row(worksheet)
    col_map = build_header_col_map(headers)
    grouped = group_standard_rows(rows)
    ordered_group_keys = order_detail_group_keys(worksheet, col_map, grouped.keys(), header_row)
    grouped_keys_by_code: dict[str, list[tuple[str, str]]] = {}
    for group_key in ordered_group_keys:
        grouped_keys_by_code.setdefault(group_key[0], []).append(group_key)
    template_row = prepare_detail_worksheet_for_fresh_write(worksheet, col_map, rows, header_row)
    updated_count = 0
    inserted_initial_codes: set[str] = set()

    for group_key in ordered_group_keys:
        group_rows = grouped[group_key]
        if should_insert_initial_row_before_group(kind, group_key, grouped_keys_by_code, inserted_initial_codes):
            insert_detail_initial_row(worksheet, col_map, template_row or header_row + 1, group_rows, kind)
            inserted_initial_codes.add(group_key[0])
        target_rows = find_detail_rows(worksheet, col_map, group_key)
        updated_count += update_detail_group(worksheet, col_map, header_row, group_key, group_rows, target_rows, kind, review_items)
    review_items.append(
        {
            "类型": "重建项目块",
            "对象": output_path.name,
            "说明": f"已清空旧年度明细，并按本次标准化导入表重建 {len(grouped)} 个基金/项目块；具体范围请看“更新计划”。",
        }
    )

    if template_row and worksheet.max_row > template_row:
        worksheet.delete_rows(template_row)
    repair_detail_formulas_after_template_delete(worksheet, col_map, header_row, kind)
    repair_initial_row_formulas(worksheet, col_map, header_row, kind)
    if kind == "concession":
        ensure_min_column_count(worksheet, 46)
    if format_reference_path and format_reference_path.exists():
        apply_detail_reference_format_if_available(worksheet, source_path, kind, format_reference_path)
        review_items.append(
            {
                "类型": "格式参考表",
                "对象": output_path.name,
                "说明": f"已优先使用 {format_reference_path.name} 作为{'产权' if kind == 'property' else '特许经营权'}格式参考表；数据仍来自本轮标准化导入和辅助表。",
            }
        )
    ensure_detail_visible_format_area(worksheet, col_map, header_row)
    normalize_detail_output_styles(
        worksheet,
        col_map,
        kind,
        header_row,
        preserve_existing_fonts=preserve_source_visuals and not (format_reference_path and format_reference_path.exists()),
    )
    if kind == "property" and not (format_reference_path and format_reference_path.exists()):
        normalize_property_group_header_styles(worksheet)
    if format_reference_path and format_reference_path.exists():
        apply_detail_reference_format_if_available(worksheet, source_path, kind, format_reference_path)
    clear_yellow_fills(worksheet)
    normalize_detail_autofilter_range(worksheet, header_row)
    finalize_worksheet_view(worksheet)
    workbook.save(output_path)
    workbook.close()
    review_items.append({"类型": "输出", "对象": output_path.name, "说明": f"{'产权' if kind == 'property' else '特许经营权'}明细表已输出。"})
    return output_path, updated_count


def keep_only_active_worksheet(workbook: Workbook, active_worksheet: Any) -> None:
    for worksheet in list(workbook.worksheets):
        if worksheet is active_worksheet:
            continue
        del workbook[worksheet.title]


def strip_unstable_template_artifacts(workbook: Workbook) -> None:
    """Drop template artifacts that openpyxl cannot safely preserve after row rebuilds."""
    if hasattr(workbook, "_external_links"):
        workbook._external_links = []
    for worksheet in workbook.worksheets:
        if hasattr(worksheet, "legacy_drawing"):
            worksheet.legacy_drawing = None


def apply_detail_reference_format_if_available(
    worksheet: Any,
    source_path: Path,
    kind: str,
    explicit_reference_path: Path | None = None,
) -> None:
    """Use the current-year checked workbook as a style reference when it is available.

    The annual-update writer rebuilds rows from last year's workbook, so the safest
    way to avoid style drift is to keep values/formulas from the generated output
    and copy only visual formatting from the checked reference workbook.
    """
    reference_path = explicit_reference_path or find_detail_reference_workbook(source_path, kind)
    if not reference_path:
        return

    try:
        reference_workbook = load_workbook(reference_path, keep_links=False)
    except Exception:
        return
    try:
        reference_worksheet = reference_workbook.active
        trim_worksheet_to_reference_columns(worksheet, reference_worksheet)
        copy_detail_reference_static_format(worksheet, reference_worksheet)
        copy_detail_reference_data_format(worksheet, reference_worksheet, kind)
        copy_detail_reference_first_data_style(worksheet, reference_worksheet)
    finally:
        reference_workbook.close()


def find_detail_reference_workbook(source_path: Path, kind: str) -> Path | None:
    wanted = "产权" if kind == "property" else "特许"
    reference_dirs: list[Path] = []
    for base in [source_path.parent, *source_path.parents]:
        if not base.exists() or not base.is_dir():
            continue
        reference_dirs.append(base)
        direct = base / "对比今年参考样表"
        if direct.exists() and direct.is_dir():
            reference_dirs.append(direct)
        try:
            for child in base.iterdir():
                if child.is_dir() and is_checked_reference_dir(child):
                    reference_dirs.append(child)
        except OSError:
            continue
        if (base / "reit_excel_auditor").exists() and (base / "pyproject.toml").exists():
            break

    seen: set[Path] = set()
    candidates: list[Path] = []
    for directory in reference_dirs:
        if directory in seen:
            continue
        seen.add(directory)
        try:
            candidates.extend(
                path
                for path in sorted(directory.glob("*.xlsx"))
                if not path.name.startswith("~$")
                and wanted in path.stem
                and is_likely_checked_reference_path(path)
                and "自动" not in path.stem
                and "过程" not in path.stem
            )
        except OSError:
            continue
    return candidates[0] if candidates else None


def trim_worksheet_to_reference_columns(worksheet: Any, reference_worksheet: Any) -> None:
    """Remove trailing columns beyond the checked reference workbook."""
    if worksheet.max_column > reference_worksheet.max_column:
        worksheet.delete_cols(reference_worksheet.max_column + 1, worksheet.max_column - reference_worksheet.max_column)


def copy_detail_reference_static_format(worksheet: Any, reference_worksheet: Any) -> None:
    max_col = min(worksheet.max_column, reference_worksheet.max_column)
    max_header_row = min(find_header_row(worksheet)[0], find_header_row(reference_worksheet)[0])

    worksheet.freeze_panes = reference_worksheet.freeze_panes
    worksheet.auto_filter.ref = reference_worksheet.auto_filter.ref
    worksheet.sheet_format.defaultRowHeight = reference_worksheet.sheet_format.defaultRowHeight
    worksheet.sheet_format.defaultColWidth = reference_worksheet.sheet_format.defaultColWidth
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        source_dimension = reference_worksheet.column_dimensions[letter]
        target_dimension = worksheet.column_dimensions[letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row <= max_header_row:
            worksheet.unmerge_cells(str(merged_range))
    for merged_range in reference_worksheet.merged_cells.ranges:
        if merged_range.max_row <= max_header_row and merged_range.max_col <= max_col:
            worksheet.merge_cells(str(merged_range))

    for row_idx in range(1, max_header_row + 1):
        worksheet.row_dimensions[row_idx].height = reference_worksheet.row_dimensions[row_idx].height
        for col_idx in range(1, max_col + 1):
            source = reference_worksheet.cell(row_idx, col_idx)
            target = worksheet.cell(row_idx, col_idx)
            if type(target).__name__ == "MergedCell":
                continue
            copy_cell_format(source, target)
            target.value = source.value

    # Merged section anchors on the first rows can lose their style when the
    # target sheet has been rebuilt or had columns inserted. Reapply them from
    # the checked reference explicitly.
    for merged_range in reference_worksheet.merged_cells.ranges:
        if merged_range.max_row > max_header_row or merged_range.max_col > max_col:
            continue
        source = reference_worksheet.cell(merged_range.min_row, merged_range.min_col)
        target = worksheet.cell(merged_range.min_row, merged_range.min_col)
        copy_cell_format(source, target)
        target.value = source.value


def copy_detail_reference_first_data_style(worksheet: Any, reference_worksheet: Any) -> None:
    """Make the row immediately after the header match the checked reference."""
    header_row, _headers = find_header_row(worksheet)
    reference_header_row, _reference_headers = find_header_row(reference_worksheet)
    if worksheet.max_row <= header_row or reference_worksheet.max_row <= reference_header_row:
        return
    target_row = header_row + 1
    source_row = reference_header_row + 1
    max_col = min(worksheet.max_column, reference_worksheet.max_column)
    worksheet.row_dimensions[target_row].height = reference_worksheet.row_dimensions[source_row].height
    for col_idx in range(1, max_col + 1):
        copy_cell_format(reference_worksheet.cell(source_row, col_idx), worksheet.cell(target_row, col_idx))


def copy_detail_reference_data_format(worksheet: Any, reference_worksheet: Any, kind: str) -> None:
    header_row, headers = find_header_row(worksheet)
    col_map = build_header_col_map(headers)
    reference_samples = find_detail_reference_style_rows(reference_worksheet, kind)
    if not reference_samples:
        return

    max_col = min(worksheet.max_column, reference_worksheet.max_column)
    code_col = find_col_by_header(col_map, "代码")
    marker_header = "中诚信补充计算现金流年份" if kind == "property" else "年份"
    marker_col = find_col_by_header(col_map, marker_header)
    residual_year_col = find_col_by_header(col_map, "残值年度")
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        source_row = reference_samples.get("data")
        marker_value = str(worksheet.cell(row_idx, marker_col).value or "").strip() if marker_col else ""
        residual_value = str(worksheet.cell(row_idx, residual_base_col).value or "").strip() if residual_base_col else ""
        has_code = bool(code_col and normalize_code(worksheet.cell(row_idx, code_col).value))
        if marker_value == "期初" and reference_samples.get("initial"):
            source_row = reference_samples["initial"]
        elif kind == "concession" and "期末" in marker_value and reference_samples.get("terminal"):
            source_row = reference_samples["terminal"]
        elif kind == "property" and residual_value == "残值" and reference_samples.get("residual"):
            source_row = reference_samples["residual"]
        elif (
            kind == "property"
            and not has_code
            and residual_year_col
            and worksheet.cell(row_idx, residual_year_col).value not in (None, "")
            and reference_samples.get("residual_helper")
        ):
            source_row = reference_samples["residual_helper"]
        elif has_code:
            source_row = reference_samples.get("data")
        if not source_row:
            continue
        worksheet.row_dimensions[row_idx].height = reference_worksheet.row_dimensions[source_row].height
        for col_idx in range(1, max_col + 1):
            copy_cell_format(reference_worksheet.cell(source_row, col_idx), worksheet.cell(row_idx, col_idx))


def find_detail_reference_style_rows(reference_worksheet: Any, kind: str) -> dict[str, int]:
    header_row, headers = find_header_row(reference_worksheet)
    col_map = build_header_col_map(headers)
    code_col = find_col_by_header(col_map, "代码")
    marker_header = "中诚信补充计算现金流年份" if kind == "property" else "年份"
    marker_col = find_col_by_header(col_map, marker_header)
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")
    samples: dict[str, int] = {}
    data_offset = 0

    for row_idx in range(header_row + 1, reference_worksheet.max_row + 1):
        marker_value = str(reference_worksheet.cell(row_idx, marker_col).value or "").strip() if marker_col else ""
        residual_value = (
            str(reference_worksheet.cell(row_idx, residual_base_col).value or "").strip() if residual_base_col else ""
        )
        has_code = bool(code_col and normalize_code(reference_worksheet.cell(row_idx, code_col).value))
        if marker_value == "期初":
            samples.setdefault("initial", row_idx)
            data_offset = 0
        elif kind == "concession" and "期末" in marker_value:
            samples.setdefault("terminal", row_idx)
        elif kind == "property" and residual_value == "残值":
            samples.setdefault("residual", row_idx)
            if row_idx + 1 <= reference_worksheet.max_row:
                samples.setdefault("residual_helper", row_idx + 1)
        elif has_code:
            data_offset += 1
            samples.setdefault("data", row_idx)
            if data_offset <= 30:
                samples.setdefault(f"data_{data_offset}", row_idx)
        if (
            "data" in samples
            and "initial" in samples
            and (kind != "property" or "residual_helper" in samples)
            and (kind != "concession" or "terminal" in samples)
        ):
            break
    return samples


def copy_cell_format(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.protection = copy(source.protection)


def clone_font_with_name(font: Any, name: str) -> Font:
    cloned = copy(font)
    cloned.name = name
    return cloned


def normalize_detail_output_styles(
    worksheet: Any,
    col_map: dict[str, int],
    kind: str,
    header_row: int,
    preserve_existing_fonts: bool = False,
) -> None:
    """Apply final annual detail style policy after template/reference copying."""
    detail_font_name = CONCESSION_DETAIL_FONT_NAME if kind == "concession" else PROPERTY_DETAIL_FONT_NAME
    terminal_marker_col = find_col_by_header(col_map, "年份") if kind == "concession" else None
    for row_idx in range(1, worksheet.max_row + 1):
        is_data_row = row_idx > header_row
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_idx, col_idx)
            if type(cell).__name__ == "MergedCell":
                continue
            if not preserve_existing_fonts:
                cell.font = clone_font_with_name(cell.font, detail_font_name)
            if not is_data_row:
                if kind == "concession" and row_idx == header_row and is_yellow_fill(cell.fill):
                    clear_cell_fill(cell)
                continue
            if kind == "property":
                clear_cell_fill(cell)
            elif kind == "concession":
                clear_cell_fill(cell)
                if terminal_marker_col == col_idx and "期末" in str(cell.value or ""):
                    cell.fill = copy(DETAIL_TERMINAL_FILL)
    if kind == "concession":
        normalize_detail_data_borders(worksheet, col_map, header_row, kind)


def ensure_detail_visible_format_area(
    worksheet: Any,
    col_map: dict[str, int],
    header_row: int,
    min_total_rows: int = DETAIL_MIN_FORMATTED_ROWS,
) -> None:
    """Keep a visible blank review area formatted like the detail table."""
    target_row = max(worksheet.max_row, min_total_rows)
    if worksheet.max_row >= target_row:
        return
    style_row = find_detail_normal_data_style_row(worksheet, col_map, header_row) or header_row + 1
    style_row = min(style_row, worksheet.max_row)
    for row_idx in range(worksheet.max_row + 1, target_row + 1):
        worksheet.row_dimensions[row_idx].height = worksheet.row_dimensions[style_row].height
        for col_idx in range(1, worksheet.max_column + 1):
            target = worksheet.cell(row_idx, col_idx)
            source = worksheet.cell(style_row, col_idx)
            copy_cell_format(source, target)
            target.value = None


def normalize_detail_data_borders(worksheet: Any, col_map: dict[str, int], header_row: int, kind: str) -> None:
    style_row = find_detail_normal_data_style_row(worksheet, col_map, header_row)
    if not style_row:
        return
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_idx, col_idx)
            if type(cell).__name__ == "MergedCell":
                continue
            style_cell = worksheet.cell(style_row, col_idx)
            cell.border = complete_cell_border(cell.border, style_cell.border)


def find_detail_normal_data_style_row(worksheet: Any, col_map: dict[str, int], header_row: int) -> int | None:
    code_col = find_col_by_header(col_map, "代码")
    year_col = find_col_by_header(col_map, "年份")
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        year_value = worksheet.cell(row_idx, year_col).value if year_col else None
        year_text = str(year_value or "")
        if "期初" in year_text or "期末" in year_text:
            continue
        has_code = bool(code_col and normalize_code(worksheet.cell(row_idx, code_col).value))
        if has_code or isinstance(year_value, int):
            return row_idx
    return None


def complete_cell_border(current: Border, reference: Border | None = None) -> Border:
    reference = reference or Border()
    return Border(
        left=complete_border_side(reference.left),
        right=complete_border_side(reference.right),
        top=complete_border_side(reference.top),
        bottom=complete_border_side(reference.bottom),
        diagonal=copy(current.diagonal),
        diagonal_direction=current.diagonal_direction,
        vertical=copy(current.vertical),
        horizontal=copy(current.horizontal),
        diagonalUp=current.diagonalUp,
        diagonalDown=current.diagonalDown,
        outline=current.outline,
        start=copy(current.start),
        end=copy(current.end),
    )


def complete_border_side(reference: Side | None = None) -> Side:
    if reference is not None and getattr(reference, "style", None):
        return copy(reference)
    return copy(DETAIL_FALLBACK_BORDER_SIDE)


def clear_cell_fill(cell: Any) -> None:
    cell.fill = PatternFill(fill_type=None)


def clear_yellow_fills(worksheet: Any) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            if type(cell).__name__ == "MergedCell":
                continue
            if is_yellow_fill(cell.fill):
                clear_cell_fill(cell)


def is_yellow_fill(fill: Any) -> bool:
    if not fill or not getattr(fill, "fill_type", None):
        return False
    for color_attr in ("fgColor", "start_color"):
        color = getattr(fill, color_attr, None)
        if not color:
            continue
        rgb = str(getattr(color, "rgb", "") or "").upper()
        if rgb and rgb[-6:] in YELLOW_FILL_RGB_SUFFIXES:
            return True
        indexed = getattr(color, "indexed", None)
        if indexed in {13, 43}:
            return True
    return False


def prepare_detail_worksheet_for_fresh_write(
    worksheet: Any,
    col_map: dict[str, int],
    rows: list[dict[str, Any]],
    header_row: int,
) -> int | None:
    if worksheet.max_row <= header_row:
        return None

    style_row = best_detail_style_row(worksheet, col_map, rows, header_row)
    template_row = header_row + 1
    if style_row != template_row:
        copy_row_style_and_formulas(
            worksheet,
            style_row,
            template_row,
            copy_static_values=False,
            copy_formulas=False,
        )

    if worksheet.max_row > template_row:
        worksheet.delete_rows(template_row + 1, worksheet.max_row - template_row)

    # Keep one hidden-ish template row for formulas/styles, but remove old values
    # so last year's rows are not treated as current-year data.
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(template_row, col_idx)
        if not is_formula(cell.value):
            cell.value = None
    return template_row


def migrate_property_detail_layout_if_needed(worksheet: Any) -> None:
    header_row, headers = find_header_row(worksheet)
    normalized_headers = [normalize_text(header) for header in headers]
    if worksheet.max_column >= 56 or len(normalized_headers) < 20:
        return
    # 2026 target property workbook has two blank helper columns after
    # 基础资产评估价值（万元）, so residual fields start from column T.
    if normalized_headers[17] != "残值年度":
        return
    worksheet.insert_cols(18, 2)
    ensure_min_column_count(worksheet, 56)
    worksheet.cell(header_row, 18).value = None
    worksheet.cell(header_row, 19).value = None
    reset_property_merged_headers(worksheet)


def reset_property_merged_headers(worksheet: Any) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        try:
            worksheet.unmerge_cells(str(merged_range))
        except KeyError:
            try:
                worksheet.merged_cells.ranges.remove(merged_range)
            except ValueError:
                pass
    for cell_range in ("A1:H2", "I1:AH2", "AI1:AJ2", "AK1:AR2", "AS1:AU2", "AV1:BA2", "BB1:BD1"):
        worksheet.merge_cells(cell_range)
    labels = {
        1: "基础信息",
        9: "基础资产预测现金流",
        35: "项目公司调整项",
        37: "资管计划及基金调整项",
        45: "经调基金预测现金流",
        48: "ccxIRR（历史市值）",
        54: "ccxIRR（最新市值）",
    }
    for col_idx, value in labels.items():
        cell = worksheet.cell(1, col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")


def ensure_min_column_count(worksheet: Any, min_columns: int) -> None:
    if worksheet.max_column >= min_columns:
        try:
            worksheet.cell(1, min_columns).value = worksheet.cell(1, min_columns).value or ""
        except AttributeError:
            pass
        return
    for col_idx in range(worksheet.max_column + 1, min_columns + 1):
        source_col = max(1, col_idx - 1)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = worksheet.column_dimensions[get_column_letter(source_col)].width
        for row_idx in range(1, min(worksheet.max_row, 3) + 1):
            source = worksheet.cell(row_idx, source_col)
            target = worksheet.cell(row_idx, col_idx)
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.number_format = source.number_format
                target.protection = copy(source.protection)
        try:
            worksheet.cell(1, col_idx).value = worksheet.cell(1, col_idx).value or ""
        except AttributeError:
            pass


def order_detail_group_keys(
    worksheet: Any,
    col_map: dict[str, int],
    group_keys: Iterable[tuple[str, str]],
    header_row: int,
) -> list[tuple[str, str]]:
    pending = list(group_keys)
    ordered: list[tuple[str, str]] = []
    code_col = col_map.get("代码")
    project_col = col_map.get("项目名称")
    if code_col and project_col:
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            row_code = normalize_code(worksheet.cell(row_idx, code_col).value)
            row_project = normalize_project(worksheet.cell(row_idx, project_col).value)
            if not row_code or not row_project:
                continue
            for key in list(pending):
                code, project = key
                if code == row_code and projects_match(row_project, project):
                    ordered.append(key)
                    pending.remove(key)
                    break
    ordered.extend(pending)
    return ordered


def is_whole_project_key(group_key: tuple[str, str]) -> bool:
    return normalize_project(group_key[1]) == normalize_project("项目整体")


def should_insert_initial_row_before_group(
    kind: str,
    group_key: tuple[str, str],
    grouped_keys_by_code: dict[str, list[tuple[str, str]]],
    inserted_initial_codes: set[str],
) -> bool:
    code, _project = group_key
    if code in inserted_initial_codes:
        return False
    group_keys = grouped_keys_by_code.get(code, [])
    has_whole_project = any(is_whole_project_key(item) for item in group_keys)
    if has_whole_project:
        return is_whole_project_key(group_key)
    return kind == "concession"


def insert_detail_initial_row(
    worksheet: Any,
    col_map: dict[str, int],
    style_row: int,
    group_rows: list[dict[str, Any]],
    kind: str,
) -> int:
    row_idx = append_detail_row(worksheet, style_row)
    first_year = min((row.get("年份") for row in group_rows if isinstance(row.get("年份"), int)), default=None)
    initial_date = date(first_year - 1, 12, 31) if isinstance(first_year, int) else None
    if kind == "property":
        set_by_header(worksheet, col_map, row_idx, "中诚信补充计算现金流年份", "期初")
        set_by_header(worksheet, col_map, row_idx, "现金流折现日期", initial_date)
    else:
        set_by_header(worksheet, col_map, row_idx, "年份", "期初")
        set_by_header(worksheet, col_map, row_idx, "现金流折现日期", initial_date)
        ensure_concession_interest_vat_rate_cell(worksheet, col_map, row_idx)
    initial_net_asset = adjusted_initial_net_asset(group_rows, first_year)
    if initial_net_asset is not None:
        set_by_header(worksheet, col_map, row_idx, "基金净资产（万元）", initial_net_asset)
    return row_idx


def ensure_concession_interest_vat_rate_cell(worksheet: Any, col_map: dict[str, int], row_idx: int) -> None:
    vat_col = find_col_by_header(col_map, "利息增值税3.26%")
    if not vat_col:
        return
    for existing_row in range(1, worksheet.max_row + 1):
        if worksheet.cell(existing_row, vat_col).value == CONCESSION_INTEREST_VAT_RATE:
            return
    cell = worksheet.cell(row_idx, vat_col)
    cell.value = CONCESSION_INTEREST_VAT_RATE
    cell.number_format = "0.00%"


def adjusted_initial_net_asset(group_rows: list[dict[str, Any]], first_year: int | None) -> float | None:
    net_asset = first_value(group_rows, "基金净资产（万元）")
    if net_asset in (None, "") or parse_number(net_asset) <= 0:
        return None
    adjusted = parse_number(net_asset)
    if isinstance(first_year, int):
        report_year = date_year(first_value(group_rows, "评估基准日")) or parse_int(first_value(group_rows, "报告期"))
        depreciation = first_value(group_rows, "折旧及摊销（万元）")
        if report_year and depreciation not in (None, ""):
            skipped_years = max(0, (first_year - 1) - report_year)
            adjusted -= parse_number(depreciation) * skipped_years
    return adjusted if adjusted > 0 else None


def build_header_col_map(headers: list[Any]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for col_idx, header in enumerate(headers, 1):
        key = normalize_text(header)
        if key and key not in col_map:
            col_map[key] = col_idx
    return col_map


def copy_workbook_without_updates(source_path: Path, output_dir: Path, label: str, review_items: list[dict[str, Any]]) -> Path:
    output_path = output_dir / f"{source_path.stem}_自动更新{source_path.suffix}"
    workbook = load_workbook(source_path)
    for worksheet in workbook.worksheets:
        finalize_worksheet_view(worksheet)
    workbook.save(output_path)
    workbook.close()
    review_items.append(
        {
            "类型": "无可写数据",
            "对象": output_path.name,
            "说明": f"已复制{label}表并保留原有格式；未读取到对应标准化现金流数据，需人工补充或上传标准导入表。",
        }
    )
    return output_path


def update_detail_group(
    worksheet: Any,
    col_map: dict[str, int],
    header_row: int,
    group_key: tuple[str, str],
    group_rows: list[dict[str, Any]],
    target_rows: list[int],
    kind: str,
    review_items: list[dict[str, Any]],
) -> int:
    updated = 0
    rows_by_year = {row.get("年份"): row for row in group_rows if isinstance(row.get("年份"), int)}
    style_row = target_rows[0] if target_rows else best_detail_style_row(worksheet, col_map, group_rows, header_row)

    for year, item in sorted(rows_by_year.items()):
        row_idx = find_detail_row_for_year(worksheet, col_map, target_rows, year, kind)
        if row_idx is None:
            insert_before = find_terminal_row(worksheet, col_map, target_rows) if kind == "concession" else None
            row_idx = append_detail_row(worksheet, style_row, insert_before=insert_before)
            if insert_before:
                target_rows[:] = [row + 1 if row >= insert_before else row for row in target_rows]
            target_rows.append(row_idx)
        updated += fill_detail_row(worksheet, col_map, row_idx, item, kind, disclosed=True)

    if kind == "property":
        updated += fill_property_growth_rows(worksheet, col_map, group_rows, target_rows, style_row, review_items)
        if is_whole_project_key(group_key):
            updated += append_property_residual_block_if_needed(worksheet, col_map, group_rows, target_rows, style_row)
    elif kind == "concession":
        updated += fill_concession_terminal_row(worksheet, col_map, group_rows, target_rows, style_row)

    return updated


def fill_detail_row(worksheet: Any, col_map: dict[str, int], row_idx: int, row: dict[str, Any], kind: str, disclosed: bool) -> int:
    updated = 0
    common = {
        "代码": row.get("REITs代码"),
        "名称": row.get("REITs名称"),
        "基础资产类型": row.get("基础设施项目类型"),
        "底层资产性质": row.get("底层资产性质"),
        "项目名称": row.get("项目名称"),
        "公告日期": row.get("公告日期"),
        "基金上市日": row.get("上市日期"),
        "基金到期日": row.get("到期日"),
    }
    for header, value in common.items():
        updated += set_by_header(worksheet, col_map, row_idx, header, value)

    if kind == "property":
        if disclosed:
            updated += set_by_header(worksheet, col_map, row_idx, "年份", row.get("年份"))
            updated += set_by_header(worksheet, col_map, row_idx, "预测现金流金额", row.get("预测现金流金额（万元）"))
            updated += set_by_header(worksheet, col_map, row_idx, "中诚信补充计算现金流年份", row.get("年份"))
            updated += set_by_header(worksheet, col_map, row_idx, "现金流折现日期", row.get("现金流折现日期") or midyear_date(row.get("年份")))
            updated += set_by_header(worksheet, col_map, row_idx, "中诚信补充计算现金流", row.get("预测现金流金额（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "基础资产评估价值（万元）", row.get("基础资产评估价值（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "折现率", row.get("折现率"))
        updated += set_by_header(worksheet, col_map, row_idx, "报告期", row.get("报告期"))
        updated += set_by_header(worksheet, col_map, row_idx, "评估基准日", row.get("评估基准日"))
        updated += set_by_header(worksheet, col_map, row_idx, "估值基准日期", row.get("评估基准日"))
        updated += set_by_header(worksheet, col_map, row_idx, "营业收入（万元）", row.get("营业收入（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "EBITDA（万元）", row.get("EBITDA（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "运营资本披露值（万元）", row.get("运营资本披露值（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "资本支出披露值（万元）", row.get("资本支出披露值（万元）"))
    else:
        if disclosed:
            updated += set_by_header(worksheet, col_map, row_idx, "年份", row.get("年份"))
            updated += set_by_header(worksheet, col_map, row_idx, "现金流折现日期", row.get("现金流折现日期") or midyear_date(row.get("年份")))
            updated += set_by_header(worksheet, col_map, row_idx, "验证-评估报告披露预测现金流金额（万元）", row.get("预测现金流金额（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "折现率", row.get("折现率"))
        updated += set_by_header(worksheet, col_map, row_idx, "报告期", row.get("报告期"))
        updated += set_by_header(worksheet, col_map, row_idx, "估值基准日期", row.get("评估基准日"))
        updated += set_by_header(worksheet, col_map, row_idx, "评估基准日", row.get("评估基准日"))
        updated += set_by_header(worksheet, col_map, row_idx, "评估报告评估价值（万元）", row.get("基础资产评估价值（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "基础资产评估价值（万元）", row.get("基础资产评估价值（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "营业收入（万元）", row.get("营业收入（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "EBITDA（万元）", row.get("EBITDA（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "运营资本披露值（万元）", row.get("运营资本披露值（万元）"))
        updated += set_by_header(worksheet, col_map, row_idx, "资本支出披露值（万元）", row.get("资本支出披露值（万元）"))

    residual_cashflow = row.get("考虑残值现金流")
    if kind == "property" and disclosed and residual_cashflow in (None, ""):
        residual_cashflow = row.get("预测现金流金额（万元）")
    updated += set_by_header(worksheet, col_map, row_idx, "残值年度", row.get("残值年度"))
    updated += set_by_header(worksheet, col_map, row_idx, "残值基础数据预测", row.get("残值基础数据预测"))
    updated += set_by_header(worksheet, col_map, row_idx, "考虑残值现金流", residual_cashflow)
    updated += set_by_header(worksheet, col_map, row_idx, "借款本金（万元）", row.get("借款本金（万元）"))
    updated += set_by_header(worksheet, col_map, row_idx, "借款利息（万元）", row.get("借款利息（万元）"))
    net_asset = row.get("基金净资产（万元）")
    if net_asset not in (None, ""):
        updated += set_by_header(worksheet, col_map, row_idx, "基金净资产（万元）", net_asset if parse_number(net_asset) > 0 else None, allow_blank=True)
    updated += set_by_header(worksheet, col_map, row_idx, "折旧及摊销（万元）", row.get("折旧及摊销（万元）"))
    updated += set_by_header(worksheet, col_map, row_idx, "固定管理费率(%)", row.get("固定管理费率(%)"))
    updated += set_by_header(worksheet, col_map, row_idx, "托管费率(%)", row.get("托管费率(%)"))
    updated += set_by_header(worksheet, col_map, row_idx, "调整浮动管理费", row.get("调整浮动管理费"))
    apply_detail_formulas_for_row(worksheet, col_map, row_idx, kind)
    return updated


def fill_property_growth_rows(
    worksheet: Any,
    col_map: dict[str, int],
    group_rows: list[dict[str, Any]],
    target_rows: list[int],
    style_row: int,
    review_items: list[dict[str, Any]],
) -> int:
    growth_start = first_value(group_rows, "增长率预测起始年度")
    growth_rate = first_value(group_rows, "预测现金流增长率")
    if not isinstance(growth_start, int) or growth_rate in (None, ""):
        return 0

    terminal_year = date_year(first_value(group_rows, "到期日")) or max_year_in_rows(group_rows)
    if not terminal_year or terminal_year < growth_start:
        return 0

    updated = 0
    explicit_year_rows = {
        row.get("年份"): row
        for row in group_rows
        if isinstance(row.get("年份"), int) and row.get("预测现金流金额（万元）") not in (None, "")
    }
    previous_row = find_detail_row_for_year(worksheet, col_map, target_rows, growth_start - 1, "property")
    for year in range(growth_start, terminal_year + 1):
        if year in explicit_year_rows:
            previous_row = find_detail_row_for_year(worksheet, col_map, target_rows, year, "property") or previous_row
            continue
        row_idx = find_detail_row_for_year(worksheet, col_map, target_rows, year, "property")
        if row_idx is None:
            row_idx = append_detail_row(worksheet, style_row)
            target_rows.append(row_idx)
        row = {**group_rows[0], "年份": None}
        updated += fill_detail_row(worksheet, col_map, row_idx, row, "property", disclosed=False)
        updated += set_by_header(worksheet, col_map, row_idx, "年份", None, allow_blank=True)
        updated += set_by_header(worksheet, col_map, row_idx, "预测现金流金额", None, allow_blank=True)
        updated += set_by_header(worksheet, col_map, row_idx, "中诚信补充计算现金流年份", year)
        updated += set_by_header(worksheet, col_map, row_idx, "现金流折现日期", midyear_date(year))
        if year == growth_start:
            updated += set_by_header(worksheet, col_map, row_idx, "增长率预测起始年度", growth_start)
            updated += set_by_header(worksheet, col_map, row_idx, "预测现金流增长率", growth_rate)
        else:
            updated += set_by_header(worksheet, col_map, row_idx, "增长率预测起始年度", None, allow_blank=True)
            updated += set_by_header(worksheet, col_map, row_idx, "预测现金流增长率", None, allow_blank=True)
        if previous_row:
            updated += set_formula_by_header(worksheet, col_map, row_idx, "中诚信补充计算现金流", previous_row, growth_rate)
            updated += set_formula_by_header(worksheet, col_map, row_idx, "考虑残值现金流", previous_row, growth_rate)
        previous_row = row_idx

    review_items.append(
        {
            "类型": "公式预测",
            "对象": f"{first_value(group_rows, 'REITs代码')} / {first_value(group_rows, '项目名称')}",
            "说明": f"已从 {growth_start} 年起按增长率 {growth_rate} 延展至 {terminal_year} 年；请复核公式区间。",
        }
    )
    return updated


def residual_parameter(group_rows: list[dict[str, Any]], field_name: str) -> Any:
    for row in group_rows:
        parameters = row.get(RESIDUAL_PARAMETER_KEY)
        if isinstance(parameters, dict) and parameters.get(field_name) not in (None, ""):
            return parameters[field_name]
    return None


def append_property_residual_block_if_needed(
    worksheet: Any,
    col_map: dict[str, int],
    group_rows: list[dict[str, Any]],
    target_rows: list[int],
    style_row: int,
) -> int:
    cashflow_context = last_property_cashflow_context(worksheet, col_map, target_rows)
    if not cashflow_context:
        return 0
    _last_cashflow_row, last_year, last_actual_year = cashflow_context

    residual_year_value = residual_parameter(group_rows, "残值年度")
    residual_start_year = date_year(residual_year_value) or parse_int(residual_year_value) or last_year + 1
    residual_base_seed = residual_parameter(group_rows, "残值基础数据预测")
    explicit_residual_cashflow = residual_parameter(group_rows, "考虑残值现金流")

    residual_row = append_detail_row(worksheet, style_row)
    item = {**group_rows[0], "年份": None, "预测现金流金额（万元）": None}
    updated = fill_detail_row(worksheet, col_map, residual_row, item, "property", disclosed=False)
    updated += set_by_header(worksheet, col_map, residual_row, "年份", None, allow_blank=True)
    updated += set_by_header(worksheet, col_map, residual_row, "预测现金流金额", None, allow_blank=True)
    updated += set_by_header(worksheet, col_map, residual_row, "中诚信补充计算现金流年份", last_year)
    updated += set_by_header(worksheet, col_map, residual_row, "现金流折现日期", midyear_date(last_year + 1))
    updated += set_by_header(worksheet, col_map, residual_row, "中诚信补充计算现金流", None, allow_blank=True)
    updated += set_by_header(worksheet, col_map, residual_row, "残值年度", None, allow_blank=True)
    updated += set_by_header(worksheet, col_map, residual_row, "残值基础数据预测", "残值")
    if explicit_residual_cashflow not in (None, ""):
        updated += set_by_header(worksheet, col_map, residual_row, "考虑残值现金流", explicit_residual_cashflow)
    else:
        updated += set_by_header(worksheet, col_map, residual_row, "考虑残值现金流", None, allow_blank=True)
    target_rows.append(residual_row)
    normalize_property_residual_row_formats(worksheet, col_map, residual_row, helper=False)

    for offset in range(1, 11):
        helper_row = append_detail_row(worksheet, style_row)
        updated += set_by_header(worksheet, col_map, helper_row, "残值年度", midyear_date(residual_start_year + offset - 1))
        helper_value = residual_base_seed if offset == 1 and residual_base_seed not in (None, "") else None
        updated += set_by_header(worksheet, col_map, helper_row, "残值基础数据预测", helper_value, allow_blank=True)
        normalize_property_residual_row_formats(worksheet, col_map, helper_row, helper=True)
    return updated


def normalize_property_residual_row_formats(
    worksheet: Any,
    col_map: dict[str, int],
    row_idx: int,
    helper: bool,
) -> None:
    residual_year_col = find_col_by_header(col_map, "残值年度")
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")
    residual_cashflow_col = find_col_by_header(col_map, "考虑残值现金流")
    year_col = find_col_by_header(col_map, "年份")

    if year_col:
        worksheet.cell(row_idx, year_col).number_format = "General"
    if residual_year_col:
        worksheet.cell(row_idx, residual_year_col).number_format = "yyyy-mm-dd" if helper else "General"
    if residual_base_col:
        worksheet.cell(row_idx, residual_base_col).number_format = "0.00" if helper else "General"
    if residual_cashflow_col:
        worksheet.cell(row_idx, residual_cashflow_col).number_format = "0.00"


def normalize_property_group_header_styles(worksheet: Any) -> None:
    anchors = [1, 9, 35, 37, 45, 48, 54]
    source = None
    for col_idx in (35, 9, 1):
        cell = worksheet.cell(1, col_idx)
        if cell.has_style and getattr(cell.fill, "fill_type", None):
            source = cell
            break
    if source is None:
        return
    for col_idx in anchors:
        target = worksheet.cell(1, col_idx)
        if type(target).__name__ == "MergedCell":
            continue
        copy_cell_format(source, target)
        target.alignment = Alignment(horizontal="center", vertical="center")


def last_property_cashflow_context(
    worksheet: Any,
    col_map: dict[str, int],
    target_rows: list[int],
) -> tuple[int, int, int | None] | None:
    year_col = find_col_by_header(col_map, "年份")
    ccx_year_col = find_col_by_header(col_map, "中诚信补充计算现金流年份")
    cashflow_col = find_col_by_header(col_map, "预测现金流金额")
    ccx_cashflow_col = find_col_by_header(col_map, "中诚信补充计算现金流")
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")
    best: tuple[int, int, int | None] | None = None

    for row_idx in sorted(set(target_rows)):
        if residual_base_col and str(worksheet.cell(row_idx, residual_base_col).value or "").strip() == "残值":
            continue
        actual_year = worksheet.cell(row_idx, year_col).value if year_col else None
        ccx_year = worksheet.cell(row_idx, ccx_year_col).value if ccx_year_col else None
        row_year = actual_year if isinstance(actual_year, int) else ccx_year
        if not isinstance(row_year, int):
            continue
        has_cashflow = any(
            col and worksheet.cell(row_idx, col).value not in (None, "")
            for col in (cashflow_col, ccx_cashflow_col)
        )
        if has_cashflow:
            best = (row_idx, row_year, actual_year if isinstance(actual_year, int) else None)
    return best


def fill_concession_terminal_row(
    worksheet: Any,
    col_map: dict[str, int],
    group_rows: list[dict[str, Any]],
    target_rows: list[int],
    style_row: int,
) -> int:
    terminal = first_value(group_rows, "经营期末")
    if terminal in (None, ""):
        return 0
    row_idx = find_terminal_row(worksheet, col_map, target_rows)
    if row_idx is None:
        row_idx = append_detail_row(worksheet, style_row)
        target_rows.append(row_idx)
    terminal_year = max_year_in_rows(group_rows) or date_year(first_value(group_rows, "到期日"))
    item = {**group_rows[0], "年份": "期末回收", "预测现金流金额（万元）": terminal}
    updated = fill_detail_row(worksheet, col_map, row_idx, item, "concession", disclosed=False)
    updated += set_by_header(worksheet, col_map, row_idx, "年份", "期末回收")
    updated += set_by_header(worksheet, col_map, row_idx, "现金流折现日期", midyear_date(terminal_year))
    updated += set_by_header(worksheet, col_map, row_idx, "验证-评估报告披露预测现金流金额（万元）", terminal)
    for header in (
        "基金净资产（万元）",
        "折旧及摊销（万元）",
        "固定管理费率(%)",
        "托管费率(%)",
        "管理费（万元）",
        "托管费",
        "调整浮动管理费",
    ):
        updated += set_by_header(worksheet, col_map, row_idx, header, None, allow_blank=True)
    return updated


def group_standard_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        code = normalize_code(row.get("REITs代码"))
        if not code:
            continue
        project = normalize_project(row.get("项目名称") or "项目整体")
        grouped.setdefault((code, project), []).append(row)
    for key in grouped:
        grouped[key].sort(key=lambda item: (item.get("年份") if isinstance(item.get("年份"), int) else 9999))
    return grouped


def find_detail_rows(worksheet: Any, col_map: dict[str, int], group_key: tuple[str, str]) -> list[int]:
    code, project_key = group_key
    code_col = col_map.get("代码")
    project_col = col_map.get("项目名称")
    if not code_col or not project_col:
        return []
    exact_rows: list[int] = []
    fuzzy_rows: list[int] = []
    for row_idx in range(1, worksheet.max_row + 1):
        row_code = normalize_code(worksheet.cell(row_idx, code_col).value)
        if row_code != code:
            continue
        row_project = normalize_project(worksheet.cell(row_idx, project_col).value)
        if row_project == project_key:
            exact_rows.append(row_idx)
        elif projects_match(row_project, project_key):
            fuzzy_rows.append(row_idx)
    return exact_rows or fuzzy_rows


def find_detail_row_for_year(worksheet: Any, col_map: dict[str, int], target_rows: list[int], year: int, kind: str) -> int | None:
    year_col = col_map.get("年份")
    ccx_year_col = col_map.get("中诚信补充计算现金流年份")
    for row_idx in target_rows:
        if year_col and worksheet.cell(row_idx, year_col).value == year:
            return row_idx
        if kind == "property" and ccx_year_col and worksheet.cell(row_idx, ccx_year_col).value == year:
            return row_idx
    return None


def find_terminal_row(worksheet: Any, col_map: dict[str, int], target_rows: list[int]) -> int | None:
    year_col = col_map.get("年份")
    if not year_col:
        return None
    for row_idx in target_rows:
        value = worksheet.cell(row_idx, year_col).value
        if isinstance(value, str) and "期末" in value:
            return row_idx
    return None


def best_detail_style_row(worksheet: Any, col_map: dict[str, int], group_rows: list[dict[str, Any]], header_row: int) -> int:
    code = normalize_code(first_value(group_rows, "REITs代码"))
    asset_nature = normalize_text(first_value(group_rows, "底层资产性质"))
    code_col = col_map.get("代码")
    nature_col = col_map.get("底层资产性质")
    if code_col:
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            if normalize_code(worksheet.cell(row_idx, code_col).value) == code:
                return row_idx
    if nature_col:
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            if asset_nature and asset_nature in normalize_text(worksheet.cell(row_idx, nature_col).value):
                return row_idx
    return header_row + 1


def append_detail_row(worksheet: Any, style_row: int, insert_before: int | None = None) -> int:
    if insert_before:
        worksheet.insert_rows(insert_before)
        row_idx = insert_before
        style_row = style_row + 1 if style_row >= insert_before else style_row
    else:
        row_idx = worksheet.max_row + 1
    copy_row_style_and_formulas(worksheet, style_row, row_idx, copy_static_values=False, copy_formulas=False)
    return row_idx


def copy_row_style_and_formulas(
    worksheet: Any,
    source_row: int,
    target_row: int,
    copy_static_values: bool = True,
    copy_formulas: bool = True,
) -> None:
    if source_row < 1 or target_row < 1:
        return
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    max_col = getattr(worksheet, "_reit_max_column_cache", None)
    if not max_col:
        max_col = worksheet.max_column
        try:
            worksheet._reit_max_column_cache = max_col
        except Exception:
            pass
    for col_idx in range(1, max_col + 1):
        source = worksheet.cell(source_row, col_idx)
        target = worksheet.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        if is_formula(source.value) and copy_formulas:
            target.value = translate_formula(source.value, source.coordinate, target.coordinate)
        elif copy_static_values:
            target.value = source.value
        else:
            target.value = None


def translate_formula(value: Any, origin: str, target: str) -> Any:
    if not is_formula(value):
        return value
    try:
        return Translator(value, origin=origin).translate_formula(target)
    except Exception:
        return value


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def set_by_header(
    worksheet: Any,
    col_map: dict[str, int],
    row_idx: int,
    header: str,
    value: Any,
    allow_blank: bool = False,
) -> int:
    col_idx = find_col_by_header(col_map, header)
    if not col_idx:
        return 0
    if value in (None, "") and not allow_blank:
        return 0
    worksheet.cell(row_idx, col_idx).value = value
    return 1


def set_formula_by_header(
    worksheet: Any,
    col_map: dict[str, int],
    row_idx: int,
    header: str,
    previous_row: int,
    growth_rate: Any,
) -> int:
    col_idx = find_col_by_header(col_map, header)
    if not col_idx:
        return 0
    col_letter = get_column_letter(col_idx)
    rate = parse_number(growth_rate)
    worksheet.cell(row_idx, col_idx).value = f"={col_letter}{previous_row}*(1+{rate})"
    return 1


def find_col_by_header(col_map: dict[str, int], header: str) -> int | None:
    for candidate in header_candidates(header):
        col_idx = col_map.get(normalize_text(candidate))
        if col_idx:
            return col_idx
    return None


def find_all_cols_by_header(worksheet: Any, header_row: int, header: str) -> list[int]:
    cache_key = (header_row, header)
    cache = getattr(worksheet, "_reit_all_header_cols_cache", None)
    if cache is None:
        cache = {}
        try:
            worksheet._reit_all_header_cols_cache = cache
        except Exception:
            cache = {}
    if cache_key in cache:
        return list(cache[cache_key])

    candidates = {normalize_text(candidate) for candidate in header_candidates(header)}
    cols: list[int] = []
    max_col = getattr(worksheet, "_reit_max_column_cache", None)
    if not max_col:
        max_col = worksheet.max_column
        try:
            worksheet._reit_max_column_cache = max_col
        except Exception:
            pass
    for col_idx in range(1, max_col + 1):
        if normalize_text(worksheet.cell(header_row, col_idx).value) in candidates:
            cols.append(col_idx)
    cache[cache_key] = tuple(cols)
    return cols


def set_formula_cell(worksheet: Any, row_idx: int, col_idx: int | None, formula: str | None) -> int:
    if not col_idx or not formula:
        return 0
    worksheet.cell(row_idx, col_idx).value = formula
    return 1


def apply_detail_formulas_for_row(worksheet: Any, col_map: dict[str, int], row_idx: int, kind: str) -> int:
    updated = 0
    header_row, _headers = find_header_row(worksheet)

    def col(header: str) -> int | None:
        return find_col_by_header(col_map, header)

    def letter(header: str) -> str | None:
        col_idx = col(header)
        return get_column_letter(col_idx) if col_idx else None

    cashflow_header = "中诚信补充计算现金流" if kind == "property" else "验证-评估报告披露预测现金流金额（万元）"
    cashflow_col = letter(cashflow_header)
    adjusted_col = col("经调-基础资产预测现金流金额（万元）")
    operating_adjustment = letter("运营资本调整（万元）")
    capex_adjustment = letter("资本支出调整（万元）")
    reserve_fee = letter("预留管理费（万元）")
    residual = letter("残值")
    adjustment_letters = [item for item in (operating_adjustment, capex_adjustment, reserve_fee) if item]
    if cashflow_col and adjusted_col and adjustment_letters:
        formula = f"={cashflow_col}{row_idx}-SUM({adjustment_letters[0]}{row_idx}:{adjustment_letters[-1]}{row_idx})"
        if kind == "property" and residual:
            formula += f"+{residual}{row_idx}"
        updated += set_formula_cell(worksheet, row_idx, adjusted_col, formula)

    adjusted_letter = letter("经调-基础资产预测现金流金额（万元）")
    principal = letter("借款本金（万元）")
    interest = letter("借款利息（万元）")
    vat_col = col("利息增值税3.26%")
    if adjusted_letter and principal and interest and vat_col:
        vat_rate = f"${get_column_letter(vat_col)}${header_row + 1}" if kind == "concession" else "3.26%"
        updated += set_formula_cell(
            worksheet,
            row_idx,
            vat_col,
            f"=({adjusted_letter}{row_idx}-{principal}{row_idx}-{interest}{row_idx})*{vat_rate}",
        )

    net_asset = letter("基金净资产（万元）")
    net_asset_col = col("基金净资产（万元）")
    depreciation = letter("折旧及摊销（万元）")
    terminal_recovery_row = is_terminal_recovery_row(worksheet, col_map, row_idx)
    if net_asset and net_asset_col and depreciation and should_roll_forward_net_asset(worksheet, col_map, row_idx, kind):
        updated += set_formula_cell(worksheet, row_idx, net_asset_col, f"={net_asset}{row_idx - 1}-{depreciation}{row_idx}")

    fixed_fee_rate = letter("固定管理费率(%)")
    custody_fee_rate = letter("托管费率(%)")
    management_fee = col("管理费（万元）")
    custody_fee = col("托管费")
    if not terminal_recovery_row and net_asset and fixed_fee_rate and management_fee:
        updated += set_formula_cell(worksheet, row_idx, management_fee, f"={net_asset}{row_idx}*{fixed_fee_rate}{row_idx}/100")
    if not terminal_recovery_row and net_asset and custody_fee_rate and custody_fee:
        updated += set_formula_cell(worksheet, row_idx, custody_fee, f"={net_asset}{row_idx}*{custody_fee_rate}{row_idx}/100")

    total_deduction = col("项目公司、ABS及基金扣减合计（万元）")
    vat = letter("利息增值税3.26%")
    management = letter("管理费（万元）")
    custody = letter("托管费")
    floating = letter("调整浮动管理费")
    if principal and interest and vat and management and custody and floating and total_deduction:
        updated += set_formula_cell(
            worksheet,
            row_idx,
            total_deduction,
            f"=SUM({principal}{row_idx}:{vat}{row_idx})+SUM({management}{row_idx}:{floating}{row_idx})",
        )

    discount_date_cols = find_all_cols_by_header(worksheet, header_row, "现金流折现日期")
    if len(discount_date_cols) >= 2:
        first = get_column_letter(discount_date_cols[0])
        updated += set_formula_cell(worksheet, row_idx, discount_date_cols[-1], f"={first}{row_idx}")

    adjusted_fund_col = col("经调基金预测现金流金额（万元）")
    total_deduction_letter = letter("项目公司、ABS及基金扣减合计（万元）")
    if adjusted_letter and total_deduction_letter and adjusted_fund_col:
        updated += set_formula_cell(worksheet, row_idx, adjusted_fund_col, f"={adjusted_letter}{row_idx}-{total_deduction_letter}{row_idx}")
    return updated


def is_terminal_recovery_row(worksheet: Any, col_map: dict[str, int], row_idx: int) -> bool:
    year_col = find_col_by_header(col_map, "年份")
    return bool(year_col and str(worksheet.cell(row_idx, year_col).value or "").strip() == "期末回收")


def should_roll_forward_net_asset(worksheet: Any, col_map: dict[str, int], row_idx: int, kind: str) -> bool:
    if row_idx <= 1:
        return False
    if is_terminal_recovery_row(worksheet, col_map, row_idx):
        return False
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")
    if residual_base_col and str(worksheet.cell(row_idx, residual_base_col).value or "").strip() == "残值":
        return False

    net_asset_col = find_col_by_header(col_map, "基金净资产（万元）")
    if not net_asset_col or worksheet.cell(row_idx - 1, net_asset_col).value in (None, ""):
        return False

    marker_header = "中诚信补充计算现金流年份" if kind == "property" else "年份"
    marker_col = find_col_by_header(col_map, marker_header)
    if marker_col and str(worksheet.cell(row_idx - 1, marker_col).value or "").strip() == "期初":
        return True

    code_col = find_col_by_header(col_map, "代码")
    if not code_col:
        return False
    current_code = normalize_code(worksheet.cell(row_idx, code_col).value)
    previous_code = normalize_code(worksheet.cell(row_idx - 1, code_col).value)
    return bool(current_code and current_code == previous_code)


def repair_detail_formulas_after_template_delete(worksheet: Any, col_map: dict[str, int], header_row: int, kind: str) -> None:
    code_col = col_map.get("代码")
    if not code_col:
        return
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        if normalize_code(worksheet.cell(row_idx, code_col).value):
            apply_detail_formulas_for_row(worksheet, col_map, row_idx, kind)
    if kind == "property":
        repair_property_growth_formulas(worksheet, col_map, header_row)


def repair_initial_row_formulas(worksheet: Any, col_map: dict[str, int], header_row: int, kind: str) -> None:
    code_col = col_map.get("代码")
    project_col = col_map.get("项目名称")
    if not code_col or not project_col:
        return
    marker_header = "中诚信补充计算现金流年份" if kind == "property" else "年份"
    marker_col = find_col_by_header(col_map, marker_header)
    if not marker_col:
        return

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        if str(worksheet.cell(row_idx, marker_col).value or "").strip() != "期初":
            continue
        first_data_row = find_next_detail_data_row(worksheet, row_idx, code_col)
        if not first_data_row:
            continue
        data_rows = contiguous_detail_rows_for_initial(worksheet, first_data_row, code_col, project_col, kind)
        if not data_rows:
            continue
        if kind == "property":
            repair_property_initial_row(worksheet, col_map, row_idx, data_rows)
        else:
            repair_concession_initial_row(worksheet, col_map, row_idx, data_rows)

    if kind == "property":
        repair_property_residual_block_formulas(worksheet, col_map, header_row)


def find_next_detail_data_row(worksheet: Any, start_row: int, code_col: int) -> int | None:
    for row_idx in range(start_row + 1, worksheet.max_row + 1):
        if normalize_code(worksheet.cell(row_idx, code_col).value):
            return row_idx
    return None


def contiguous_detail_rows_for_initial(
    worksheet: Any,
    first_data_row: int,
    code_col: int,
    project_col: int,
    kind: str,
) -> list[int]:
    code = normalize_code(worksheet.cell(first_data_row, code_col).value)
    project = normalize_project(worksheet.cell(first_data_row, project_col).value)
    rows: list[int] = []
    residual_marker_col = None
    if kind == "property":
        header_row, headers = find_header_row(worksheet)
        residual_marker_col = find_col_by_header(build_header_col_map(headers), "残值基础数据预测")
    for row_idx in range(first_data_row, worksheet.max_row + 1):
        row_code = normalize_code(worksheet.cell(row_idx, code_col).value)
        row_project = normalize_project(worksheet.cell(row_idx, project_col).value)
        if row_code != code or row_project != project:
            break
        if residual_marker_col and str(worksheet.cell(row_idx, residual_marker_col).value or "").strip() == "残值":
            break
        rows.append(row_idx)
    return rows


def repair_property_initial_row(worksheet: Any, col_map: dict[str, int], initial_row: int, data_rows: list[int]) -> None:
    cashflow_col = find_col_by_header(col_map, "中诚信补充计算现金流")
    discount_date_col = find_col_by_header(col_map, "现金流折现日期")
    irr_col = find_col_by_header(col_map, "验证中诚信现金流模拟收益率")
    valuation_col = find_col_by_header(col_map, "基础资产评估价值（万元）")
    if not data_rows:
        return
    first_row = data_rows[0]
    last_row = data_rows[-1]
    if cashflow_col and valuation_col:
        worksheet.cell(initial_row, cashflow_col).value = f"=-{get_column_letter(valuation_col)}{first_row}"
    if cashflow_col and discount_date_col and irr_col and last_row >= first_row:
        cashflow_letter = get_column_letter(cashflow_col)
        date_letter = get_column_letter(discount_date_col)
        worksheet.cell(initial_row, irr_col).value = (
            f"=XIRR({cashflow_letter}{initial_row}:{cashflow_letter}{last_row},"
            f"{date_letter}{initial_row}:{date_letter}{last_row})"
        )


def repair_concession_initial_row(worksheet: Any, col_map: dict[str, int], initial_row: int, data_rows: list[int]) -> None:
    cashflow_col = find_col_by_header(col_map, "验证-评估报告披露预测现金流金额（万元）")
    discount_date_col = find_col_by_header(col_map, "现金流折现日期")
    irr_col = find_col_by_header(col_map, "验证中诚信现金流模拟收益率")
    valuation_col = find_col_by_header(col_map, "评估报告评估价值（万元）") or find_col_by_header(col_map, "基础资产评估价值（万元）")
    if not data_rows:
        return
    first_row = data_rows[0]
    last_row = data_rows[-1]
    if cashflow_col and valuation_col:
        worksheet.cell(initial_row, cashflow_col).value = f"=-{get_column_letter(valuation_col)}{first_row}"
    if cashflow_col and discount_date_col and irr_col and last_row >= first_row:
        cashflow_letter = get_column_letter(cashflow_col)
        date_letter = get_column_letter(discount_date_col)
        worksheet.cell(initial_row, irr_col).value = (
            f"=XIRR({cashflow_letter}{initial_row}:{cashflow_letter}{last_row},"
            f"{date_letter}{initial_row}:{date_letter}{last_row})"
        )


def repair_property_residual_block_formulas(worksheet: Any, col_map: dict[str, int], header_row: int) -> None:
    code_col = col_map.get("代码")
    project_col = col_map.get("项目名称")
    year_col = find_col_by_header(col_map, "年份")
    cashflow_col = find_col_by_header(col_map, "预测现金流金额")
    residual_year_col = find_col_by_header(col_map, "残值年度")
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")
    residual_cashflow_col = find_col_by_header(col_map, "考虑残值现金流")
    residual_col = find_col_by_header(col_map, "残值")
    discount_col = find_col_by_header(col_map, "折现率")
    ccx_year_col = find_col_by_header(col_map, "中诚信补充计算现金流年份")
    ccx_cashflow_col = find_col_by_header(col_map, "中诚信补充计算现金流")
    if (
        not all([code_col, project_col, residual_year_col, residual_base_col, residual_cashflow_col])
        or not (year_col or ccx_year_col)
        or not (cashflow_col or ccx_cashflow_col)
    ):
        return

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        if str(worksheet.cell(row_idx, residual_base_col).value or "").strip() != "残值":
            continue
        code = normalize_code(worksheet.cell(row_idx, code_col).value)
        project = normalize_project(worksheet.cell(row_idx, project_col).value)
        if not code or not project:
            continue
        previous_data_row = find_previous_property_cashflow_row(
            worksheet,
            row_idx,
            code,
            project,
            code_col,
            project_col,
            year_col,
            ccx_year_col,
            residual_base_col,
            cashflow_col,
            ccx_cashflow_col,
        )
        helper_rows = find_property_residual_helper_rows(worksheet, row_idx, code_col, residual_year_col, residual_base_col)
        if not previous_data_row or not helper_rows:
            continue
        base_letter = get_column_letter(residual_base_col)
        source_cashflow_col = cashflow_col or ccx_cashflow_col
        if cashflow_col and worksheet.cell(previous_data_row, cashflow_col).value in (None, "") and ccx_cashflow_col:
            source_cashflow_col = ccx_cashflow_col
        if not source_cashflow_col:
            continue
        cashflow_letter = get_column_letter(source_cashflow_col)
        for index, helper_row in enumerate(helper_rows):
            if worksheet.cell(helper_row, residual_base_col).value not in (None, ""):
                continue
            if index == 0:
                worksheet.cell(helper_row, residual_base_col).value = f"={cashflow_letter}{previous_data_row}*0.9"
            else:
                worksheet.cell(helper_row, residual_base_col).value = f"={base_letter}{helper_rows[index - 1]}*0.9"

        discount = parse_number(worksheet.cell(row_idx, discount_col).value) if discount_col else 0
        current_residual_cashflow = worksheet.cell(row_idx, residual_cashflow_col).value
        if discount and (current_residual_cashflow in (None, "") or is_formula(current_residual_cashflow)):
            year_letter = get_column_letter(residual_year_col)
            worksheet.cell(row_idx, residual_cashflow_col).value = (
                f"=XNPV({discount},{base_letter}{helper_rows[0]}:{base_letter}{helper_rows[-1]},"
                f"{year_letter}{helper_rows[0]}:{year_letter}{helper_rows[-1]})/(1+{discount})"
            )
        if residual_col:
            worksheet.cell(row_idx, residual_col).value = f"={get_column_letter(residual_cashflow_col)}{row_idx}"


def find_previous_property_cashflow_row(
    worksheet: Any,
    residual_row: int,
    code: str,
    project: str,
    code_col: int,
    project_col: int,
    year_col: int | None,
    ccx_year_col: int | None,
    residual_marker_col: int,
    cashflow_col: int | None,
    ccx_cashflow_col: int | None,
) -> int | None:
    for row_idx in range(residual_row - 1, 0, -1):
        if normalize_code(worksheet.cell(row_idx, code_col).value) != code:
            continue
        if normalize_project(worksheet.cell(row_idx, project_col).value) != project:
            continue
        if str(worksheet.cell(row_idx, residual_marker_col).value or "").strip() == "残值":
            continue
        actual_year = worksheet.cell(row_idx, year_col).value if year_col else None
        ccx_year = worksheet.cell(row_idx, ccx_year_col).value if ccx_year_col else None
        has_year = isinstance(actual_year, int) or isinstance(ccx_year, int)
        has_cashflow = any(
            col and worksheet.cell(row_idx, col).value not in (None, "")
            for col in (cashflow_col, ccx_cashflow_col)
        )
        if has_year and has_cashflow:
            return row_idx
    return None


def find_property_residual_helper_rows(
    worksheet: Any,
    residual_row: int,
    code_col: int,
    residual_year_col: int,
    residual_base_col: int,
) -> list[int]:
    rows: list[int] = []
    for row_idx in range(residual_row + 1, worksheet.max_row + 1):
        if normalize_code(worksheet.cell(row_idx, code_col).value):
            break
        marker = worksheet.cell(row_idx, residual_year_col).value
        if marker in (None, ""):
            break
        if str(worksheet.cell(row_idx, residual_base_col).value or "").strip() == "残值":
            break
        rows.append(row_idx)
        if len(rows) >= 10:
            break
    return rows


def repair_property_growth_formulas(worksheet: Any, col_map: dict[str, int], header_row: int) -> None:
    code_col = col_map.get("代码")
    project_col = col_map.get("项目名称")
    year_col = find_col_by_header(col_map, "年份")
    ccx_year_col = find_col_by_header(col_map, "中诚信补充计算现金流年份")
    growth_rate_col = find_col_by_header(col_map, "预测现金流增长率")
    ccx_cashflow_col = find_col_by_header(col_map, "中诚信补充计算现金流")
    residual_cashflow_col = find_col_by_header(col_map, "考虑残值现金流")
    residual_base_col = find_col_by_header(col_map, "残值基础数据预测")
    if not all([code_col, project_col, ccx_year_col, ccx_cashflow_col]):
        return

    rows_by_group: dict[tuple[str, str], list[int]] = {}
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        code = normalize_code(worksheet.cell(row_idx, code_col).value)
        project = normalize_project(worksheet.cell(row_idx, project_col).value)
        if code and project:
            rows_by_group.setdefault((code, project), []).append(row_idx)

    for row_indices in rows_by_group.values():
        year_to_row: dict[int, int] = {}
        last_rate: Any = None
        for row_idx in row_indices:
            if residual_base_col and str(worksheet.cell(row_idx, residual_base_col).value or "").strip() == "残值":
                continue
            actual_year = worksheet.cell(row_idx, year_col).value if year_col else None
            ccx_year = worksheet.cell(row_idx, ccx_year_col).value if ccx_year_col else None
            row_year = actual_year if isinstance(actual_year, int) else ccx_year
            rate = worksheet.cell(row_idx, growth_rate_col).value if growth_rate_col else None
            if rate not in (None, ""):
                last_rate = rate
            if isinstance(actual_year, int):
                year_to_row[actual_year] = row_idx
                continue
            if not isinstance(ccx_year, int):
                continue
            previous_row = year_to_row.get(ccx_year - 1)
            if previous_row and last_rate not in (None, ""):
                ccx_letter = get_column_letter(ccx_cashflow_col)
                worksheet.cell(row_idx, ccx_cashflow_col).value = f"={ccx_letter}{previous_row}*(1+{parse_number(last_rate)})"
                if residual_cashflow_col:
                    residual_letter = get_column_letter(residual_cashflow_col)
                    worksheet.cell(row_idx, residual_cashflow_col).value = f"={residual_letter}{previous_row}*(1+{parse_number(last_rate)})"
            if isinstance(row_year, int):
                year_to_row[row_year] = row_idx


@lru_cache(maxsize=None)
def header_candidates(header: str) -> tuple[str, ...]:
    candidates = [header]
    normalized_header = normalize_text(header)
    for field_name, aliases in FIELD_ALIASES.items():
        normalized_aliases = {normalize_text(field_name), *[normalize_text(alias) for alias in aliases]}
        if normalized_header in normalized_aliases:
            candidates.append(field_name)
            candidates.extend(aliases)
            break
    seen: set[str] = set()
    unique: list[str] = []
    for candidate in candidates:
        key = normalize_text(candidate)
        if key and key not in seen:
            seen.add(key)
            unique.append(candidate)
    return tuple(unique)


def write_ocr_workbook(path: Path, items: list[OcrItem]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "OCR原始识别"
    headers = ["来源文件", "页码", "方法", "是否发送AI", "提示", "识别文本"]
    worksheet.append(headers)
    for item in items:
        worksheet.append(
            [
                item.source_file.name,
                item.page,
                item.method,
                "是" if item.used_for_ai else "否",
                item.warning,
                trim_excel_text(item.text),
            ]
        )
    style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def write_standard_rows_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "标准导入表"
    worksheet.append(STANDARD_FIELDS)
    for row in rows:
        worksheet.append([row.get(field) for field in STANDARD_FIELDS])
    style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def write_ai_call_workbook(path: Path, records: list[AiCallRecord]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AI调用记录"
    headers = [
        "批次",
        "总批次",
        "状态",
        "来源文件",
        "来源页码",
        "提示词字符数",
        "输出行数",
        "耗时秒",
        "输入Token",
        "输出Token",
        "总Token",
        "错误或跳过原因",
    ]
    worksheet.append(headers)
    for record in records:
        worksheet.append(
            [
                record.batch_index,
                record.total_batches,
                record.status,
                record.source_files,
                record.source_pages,
                record.prompt_chars,
                record.row_count,
                record.elapsed_seconds,
                record.input_tokens,
                record.output_tokens,
                record.total_tokens,
                record.error,
            ]
        )
    style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def write_plan_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "更新计划"
    headers = ["REITs代码", "REITs名称", "资产性质", "项目名称", "年度数量", "最早年份", "最晚年份", "增长起始", "增长率", "经营期末"]
    worksheet.append(headers)
    for group_rows in group_standard_rows(rows).values():
        years = [row.get("年份") for row in group_rows if isinstance(row.get("年份"), int)]
        worksheet.append(
            [
                first_value(group_rows, "REITs代码"),
                first_value(group_rows, "REITs名称"),
                first_value(group_rows, "底层资产性质"),
                first_value(group_rows, "项目名称"),
                len(years),
                min(years) if years else None,
                max(years) if years else None,
                first_value(group_rows, "增长率预测起始年度"),
                first_value(group_rows, "预测现金流增长率"),
                first_value(group_rows, "经营期末"),
            ]
        )
    style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def write_review_workbook(path: Path, review_items: list[dict[str, Any]], warnings: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "人工复核清单"
    headers = ["类型", "对象", "说明"]
    worksheet.append(headers)
    for warning in warnings:
        worksheet.append(["提示", "", warning])
    for item in review_items:
        worksheet.append([item.get("类型"), item.get("对象"), item.get("说明")])
    style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def write_summary_workbook(
    path: Path,
    discovered: DiscoveredFiles,
    rows: list[dict[str, Any]],
    ocr_items: list[OcrItem],
    warnings: list[str],
    output_files: list[Path | None],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "更新结果汇总"
    worksheet.append(["项目", "内容"])
    worksheet.append(["工作文件夹", str(discovered.workspace)])
    worksheet.append(["OCR来源", str(discovered.ocr_source_path or "工作文件夹自动扫描")])
    worksheet.append(["标准化行数", len(rows)])
    worksheet.append(["OCR记录数", len(ocr_items)])
    worksheet.append(["产权表", str(discovered.property_workbook or "")])
    worksheet.append(["特许经营权表", str(discovered.concession_workbook or "")])
    worksheet.append(["产权格式参考表", display_path(discovered.property_format_reference)])
    worksheet.append(["特许经营权格式参考表", display_path(discovered.concession_format_reference)])
    worksheet.append(["未来现金流格式参考表", display_path(discovered.future_cashflow_format_reference)])
    worksheet.append(["残值参数辅助表", str(discovered.residual_workbook or "")])
    worksheet.append(["未来现金流参考表（未作为默认输入）", str(discovered.future_cashflow_workbook or "")])
    worksheet.append(["输出文件", "\n".join(str(path) for path in output_files if path)])
    worksheet.append(["提示", "\n".join(warnings)])
    style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def write_process_workbook(
    path: Path,
    discovered: DiscoveredFiles,
    rows: list[dict[str, Any]],
    ocr_items: list[OcrItem],
    ai_records: list[AiCallRecord],
    review_items: list[dict[str, Any]],
    warnings: list[str],
    output_files: list[Path | None],
    comparison_pairs: list[tuple[Any, ...]],
    annual_report_financial_rows: list[dict[str, Any]] | None = None,
    annual_report_reference_rows: dict[str, dict[str, Any]] | None = None,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "更新结果汇总"
    populate_summary_sheet(summary, discovered, rows, ocr_items, warnings, output_files)
    populate_standard_rows_sheet(workbook.create_sheet("标准化导入表"), rows)
    populate_review_sheet(workbook.create_sheet("人工复核清单"), review_items, warnings)
    populate_plan_sheet(workbook.create_sheet("更新计划"), rows)
    populate_difference_reason_sheet(workbook.create_sheet("字段差异原因说明"))
    if annual_report_financial_rows:
        populate_annual_report_financial_sheets(
            workbook,
            annual_report_financial_rows,
            annual_report_reference_rows or {},
        )
    if ocr_items:
        populate_ocr_sheet(workbook.create_sheet("OCR原始识别"), ocr_items)
    if ai_records:
        populate_ai_call_sheet(workbook.create_sheet("AI调用记录"), ai_records)
    append_comparison_sheets(workbook, comparison_pairs)
    for worksheet in workbook.worksheets:
        style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def populate_summary_sheet(
    worksheet: Any,
    discovered: DiscoveredFiles,
    rows: list[dict[str, Any]],
    ocr_items: list[OcrItem],
    warnings: list[str],
    output_files: list[Path | None],
) -> None:
    worksheet.append(["项目", "内容"])
    worksheet.append(["工作文件夹", str(discovered.workspace)])
    worksheet.append(["OCR来源", str(discovered.ocr_source_path or "工作文件夹自动扫描")])
    worksheet.append(["标准化行数", len(rows)])
    worksheet.append(["OCR记录数", len(ocr_items)])
    worksheet.append(["产权表", str(discovered.property_workbook or "")])
    worksheet.append(["特许经营权表", str(discovered.concession_workbook or "")])
    worksheet.append(["产权格式参考表", display_path(discovered.property_format_reference)])
    worksheet.append(["特许经营权格式参考表", display_path(discovered.concession_format_reference)])
    worksheet.append(["未来现金流格式参考表", display_path(discovered.future_cashflow_format_reference)])
    worksheet.append(["残值参数辅助表", str(discovered.residual_workbook or "")])
    worksheet.append(["未来现金流参考表（未作为默认输入）", str(discovered.future_cashflow_workbook or "")])
    worksheet.append(["输出文件", "\n".join(str(item) for item in output_files if item)])
    worksheet.append(["提示", "\n".join(warnings)])
    worksheet.append(["说明", "本工作簿已合并 OCR 原始识别、AI 调用记录、更新计划、人工复核清单、字段差异原因和输出对比，减少中间文件数量。"])


def populate_standard_rows_sheet(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    worksheet.append(STANDARD_FIELDS)
    for row in rows:
        worksheet.append([row.get(field) for field in STANDARD_FIELDS])


def populate_ocr_sheet(worksheet: Any, items: list[OcrItem]) -> None:
    headers = ["来源文件", "页码", "方法", "是否发送AI", "提示", "识别文本"]
    worksheet.append(headers)
    for item in items:
        worksheet.append(
            [
                item.source_file.name,
                item.page,
                item.method,
                "是" if item.used_for_ai else "否",
                item.warning,
                trim_excel_text(item.text),
            ]
        )


def populate_ai_call_sheet(worksheet: Any, records: list[AiCallRecord]) -> None:
    headers = [
        "批次",
        "总批次",
        "状态",
        "来源文件",
        "来源页码",
        "提示词字符数",
        "输出行数",
        "耗时秒",
        "输入Token",
        "输出Token",
        "总Token",
        "错误或跳过原因",
    ]
    worksheet.append(headers)
    for record in records:
        worksheet.append(
            [
                record.batch_index,
                record.total_batches,
                record.status,
                record.source_files,
                record.source_pages,
                record.prompt_chars,
                record.row_count,
                record.elapsed_seconds,
                record.input_tokens,
                record.output_tokens,
                record.total_tokens,
                record.error,
            ]
        )


def populate_annual_report_financial_sheets(
    workbook: Workbook,
    rows: list[dict[str, Any]],
    reference_rows: dict[str, dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet("年报净资产折旧提取")
    headers = [
        "代码",
        "基金名称",
        "匹配PDF",
        "匹配方式",
        "基金净资产(元)",
        "基金净资产(万元)",
        "折旧及摊销(元)",
        "折旧及摊销(万元)",
        "数据来源",
        "备注",
    ]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    comparison = workbook.create_sheet("年报净资产折旧对比")
    comparison.append(["代码", "基金名称", "字段", "程序值", "参考值", "差额", "判断", "程序PDF", "参考PDF"])
    append_annual_report_financial_comparison_rows(comparison, rows, reference_rows)


def populate_plan_sheet(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = ["REITs代码", "REITs名称", "资产性质", "项目名称", "年度数量", "最早年份", "最晚年份", "增长起始", "增长率", "经营期末"]
    worksheet.append(headers)
    for group_rows in group_standard_rows(rows).values():
        years = [row.get("年份") for row in group_rows if isinstance(row.get("年份"), int)]
        worksheet.append(
            [
                first_value(group_rows, "REITs代码"),
                first_value(group_rows, "REITs名称"),
                first_value(group_rows, "底层资产性质"),
                first_value(group_rows, "项目名称"),
                len(years),
                min(years) if years else None,
                max(years) if years else None,
                first_value(group_rows, "增长率预测起始年度"),
                first_value(group_rows, "预测现金流增长率"),
                first_value(group_rows, "经营期末"),
            ]
        )


def populate_review_sheet(worksheet: Any, review_items: list[dict[str, Any]], warnings: list[str]) -> None:
    headers = ["类型", "对象", "说明"]
    worksheet.append(headers)
    for warning in warnings:
        worksheet.append(["提示", "", warning])
    for item in review_items:
        worksheet.append([item.get("类型"), item.get("对象"), item.get("说明")])


def populate_difference_reason_sheet(worksheet: Any) -> None:
    worksheet.append(["差异类型", "常见原因", "建议检查方式"])
    rows = [
        ("字段为空", "截图、标准导入表或补全表没有提供该字段；程序不会猜测不存在的内容。", "检查标准化导入表对应列，必要时人工补充或提供管理费率、评估价值、公告日期等辅助表。"),
        ("金额不同", "OCR 识别错位、AI 把收入/净收益/合计行取错、标准表做过人工修正，或小数精度不同。", "优先回看 OCR 原始识别和截图；金额差异很小通常是四舍五入。"),
        ("项目名称/颗粒度不同", "报告截图按估价对象、路段、项目包披露，而原表可能按项目整体或子项目拆分。", "如果代码+年份合计一致，通常是颗粒度问题；如果合计也不同，再检查是否漏行或多取行。"),
        ("标准表无对应年份", "报告披露周期长于当前标准表，或截图含有标准表没有保留的远期年份。", "明年正式使用时不属于错误；可按基金到期日判断是否需要保留。"),
        ("未生成产权/特许经营权表", "标准化数据中没有对应的底层资产性质，或辅助资产性质表缺失。", "在标准化导入表补充“底层资产性质”为“产权”或“特许经营权”，或提供评估价值+资产性质表。"),
        ("公式列差异", "正式表中的 IRR、市值、管理费等列通常来自模板公式，不由 OCR 直接填写。", "保留模板公式；只核对程序负责填写的现金流、日期、评估值、费率等来源字段。"),
        ("格式差异", "新增行复制模板行样式；如果模板存在重复表头或隐藏列，差异报告会提示。", "重点看结构格式差异是否影响打开和公式，不要把正常新增内容变化都视为错误。"),
    ]
    for row in rows:
        worksheet.append(row)


def write_comparison_workbook(path: Path, pairs: list[tuple[Any, ...]]) -> None:
    workbook = Workbook()
    append_comparison_sheets(workbook, pairs, workbook.active)
    for worksheet in workbook.worksheets:
        style_simple_table(worksheet)
    workbook.save(path)
    workbook.close()


def append_comparison_sheets(workbook: Workbook, pairs: list[tuple[Any, ...]], summary: Any | None = None) -> None:
    summary = summary or workbook.create_sheet("对比汇总")
    summary.title = "对比汇总"
    summary.append(["表类型", "数据源表", "格式参考表", "输出表", "结构格式差异数", "内容变化数", "说明"])
    structure_sheet = workbook.create_sheet("结构格式差异")
    structure_sheet.append(["表类型", "位置", "项目", "标准表", "输出表", "级别", "差异原因", "建议处理"])
    content_sheet = workbook.create_sheet("内容变化")
    content_sheet.append(["表类型", "判断级别", "行标识", "字段", "标准表", "输出表", "说明", "差异原因", "建议处理"])

    for pair in pairs:
        label, source_path, output_path = pair[:3]
        format_reference_path = pair[3] if len(pair) >= 4 else None
        if not output_path or not output_path.exists():
            summary.append([label, display_path(source_path), display_path(format_reference_path), display_path(output_path), "", "", "输出表不存在，已跳过对比。"])
            continue

        structure_source = format_reference_path if format_reference_path and format_reference_path.exists() else source_path
        if not structure_source or not structure_source.exists():
            summary.append([label, display_path(source_path), display_path(format_reference_path), display_path(output_path), "", "", "没有可用的数据源表或格式参考表，已跳过对比。"])
            continue

        structure_diffs, _format_content_diffs = compare_workbooks_for_report(label, structure_source, output_path)
        if source_path and source_path.exists() and structure_source == source_path:
            content_diffs = _format_content_diffs
        elif source_path and source_path.exists():
            _source_structure_diffs, content_diffs = compare_workbooks_for_report(label, source_path, output_path)
        else:
            content_diffs = []
        for item in structure_diffs:
            append_report_row(structure_sheet, item + list(explain_structure_difference(item)))
        for item in content_diffs:
            append_report_row(content_sheet, item)
        if structure_source == source_path:
            explanation = "输出基于数据源表复制后更新；内容变化不一定是错误，可能是本次自动更新写入或源表原始空缺。"
        else:
            explanation = "结构格式按格式参考表对比，内容变化按数据源表对比；格式参考表只提供布局和样式，不作为数据来源。"
        summary.append([label, display_path(source_path), display_path(format_reference_path), display_path(output_path), len(structure_diffs), len(content_diffs), explanation])


def append_report_row(worksheet: Any, values: list[Any]) -> None:
    worksheet.append([safe_report_value(value) for value in values])


def explain_structure_difference(item: list[Any]) -> tuple[str, str]:
    diff_item = str(item[2] if len(item) > 2 else "")
    level = str(item[5] if len(item) > 5 else "")
    if diff_item in {"列宽", "首行数据样式"}:
        return (
            "通常是新增行复制模板样式、Excel 自动修复另存或列宽精度造成的显示层差异。",
            "优先人工打开输出表检查是否能正常显示和计算；若不影响阅读，可作为提示项保留。",
        )
    if diff_item in {"冻结窗格", "筛选范围"}:
        return (
            "输出表会尽量继承原表设置，但模板、Excel 兼容修复或新增行范围可能导致视图设置变化。",
            "如果审核习惯需要固定视图，可在模板中统一冻结窗格和筛选范围后重新生成。",
        )
    if diff_item in {"合并单元格", "最大列数", "表头行号", "表头"}:
        return (
            "模板结构和输出结构不完全一致，可能来自模板版本差异、隐藏列、重复表头或工作表结构变化。",
            "需要重点复核；正式使用前建议统一今年/去年模板结构，再用同一模板跑更新。",
        )
    if diff_item == "名称":
        return (
            "工作表名称不同，只影响标签展示，不一定影响数据。",
            "如需完全一致，可在源模板中固定工作表名称，或人工改名。",
        )
    if level == "需确认":
        return (
            "该差异可能影响公式引用、筛选或审核阅读。",
            "请打开源表和输出表并按位置复核。",
        )
    return (
        "多为格式展示或模板继承差异。",
        "若输出表能正常打开、字段和公式正确，可作为低风险提示处理。",
    )


def safe_report_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def display_path(path: Path | str | None) -> str:
    if not path:
        return ""
    candidate = Path(path)
    try:
        resolved = candidate.resolve()
    except OSError:
        return str(path)

    return str(path)


def repair_workbooks_with_excel_if_needed(paths: Iterable[Path | None]) -> list[str]:
    existing_paths = [Path(path).resolve() for path in paths if path and Path(path).exists()]
    if not existing_paths or os.name != "nt":
        return []
    try:
        import win32com.client  # type: ignore
    except Exception:
        return ["未安装 pywin32，已跳过本机 Excel 打开校验；输出文件仍已通过 openpyxl 写入。"]

    messages: list[str] = []
    excel = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        for path in existing_paths:
            workbook = None
            try:
                workbook = excel.Workbooks.Open(str(path), 0, True)
            except Exception:
                if workbook is not None:
                    try:
                        workbook.Close(False)
                    except Exception:
                        pass
                repair_message = repair_one_workbook_with_excel(excel, path)
                if repair_message:
                    messages.append(repair_message)
            else:
                workbook.Close(False)
    except Exception as exc:
        messages.append(f"未能启动或调用本机 Excel 进行打开校验：{exc}")
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
    return messages


def repair_one_workbook_with_excel(excel: Any, path: Path) -> str:
    workbook = None
    temp_path = path.with_name(f"{path.stem}_excel_compat_tmp{path.suffix}")
    safe_unlink(temp_path)
    try:
        workbook = excel.Workbooks.Open(
            str(path),
            0,
            False,
            5,
            "",
            "",
            True,
            2,
            "",
            False,
            False,
            0,
            False,
            False,
            1,
        )
        workbook.SaveAs(str(temp_path), 51)
        workbook.Close(False)
        workbook = None
        os.replace(temp_path, path)
        verify = excel.Workbooks.Open(str(path), 0, True)
        verify.Close(False)
        return f"Excel 打开校验：{path.name} 需要兼容修复，已由本机 Excel 自动另存为可打开版本。"
    except Exception as exc:
        return f"Excel 打开校验失败：{path.name}，原因：{exc}。请人工打开复核；如 Excel 提示修复，可另存后继续使用。"
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        safe_unlink(temp_path)


def compare_workbooks_for_report(label: str, source_path: Path, output_path: Path) -> tuple[list[list[Any]], list[list[Any]]]:
    source_wb = load_workbook(source_path, data_only=False)
    output_wb = load_workbook(output_path, data_only=False)
    try:
        source_ws = source_wb.active
        output_ws = output_wb.active
        structure_diffs: list[list[Any]] = []
        content_diffs: list[list[Any]] = []

        add_diff = structure_diffs.append
        if source_ws.title != output_ws.title:
            add_diff([label, "工作表", "名称", source_ws.title, output_ws.title, "提示"])
        if source_ws.max_column != output_ws.max_column:
            add_diff([label, "工作表", "最大列数", source_ws.max_column, output_ws.max_column, "需确认"])
        if source_ws.freeze_panes != output_ws.freeze_panes:
            add_diff([label, "工作表", "冻结窗格", source_ws.freeze_panes, output_ws.freeze_panes, "需确认"])
        if source_ws.auto_filter.ref != output_ws.auto_filter.ref:
            add_diff([label, "工作表", "筛选范围", source_ws.auto_filter.ref, output_ws.auto_filter.ref, "提示"])

        source_merges = sorted(str(item) for item in source_ws.merged_cells.ranges)
        output_merges = sorted(str(item) for item in output_ws.merged_cells.ranges)
        if source_merges != output_merges:
            add_diff([label, "工作表", "合并单元格", "\n".join(source_merges[:20]), "\n".join(output_merges[:20]), "需确认"])

        compare_column_dimensions(label, source_ws, output_ws, structure_diffs)
        compare_header_and_style(label, source_ws, output_ws, structure_diffs)
        content_diffs.extend(compare_content_changes(label, source_ws, output_ws))
        return structure_diffs, content_diffs
    finally:
        source_wb.close()
        output_wb.close()


def compare_column_dimensions(label: str, source_ws: Any, output_ws: Any, diffs: list[list[Any]]) -> None:
    for col_idx in range(1, max(source_ws.max_column, output_ws.max_column) + 1):
        letter = get_column_letter(col_idx)
        source_dim = source_ws.column_dimensions[letter]
        output_dim = output_ws.column_dimensions[letter]
        if normalize_dimension_width(source_dim.width) != normalize_dimension_width(output_dim.width):
            diffs.append([label, letter, "列宽", source_dim.width, output_dim.width, "提示"])
        if bool(source_dim.hidden) != bool(output_dim.hidden):
            diffs.append([label, letter, "隐藏列", source_dim.hidden, output_dim.hidden, "需确认"])


def compare_header_and_style(label: str, source_ws: Any, output_ws: Any, diffs: list[list[Any]]) -> None:
    source_header_row, source_headers = find_header_row(source_ws)
    output_header_row, output_headers = find_header_row(output_ws)
    if source_header_row != output_header_row:
        diffs.append([label, "表头", "表头行号", source_header_row, output_header_row, "需确认"])

    max_col = max(source_ws.max_column, output_ws.max_column)
    for col_idx in range(1, max_col + 1):
        source_header = source_headers[col_idx - 1] if col_idx <= len(source_headers) else None
        output_header = output_headers[col_idx - 1] if col_idx <= len(output_headers) else None
        if source_header != output_header:
            diffs.append([label, get_column_letter(col_idx), "表头", source_header, output_header, "需确认"])

        source_style_cell = source_ws.cell(min(source_ws.max_row, source_header_row + 1), col_idx)
        output_style_cell = output_ws.cell(min(output_ws.max_row, output_header_row + 1), col_idx)
        source_signature = style_signature(source_style_cell)
        output_signature = style_signature(output_style_cell)
        if source_signature != output_signature:
            diffs.append([label, get_column_letter(col_idx), "首行数据样式", source_signature, output_signature, "提示"])


def compare_content_changes(label: str, source_ws: Any, output_ws: Any) -> list[list[Any]]:
    source_header_row, source_headers = find_header_row(source_ws)
    output_header_row, output_headers = find_header_row(output_ws)
    source_cols = build_normalized_header_col_map(source_headers)
    output_cols = build_normalized_header_col_map(output_headers)
    common_headers = [header for header in source_cols if header in output_cols]
    key_headers = [normalize_text(item) for item in ("代码", "项目名称", "年份", "中诚信补充计算现金流年份")]

    source_rows = build_keyed_rows(source_ws, source_header_row, source_cols, key_headers)
    output_rows = build_keyed_rows(output_ws, output_header_row, output_cols, key_headers)
    diffs: list[list[Any]] = []
    for key, output_row_idx in output_rows.items():
        source_row_idx = source_rows.get(key)
        if not source_row_idx:
            diffs.append(
                [
                    label,
                    "需复核",
                    key,
                    "整行",
                    "",
                    f"输出新增行 {output_row_idx}",
                    "输出新增；请复核是否为新项目或新年份。",
                    "标准导入表包含源表没有的基金、项目或年份，程序按本次材料新增了明细行。",
                    "确认该基金/项目/年份是否确实应进入今年工作底稿；如果不需要，从标准导入表删除后重跑。",
                ]
            )
            if len(diffs) >= 5000:
                return diffs
            continue
        for header in common_headers:
            source_value = source_ws.cell(source_row_idx, source_cols[header]).value
            output_value = output_ws.cell(output_row_idx, output_cols[header]).value
            if values_equivalent(source_value, output_value):
                continue
            field_name = source_ws.cell(source_header_row, source_cols[header]).value
            level = classify_content_difference(field_name, source_value, output_value)
            reason, action = explain_content_difference(field_name, source_value, output_value)
            diffs.append(
                [
                    label,
                    level,
                    key,
                    field_name,
                    source_value,
                    output_value,
                    "内容变化；可能是本次更新写入，也可能是标准表原本为空或人工留存内容。",
                    reason,
                    action,
                ]
            )
            if len(diffs) >= 5000:
                return diffs
    return diffs


def build_normalized_header_col_map(headers: Sequence[Any]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for idx, header in enumerate(headers, 1):
        key = normalize_text(header)
        if key and key not in cols:
            cols[key] = idx
    return cols


def classify_content_difference(field_name: Any, source_value: Any, output_value: Any) -> str:
    field = normalize_text(field_name)
    if output_value in (None, "") and source_value not in (None, ""):
        if "公式" in str(source_value) or is_formula(source_value):
            return "需复核"
        if "代码" in field or "项目名称" in field or "年份" in field or "预测现金流" in field:
            return "风险差异"
        return "需复核"
    if source_value in (None, "") and output_value not in (None, ""):
        return "正常更新"
    if is_formula(source_value) or is_formula(output_value):
        return "需复核"
    if "项目名称" in field:
        return "风险差异"
    if "上市" in field or "到期" in field:
        return "风险差异"
    if "公告" in field or "报告期" in field or "评估基准" in field or "估值基准" in field or "日期" in field:
        return "需复核"
    if "费率" in field or "折现率" in field or "增长率" in field:
        return "风险差异"
    if "现金流" in field or "金额" in field or "价值" in field or "净资产" in field or "折旧" in field:
        return "正常更新"
    return "需复核"


def explain_content_difference(field_name: Any, source_value: Any, output_value: Any) -> tuple[str, str]:
    field = normalize_text(field_name)
    if output_value in (None, ""):
        if source_value in (None, ""):
            return ("两边均为空或等价，通常不会进入差异；如出现请检查格式化值。", "可忽略或复核单元格格式。")
        return (
            "本次截图、标准导入表或辅助表没有提供该字段；程序不会把去年旧值强行沿用到需要更新的字段。",
            "如果今年仍需该字段，请在标准导入表或对应辅助表中补充后重跑；否则保持为空是安全做法。",
        )
    if source_value in (None, ""):
        return (
            "源表该位置为空，本次自动更新根据截图、标准导入表或辅助表写入了新值。",
            "回看标准化导入表和来源截图，确认该值是否为今年应填内容。",
        )
    if is_formula(source_value) or is_formula(output_value):
        return (
            "该字段涉及模板公式。程序会尽量保留或平移公式，不会用 OCR 文本随意覆盖公式列。",
            "重点检查公式引用行号是否跟随新增行正确平移。",
        )
    if "日期" in field or "公告" in field or "上市" in field or "到期" in field or "报告期" in field or "评估基准" in field or "估值基准" in field:
        return (
            "日期类字段来自标准导入表、公告日期表、OCR/AI 标准化结果、辅助表或源模板既有信息；不同表可能存在披露日、基准日、到期日口径差异。",
            "已有项目会优先按标准审核表保留；如果仍出现差异，说明该项目可能是新增、项目名未匹配，或辅助表与标准审核表口径不同。",
        )
    if "费率" in field or "折现率" in field or "增长率" in field:
        return (
            "比例/费率类字段来自辅助表或标准导入表；若标准审核表已有同项目口径，程序会优先保留标准审核表值。",
            "若仍有差异，请检查该项目是否未匹配到标准审核表，或辅助表中百分比是否误填。",
        )
    if "现金流" in field or "金额" in field or "价值" in field or "净资产" in field or "折旧" in field:
        return (
            "金额类字段变化通常来自今年报告披露的新现金流、评估价值、净资产/折旧辅助表，或 OCR/AI 识别口径不同。",
            "优先核对来源截图和标准化导入表；如果只是项目颗粒度不同，再看代码+年份合计是否一致。",
        )
    return (
        "字段值与源表不同，通常是本次自动更新写入、源表旧值被清理，或今年材料没有提供完整字段。",
        "结合“更新计划”“标准化导入表”和来源截图判断是否需要人工补充。",
    )


def build_keyed_rows(worksheet: Any, header_row: int, cols: dict[str, int], key_headers: list[str]) -> dict[str, int]:
    code_col = cols.get("代码")
    project_col = cols.get("项目名称")
    year_col = cols.get("年份") or cols.get("中诚信补充计算现金流年份")
    keyed: dict[str, int] = {}
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        code = normalize_code(worksheet.cell(row_idx, code_col).value) if code_col else ""
        project = normalize_project(worksheet.cell(row_idx, project_col).value) if project_col else ""
        year = worksheet.cell(row_idx, year_col).value if year_col else row_idx
        if not code and not project and year in (None, ""):
            continue
        keyed[f"{code}|{project}|{year}"] = row_idx
    return keyed


def normalize_dimension_width(width: Any) -> float | None:
    if width is None:
        return None
    try:
        return round(float(width), 4)
    except (TypeError, ValueError):
        return None


def style_signature(cell: Any) -> str:
    fill_color = cell.fill.fgColor.rgb or cell.fill.fgColor.indexed or cell.fill.fgColor.theme
    border = ",".join(
        str(getattr(side, "style", None) or "")
        for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom)
    )
    return "|".join(
        [
            str(cell.number_format),
            str(cell.font.name),
            str(cell.font.sz),
            str(bool(cell.font.bold)),
            str(fill_color),
            str(cell.alignment.horizontal),
            str(cell.alignment.vertical),
            border,
        ]
    )


def values_equivalent(left: Any, right: Any) -> bool:
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= 0.0001
    return str(left) == str(right)


def style_simple_table(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(name="Microsoft YaHei", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col_idx in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 18
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]


def finalize_worksheet_view(worksheet: Any) -> None:
    worksheet.sheet_view.topLeftCell = "A1"
    if worksheet.freeze_panes is None:
        worksheet.sheet_view.pane = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        return

    selections = list(worksheet.sheet_view.selection or [])
    if not selections:
        selections = [Selection()]
    pane = getattr(worksheet.sheet_view, "pane", None)
    for selection in selections:
        selection.activeCell = "A1"
        selection.sqref = "A1"
        if pane is not None and not selection.pane:
            selection.pane = pane.activePane
    worksheet.sheet_view.selection = selections


def first_value(rows: list[dict[str, Any]], field_name: str) -> Any:
    for row in rows:
        value = row.get(field_name)
        if value not in (None, ""):
            return value
    return None


def max_year_in_rows(rows: list[dict[str, Any]]) -> int | None:
    years = [row.get("年份") for row in rows if isinstance(row.get("年份"), int)]
    return max(years) if years else None


def midyear_date(year: Any) -> date | None:
    if not isinstance(year, int):
        return None
    return date(year, 6, 30)


def date_year(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    parsed = parse_date_like(value)
    if isinstance(parsed, datetime):
        return parsed.year
    if isinstance(parsed, date):
        return parsed.year
    if isinstance(value, int) and value > 1900:
        text = str(value)
        if len(text) >= 4:
            return int(text[:4])
    return None


def to_yyyymmdd(value: Any) -> Any:
    parsed = parse_date_like(value)
    if isinstance(parsed, datetime):
        return int(parsed.strftime("%Y%m%d"))
    if isinstance(parsed, date):
        return int(parsed.strftime("%Y%m%d"))
    return value


def parse_date_like(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, (int, float)) and 19000101 <= int(value) <= 22001231:
        text = str(int(value))
        try:
            return datetime.strptime(text, "%Y%m%d").date()
        except ValueError:
            return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return value


def parse_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    multiplier = 1.0
    if "%" in text:
        multiplier = 0.01
    text = (
        text.replace(",", "")
        .replace("，", "")
        .replace("%", "")
        .replace("万元", "")
        .replace("万", "")
        .replace("元", "")
        .replace("人民币", "")
        .replace("约", "")
        .strip()
    )
    try:
        number = float(Decimal(text) * Decimal(str(multiplier)))
    except (InvalidOperation, ValueError):
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        number = float(match.group()) * multiplier if match else 0.0
    return -number if negative else number


def parse_fee_rate_value(value: Any) -> float:
    """Fee-rate columns are stored as percentage points because formulas divide by 100."""
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    if "%" not in text and "％" not in text:
        return parse_number(text)
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("%", "").replace("％", "").replace(",", "").strip()
    try:
        number = float(Decimal(text))
    except (InvalidOperation, ValueError):
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        number = float(match.group()) if match else 0.0
    return -number if negative else number


def parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    match = re.search(r"\d{4}", str(value))
    return int(match.group()) if match else None


def normalize_code(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip().upper()
    text = text.replace("。", ".").replace("．", ".")
    text = re.sub(r"\s+", "", text)
    match = re.fullmatch(r"(\d{6})(?:\.(SH|SZ|SS))?", text)
    if not match:
        return ""
    code, suffix = match.groups()
    if suffix == "SS":
        suffix = "SH"
    if not suffix:
        if code.startswith("508"):
            suffix = "SH"
        elif code.startswith("180"):
            suffix = "SZ"
        else:
            return ""
    return f"{code}.{suffix}"


def is_valid_reits_code(value: Any) -> bool:
    return bool(normalize_code(value))


def normalize_project(value: Any) -> str:
    text = str(value or "项目整体")
    text = text.replace("\u200c", "").replace("\u200b", "")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[（）()《》<>【】\[\]、，,。.\-—_]", "", text)
    return text or "项目整体"


def clean_project_name(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "项目整体"
    normalized = normalize_project(text)
    if normalized == "项目整体":
        return "项目整体"
    if any(term in normalized for term in METRIC_PROJECT_TERMS):
        residue = normalized
        for term in METRIC_PROJECT_TERMS:
            residue = residue.replace(term, "")
        generic_residue = residue
        for token in ("估价对象", "评估对象", "全周期", "合计", "及", "至", "一", "二", "三", "四", "五", "六", "七", "八", "九", "十"):
            generic_residue = generic_residue.replace(token, "")
        if not generic_residue:
            return "项目整体"
    return text


def infer_missing_asset_nature(row: dict[str, Any]) -> None:
    if row.get("底层资产性质") not in (None, ""):
        return
    combined = normalize_text(
        " ".join(
            str(row.get(field) or "")
            for field in ("REITs名称", "基础设施项目类型", "项目名称", "来源文件", "备注")
        )
    )
    if not combined:
        return
    if any(normalize_text(keyword) in combined for keyword in CONCESSION_HINTS):
        row["底层资产性质"] = "特许经营权"
    elif any(normalize_text(keyword) in combined for keyword in PROPERTY_HINTS):
        row["底层资产性质"] = "产权"


def normalize_fund_name(value: Any) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = re.sub(r"\.(pdf|docx?|xlsx?|png|jpe?g)$", "", text, flags=re.I)
    text = re.sub(r"\d{4}年(度)?", "", text)
    text = re.sub(r"(年度报告|年报|评估报告|市场价值|持有的全部不动产项目的)", "", text)
    text = re.sub(r"年度?$", "", text)
    return normalize_text(text)


def normalize_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip()
    text = text.replace("\u200c", "").replace("\u200b", "")
    return re.sub(r"[\s（）()《》<>【】\[\]、，,。.:：;；%％\-—_/\\]+", "", text)


def projects_match(left: str, right: str) -> bool:
    if not left or not right:
        return False
    if left == right:
        return True
    if left == "项目整体" or right == "项目整体":
        return left == right
    return left in right or right in left or SequenceMatcher(None, left, right).ratio() >= 0.82


def trim_excel_text(text: str) -> str:
    if len(text) <= MAX_EXCEL_TEXT_LENGTH:
        return text
    return text[: MAX_EXCEL_TEXT_LENGTH - 20] + "\n...已截断"
