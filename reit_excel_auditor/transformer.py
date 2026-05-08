from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
import json
import re
import sys
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.views import Selection


AUTO_TYPE = "auto"
CUSTOM_TEMPLATE_TYPE = "custom_template"

TABLE_TYPES: dict[str, str] = {
    AUTO_TYPE: "自动识别",
    CUSTOM_TEMPLATE_TYPE: "自定义模板输出",
    "valuation": "1-基础资产评估价值表",
    "traffic": "2-交通基础设施经营指标表",
    "finance": "3-基础资产财务指标表",
    "property": "4-产权类经营指标表",
    "energy": "5-能源基础设施经营指标表",
}

TARGET_HEADERS: dict[str, list[str]] = {
    "valuation": ["REITs代码", "REITs名称", "上市日期", "公告日期", "数据来源", "评估基准日", "评估价值(元)"],
    "traffic": ["REITs代码", "REITs简称", "STARTDATE", "ENDDATE", "公告日期", "本期时间范围", "资产项目名称", "日均车流量(辆次)", "当月路费收入(万元)"],
    "finance": ["INFOPUBLDATE", "STARTDATE", "ENDDATE", "REITs代码", "REITs名称", "资产项目名称", "数据来源", "主要资产（元）", "主要负债（元）", "净资产（元）", "营业收入（元）", "营业成本及费用（元）", "净利润（元）"],
    "property": ["REITs代码", "REITs名称", "公告日期", "STARTDATE", "ENDDATE", "资产项目名称", "出租率", "租金单价(元/月/平方米)", "租金收缴率", "数据来源"],
    "energy": ["REITs代码", "REITs名称", "STARTDATE", "ENDDATE", "公告日期", "资产项目名称", "发电量(万千瓦时)", "结算电量(万千瓦时)", "结算电价(元/千瓦时)", "发电收入(元)"],
}

SHEET_NAMES: dict[str, str] = {
    "valuation": "reits估值",
    "traffic": "202604",
    "finance": "202604",
    "property": "202604",
    "energy": "能源经营数据",
}

STANDARD_TEMPLATE_DIR_NAMES = ("已审核标准模板", "standard_templates")
CONFIG_DIR_NAMES = ("config",)
DEFAULT_STANDARD_TEMPLATE_NAMES: dict[str, str] = {
    "valuation": "01-基础资产估值标准模板.xlsx",
    "traffic": "02-高速经营数据标准模板.xlsx",
    "finance": "03-资产负债收入成本标准模板.xlsx",
    "property": "04-产权经营数据标准模板.xlsx",
    "energy": "05-能源经营数据标准模板.xlsx",
}

DATE_COLUMNS = {"上市日期", "公告日期", "INFOPUBLDATE"}
FINANCE_DATE_COLUMNS = {"STARTDATE", "ENDDATE"}
PERCENT_COLUMNS = {"出租率", "租金收缴率"}
GENERATED_MARKERS = ("_自动审核", "自动审核_批量汇总")
CUSTOM_TEMPLATE_FOLDER_SIMILARITY_THRESHOLD = 0.70


def clone_config(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): clone_config(item) for key, item in value.items()}
    if isinstance(value, list):
        return [clone_config(item) for item in value]
    return value


def candidate_resource_bases() -> list[Path]:
    bases: list[Path] = []
    bundled_base = getattr(sys, "_MEIPASS", None)
    if bundled_base:
        bases.append(Path(bundled_base))
    bases.extend(
        [
            Path(__file__).resolve().parents[1],
            Path.cwd(),
            Path(sys.executable).resolve().parent,
        ]
    )

    unique_bases: list[Path] = []
    seen: set[Path] = set()
    for base in bases:
        resolved_base = base.resolve()
        if resolved_base in seen:
            continue
        seen.add(resolved_base)
        unique_bases.append(resolved_base)
    return unique_bases


def load_json_config(file_name: str, default: dict[str, Any]) -> dict[str, Any]:
    for base in candidate_resource_bases():
        for config_dir_name in CONFIG_DIR_NAMES:
            path = base / config_dir_name / file_name
            if not path.exists():
                continue
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            if isinstance(data, dict):
                return data
    return clone_config(default)


def load_standard_template_names() -> dict[str, str]:
    raw_config = load_json_config("table_templates.json", DEFAULT_STANDARD_TEMPLATE_NAMES)
    names = dict(DEFAULT_STANDARD_TEMPLATE_NAMES)
    for table_type, template_name in raw_config.items():
        if isinstance(template_name, str) and template_name.strip():
            names[str(table_type)] = template_name.strip()
    return names


def load_field_aliases() -> dict[str, list[str]]:
    raw_config = load_json_config("field_aliases.json", DEFAULT_FIELD_ALIASES)
    aliases = {field_name: list(values) for field_name, values in DEFAULT_FIELD_ALIASES.items()}
    for field_name, values in raw_config.items():
        if isinstance(values, str):
            normalized_values = [values]
        elif isinstance(values, list):
            normalized_values = [str(value) for value in values if value not in (None, "")]
        else:
            continue

        target_values = aliases.setdefault(str(field_name), [])
        for value in normalized_values:
            if value and value not in target_values:
                target_values.append(value)
    return aliases


def standard_template_names_for(table_type: str) -> list[str]:
    names: list[str] = []
    template_name = STANDARD_TEMPLATE_NAMES.get(table_type)
    if template_name:
        names.append(template_name)
    return names


STANDARD_TEMPLATE_NAMES: dict[str, str] = load_standard_template_names()

DEFAULT_FIELD_ALIASES: dict[str, list[str]] = {
    "REITs代码": ["ReitsCode", "代码", "证券代码", "基金代码", "产品代码"],
    "ReitsCode": ["REITs代码", "代码", "证券代码", "基金代码", "产品代码"],
    "REITs名称": ["REITs简称", "基金名称", "基金简称", "证券简称", "证券名称", "产品名称"],
    "REITs简称": ["REITs名称", "基金名称", "基金简称", "证券简称", "证券名称", "产品名称"],
    "上市日期": ["发行日期"],
    "资产项目名称": ["基础设施项目名称", "基础设施项目公司名称", "项目名称"],
    "项目名称": ["资产项目名称", "基础设施项目名称", "基础设施项目公司名称"],
    "STARTDATE": ["开始日期", "报告期开始日期", "报告期起始日期", "区间开始日期"],
    "ENDDATE": ["结束日期", "报告期结束日期", "报告期截止日期", "区间结束日期"],
    "INFOPUBLDATE": ["公告日期", "信息披露日期", "披露日期"],
    "公告日期": ["INFOPUBLDATE", "信息披露日期", "披露日期"],
    "主要资产（元）": ["资产合计（元）"],
    "主要负债（元）": ["负债合计（元）"],
    "营业成本及费用（元）": ["营业成本（元）"],
    "日均车流量(辆次)": ["日均收费车流量(辆次)"],
    "租金单价(元/月/平方米)": ["租金单价(单位:元/月/平方米or元/月/个)"],
}

FIELD_ALIASES: dict[str, list[str]] = load_field_aliases()

OUTPUT_NUMBER_FORMATS: dict[str, dict[str, str]] = {
    "valuation": {
        "上市日期": "yyyymmdd",
        "公告日期": "yyyymmdd",
        "评估价值(元)": "#,##0.00",
    },
    "traffic": {
        "公告日期": "yyyymmdd",
        "本期时间范围": "@",
    },
    "finance": {
        "INFOPUBLDATE": "yyyymmdd",
        "STARTDATE": "yyyymmdd",
        "ENDDATE": "yyyymmdd",
    },
    "property": {
        "公告日期": "yyyymmdd",
        "出租率": "0.00%",
        "租金收缴率": "0.00%",
    },
    "energy": {
        "公告日期": "yyyymmdd",
    },
}

PROPERTY_PROCESSED_PERCENT_HEADERS = {"出租率", "租金收缴率"}
PROPERTY_PROCESSED_YYYYMMDD_DATE_HEADERS = {"公告日期"}
PROPERTY_PROCESSED_NUMERIC_HEADERS = {
    "开始日期",
    "结束日期",
    "主配套资产单项可出租面积(平方米)",
    "主配套资产合计的可出租面积(平方米)",
    "主配套资产可出租数量(个/间/套)",
    "主配套资产单项实际出租面积(平方米)",
    "主配套资产合计的实际出租面积(平方米)",
    "主配套资产实际出租数量(个/间/套)",
    "租金单价(单位:元/月/平方米or元/月/个)",
    "报告期末平均剩余租期(天)",
}


@dataclass
class SourceTable:
    path: Path
    sheet_name: str
    header_row: int
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass
class MetadataRecord:
    code: str
    values: dict[str, Any] = field(default_factory=dict)


@dataclass
class ConversionResult:
    detected_type: str
    output_files: list[Path]
    warnings: list[str]
    row_count: int


@dataclass
class BatchItemResult:
    input_file: Path
    status: str
    detected_type: str | None = None
    row_count: int = 0
    output_files: list[Path] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    error: str | None = None


@dataclass
class BatchResult:
    input_path: Path
    output_dir: Path
    summary_file: Path
    items: list[BatchItemResult]

    @property
    def total_count(self) -> int:
        return len(self.items)

    @property
    def success_count(self) -> int:
        return sum(1 for item in self.items if item.status == "成功")

    @property
    def failed_count(self) -> int:
        return sum(1 for item in self.items if item.status == "失败")


class ConversionError(Exception):
    pass


@dataclass
class CustomTemplateLayout:
    path: Path
    header_row: int
    headers: list[str]
    data_number_formats: list[str]
    data_formulas: list[str | None] = field(default_factory=list)


def convert_input_path(
    input_path: str | Path,
    selected_type: str = AUTO_TYPE,
    metadata_path: str | Path | None = None,
    custom_template_path: str | Path | None = None,
    output_dir: str | Path | None = None,
    generate_property_processed: bool = False,
) -> BatchResult:
    """Convert a single Excel file or every eligible Excel file in a folder."""
    input_path = Path(input_path)
    if not input_path.exists():
        raise ConversionError(f"输入路径不存在：{input_path}")

    excluded_paths = [path for path in (metadata_path, custom_template_path) if path]
    files = collect_input_files(input_path, excluded_paths=excluded_paths)
    if not files:
        raise ConversionError("未找到可转换的 .xlsx 文件。请确认文件不是 Excel 临时文件，也不是已生成的 _自动审核 文件。")

    output_dir_path = Path(output_dir) if output_dir else (input_path.parent if input_path.is_file() else input_path)
    output_dir_path.mkdir(parents=True, exist_ok=True)

    reference_headers = None
    if custom_template_path and input_path.is_dir() and len(files) > 1:
        try:
            reference_headers = read_source_table(files[0]).headers
        except Exception:
            reference_headers = None

    items: list[BatchItemResult] = []
    for file_path in files:
        try:
            result = convert_file(
                input_path=file_path,
                selected_type=selected_type,
                metadata_path=metadata_path,
                custom_template_path=custom_template_path,
                output_dir=output_dir_path,
                reference_headers=reference_headers,
                generate_property_processed=generate_property_processed,
            )
        except Exception as exc:
            items.append(
                BatchItemResult(
                    input_file=file_path,
                    status="失败",
                    error=str(exc),
                )
            )
        else:
            items.append(
                BatchItemResult(
                    input_file=file_path,
                    status="成功",
                    detected_type=result.detected_type,
                    row_count=result.row_count,
                    output_files=result.output_files,
                    warnings=result.warnings,
                )
            )

    summary_file = output_dir_path / "自动审核_批量汇总.xlsx"
    write_batch_summary(summary_file, items)
    return BatchResult(input_path=input_path, output_dir=output_dir_path, summary_file=summary_file, items=items)


def collect_input_files(input_path: Path, excluded_paths: Iterable[str | Path] | None = None) -> list[Path]:
    excluded = {Path(path).resolve() for path in (excluded_paths or []) if path}
    if input_path.is_file():
        return [input_path] if is_convertible_excel(input_path) and input_path.resolve() not in excluded else []
    return sorted(path for path in input_path.rglob("*.xlsx") if is_convertible_excel(path) and path.resolve() not in excluded)


def is_convertible_excel(path: Path) -> bool:
    if path.name.startswith("~$"):
        return False
    if path.suffix.lower() != ".xlsx":
        return False
    return not any(marker in path.stem for marker in GENERATED_MARKERS)


def convert_file(
    input_path: str | Path,
    selected_type: str = AUTO_TYPE,
    metadata_path: str | Path | None = None,
    custom_template_path: str | Path | None = None,
    output_dir: str | Path | None = None,
    reference_headers: list[str] | None = None,
    generate_property_processed: bool = False,
) -> ConversionResult:
    input_path = Path(input_path)
    if not input_path.exists():
        raise ConversionError(f"输入文件不存在：{input_path}")
    if not is_convertible_excel(input_path):
        raise ConversionError("请选择普通 .xlsx 文件，不要选择 Excel 临时文件或已经生成的 _自动审核 文件。")

    table = read_source_table(input_path)
    if custom_template_path:
        return convert_file_with_custom_template(
            input_path=input_path,
            table=table,
            custom_template_path=custom_template_path,
            metadata_path=metadata_path,
            output_dir=output_dir,
            reference_headers=reference_headers,
        )

    metadata = read_metadata(metadata_path) if metadata_path else {}
    warnings: list[str] = []

    table_type = detect_table_type(table.headers) if selected_type == AUTO_TYPE else selected_type
    if table_type not in TARGET_HEADERS:
        raise ConversionError("无法识别表类型，请手动选择转换逻辑后重试。")

    transformer = {
        "valuation": transform_valuation,
        "traffic": transform_traffic,
        "finance": transform_finance,
        "property": transform_property,
        "energy": transform_energy,
    }[table_type]
    output_rows = transformer(table, metadata, warnings)

    output_dir_path = Path(output_dir) if output_dir else input_path.parent
    output_dir_path.mkdir(parents=True, exist_ok=True)

    main_output = make_output_path(input_path, output_dir_path, "_自动审核", ".xlsx")
    write_standard_workbook(table_type, output_rows, main_output)

    output_files = [main_output]
    if table_type == "property" and generate_property_processed:
        processed_output = make_output_path(input_path, output_dir_path, "_自动审核_处理版", ".xlsx")
        changed = write_property_processed_workbook(input_path, processed_output)
        output_files.append(processed_output)
        warnings.append(f"第4类表已按选项输出处理版：修复/补全 {changed} 个面积相关单元格。")

    return ConversionResult(
        detected_type=table_type,
        output_files=output_files,
        warnings=warnings,
        row_count=len(output_rows),
    )


def convert_file_with_custom_template(
    input_path: Path,
    table: SourceTable,
    custom_template_path: str | Path,
    metadata_path: str | Path | None = None,
    output_dir: str | Path | None = None,
    reference_headers: list[str] | None = None,
) -> ConversionResult:
    template_path = Path(custom_template_path)
    if not template_path.exists():
        raise ConversionError(f"自定义模板表不存在：{template_path}")
    if template_path.suffix.lower() != ".xlsx" or template_path.name.startswith("~$"):
        raise ConversionError("自定义模板表必须是普通 .xlsx 文件，不能是 Excel 临时文件。")

    if reference_headers and not source_headers_are_compatible(reference_headers, table.headers):
        similarity = headers_similarity(reference_headers, table.headers)
        raise ConversionError(
            f"自定义模板文件夹模式要求来源表格式一致；当前文件表头相似度 {similarity:.0%}，低于 "
            f"{CUSTOM_TEMPLATE_FOLDER_SIMILARITY_THRESHOLD:.0%}，请单独处理或拆分文件夹。"
        )

    warnings: list[str] = []
    metadata = read_metadata(metadata_path) if metadata_path else {}
    layout = read_custom_template_layout(template_path)
    output_rows = build_custom_template_rows(table, layout, warnings, metadata=metadata)

    output_dir_path = Path(output_dir) if output_dir else input_path.parent
    output_dir_path.mkdir(parents=True, exist_ok=True)
    main_output = make_output_path(input_path, output_dir_path, "_自动审核", ".xlsx")
    write_custom_template_workbook(layout, output_rows, main_output)

    return ConversionResult(
        detected_type=CUSTOM_TEMPLATE_TYPE,
        output_files=[main_output],
        warnings=warnings,
        row_count=len(output_rows),
    )


def read_source_table(path: Path) -> SourceTable:
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook.active
        header_row = find_header_row(worksheet)
        headers = [clean_text(worksheet.cell(header_row, col).value) for col in range(1, worksheet.max_column + 1)]

        rows: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            record: dict[str, Any] = {"__row_number__": row_idx}
            has_value = False
            for col_idx, header in enumerate(headers, 1):
                if not header:
                    continue
                value = worksheet.cell(row_idx, col_idx).value
                if not is_blank(value):
                    has_value = True
                normalized_header = normalize_header(header)
                record[normalized_header] = value
                record[format_key(normalized_header)] = worksheet.cell(row_idx, col_idx).number_format
            if has_value:
                rows.append(record)

        return SourceTable(path=path, sheet_name=worksheet.title, header_row=header_row, headers=headers, rows=rows)
    finally:
        workbook.close()


def find_header_row(worksheet: Any) -> int:
    best_row = 1
    best_count = 0
    for row_idx in range(1, min(worksheet.max_row, 20) + 1):
        count = sum(1 for col_idx in range(1, worksheet.max_column + 1) if not is_blank(worksheet.cell(row_idx, col_idx).value))
        if count > best_count:
            best_row = row_idx
            best_count = count
    if best_count < 1:
        raise ConversionError("未能找到有效表头行。")
    return best_row


def detect_table_type(headers: Iterable[str]) -> str:
    normalized = [normalize_header(header) for header in headers if header]
    rules = {
        "valuation": ["评估基准日", "评估价值"],
        "traffic": ["日均收费车流量", "日均车流量", "当月路费收入"],
        "finance": ["资产合计", "主要资产", "负债合计", "主要负债", "净利润", "营业收入"],
        "property": ["主配套资产类别", "出租率", "租金收缴率", "租金单价"],
        "energy": ["发电量", "结算电量", "结算电价"],
    }

    scored: list[tuple[str, int]] = []
    for table_type, tokens in rules.items():
        score = 0
        for token in tokens:
            normalized_token = normalize_header(token)
            if any(normalized_token in header for header in normalized):
                score += 1
        scored.append((table_type, score))

    scored.sort(key=lambda item: item[1], reverse=True)
    if not scored or scored[0][1] < 2:
        raise ConversionError("自动识别失败：表头特征不足。")
    if len(scored) > 1 and scored[0][1] == scored[1][1]:
        raise ConversionError("自动识别结果不唯一，请手动选择转换逻辑。")
    return scored[0][0]


def read_metadata(path: str | Path | None) -> dict[str, MetadataRecord]:
    if not path:
        return {}
    metadata_path = Path(path)
    if not metadata_path.exists():
        raise ConversionError(f"补全信息表不存在：{metadata_path}")

    table = read_source_table(metadata_path)
    records: dict[str, MetadataRecord] = {}
    for row in table.rows:
        raw_code = get_value(row, "REITs代码", "ReitsCode", "代码", "证券代码", "基金代码", "产品代码")
        code_key = code_digits(raw_code)
        if not code_key:
            continue
        records[code_key] = MetadataRecord(
            code=code_key,
            values={
                "REITs代码": raw_code,
                "REITs名称": get_value(row, "REITs名称", "REITs简称", "基金名称", "基金简称", "证券简称", "证券名称", "产品名称"),
                "上市日期": get_value(row, "上市日期", "发行日期"),
                "公告日期": get_value(row, "公告日期", "INFOPUBLDATE", "信息披露日期", "披露日期"),
                "开始日期": get_value(row, "开始日期", "STARTDATE", "报告期开始日期", "报告期起始日期", "区间开始日期"),
                "结束日期": get_value(row, "结束日期", "ENDDATE", "报告期结束日期", "报告期截止日期", "区间结束日期"),
            },
        )
    return records


def lookup_metadata(metadata: dict[str, MetadataRecord], code: Any) -> MetadataRecord | None:
    return metadata.get(code_digits(code))


def read_custom_template_layout(path: Path) -> CustomTemplateLayout:
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook.active
        header_row = find_header_row(worksheet)
        headers = [clean_text(worksheet.cell(header_row, col_idx).value) for col_idx in range(1, worksheet.max_column + 1)]
        style_row = header_row + 1 if worksheet.max_row >= header_row + 1 else header_row
        data_number_formats = [worksheet.cell(style_row, col_idx).number_format for col_idx in range(1, worksheet.max_column + 1)]
        data_formulas = [cell_formula(worksheet.cell(style_row, col_idx).value) for col_idx in range(1, worksheet.max_column + 1)]
        return CustomTemplateLayout(
            path=path,
            header_row=header_row,
            headers=headers,
            data_number_formats=data_number_formats,
            data_formulas=data_formulas,
        )
    finally:
        workbook.close()


def build_custom_template_rows(
    table: SourceTable,
    layout: CustomTemplateLayout,
    warnings: list[str],
    metadata: dict[str, MetadataRecord] | None = None,
) -> list[list[Any]]:
    source_headers = [header for header in table.headers if header]
    matches: dict[int, str] = {}
    missing_headers: list[str] = []
    ambiguous_headers: list[str] = []
    formula_headers: list[str] = []
    metadata_headers: list[str] = []
    metadata = metadata or {}

    for col_idx, template_header in enumerate(layout.headers):
        if not template_header:
            continue
        match = match_source_header(template_header, source_headers)
        if match.status == "matched" and match.header:
            matches[col_idx] = match.header
        elif match.status == "ambiguous":
            ambiguous_headers.append(template_header)
        elif custom_template_formula_for_col(layout, col_idx):
            formula_headers.append(template_header)
        elif metadata and metadata_candidates_for_header(template_header):
            metadata_headers.append(template_header)
        else:
            missing_headers.append(template_header)

    if missing_headers:
        warnings.append("自定义模板缺少来源字段，已留空：" + "、".join(unique_texts(missing_headers)))
    if ambiguous_headers:
        warnings.append("自定义模板字段匹配不唯一，已留空：" + "、".join(unique_texts(ambiguous_headers)))
    if metadata_headers:
        warnings.append("自定义模板字段已尝试通过补全信息表填充：" + "、".join(unique_texts(metadata_headers)))
    if formula_headers:
        warnings.append("自定义模板字段将按模板公式自动生成：" + "、".join(unique_texts(formula_headers)))

    used_source_headers = {normalize_header(source_header) for source_header in matches.values()}
    unused_headers = [header for header in source_headers if normalize_header(header) not in used_source_headers]
    if unused_headers:
        preview = "、".join(unique_texts(unused_headers)[:20])
        suffix = "等" if len(unique_texts(unused_headers)) > 20 else ""
        warnings.append(f"自定义模板未输出的来源字段：{preview}{suffix}")

    output_rows: list[list[Any]] = []
    for row in table.rows:
        output_row: list[Any] = []
        for col_idx, template_header in enumerate(layout.headers):
            source_header = matches.get(col_idx)
            value = row.get(normalize_header(source_header)) if source_header else None
            source_number_format = get_source_number_format(row, source_header) if source_header else None
            if is_blank(value):
                value = custom_template_metadata_value(row, template_header, metadata)
                source_number_format = None
            number_format = layout.data_number_formats[col_idx] if col_idx < len(layout.data_number_formats) else "General"
            output_row.append(coerce_custom_template_value(value, template_header, number_format, source_number_format))
        output_rows.append(output_row)
    return output_rows


def cell_formula(value: Any) -> str | None:
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def custom_template_formula_for_col(layout: CustomTemplateLayout, col_idx: int) -> str | None:
    if col_idx >= len(layout.data_formulas):
        return None
    return layout.data_formulas[col_idx]


def metadata_candidates_for_header(template_header: str) -> list[str]:
    alias_groups = {
        "REITs代码": ("REITs代码", "ReitsCode", "代码", "证券代码", "基金代码", "产品代码"),
        "REITs名称": ("REITs名称", "REITs简称", "基金名称", "基金简称", "证券简称", "证券名称", "产品名称"),
        "上市日期": ("上市日期", "发行日期"),
        "公告日期": ("公告日期", "INFOPUBLDATE", "信息披露日期", "披露日期"),
        "开始日期": ("开始日期", "STARTDATE", "报告期开始日期", "报告期起始日期", "区间开始日期"),
        "结束日期": ("结束日期", "ENDDATE", "报告期结束日期", "报告期截止日期", "区间结束日期"),
    }
    for field_name, aliases in alias_groups.items():
        if header_matches_alias_group(template_header, aliases):
            return [field_name]
    return []


def header_matches_alias_group(template_header: str, aliases: Iterable[str]) -> bool:
    normalized_header = normalize_header(template_header)
    compact_header = compact_header_key(template_header)
    for alias in aliases:
        normalized_alias = normalize_header(alias)
        compact_alias = compact_header_key(alias)
        if normalized_header == normalized_alias or compact_header == compact_alias:
            return True
        if len(compact_alias) >= 4 and compact_alias in compact_header:
            return True
    return False


def custom_template_metadata_value(
    row: dict[str, Any],
    template_header: str,
    metadata: dict[str, MetadataRecord],
) -> Any:
    candidates = metadata_candidates_for_header(template_header)
    if not candidates or not metadata:
        return None
    record = metadata_record_for_row(row, metadata)
    if not record:
        return None
    return first_non_blank(*(meta_value(record, candidate) for candidate in candidates))


def metadata_record_for_row(row: dict[str, Any], metadata: dict[str, MetadataRecord]) -> MetadataRecord | None:
    code = get_value(row, "REITs代码", "ReitsCode", "代码", "证券代码", "基金代码", "产品代码")
    record = lookup_metadata(metadata, code)
    if record:
        return record
    if len(metadata) == 1:
        return next(iter(metadata.values()))
    return None


@dataclass
class HeaderMatch:
    status: str
    header: str | None = None


def match_source_header(template_header: str, source_headers: list[str]) -> HeaderMatch:
    normalized_sources = {normalize_header(header): header for header in source_headers}
    normalized_template = normalize_header(template_header)
    if normalized_template in normalized_sources:
        return HeaderMatch(status="matched", header=normalized_sources[normalized_template])

    for alias in FIELD_ALIASES.get(template_header, []):
        normalized_alias = normalize_header(alias)
        if normalized_alias in normalized_sources:
            return HeaderMatch(status="matched", header=normalized_sources[normalized_alias])

    for canonical, aliases in FIELD_ALIASES.items():
        normalized_aliases = {normalize_header(canonical), *(normalize_header(alias) for alias in aliases)}
        if normalized_template in normalized_aliases:
            for candidate in normalized_aliases:
                if candidate in normalized_sources:
                    return HeaderMatch(status="matched", header=normalized_sources[candidate])

    compact_sources = {compact_header_key(header): header for header in source_headers if compact_header_key(header)}
    compact_template = compact_header_key(template_header)
    if compact_template and compact_template in compact_sources:
        return HeaderMatch(status="matched", header=compact_sources[compact_template])

    containment_matches = [
        header
        for header in source_headers
        if header_containment_matches(normalized_template, normalize_header(header))
        or header_containment_matches(compact_template, compact_header_key(header))
    ]
    if len(containment_matches) == 1:
        return HeaderMatch(status="matched", header=containment_matches[0])
    if len(containment_matches) > 1:
        return HeaderMatch(status="ambiguous")

    scored = [
        (header_match_score(template_header, header), header)
        for header in source_headers
    ]
    scored = [(score, header) for score, header in scored if score >= 0.82]
    scored.sort(key=lambda item: item[0], reverse=True)
    if not scored:
        return HeaderMatch(status="missing")
    if len(scored) > 1 and scored[0][0] - scored[1][0] < 0.06:
        return HeaderMatch(status="ambiguous")
    return HeaderMatch(status="matched", header=scored[0][1])


def header_containment_matches(template_header: str, source_header: str) -> bool:
    if len(template_header) < 4 or len(source_header) < 4:
        return False
    return template_header in source_header or source_header in template_header


def compact_header_key(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalize_header(value))


def header_match_score(template_header: str, source_header: str) -> float:
    normalized_template = normalize_header(template_header)
    normalized_source = normalize_header(source_header)
    compact_template = compact_header_key(template_header)
    compact_source = compact_header_key(source_header)
    scores = [
        SequenceMatcher(None, normalized_template, normalized_source).ratio(),
        SequenceMatcher(None, compact_template, compact_source).ratio(),
    ]
    token_score = header_token_overlap_score(compact_template, compact_source)
    if token_score:
        scores.append(token_score)
    return max(scores)


def header_token_overlap_score(template_key: str, source_key: str) -> float:
    if not template_key or not source_key:
        return 0.0
    template_tokens = header_tokens(template_key)
    source_tokens = header_tokens(source_key)
    if not template_tokens or not source_tokens:
        return 0.0
    overlap = template_tokens & source_tokens
    return (2 * len(overlap)) / (len(template_tokens) + len(source_tokens))


def header_tokens(value: str) -> set[str]:
    tokens = set(re.findall(r"[a-z]+|\d+|[\u4e00-\u9fff]{2,}", value))
    # Add short meaningful Chinese words that often identify financial fields.
    for token in ("收入", "成本", "费用", "利润", "资产", "负债", "面积", "电价", "电量", "出租率", "收缴率"):
        if token in value:
            tokens.add(token)
    return tokens


def source_headers_are_compatible(reference_headers: list[str], current_headers: list[str]) -> bool:
    return headers_similarity(reference_headers, current_headers) >= CUSTOM_TEMPLATE_FOLDER_SIMILARITY_THRESHOLD


def headers_similarity(reference_headers: list[str], current_headers: list[str]) -> float:
    reference = {normalize_header(header) for header in reference_headers if header}
    current = {normalize_header(header) for header in current_headers if header}
    compact_reference = {compact_header_key(header) for header in reference_headers if compact_header_key(header)}
    compact_current = {compact_header_key(header) for header in current_headers if compact_header_key(header)}
    return max(jaccard_similarity(reference, current), jaccard_similarity(compact_reference, compact_current))


def jaccard_similarity(reference: set[str], current: set[str]) -> float:
    if not reference and not current:
        return 1.0
    if not reference or not current:
        return 0.0
    return len(reference & current) / len(reference | current)


def coerce_custom_template_value(
    value: Any,
    template_header: str,
    number_format: str,
    source_number_format: str | None = None,
) -> Any:
    if is_blank(value):
        return None

    normalized_header = normalize_header(template_header)
    normalized_format = clean_text(number_format).lower()
    if "%" in normalized_format or "率" in template_header:
        return to_ratio(value, source_number_format=source_number_format)
    if is_date_like_header(normalized_header) or is_date_like_format(normalized_format):
        if normalized_header in {"startdate", "enddate"} and not is_date_like_format(normalized_format):
            return to_yyyymmdd_number(value)
        return to_excel_date(value)
    if is_numeric_like_header(template_header) or is_numeric_like_format(normalized_format):
        return to_excel_number(value)
    return value


def is_date_like_header(normalized_header: str) -> bool:
    return "日期" in normalized_header or normalized_header in {"startdate", "enddate", "infopubldate"}


def is_date_like_format(normalized_format: str) -> bool:
    return "yy" in normalized_format or "年" in normalized_format


def is_numeric_like_header(header: str) -> bool:
    numeric_tokens = ("元", "万元", "面积", "数量", "车流量", "发电量", "电量", "电价", "收入", "成本", "费用", "利润", "资产", "负债")
    return any(token in header for token in numeric_tokens)


def is_numeric_like_format(normalized_format: str) -> bool:
    if normalized_format in {"general", "@"}:
        return False
    return any(token in normalized_format for token in ("0", "#"))


def unique_texts(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = normalize_header(text)
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    return result


def transform_valuation(table: SourceTable, metadata: dict[str, MetadataRecord], warnings: list[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    missing_name = False
    missing_list_date = False

    for row in table.rows:
        code = first_non_blank(get_value(row, "ReitsCode"), get_value(row, "REITs代码"))
        meta = lookup_metadata(metadata, code)
        name = first_non_blank(get_value(row, "REITs名称", "REITs简称"), meta_value(meta, "REITs名称"))
        listing_date = first_non_blank(get_value(row, "上市日期"), meta_value(meta, "上市日期"))
        if is_blank(name):
            missing_name = True
        if is_blank(listing_date):
            missing_list_date = True

        output.append(
            {
                "REITs代码": normalize_code(code, keep_suffix=True),
                "REITs名称": name,
                "上市日期": to_excel_date(listing_date),
                "公告日期": to_excel_date(first_non_blank(get_value(row, "公告日期"), meta_value(meta, "公告日期"))),
                "数据来源": get_value(row, "数据来源"),
                "评估基准日": to_yyyymmdd_number(get_value(row, "评估基准日")),
                "评估价值(元)": to_excel_number(get_value(row, "评估价值(元)")),
            }
        )

    if missing_name:
        warnings.append("第1类表缺少 REITs名称：已保留空列，可通过补全信息表自动填写。")
    if missing_list_date:
        warnings.append("第1类表缺少 上市日期：已保留空列，可通过补全信息表自动填写。")
    return output


def transform_traffic(table: SourceTable, metadata: dict[str, MetadataRecord], warnings: list[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in table.rows:
        code = first_non_blank(get_value(row, "ReitsCode"), get_value(row, "REITs代码"))
        start_date = first_non_blank(get_value(row, "开始日期"), get_value(row, "STARTDATE"))
        end_date = first_non_blank(get_value(row, "结束日期"), get_value(row, "ENDDATE"))
        announcement = get_value(row, "公告日期")
        meta = lookup_metadata(metadata, code)

        output.append(
            {
                "REITs代码": code_without_suffix(code),
                "REITs简称": first_non_blank(get_value(row, "REITs名称", "REITs简称"), meta_value(meta, "REITs名称")),
                "STARTDATE": to_yyyymmdd_number(first_non_blank(start_date, meta_value(meta, "开始日期"))),
                "ENDDATE": to_yyyymmdd_number(first_non_blank(end_date, meta_value(meta, "结束日期"))),
                "公告日期": to_excel_date(first_non_blank(announcement, meta_value(meta, "公告日期"))),
                "本期时间范围": first_non_blank(get_value(row, "本期时间范围"), month_range_text(start_date)),
                "资产项目名称": first_non_blank(get_value(row, "资产项目名称"), get_value(row, "基础设施项目名称")),
                "日均车流量(辆次)": to_excel_number(first_non_blank(get_value(row, "日均车流量(辆次)"), get_value(row, "日均收费车流量(辆次)"))),
                "当月路费收入(万元)": to_excel_number(get_value(row, "当月路费收入(万元)")),
            }
        )
    return output


def transform_finance(table: SourceTable, metadata: dict[str, MetadataRecord], warnings: list[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    missing_announcement = False

    for row in table.rows:
        code = first_non_blank(get_value(row, "ReitsCode"), get_value(row, "REITs代码"))
        meta = lookup_metadata(metadata, code)
        announcement = first_non_blank(get_value(row, "公告日期", "INFOPUBLDATE"), meta_value(meta, "公告日期"))
        if is_blank(announcement):
            missing_announcement = True

        output.append(
            {
                "INFOPUBLDATE": to_excel_date(announcement),
                "STARTDATE": to_excel_date(first_non_blank(get_value(row, "开始日期", "STARTDATE"), meta_value(meta, "开始日期"))),
                "ENDDATE": to_excel_date(first_non_blank(get_value(row, "结束日期", "ENDDATE"), meta_value(meta, "结束日期"))),
                "REITs代码": code_without_suffix(code),
                "REITs名称": first_non_blank(get_value(row, "REITs名称", "REITs简称"), meta_value(meta, "REITs名称")),
                "资产项目名称": first_non_blank(get_value(row, "基础设施项目公司名称"), get_value(row, "基础设施项目名称"), get_value(row, "资产项目名称")),
                "数据来源": get_value(row, "数据来源"),
                "主要资产（元）": to_excel_number(first_non_blank(get_value(row, "资产合计（元）"), get_value(row, "主要资产（元）"))),
                "主要负债（元）": to_excel_number(first_non_blank(get_value(row, "负债合计（元）"), get_value(row, "主要负债（元）"))),
                "净资产（元）": to_excel_number(get_value(row, "净资产（元）")),
                "营业收入（元）": to_excel_number(get_value(row, "营业收入（元）")),
                "营业成本及费用（元）": to_excel_number(first_non_blank(get_value(row, "营业成本（元）"), get_value(row, "营业成本及费用（元）"))),
                "净利润（元）": to_excel_number(get_value(row, "净利润（元）")),
            }
        )

    if missing_announcement:
        warnings.append("第3类表存在空公告日期：已保留空值，可通过补全信息表填写。")
    return output


def transform_property(table: SourceTable, metadata: dict[str, MetadataRecord], warnings: list[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    missing_name = False

    for row in table.rows:
        asset_type = clean_text(get_value(row, "主配套资产类别"))
        if asset_type != "主要资产":
            continue

        code = first_non_blank(get_value(row, "ReitsCode"), get_value(row, "REITs代码"))
        meta = lookup_metadata(metadata, code)
        name = first_non_blank(get_value(row, "REITs名称", "REITs简称"), meta_value(meta, "REITs名称"))
        if is_blank(name):
            missing_name = True

        output.append(
            {
                "REITs代码": code_without_suffix(code),
                "REITs名称": name,
                "公告日期": to_excel_date(first_non_blank(get_value(row, "公告日期"), meta_value(meta, "公告日期"))),
                "STARTDATE": to_yyyymmdd_number(first_non_blank(get_value(row, "开始日期", "STARTDATE"), meta_value(meta, "开始日期"))),
                "ENDDATE": to_yyyymmdd_number(first_non_blank(get_value(row, "结束日期", "ENDDATE"), meta_value(meta, "结束日期"))),
                "资产项目名称": first_non_blank(get_value(row, "资产项目名称"), get_value(row, "基础设施项目名称")),
                "出租率": to_ratio(get_value(row, "出租率"), source_number_format=get_source_number_format(row, "出租率")),
                "租金单价(元/月/平方米)": to_excel_number(first_non_blank(get_value(row, "租金单价(元/月/平方米)"), get_value(row, "租金单价(单位:元/月/平方米or元/月/个)"))),
                "租金收缴率": to_ratio(get_value(row, "租金收缴率"), source_number_format=get_source_number_format(row, "租金收缴率")),
                "数据来源": get_value(row, "数据来源"),
            }
        )

    if missing_name:
        warnings.append("第4类表缺少 REITs名称：已保留空列，可通过补全信息表自动填写。")
    return output


def transform_energy(table: SourceTable, metadata: dict[str, MetadataRecord], warnings: list[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in table.rows:
        code = first_non_blank(get_value(row, "ReitsCode"), get_value(row, "REITs代码"))
        meta = lookup_metadata(metadata, code)

        output.append(
            {
                "REITs代码": code_without_suffix(code),
                "REITs名称": first_non_blank(get_value(row, "REITs名称", "REITs简称"), meta_value(meta, "REITs名称")),
                "STARTDATE": to_yyyymmdd_number(first_non_blank(get_value(row, "开始日期", "STARTDATE"), meta_value(meta, "开始日期"))),
                "ENDDATE": to_yyyymmdd_number(first_non_blank(get_value(row, "结束日期", "ENDDATE"), meta_value(meta, "结束日期"))),
                "公告日期": to_excel_date(first_non_blank(get_value(row, "公告日期"), meta_value(meta, "公告日期"))),
                "资产项目名称": first_non_blank(get_value(row, "资产项目名称"), get_value(row, "基础设施项目名称")),
                "发电量(万千瓦时)": to_excel_number(get_value(row, "发电量(万千瓦时)")),
                "结算电量(万千瓦时)": to_excel_number(get_value(row, "结算电量(万千瓦时)")),
                "结算电价(元/千瓦时)": to_excel_number(get_value(row, "结算电价(元/千瓦时)")),
                "发电收入(元)": to_excel_number(get_value(row, "发电收入(元)")),
            }
        )
    return output


def write_standard_workbook(table_type: str, rows: list[dict[str, Any]], output_path: Path) -> None:
    template_path = find_standard_template_path(table_type)
    if template_path:
        write_workbook_from_template(table_type, rows, output_path, template_path)
        return

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAMES[table_type]
    headers = TARGET_HEADERS[table_type]
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    style_plain_sheet(worksheet, row_count=len(rows), column_count=len(headers), table_type=table_type)
    reset_sheet_view_to_a1(worksheet)
    workbook.save(output_path)
    workbook.close()


def find_standard_template_path(table_type: str) -> Path | None:
    template_names = standard_template_names_for(table_type)
    if not template_names:
        return None

    for base in candidate_resource_bases():
        for template_dir_name in STANDARD_TEMPLATE_DIR_NAMES:
            for template_name in template_names:
                path = base / template_dir_name / template_name
                if path.exists():
                    return path
    return None


def write_workbook_from_template(table_type: str, rows: list[dict[str, Any]], output_path: Path, template_path: Path) -> None:
    workbook = load_workbook(template_path)
    worksheet = workbook.active
    headers = TARGET_HEADERS[table_type]
    template_max_column = len(headers)
    header_styles = {col_idx: capture_cell_format(worksheet.cell(1, col_idx)) for col_idx in range(1, template_max_column + 1)}
    header_row_height = worksheet.row_dimensions[1].height
    style_row = 2 if worksheet.max_row >= 2 else 1
    cell_styles = {col_idx: capture_cell_format(worksheet.cell(style_row, col_idx)) for col_idx in range(1, template_max_column + 1)}
    data_row_height = worksheet.row_dimensions[style_row].height
    freeze_panes = worksheet.freeze_panes
    auto_filter_ref = worksheet.auto_filter.ref

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    worksheet.row_dimensions[1].height = header_row_height
    for col_idx in range(1, template_max_column + 1):
        apply_cell_format(worksheet.cell(1, col_idx), header_styles[col_idx])

    for row_offset, row in enumerate(rows, start=2):
        worksheet.row_dimensions[row_offset].height = data_row_height
        for col_idx in range(1, template_max_column + 1):
            cell = worksheet.cell(row_offset, col_idx)
            apply_cell_format(cell, cell_styles[col_idx])
            if col_idx <= len(headers):
                cell.value = row.get(headers[col_idx - 1])
            else:
                cell.value = None

    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = standard_auto_filter_ref(table_type, auto_filter_ref, len(headers), len(rows))
    reset_sheet_view_to_a1(worksheet)
    workbook.save(output_path)
    workbook.close()


def write_custom_template_workbook(layout: CustomTemplateLayout, rows: list[list[Any]], output_path: Path) -> None:
    workbook = load_workbook(layout.path)
    worksheet = workbook.active
    header_row = find_header_row(worksheet)
    template_max_column = max(worksheet.max_column, len(layout.headers))
    header_styles = {col_idx: capture_cell_format(worksheet.cell(header_row, col_idx)) for col_idx in range(1, template_max_column + 1)}
    header_row_height = worksheet.row_dimensions[header_row].height
    style_row = header_row + 1 if worksheet.max_row >= header_row + 1 else header_row
    cell_styles = {col_idx: capture_cell_format(worksheet.cell(style_row, col_idx)) for col_idx in range(1, template_max_column + 1)}
    data_row_height = worksheet.row_dimensions[style_row].height
    freeze_panes = worksheet.freeze_panes
    auto_filter_ref = worksheet.auto_filter.ref

    if worksheet.max_row > header_row:
        worksheet.delete_rows(header_row + 1, worksheet.max_row - header_row)

    worksheet.row_dimensions[header_row].height = header_row_height
    for col_idx in range(1, template_max_column + 1):
        apply_cell_format(worksheet.cell(header_row, col_idx), header_styles[col_idx])

    for row_offset, row_values in enumerate(rows, start=header_row + 1):
        worksheet.row_dimensions[row_offset].height = data_row_height
        for col_idx in range(1, template_max_column + 1):
            cell = worksheet.cell(row_offset, col_idx)
            apply_cell_format(cell, cell_styles[col_idx])
            value = row_values[col_idx - 1] if col_idx <= len(row_values) else None
            formula = layout.data_formulas[col_idx - 1] if col_idx <= len(layout.data_formulas) else None
            cell.value = translate_formula(formula, worksheet.cell(style_row, col_idx).coordinate, cell.coordinate) if value is None and formula else value

    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = adjusted_auto_filter_ref(auto_filter_ref, len(rows))
    reset_sheet_view_to_a1(worksheet)
    workbook.save(output_path)
    workbook.close()


def translate_formula(formula: str, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return formula


def adjusted_auto_filter_ref(auto_filter_ref: str | None, row_count: int) -> str | None:
    if not auto_filter_ref:
        return None
    try:
        min_col, min_row, max_col, _ = range_boundaries(auto_filter_ref)
    except ValueError:
        return auto_filter_ref
    max_row = max(min_row, min_row + row_count)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def standard_auto_filter_ref(table_type: str, auto_filter_ref: str | None, header_count: int, row_count: int) -> str | None:
    if not auto_filter_ref:
        return None
    try:
        min_col, min_row, max_col, _ = range_boundaries(auto_filter_ref)
    except ValueError:
        return auto_filter_ref
    max_col = min(max_col, header_count)
    max_row = max(min_row, min_row + row_count)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def reset_sheet_view_to_a1(worksheet: Any) -> None:
    try:
        worksheet.sheet_view.topLeftCell = "A1"
    except Exception:
        pass
    if worksheet.freeze_panes is None:
        # Excel may repair files that keep frozen-pane selections after panes
        # have been removed, so reset to a single plain A1 selection.
        worksheet.sheet_view.pane = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        return
    for selection in worksheet.sheet_view.selection:
        selection.activeCell = "A1"
        selection.sqref = "A1"


def capture_cell_format(cell: Any) -> dict[str, Any]:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "number_format": cell.number_format,
    }


def apply_cell_format(cell: Any, style: dict[str, Any]) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.protection = copy(style["protection"])
    cell.number_format = style["number_format"]


def write_batch_summary(summary_file: Path, items: list[BatchItemResult]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "批量汇总"
    headers = ["文件名", "完整路径", "状态", "识别类型", "输出行数", "输出文件", "问题和提示", "失败原因"]
    worksheet.append(headers)

    for item in items:
        worksheet.append(
            [
                item.input_file.name,
                str(item.input_file),
                item.status,
                TABLE_TYPES.get(item.detected_type or "", item.detected_type or ""),
                item.row_count,
                "\n".join(str(path) for path in item.output_files),
                "\n".join(item.warnings) if item.warnings else "",
                item.error or "",
            ]
        )

    style_summary_sheet(worksheet, len(items), len(headers))
    reset_sheet_view_to_a1(worksheet)
    workbook.save(summary_file)
    workbook.close()


def style_plain_sheet(worksheet: Any, row_count: int, column_count: int, table_type: str | None = None) -> None:
    standard_font = Font(name="等线", bold=False, color="000000", size=11)
    no_border = Border()
    no_fill = PatternFill(fill_type=None)

    for row in worksheet.iter_rows(min_row=1, max_row=max(row_count + 1, 1), max_col=column_count):
        for cell in row:
            cell.fill = no_fill
            cell.font = standard_font
            cell.border = no_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx in range(1, column_count + 1):
        cell = worksheet.cell(1, col_idx)
        cell.font = standard_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{max(row_count + 1, 1)}"

    if table_type:
        headers = TARGET_HEADERS[table_type]
    else:
        headers = [worksheet.cell(1, col_idx).value for col_idx in range(1, column_count + 1)]

    for col_idx, header in enumerate(headers, 1):
        header_text = str(header or "")
        width = max(10, min(42, len(header_text) * 2 + 4))
        if "名称" in header_text or "资产项目" in header_text or header_text == "REITs简称":
            width = 38
        if header_text in {"REITs代码", "STARTDATE", "ENDDATE"}:
            width = 13
        if header_text in {"公告日期", "INFOPUBLDATE", "上市日期"}:
            width = 14
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
        for cell in row:
            header = worksheet.cell(1, cell.column).value
            cell.number_format = get_number_format(table_type, header)


def get_number_format(table_type: str | None, header: Any) -> str:
    if table_type and header in OUTPUT_NUMBER_FORMATS.get(table_type, {}):
        return OUTPUT_NUMBER_FORMATS[table_type][header]
    return "General"


def style_summary_sheet(worksheet: Any, row_count: int, column_count: int) -> None:
    style_plain_sheet(worksheet, row_count=row_count, column_count=column_count)
    widths = [24, 56, 10, 28, 10, 56, 56, 56]
    for col_idx, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width
    for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1, max_col=column_count):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_property_processed_workbook(input_path: Path, output_path: Path) -> int:
    workbook = load_workbook(input_path, data_only=False)
    try:
        worksheet = workbook.active
        header_row = find_header_row(worksheet)
        headers = [clean_text(worksheet.cell(header_row, col).value) for col in range(1, worksheet.max_column + 1)]

        required = {
            "code": ["REITs代码", "ReitsCode"],
            "project": ["基础设施项目名称"],
            "asset_type": ["主配套资产类别"],
            "asset_name": ["主配套资产名称"],
            "rent_single": ["主配套资产单项可出租面积(平方米)"],
            "rent_total": ["主配套资产合计的可出租面积(平方米)"],
            "actual_single": ["主配套资产单项实际出租面积(平方米)"],
            "actual_total": ["主配套资产合计的实际出租面积(平方米)"],
        }
        columns = {key: find_column(headers, candidates) for key, candidates in required.items()}
        missing = [key for key, col in columns.items() if col is None]
        if missing:
            raise ConversionError("第4类处理版缺少必要字段，无法修复：" + "、".join(missing))

        groups: dict[tuple[str, str], list[int]] = {}
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            code = worksheet.cell(row_idx, columns["code"]).value
            project = worksheet.cell(row_idx, columns["project"]).value
            if is_blank(project):
                continue
            key = (code_digits(code), clean_text(project))
            groups.setdefault(key, []).append(row_idx)

        changed_count = 0
        for row_indices in groups.values():
            categories = [clean_text(worksheet.cell(row_idx, columns["asset_type"]).value) for row_idx in row_indices]
            has_main = any("主要" in category for category in categories)
            has_accessory = any("配套" in category for category in categories)

            if has_main and not has_accessory:
                for row_idx in row_indices:
                    if "主要" in clean_text(worksheet.cell(row_idx, columns["asset_type"]).value):
                        changed_count += set_cell_value(worksheet.cell(row_idx, columns["asset_name"]), "未分类")

            changed_count += clean_area_pair(worksheet, row_indices, columns["rent_single"], columns["rent_total"], has_accessory)
            changed_count += clean_area_pair(worksheet, row_indices, columns["actual_single"], columns["actual_total"], has_accessory)

        normalize_property_processed_values(worksheet, header_row, headers)
        style_property_processed_sheet(worksheet, header_row, headers)
        reset_sheet_view_to_a1(worksheet)
        workbook.save(output_path)
        return changed_count
    finally:
        workbook.close()


def normalize_property_processed_values(worksheet: Any, header_row: int, headers: list[str]) -> None:
    header_to_col = {normalize_header(header): idx + 1 for idx, header in enumerate(headers) if header}
    reits_code_col = header_to_col.get(normalize_header("REITs代码"))
    reitscode_col = header_to_col.get(normalize_header("ReitsCode"))

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        raw_code = None
        if reits_code_col:
            raw_code = worksheet.cell(row_idx, reits_code_col).value
        if is_blank(raw_code) and reitscode_col:
            raw_code = worksheet.cell(row_idx, reitscode_col).value

        if reitscode_col and is_blank(worksheet.cell(row_idx, reitscode_col).value) and not is_blank(raw_code):
            worksheet.cell(row_idx, reitscode_col).value = normalize_code(raw_code, keep_suffix=True)
        if reits_code_col and not is_blank(raw_code):
            worksheet.cell(row_idx, reits_code_col).value = code_without_suffix(raw_code)

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row_idx, col_idx)
            if is_blank(cell.value):
                continue
            if header in PROPERTY_PROCESSED_YYYYMMDD_DATE_HEADERS:
                cell.value = to_excel_date(cell.value)
            elif header in {"开始日期", "结束日期"}:
                cell.value = to_yyyymmdd_number(cell.value)
            elif header == "出租率":
                cell.value = to_ratio(cell.value, source_number_format=cell.number_format)
            elif header == "租金收缴率":
                cell.value = to_ratio(cell.value, source_number_format=cell.number_format)
            elif header in PROPERTY_PROCESSED_NUMERIC_HEADERS:
                cell.value = to_excel_number(cell.value)


def style_property_processed_sheet(worksheet: Any, header_row: int, headers: list[str]) -> None:
    clear_workbook_fill(worksheet)
    worksheet.freeze_panes = None
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{worksheet.max_row}"

    for col_idx, header in enumerate(headers, 1):
        width = max(10, min(34, len(str(header or "")) * 2 + 4))
        if "名称" in str(header) or "项目" in str(header):
            width = 28
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row, max_col=len(headers)):
        for cell in row:
            header = headers[cell.column - 1]
            if header in PROPERTY_PROCESSED_YYYYMMDD_DATE_HEADERS:
                cell.number_format = "yyyymmdd"
            elif header in PROPERTY_PROCESSED_PERCENT_HEADERS:
                cell.number_format = "0.00%"
            else:
                cell.number_format = "General"


def clear_workbook_fill(worksheet: Any) -> None:
    no_fill = PatternFill(fill_type=None)
    standard_font = Font(name="等线", bold=False, color="000000", size=11)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.fill = no_fill
            cell.font = standard_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = Border()


def clean_area_pair(worksheet: Any, row_indices: list[int], single_col: int, total_col: int, has_accessory: bool) -> int:
    changed = 0

    if not has_accessory:
        for row_idx in row_indices:
            single_value = parse_decimal(worksheet.cell(row_idx, single_col).value)
            total_value = parse_decimal(worksheet.cell(row_idx, total_col).value)
            if single_value is None and total_value is not None:
                changed += set_cell_decimal(worksheet.cell(row_idx, single_col), total_value)
                single_value = total_value
            if single_value is not None:
                changed += set_cell_decimal(worksheet.cell(row_idx, total_col), single_value)
        return changed

    total_values = [(row_idx, parse_decimal(worksheet.cell(row_idx, total_col).value)) for row_idx in row_indices]
    non_blank_totals = [value for _, value in total_values if value is not None]
    if len(unique_decimals(non_blank_totals)) > 1:
        for row_idx, total_value in total_values:
            if total_value is not None:
                changed += set_cell_decimal(worksheet.cell(row_idx, single_col), total_value)
                changed += set_cell_value(worksheet.cell(row_idx, total_col), None)

    single_values = [(row_idx, parse_decimal(worksheet.cell(row_idx, single_col).value)) for row_idx in row_indices]
    non_blank_singles = [value for _, value in single_values if value is not None]
    unique_singles = unique_decimals(non_blank_singles)

    if non_blank_singles:
        if len(unique_singles) == 1 and len(non_blank_singles) >= 2:
            group_total = non_blank_singles[0]
            for row_idx, single_value in single_values:
                if single_value is not None:
                    changed += set_cell_value(worksheet.cell(row_idx, single_col), None)
            for row_idx in row_indices:
                changed += set_cell_decimal(worksheet.cell(row_idx, total_col), group_total)
        else:
            group_total = sum(non_blank_singles, Decimal("0"))
            for row_idx in row_indices:
                changed += set_cell_decimal(worksheet.cell(row_idx, total_col), group_total)
    elif non_blank_totals and len(unique_decimals(non_blank_totals)) == 1:
        group_total = non_blank_totals[0]
        for row_idx in row_indices:
            changed += set_cell_decimal(worksheet.cell(row_idx, total_col), group_total)

    return changed


def find_column(headers: list[str], candidates: Iterable[str]) -> int | None:
    normalized = {normalize_header(header): idx + 1 for idx, header in enumerate(headers)}
    for candidate in candidates:
        column = normalized.get(normalize_header(candidate))
        if column is not None:
            return column
    return None


def get_value(row: dict[str, Any], *candidates: str) -> Any:
    for candidate in candidates:
        value = row.get(normalize_header(candidate))
        if not is_blank(value):
            return value
    return None


def get_source_number_format(row: dict[str, Any], *candidates: str | None) -> str | None:
    for candidate in candidates:
        if not candidate:
            continue
        value = row.get(format_key(normalize_header(candidate)))
        if not is_blank(value):
            return clean_text(value)
    return None


def format_key(normalized_header: str) -> str:
    return f"__format__:{normalized_header}"


def meta_value(record: MetadataRecord | None, field: str) -> Any:
    if not record:
        return None
    value = record.values.get(field)
    return None if is_blank(value) else value


def first_non_blank(*values: Any) -> Any:
    for value in values:
        if not is_blank(value):
            return value
    return None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def normalize_header(value: Any) -> str:
    text = clean_text(value).lower()
    replacements = {
        "（": "(",
        "）": ")",
        "：": ":",
        "　": "",
        " ": "",
        "\t": "",
        "\r": "",
        "\n": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_code(value: Any, keep_suffix: bool = False) -> Any:
    if is_blank(value):
        return None
    text = clean_text(value).upper()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    if keep_suffix:
        return text
    return code_without_suffix(text)


def code_digits(value: Any) -> str:
    if is_blank(value):
        return ""
    text = clean_text(value).upper()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    match = re.search(r"\d{3,}", text)
    return match.group(0) if match else ""


def code_without_suffix(value: Any) -> Any:
    digits = code_digits(value)
    if not digits:
        return clean_text(value) if not is_blank(value) else None
    try:
        return int(digits)
    except ValueError:
        return digits


def parse_date(value: Any) -> datetime | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float, Decimal)):
        text = str(int(value)) if float(value).is_integer() else str(value)
        if re.fullmatch(r"\d{8}", text):
            try:
                return datetime.strptime(text, "%Y%m%d")
            except ValueError:
                return None
        return None

    text = clean_text(value)
    text = text.replace("/", "-").replace(".", "-")
    if re.fullmatch(r"\d{8}", text):
        try:
            return datetime.strptime(text, "%Y%m%d")
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def to_excel_date(value: Any) -> Any:
    parsed = parse_date(value)
    if parsed:
        return parsed
    return None if is_blank(value) else value


def to_yyyymmdd_number(value: Any) -> Any:
    parsed = parse_date(value)
    if parsed:
        return int(parsed.strftime("%Y%m%d"))
    if is_blank(value):
        return None
    digits = re.sub(r"\D", "", clean_text(value))
    if re.fullmatch(r"\d{8}", digits):
        return int(digits)
    return value


def month_range_text(start_date: Any) -> str | None:
    parsed = parse_date(start_date)
    if not parsed:
        digits = re.sub(r"\D", "", clean_text(start_date))
        if re.fullmatch(r"\d{8}", digits):
            return f"{digits[:4]}年{int(digits[4:6])}月"
        return None
    return f"{parsed.year}年{parsed.month}月"


def parse_decimal(value: Any) -> Decimal | None:
    if is_blank(value):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = clean_text(value).replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def decimal_to_excel(value: Decimal | None) -> Any:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def to_excel_number(value: Any) -> Any:
    number = parse_decimal(value)
    if number is None:
        return None if is_blank(value) else value
    return decimal_to_excel(number)


def to_ratio(value: Any, cap_at_one: bool = False, source_number_format: str | None = None) -> Any:
    number = parse_decimal(value)
    if number is None:
        return None if is_blank(value) else value
    if should_scale_ratio_by_100(value, number, source_number_format):
        number = number / Decimal("100")
    if cap_at_one and number > Decimal("1"):
        number = Decimal("1")
    return decimal_to_excel(number)


def should_scale_ratio_by_100(value: Any, number: Decimal, source_number_format: str | None = None) -> bool:
    if value_is_percent_formatted(value, source_number_format):
        return isinstance(value, str) and clean_text(value).endswith("%")
    return abs(number) > Decimal("1")


def value_is_percent_formatted(value: Any, source_number_format: str | None = None) -> bool:
    return bool(source_number_format and "%" in source_number_format)


def unique_decimals(values: Iterable[Decimal]) -> set[Decimal]:
    return {value.normalize() for value in values}


def set_cell_value(cell: Any, value: Any) -> int:
    old_value = cell.value
    if old_value == value or (is_blank(old_value) and is_blank(value)):
        return 0
    cell.value = value
    return 1


def set_cell_decimal(cell: Any, value: Decimal | None) -> int:
    return set_cell_value(cell, decimal_to_excel(value))


def make_output_path(input_path: Path, output_dir: Path, suffix: str, extension: str) -> Path:
    output_path = output_dir / f"{input_path.stem}{suffix}{extension}"
    if not output_path.exists():
        return output_path
    for index in range(1, 1000):
        candidate = output_dir / f"{input_path.stem}{suffix}_{index}{extension}"
        if not candidate.exists():
            return candidate
    raise ConversionError(f"输出文件已存在且序号过多，请清理后重试：{output_path}")


def write_report(
    report_path: Path,
    input_path: Path,
    table_type: str,
    output_files: list[Path],
    warnings: list[str],
) -> None:
    lines = [
        "REITs Excel 自动审核转换提示",
        f"输入文件：{input_path}",
        f"识别类型：{TABLE_TYPES.get(table_type, table_type)}",
        "",
        "输出文件：",
    ]
    lines.extend(f"- {path}" for path in output_files)
    lines.extend(["", "提示："])
    lines.extend(f"- {warning}" for warning in warnings)
    report_path.write_text("\n".join(lines), encoding="utf-8-sig")
