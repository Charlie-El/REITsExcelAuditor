from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from reit_excel_auditor.transformer import convert_file, find_standard_template_path
from scripts.check_private_files import find_private_files


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def test_metadata_helper_template_headers_are_readable() -> None:
    workbook = load_workbook(PROJECT_ROOT / "examples" / "自动审核补全信息表模板.xlsx")
    worksheet = workbook.active
    headers = [worksheet.cell(1, col_idx).value for col_idx in range(1, 7)]
    creator = workbook.properties.creator
    active_cell = worksheet.sheet_view.selection[0].activeCell
    workbook.close()

    assert headers == ["REITs代码", "REITs名称", "上市日期", "公告日期", "开始日期", "结束日期"]
    assert all("?" not in str(header) for header in headers)
    assert creator == "REITsExcelAuditor"
    assert active_cell == "A1"


def test_annual_update_helper_templates_are_readable() -> None:
    expected_headers = {
        "年度更新_标准导入表模板.xlsx": (1, ["REITs代码", "REITs名称", "基础设施项目类型", "底层资产性质"]),
        "年度更新_统一补充大表模板.xlsx": (2, ["REITs代码", "REITs名称", "基础设施项目类型", "底层资产性质"]),
        "年度更新_管理费率辅助表模板.xlsx": (1, ["REITs代码", "REITs名称", "项目名称", "底层资产性质"]),
        "年度更新_评估价值与资产性质辅助表模板.xlsx": (1, ["REITs代码", "REITs名称", "项目名称", "底层资产性质"]),
        "年度更新_残值参数辅助表模板.xlsx": (1, ["REITs代码", "REITs名称", "项目名称", "残值年度"]),
        "年度更新_公告日期辅助表模板.xlsx": (1, ["REITs代码", "REITs名称", "项目名称", "公告日期"]),
        "年度更新_项目别名映射表模板.xlsx": (1, ["REITs代码", "REITs名称", "原项目名称", "标准项目名称"]),
        "年度更新_基金净资产与折旧摊销参考表模板.xlsx": (1, ["代码", "基金名称", "匹配PDF", "基金净资产(万元)"]),
    }
    template_dir = PROJECT_ROOT / "examples" / "annual_update_helper_templates"

    for filename, (header_row, leading_headers) in expected_headers.items():
        workbook = load_workbook(template_dir / filename)
        worksheet = workbook["填写模板"]
        headers = [worksheet.cell(header_row, col_idx).value for col_idx in range(1, len(leading_headers) + 1)]
        creator = workbook.properties.creator
        active_cell = worksheet.sheet_view.selection[0].activeCell
        workbook.close()

        assert headers == leading_headers
        assert creator == "REITsExcelAuditor"
        assert active_cell == "A1"


def test_annual_update_output_templates_are_readable() -> None:
    template_dir = PROJECT_ROOT / "standard_templates" / "annual_update"
    expected = {
        "产权年报提取模板.xlsx": ("sheet", "A4"),
        "特许经营权年报提取模板.xlsx": ("Sheet1", "A4"),
        "未来现金流模板.xlsx": ("Sheet1", "A2"),
    }

    for filename, (sheet_name, freeze_panes) in expected.items():
        workbook = load_workbook(template_dir / filename)
        worksheet = workbook[sheet_name]
        workbook.close()

        assert worksheet.freeze_panes == freeze_panes
        assert worksheet.max_row >= 2
        assert worksheet.max_column >= 14


def test_config_template_files_exist() -> None:
    config = json.loads((PROJECT_ROOT / "config" / "table_templates.json").read_text(encoding="utf-8"))

    assert config == {
        "valuation": "01-基础资产估值标准模板.xlsx",
        "traffic": "02-高速经营数据标准模板.xlsx",
        "finance": "03-资产负债收入成本标准模板.xlsx",
        "property": "04-产权经营数据标准模板.xlsx",
        "energy": "05-能源经营数据标准模板.xlsx",
    }
    for table_type, template_name in config.items():
        template_path = find_standard_template_path(table_type)
        assert template_path is not None
        assert template_path.name == template_name


def test_standard_template_filters_only_cover_header_columns() -> None:
    expected_filters = {
        "valuation": (None, 7),
        "traffic": ("A1:I1", 9),
        "finance": ("A1:M1", 13),
        "property": (None, 10),
        "energy": ("A1:J1", 10),
    }

    for table_type, (expected_filter, expected_column_count) in expected_filters.items():
        template_path = find_standard_template_path(table_type)
        assert template_path is not None
        workbook = load_workbook(template_path)
        worksheet = workbook.active
        creator = workbook.properties.creator
        actual_filter = worksheet.auto_filter.ref
        active_cell = worksheet.sheet_view.selection[0].activeCell
        headers = [worksheet.cell(1, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
        workbook.close()

        assert creator == "REITsExcelAuditor"
        assert actual_filter == expected_filter
        assert active_cell == "A1"
        assert worksheet.max_column == expected_column_count
        assert all(header is not None for header in headers)


def test_field_alias_config_contains_core_aliases() -> None:
    aliases = json.loads((PROJECT_ROOT / "config" / "field_aliases.json").read_text(encoding="utf-8"))

    assert "基金名称" in aliases["REITs名称"]
    assert "基础设施项目名称" in aliases["资产项目名称"]
    assert "基础设施项目公司名称" in aliases["资产项目名称"]


def test_private_file_scan_flags_local_desktop_paths(tmp_path: Path) -> None:
    notes = tmp_path / "notes.md"
    local_path = "D:" + "\\" + "work" + "\\" + "待审核数据.xlsx"
    notes.write_text(f"local file is {local_path}", encoding="utf-8")

    assert find_private_files(tmp_path) == [notes]


def test_private_file_scan_flags_function_workspace(tmp_path: Path) -> None:
    private_dir = tmp_path / "function"
    private_dir.mkdir()
    private_workbook = private_dir / "年报样例.xlsx"
    private_workbook.write_text("private", encoding="utf-8")

    assert find_private_files(tmp_path) == [private_workbook]


def create_property_source_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "REITs代码",
            "基础设施项目名称",
            "主配套资产类别",
            "主配套资产名称",
            "主配套资产单项可出租面积(平方米)",
            "主配套资产合计的可出租面积(平方米)",
            "主配套资产可出租数量(个/间/套)",
            "主配套资产单项实际出租面积(平方米)",
            "主配套资产合计的实际出租面积(平方米)",
            "出租率",
            "租金收缴率",
            "租金单价(单位:元/月/平方米or元/月/个)",
        ]
    )
    worksheet.append(["180301.SZ", "项目A", "主要资产", None, 100, None, "文档章节未提及", 80, None, 0.9, 1.05, 10])
    worksheet.freeze_panes = "A2"
    workbook.save(path)
    workbook.close()


def test_property_processed_output_is_optional(tmp_path: Path) -> None:
    source = tmp_path / "property_source.xlsx"
    create_property_source_workbook(source)

    result = convert_file(source, selected_type="property", output_dir=tmp_path)

    assert len(result.output_files) == 1
    assert not any("处理版" in path.name for path in result.output_files)


def test_property_processed_output_can_be_enabled(tmp_path: Path) -> None:
    source = tmp_path / "property_source.xlsx"
    create_property_source_workbook(source)

    result = convert_file(source, selected_type="property", output_dir=tmp_path, generate_property_processed=True)

    processed_files = [path for path in result.output_files if "处理版" in path.name]
    assert len(processed_files) == 1
    assert any("已按选项输出处理版" in warning for warning in result.warnings)
    workbook = load_workbook(processed_files[0])
    worksheet = workbook.active
    headers = [worksheet.cell(1, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    rentable_count_col = headers.index("主配套资产可出租数量(个/间/套)") + 1
    rentable_count = worksheet.cell(2, rentable_count_col).value
    workbook.close()

    with ZipFile(processed_files[0]) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert worksheet.freeze_panes is None
    assert rentable_count is None
    assert "<pane" not in sheet_xml
    assert "pane=" not in sheet_xml


def test_custom_template_can_use_metadata_fallback(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "项目名称"])
    worksheet.append(["180301.SZ", "项目A"])
    workbook.save(source)
    workbook.close()

    metadata = tmp_path / "metadata.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "REITs名称", "开始日期"])
    worksheet.append(["180301.SZ", "测试REITs", 20260401])
    workbook.save(metadata)
    workbook.close()

    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "REITs名称", "STARTDATE", "项目名称"])
    worksheet.append([None, None, None, None])
    workbook.save(template)
    workbook.close()

    result = convert_file(source, custom_template_path=template, metadata_path=metadata, output_dir=tmp_path)
    output = result.output_files[0]
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    values = [worksheet.cell(2, col_idx).value for col_idx in range(1, 5)]
    workbook.close()

    assert values == ["180301.SZ", "测试REITs", 20260401, "项目A"]
    assert any("补全信息表填充" in warning for warning in result.warnings)


def test_custom_template_metadata_fallback_accepts_common_aliases(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["项目名称"])
    worksheet.append(["项目A"])
    workbook.save(source)
    workbook.close()

    metadata = tmp_path / "metadata.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["证券代码", "基金简称", "报告期开始日期"])
    worksheet.append(["180301.SZ", "测试REITs", 20260401])
    workbook.save(metadata)
    workbook.close()

    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "基金简称", "STARTDATE", "项目名称"])
    worksheet.append([None, None, None, None])
    workbook.save(template)
    workbook.close()

    result = convert_file(source, custom_template_path=template, metadata_path=metadata, output_dir=tmp_path)
    workbook = load_workbook(result.output_files[0], data_only=False)
    worksheet = workbook.active
    values = [worksheet.cell(2, col_idx).value for col_idx in range(1, 5)]
    workbook.close()

    assert values == ["180301.SZ", "测试REITs", 20260401, "项目A"]


def test_custom_template_preserves_and_translates_formula_columns(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["字段A", "字段B"])
    worksheet.append([2, 3])
    worksheet.append([4, 5])
    workbook.save(source)
    workbook.close()

    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["字段A", "字段B", "合计"])
    worksheet.append([None, None, "=A2+B2"])
    workbook.save(template)
    workbook.close()

    result = convert_file(source, custom_template_path=template, output_dir=tmp_path)
    output = result.output_files[0]
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    formulas = [worksheet.cell(row_idx, 3).value for row_idx in range(2, 4)]
    workbook.close()

    assert formulas == ["=A2+B2", "=A3+B3"]
    assert any("模板公式自动生成" in warning for warning in result.warnings)
