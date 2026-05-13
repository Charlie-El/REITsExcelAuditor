from __future__ import annotations

from pathlib import Path
from datetime import date
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
import pytest

from reit_excel_auditor.annual_update import (
    AnnualUpdateOptions,
    AnnualUpdateError,
    OcrItem,
    apply_annual_period_defaults_to_rows,
    apply_existing_detail_context_to_rows,
    canonicalize_project_names_from_existing_workbooks,
    normalize_one_standard_row,
    discover_annual_files,
    enrich_rows_from_lookups,
    extract_annual_report_financial_rows,
    extract_amount_by_labels,
    extract_docx_text,
    fill_detail_row,
    filter_ai_rows_for_unparsed_ocr_sources,
    find_internal_annual_template_path,
    read_lookup_rows,
    read_standard_rows,
    read_future_cashflow_rows,
    run_annual_update,
    standardize_ocr_with_ai,
    update_detail_workbook,
    update_future_cashflow_workbook,
)
import reit_excel_auditor.annual_update as annual_update


def create_future_cashflow(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(
        [
            "名称",
            "代码",
            "公告日期",
            "基础设施项目类型",
            "底层资产性质",
            "股债",
            "上市日期",
            "上市年份",
            "到期日",
            "报告期",
            "项目名称",
            "自年份",
            "未来增长率",
            "经营期末",
            2026,
            2027,
        ]
    )
    worksheet.append(["测试产权基金", "180000.SZ", 20260325, "园区基础设施", "产权", "股", 20210101, 2021, 20301231, "2025年评估报告", "项目整体", 2028, 0.03, None, 100, 110])
    worksheet.append(["测试特许基金", "508000.SH", 20260326, "交通基础设施", "特许经营权", "股", 20210101, 2021, 20351231, "2025年评估报告", "项目整体", None, None, 50, 200, 210])
    workbook.save(path)
    workbook.close()


def create_property_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(
        [
            "代码",
            "名称",
            "基础资产类型",
            "底层资产性质",
            "项目名称",
            "公告日期",
            "基金上市日",
            "基金到期日",
            "年份",
            "预测现金流金额",
            "增长率预测起始年度",
            "预测现金流增长率",
            "中诚信补充计算现金流年份",
            "现金流折现日期",
            "中诚信补充计算现金流",
            "验证中诚信现金流模拟收益率",
            "基础资产评估价值（万元）",
            "折现率",
            "报告期",
            "评估基准日",
        ]
    )
    worksheet.append(["180000.SZ", "旧名称", "园区基础设施", "产权", "项目整体", None, None, None, 2026, 1, None, None, 2026, None, 1, None, None, None, None, None])
    workbook.save(path)
    workbook.close()


def create_property_checked_reference(path: Path) -> None:
    create_property_workbook(path)
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.column_dimensions["B"].width = 66
    workbook.save(path)
    workbook.close()


def create_standard_context_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(
        [
            "代码",
            "名称",
            "底层资产性质",
            "项目名称",
            "基金上市日",
            "基金到期日",
            "固定管理费率(%)",
            "托管费率(%)",
            "年份",
            "预测现金流金额（万元）",
        ]
    )
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def create_concession_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(
        [
            "代码",
            "名称",
            "基础资产类型",
            "底层资产性质",
            "项目名称",
            "公告日期",
            "基金上市日",
            "基金到期日",
            "年份",
            "现金流折现日期",
            "验证-评估报告披露预测现金流金额（万元）",
            "折现率",
            "报告期",
            "估值基准日期",
            "验证中诚信现金流模拟收益率",
            "评估报告评估价值（万元）",
            "现金流折现日期",
        ]
    )
    worksheet.append(["508000.SH", "旧名称", "交通基础设施", "特许经营权", "项目整体", None, None, None, 2026, None, 1, None, None, None, None, None])
    worksheet.append(["508000.SH", "旧名称", "交通基础设施", "特许经营权", "项目整体", None, None, None, "期末回收", None, 1, None, None, None, None, None])
    workbook.save(path)
    workbook.close()


def create_full_property_workbook(path: Path) -> None:
    headers = [
        "代码",
        "名称",
        "基础资产类型",
        "底层资产性质",
        "项目名称",
        "公告日期",
        "基金上市日",
        "基金到期日",
        "年份",
        "预测现金流金额",
        "增长率预测起始年度",
        "预测现金流增长率",
        "中诚信补充计算现金流年份",
        "现金流折现日期",
        "中诚信补充计算现金流",
        "验证中诚信现金流模拟收益率",
        "基础资产评估价值（万元）",
        None,
        None,
        "残值年度",
        "残值基础数据预测",
        "考虑残值现金流",
        "折现率",
        "报告期",
        "评估基准日",
        "营业收入（万元）",
        "EBITDA（万元）",
        "运营资本披露值（万元）",
        "资本支出披露值（万元）",
        "运营资本调整（万元）",
        "资本支出调整（万元）",
        "预留管理费（万元）",
        "残值",
        "经调-基础资产预测现金流金额（万元）",
        "借款本金（万元）",
        "借款利息（万元）",
        "利息增值税3.26%",
        "基金净资产（万元）",
        "折旧及摊销（万元）",
        "固定管理费率(%)",
        "托管费率(%)",
        "管理费（万元）",
        "托管费",
        "调整浮动管理费",
        "项目公司、ABS及基金扣减合计（万元）",
        "现金流折现日期",
        "经调基金预测现金流金额（万元）",
        "历史市值基准日期",
        "基金市值（万元）",
        "期末基金份额公允价值参考净值（元）",
        "基金份额（万份）",
        "基金预测现金流金额（万元）-市值",
        "ccxIRR(市值)",
        "最新基金市值（万元）",
        "基金预测现金流金额（万元）-最新市值",
        "最新ccxIRR（市值）",
    ]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(headers)
    worksheet.append(["180000.SZ", "旧名称", "园区基础设施", "产权", "项目整体", None, None, None, 2026, 1])
    workbook.save(path)
    workbook.close()


def test_future_cashflow_reader_turns_wide_years_into_standard_rows(tmp_path: Path) -> None:
    future = tmp_path / "future.xlsx"
    create_future_cashflow(future)

    rows = read_future_cashflow_rows(future)

    assert len(rows) == 4
    first = rows[0]
    assert first["REITs代码"] == "180000.SZ"
    assert first["年份"] == 2026
    assert first["预测现金流金额（万元）"] == 100


def test_annual_update_outputs_process_and_updated_workbooks(tmp_path: Path) -> None:
    create_future_cashflow(tmp_path / "未来现金流.xlsx")
    create_property_workbook(tmp_path / "产权表.xlsx")
    create_concession_workbook(tmp_path / "特许经营权表.xlsx")
    output_dir = tmp_path / "out"

    result = run_annual_update(
        AnnualUpdateOptions(
            workspace_path=tmp_path,
            output_dir=output_dir,
            standard_input_path=tmp_path / "未来现金流.xlsx",
            ocr_engine="pdf_text",
            output_start_year=2026,
            excel_open_check=False,
        )
    )

    assert result.standard_row_count == 4
    assert result.ocr_item_count == 0
    assert result.standard_file.exists()
    assert result.standard_file == result.summary_file
    assert result.plan_file.exists()
    assert result.review_file.exists()
    assert result.comparison_file.exists()
    assert result.future_cashflow_file and result.future_cashflow_file.exists()
    assert result.future_cashflow_file.name == "年度更新_未来现金流汇总表.xlsx"
    assert result.property_file and result.property_file.exists()
    assert result.concession_file and result.concession_file.exists()

    workbook = load_workbook(result.property_file, data_only=False)
    worksheet = workbook.active
    assert worksheet.cell(4, 13).value == "期初"
    assert worksheet.cell(5, 2).value == "测试产权基金"
    assert worksheet.cell(5, 10).value == 100
    assert worksheet.cell(6, 9).value == 2027
    assert worksheet.cell(6, 10).value == 110
    assert worksheet.freeze_panes == "A4"
    workbook.close()

    workbook = load_workbook(result.summary_file, data_only=False)
    assert "标准化导入表" in workbook.sheetnames
    assert workbook["标准化导入表"].max_row == 5
    workbook.close()

    workbook = load_workbook(result.concession_file, data_only=False)
    worksheet = workbook.active
    assert worksheet.cell(4, 9).value == "期初"
    assert worksheet.cell(5, 2).value == "测试特许基金"
    assert worksheet.cell(5, 10).value is not None
    assert worksheet.cell(5, 11).value == 200
    assert worksheet.cell(7, 9).value == "期末回收"
    assert worksheet.cell(7, 10).value is not None
    assert worksheet.cell(7, 11).value == 50
    workbook.close()


def test_annual_update_prefers_internal_template_format_for_detail_output(tmp_path: Path) -> None:
    create_future_cashflow(tmp_path / "未来现金流.xlsx")
    create_property_workbook(tmp_path / "产权表.xlsx")
    create_concession_workbook(tmp_path / "特许经营权表.xlsx")
    create_property_checked_reference(tmp_path / "产权-已核_年报提取.xlsx")
    output_dir = tmp_path / "out"

    result = run_annual_update(
        AnnualUpdateOptions(
            workspace_path=tmp_path,
            output_dir=output_dir,
            standard_input_path=tmp_path / "未来现金流.xlsx",
            ocr_engine="pdf_text",
            output_start_year=2026,
            excel_open_check=False,
        )
    )

    assert result.property_file and result.property_file.exists()
    assert result.property_file.name == "产权表_自动更新.xlsx"
    workbook = load_workbook(result.property_file, data_only=False)
    worksheet = workbook.active
    assert worksheet.title == "Sheet"
    assert worksheet.freeze_panes == "A4"
    assert worksheet.column_dimensions["B"].width != 66
    assert worksheet.cell(5, 2).value == "测试产权基金"
    workbook.close()

    summary = load_workbook(result.summary_file, data_only=False)
    assert "产权格式参考表" in [summary["更新结果汇总"].cell(row_idx, 1).value for row_idx in range(1, 20)]
    summary.close()


def test_annual_update_skips_target_workbooks_when_no_matching_rows(tmp_path: Path) -> None:
    create_property_workbook(tmp_path / "产权表.xlsx")
    create_concession_workbook(tmp_path / "特许经营权表.xlsx")
    output_dir = tmp_path / "out"

    result = run_annual_update(
        AnnualUpdateOptions(
            workspace_path=tmp_path,
            output_dir=output_dir,
            ocr_engine="pdf_text",
            max_ocr_pages_per_file=-1,
            excel_open_check=False,
        )
    )

    assert result.standard_row_count == 0
    assert result.property_file is None
    assert result.concession_file is None
    assert any("未读取到可用于填表的标准化现金流数据" in warning for warning in result.warnings)
    assert any("未生成产权正式输出" in warning for warning in result.warnings)
    assert any("未生成特许经营权正式输出" in warning for warning in result.warnings)


def test_future_cashflow_workbook_is_output_not_default_input(tmp_path: Path) -> None:
    create_future_cashflow(tmp_path / "202604reits未来现金流.xlsx")
    output_dir = tmp_path / "out"

    result = run_annual_update(
        AnnualUpdateOptions(
            workspace_path=tmp_path,
            output_dir=output_dir,
            ocr_engine="pdf_text",
            max_ocr_pages_per_file=-1,
            excel_open_check=False,
        )
    )

    assert result.standard_row_count == 0
    assert result.future_cashflow_file and result.future_cashflow_file.exists()
    assert result.future_cashflow_file.name == "年度更新_未来现金流汇总表.xlsx"
    assert read_future_cashflow_rows(result.future_cashflow_file) == []
    assert any("不会把它作为默认输入" in warning for warning in result.warnings)


def test_clean_report_fund_name_strips_image_suffix() -> None:
    name = annual_update.clean_report_fund_name_from_filename(Path("模拟产权基金封闭式基础设施证券投资基金2025年评估报告.png"))

    assert name == "模拟产权基金封闭式基础设施证券投资基金"


def test_cloud_vision_ocr_requires_api_key_for_remote_endpoint(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from PIL import Image

    monkeypatch.delenv("REIT_TEST_MISSING_OCR_KEY", raising=False)
    image_path = tmp_path / "ocr.png"
    Image.new("RGB", (8, 8), "white").save(image_path)

    with pytest.raises(AnnualUpdateError, match="云端视觉 OCR 需要 API Key"):
        annual_update.ocr_image_file(
            image_path,
            "vision_api",
            AnnualUpdateOptions(
                workspace_path=tmp_path,
                ocr_api_key_env="REIT_TEST_MISSING_OCR_KEY",
                ocr_base_url="https://example.com/v1",
            ),
        )


def test_discovery_prefers_manual_cashflow_ocr_folder(tmp_path: Path) -> None:
    create_future_cashflow(tmp_path / "未来现金流.xlsx")
    annual_report_dir = tmp_path / "公募reits年报"
    annual_report_dir.mkdir()
    (annual_report_dir / "完整年报.pdf").write_bytes(b"%PDF-1.4\n")
    manual_dir = tmp_path / "人工出的pdf示例" / "现金流"
    manual_dir.mkdir(parents=True)
    (manual_dir / "现金流截图.pdf").write_bytes(b"%PDF-1.4\n")

    discovered = discover_annual_files(tmp_path)

    assert discovered.ocr_source_path == manual_dir
    assert discovered.pdf_files == [manual_dir / "现金流截图.pdf"]


def test_discovery_does_not_use_full_annual_reports_as_cashflow_ocr_source(tmp_path: Path) -> None:
    annual_report_dir = tmp_path / "公募reits年报"
    annual_report_dir.mkdir()
    annual_pdf = annual_report_dir / "测试基金2025年年度报告.pdf"
    annual_pdf.write_bytes(b"%PDF-1.4\n")

    discovered = discover_annual_files(tmp_path)

    assert annual_pdf in discovered.annual_report_pdf_files
    assert annual_pdf not in discovered.pdf_files


def test_discovery_recognizes_generic_user_named_annual_report_and_ocr_folders(tmp_path: Path) -> None:
    annual_report_dir = tmp_path / "03_公募年报PDF"
    annual_report_dir.mkdir()
    annual_pdf = annual_report_dir / "测试基金2025年年度报告.pdf"
    annual_pdf.write_bytes(b"%PDF-1.4\n")
    ocr_dir = tmp_path / "04_人工OCR摘页"
    ocr_dir.mkdir()
    ocr_pdf = ocr_dir / "基金A_现金流摘页.pdf"
    ocr_pdf.write_bytes(b"%PDF-1.4\n")

    discovered = discover_annual_files(tmp_path)

    assert annual_pdf in discovered.annual_report_pdf_files
    assert discovered.ocr_source_path == ocr_dir
    assert discovered.pdf_files == [ocr_pdf]


def test_discovery_can_use_separate_annual_report_folder(tmp_path: Path) -> None:
    workspace = tmp_path / "工作材料"
    workspace.mkdir()
    create_property_workbook(workspace / "产权表.xlsx")
    report_dir = tmp_path / "单独年报"
    report_dir.mkdir()
    annual_pdf = report_dir / "任意命名.pdf"
    annual_pdf.write_bytes(b"%PDF-1.4\n")

    discovered = discover_annual_files(workspace, annual_report_source_path=report_dir)

    assert discovered.annual_report_source_path == report_dir
    assert annual_pdf in discovered.annual_report_pdf_files


def test_discovery_skips_generated_output_folders(tmp_path: Path) -> None:
    original = tmp_path / "产权表.xlsx"
    create_property_workbook(original)
    generated_dir = tmp_path / "年度更新_输出结果_验收"
    generated_dir.mkdir()
    generated_file = generated_dir / "产权-已核_年报提取_自动更新.xlsx"
    create_property_workbook(generated_file)
    common_output_dir = tmp_path / "输出结果"
    common_output_dir.mkdir()
    generated_future = common_output_dir / "202604reits未来现金流_自动更新.xlsx"
    create_future_cashflow(generated_future)

    discovered = discover_annual_files(tmp_path)

    assert discovered.property_workbook == original
    assert generated_file not in discovered.excel_files
    assert generated_future not in discovered.excel_files
    assert discovered.future_cashflow_workbook is None


def test_discovery_skips_custom_output_directory(tmp_path: Path) -> None:
    original = tmp_path / "产权表.xlsx"
    create_property_workbook(original)
    custom_output_dir = tmp_path / "验收输出"
    custom_output_dir.mkdir()
    generated_file = custom_output_dir / "产权-已核_年报提取_自动更新.xlsx"
    create_property_workbook(generated_file)

    discovered = discover_annual_files(tmp_path, skip_paths=[custom_output_dir])

    assert discovered.property_workbook == original
    assert generated_file not in discovered.excel_files


def test_extract_docx_text_reads_manual_ocr_source(tmp_path: Path) -> None:
    docx = tmp_path / "现金流.docx"
    with ZipFile(docx, "w") as archive:
        archive.writestr(
            "word/document.xml",
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            "<w:body><w:p><w:r><w:t>预测现金流</w:t></w:r></w:p>"
            "<w:p><w:r><w:t>2027 100.5</w:t></w:r></w:p></w:body></w:document>",
        )

    assert "预测现金流" in extract_docx_text(docx)


def test_extract_annual_report_financial_rows_reads_pdf_values(tmp_path: Path) -> None:
    import fitz

    report_dir = tmp_path / "公募reits年报"
    report_dir.mkdir()
    pdf = report_dir / "测试产业园封闭式基础设施证券投资基金2025年年度报告.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text(
        (72, 72),
        "3.1 Key financial data\n"
        "Fund net assets\n"
        "1,234,567,890.12\n"
        "3.3.2 Calculation process\n"
        "Depreciation and amortization\n"
        "12,345,678.90\n",
    )
    document.save(pdf)
    document.close()

    rows, warnings = extract_annual_report_financial_rows(
        [pdf],
        [{"REITs代码": "180999.SZ", "REITs名称": "测试产业园封闭式基础设施证券投资基金"}],
    )

    assert warnings == []
    assert rows[0]["代码"] == "180999.SZ"
    assert rows[0]["基金净资产(万元)"] == 123456.789012
    assert rows[0]["折旧及摊销(万元)"] == 1234.56789


def test_extract_amount_by_labels_handles_split_pdf_headers() -> None:
    text = "\n".join(
        [
            "期末不动",
            "产基金净",
            "资产",
            "1,986,859,841.63",
            "2,086,787,467.86",
        ]
    )

    value = extract_amount_by_labels(text, ["期末不动产基金净资产"], min_abs_value=50_000_000)

    assert value == 1986859841.63


def test_extract_terminal_recovery_from_cashflow_table_terminal_column() -> None:
    text = "\n".join(
        [
            "资产组（含收费权益)现金流的预测表",
            "2026年",
            "2027年",
            "2028年",
            "期末",
            "四、现金流量",
            "47,355.51",
            "56,664.82",
            "66,106.91",
            "8,266.15",
            "折现期",
        ]
    )

    assert annual_update.extract_terminal_recovery(text) == 8266.15


def test_fill_detail_row_updates_property_and_concession_financial_fields() -> None:
    row = {
        "REITs代码": "180000.SZ",
        "REITs名称": "测试基金",
        "基础设施项目类型": "产业园",
        "底层资产性质": "产权",
        "项目名称": "测试项目",
        "公告日期": 20260430,
        "上市日期": 20210101,
        "到期日": 20710101,
        "年份": 2027,
        "预测现金流金额（万元）": 100,
        "现金流折现日期": 20271230,
        "折现率": 0.06,
        "报告期": "2026年报",
        "评估基准日": 20251231,
        "基础资产评估价值（万元）": 12345,
        "残值年度": 2071,
        "残值基础数据预测": 88,
        "考虑残值现金流": 99,
        "基金净资产（万元）": 45678,
        "折旧及摊销（万元）": 789,
        "固定管理费率(%)": 0.01,
        "托管费率(%)": 0.002,
    }
    property_headers = [
        "代码",
        "名称",
        "项目名称",
        "年份",
        "预测现金流金额（万元）",
        "中诚信补充计算现金流年份",
        "现金流折现日期",
        "中诚信补充计算现金流",
        "基础资产评估价值（万元）",
        "评估基准日",
        "残值年度",
        "残值基础数据预测",
        "考虑残值现金流",
        "基金净资产（万元）",
        "折旧及摊销（万元）",
    ]
    concession_headers = [
        "代码",
        "名称",
        "项目名称",
        "年份",
        "现金流折现日期",
        "验证-评估报告披露预测现金流金额（万元）",
        "评估报告评估价值（万元）",
        "估值基准日期",
        "残值年度",
        "残值基础数据预测",
        "考虑残值现金流",
        "基金净资产（万元）",
        "折旧及摊销（万元）",
    ]

    for kind, headers in (("property", property_headers), ("concession", concession_headers)):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(headers)
        worksheet.append([None] * len(headers))
        col_map = {annual_update.normalize_text(header): idx for idx, header in enumerate(headers, 1)}

        fill_detail_row(worksheet, col_map, 2, row, kind, disclosed=True)

        values = {header: worksheet.cell(2, idx).value for idx, header in enumerate(headers, 1)}
        assert values["代码"] == "180000.SZ"
        assert values["现金流折现日期"] == 20271230
        assert values["残值年度"] == 2071
        assert values["残值基础数据预测"] == 88
        assert values["考虑残值现金流"] == 99
        assert values["基金净资产（万元）"] == 45678
        assert values["折旧及摊销（万元）"] == 789
        workbook.close()


def test_update_detail_workbook_does_not_carry_optional_disclosure_formula_when_missing(tmp_path: Path) -> None:
    source = tmp_path / "产权表.xlsx"
    create_property_workbook(source)
    workbook = load_workbook(source)
    worksheet = workbook.active
    worksheet.cell(3, 21).value = "营业收入（万元）"
    worksheet.cell(4, 21).value = "=J4*2"
    workbook.save(source)
    workbook.close()

    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]
    output, _count = update_detail_workbook(source, rows, "property", tmp_path, [])

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    assert worksheet.cell(4, 21).value is None
    workbook.close()


def test_property_whole_project_adds_initial_and_residual_formula_block(tmp_path: Path) -> None:
    source = tmp_path / "产权表.xlsx"
    create_full_property_workbook(source)
    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
        },
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2028,
            "预测现金流金额（万元）": 120,
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
        },
    ]

    output, _count = update_detail_workbook(source, rows, "property", tmp_path, [])

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    assert worksheet.cell(4, 13).value == "期初"
    assert worksheet.cell(4, 15).value == "=-Q5"
    assert worksheet.cell(4, 16).value == "=XIRR(O4:O6,N4:N6)"
    assert worksheet.cell(7, 20).value is None
    assert worksheet.cell(7, 21).value == "残值"
    assert worksheet.cell(7, 22).value == "=XNPV(0.06,U8:U17,T8:T17)/(1+0.06)"
    assert worksheet.cell(7, 33).value == "=V7"
    assert worksheet.cell(8, 21).value == "=J6*0.9"
    assert worksheet.cell(9, 21).value == "=U8*0.9"
    workbook.close()


def test_ai_standardization_batches_and_continues_after_failed_batch(monkeypatch) -> None:
    calls: list[list[dict[str, str]]] = []

    def fake_chat(**kwargs):
        calls.append(kwargs["messages"])
        if len(calls) == 2:
            raise AnnualUpdateError("timeout")
        return '{"rows":[{"REITs名称":"测试基金","年份":2027,"预测现金流金额（万元）":100}]}'

    monkeypatch.setattr(annual_update, "call_openai_compatible_chat", fake_chat)
    items = [
        OcrItem(source_file=Path("a.pdf"), page=1, method="pdf-page-ocr", text="预测现金流 " + "1" * 900),
        OcrItem(source_file=Path("b.pdf"), page=1, method="pdf-page-ocr", text="预测现金流 " + "2" * 900),
    ]

    rows, warnings, records = standardize_ocr_with_ai(
        items,
        AnnualUpdateOptions(workspace_path=Path("."), use_ai=True, api_key="test", max_ai_chars=1000),
    )

    assert len(calls) == 2
    assert len(records) == 2
    assert len(rows) == 1
    assert rows[0]["REITs名称"] == "测试基金"
    assert items[0].used_for_ai is True
    assert items[1].used_for_ai is False
    assert "AI 第 2/2 批标准化失败" in warnings[0]
    assert "timeout" in items[1].warning


def test_lookup_enrichment_can_match_ai_rows_by_reits_name(tmp_path: Path) -> None:
    lookup_path = tmp_path / "评估价值.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["代码", "名称", "底层资产性质", "基础资产评估价值（万元）"])
    worksheet.append(["180000.SZ", "中信建投沈阳国际软件园封闭式基础设施证券投资基金", "产权", 12345])
    workbook.save(lookup_path)
    workbook.close()

    rows = [
        {
            "REITs名称": "中信建投沈阳国际软件园封闭式基础设施证券投资基金",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]

    enrich_rows_from_lookups(rows, {}, read_lookup_rows(lookup_path))

    assert rows[0]["REITs代码"] == "180000.SZ"
    assert rows[0]["底层资产性质"] == "产权"
    assert rows[0]["基础资产评估价值（万元）"] == 12345


def test_lookup_enrichment_converts_valuation_yuan_to_ten_thousand_yuan(tmp_path: Path) -> None:
    lookup_path = tmp_path / "评估价值.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "基金名称", "评估价值(元)", "资产性质"])
    worksheet.append(["180000.SZ", "测试基金", 123_450_000, "产权"])
    workbook.save(lookup_path)
    workbook.close()

    rows = [{"REITs代码": "180000.SZ", "项目名称": "项目整体", "年份": 2027, "预测现金流金额（万元）": 100}]

    enrich_rows_from_lookups(rows, {}, read_lookup_rows(lookup_path))

    assert rows[0]["基础资产评估价值（万元）"] == 12345


def test_discover_annual_files_recognizes_standard_helper_templates(tmp_path: Path) -> None:
    valuation_path = tmp_path / "评估价值资产性质辅助表模板.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "资产性质", "基础资产评估价值（万元）"])
    worksheet.append(["180000.SZ", "产权", 12345])
    workbook.save(valuation_path)
    workbook.close()

    announcement_path = tmp_path / "公告日期辅助表模板.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "REITs名称", "公告日期"])
    worksheet.append(["180000.SZ", "测试基金", 20260430])
    workbook.save(announcement_path)
    workbook.close()

    discovered = discover_annual_files(tmp_path)

    assert discovered.valuation_workbook == valuation_path
    assert discovered.announcement_workbook == announcement_path


def test_discover_annual_files_prefers_filled_helper_tables_over_validation_baseline(tmp_path: Path) -> None:
    baseline_dir = tmp_path / "validation_baseline"
    baseline_dir.mkdir()
    helper_dir = tmp_path / "辅助表"
    helper_dir.mkdir()

    baseline_path = baseline_dir / "年度更新_评估价值与资产性质辅助表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "资产性质", "基础资产评估价值（万元）"])
    worksheet.append(["180000.SZ", "产权", 0])
    workbook.save(baseline_path)
    workbook.close()

    helper_path = helper_dir / "评估价值+资产性质.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "资产性质", "基础资产评估价值（万元）"])
    worksheet.append(["180000.SZ", "产权", 12345])
    workbook.save(helper_path)
    workbook.close()

    discovered = discover_annual_files(tmp_path)

    assert discovered.valuation_workbook == helper_path


def test_discover_annual_files_prefers_generic_helper_input_folder(tmp_path: Path) -> None:
    helper_dir = tmp_path / "02_补充资料"
    helper_dir.mkdir()

    helper_path = helper_dir / "评估价值+资产性质.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "资产性质", "基础资产评估价值（万元）"])
    worksheet.append(["180000.SZ", "产权", 12345])
    workbook.save(helper_path)
    workbook.close()

    discovered = discover_annual_files(tmp_path)

    assert discovered.valuation_workbook == helper_path


def test_is_cashflow_metric_label_recognizes_concession_cashflow_labels() -> None:
    assert annual_update.is_cashflow_metric_label("四、税前净现金流")
    assert annual_update.is_cashflow_metric_label("四、现金流量")


def test_discovery_does_not_misclassify_standard_import_as_fee_table(tmp_path: Path) -> None:
    standard_path = tmp_path / "标准导入表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "项目名称", "年份", "预测现金流金额（万元）", "固定管理费率(%)", "托管费率(%)"])
    worksheet.append(["180000.SZ", "项目整体", 2027, 100, 0.03, 0.0002])
    workbook.save(standard_path)
    workbook.close()

    discovered = discover_annual_files(tmp_path)

    assert discovered.fee_workbook is None


def test_discovery_recognizes_unified_helper_table_as_standard_input(tmp_path: Path) -> None:
    unified_path = tmp_path / "统一补充大表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "填写模板"
    worksheet.append(["基础信息", None, None, None, None, None, None, None, None, None])
    worksheet.append(
        [
            "REITs代码",
            "REITs名称",
            "基础设施项目类型",
            "底层资产性质",
            "项目名称",
            "公告日期",
            "上市日期",
            "到期日",
            "年份",
            "预测现金流金额（万元）",
        ]
    )
    worksheet.append(["180000.SZ", "测试基金", "园区基础设施", "产权", "项目整体", 20260430, 20210531, 20710606, 2027, 100])
    workbook.save(unified_path)
    workbook.close()

    discovered = discover_annual_files(tmp_path)

    assert discovered.fee_workbook is None
    assert discovered.standard_input_workbook == unified_path
    rows = read_standard_rows(unified_path)
    assert rows[0]["REITs代码"] == "180000.SZ"
    assert rows[0]["年份"] == 2027
    assert rows[0]["预测现金流金额（万元）"] == 100


def test_announcement_lookup_can_fill_announcement_date(tmp_path: Path) -> None:
    lookup_path = tmp_path / "公告日期辅助表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "公告日期"])
    worksheet.append(["180000.SZ", 20260430])
    workbook.save(lookup_path)
    workbook.close()

    rows = [{"REITs代码": "180000.SZ", "项目名称": "项目整体", "年份": 2027, "预测现金流金额（万元）": 100}]

    enrich_rows_from_lookups(rows, {}, {}, announcement_rows=read_lookup_rows(lookup_path))

    assert rows[0]["公告日期"] == date(2026, 4, 30)


def test_annual_period_defaults_refresh_report_and_valuation_dates(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(["代码", "项目名称", "公告日期"])
    worksheet.append(["180000.SZ", "项目整体", date(2024, 3, 30)])
    workbook.save(detail)
    workbook.close()
    rows = [
        {
            "REITs代码": "180000.SZ",
            "项目名称": "项目整体",
            "年份": 2027,
            "公告日期": date(2024, 4, 15),
            "报告期": "2023年评估报告",
            "评估基准日": date(2023, 12, 31),
        }
    ]
    review_items: list[dict[str, object]] = []

    apply_annual_period_defaults_to_rows(rows, [detail], 2027, review_items)

    assert rows[0]["公告日期"] == date(2027, 4, 15)
    assert rows[0]["报告期"] == "2026年评估报告"
    assert rows[0]["评估基准日"] == date(2026, 12, 31)
    assert any(item["类型"] == "年度字段更新" for item in review_items)


def test_annual_period_defaults_uses_existing_announcement_month_day_when_missing(tmp_path: Path) -> None:
    detail = tmp_path / "特许经营权表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(["代码", "项目名称", "公告日期"])
    worksheet.append(["508000.SH", "项目整体", date(2025, 3, 26)])
    workbook.save(detail)
    workbook.close()
    rows = [
        {
            "REITs代码": "508000.SH",
            "项目名称": "项目整体",
            "年份": 2027,
            "报告期": "2024年评估报告",
        }
    ]

    apply_annual_period_defaults_to_rows(rows, [detail], 2027, [])

    assert rows[0]["公告日期"] == date(2027, 3, 26)
    assert rows[0]["报告期"] == "2026年评估报告"
    assert rows[0]["评估基准日"] == date(2026, 12, 31)


def test_expand_standard_rows_with_year_skeletons_computes_growth_years(tmp_path: Path) -> None:
    skeleton = tmp_path / "产权表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["代码", "项目名称", "年份"])
    for year in range(2026, 2031):
        worksheet.append(["180000.SZ", "项目整体", year])
    workbook.save(skeleton)
    workbook.close()

    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试基金",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2026,
            "预测现金流金额（万元）": 100.0,
            "到期日": date(2030, 12, 31),
        },
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试基金",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 110.0,
            "增长率预测起始年度": 2028,
            "预测现金流增长率": 0.1,
            "到期日": date(2030, 12, 31),
        },
    ]

    review_items: list[dict[str, object]] = []
    annual_update.expand_standard_rows_with_year_skeletons(rows, [skeleton], review_items)

    amounts = {row["年份"]: row.get("预测现金流金额（万元）") for row in rows if isinstance(row.get("年份"), int)}
    assert amounts[2028] == 121.0
    assert amounts[2029] == 133.1
    assert amounts[2030] == 146.41
    assert any(item["类型"] == "年份骨架补齐" for item in review_items)


def test_expand_standard_rows_with_year_skeletons_keeps_missing_years_blank_without_growth(tmp_path: Path) -> None:
    rows = [
        {
            "REITs代码": "180001.SZ",
            "REITs名称": "测试基金",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2026,
            "预测现金流金额（万元）": 100.0,
            "到期日": date(2029, 12, 31),
        },
        {
            "REITs代码": "180001.SZ",
            "REITs名称": "测试基金",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 105.0,
            "到期日": date(2029, 12, 31),
        },
    ]

    annual_update.expand_standard_rows_with_year_skeletons(rows, [], [])

    years = sorted(row["年份"] for row in rows if isinstance(row.get("年份"), int))
    amounts = {row["年份"]: row.get("预测现金流金额（万元）") for row in rows if isinstance(row.get("年份"), int)}
    assert years == [2026, 2027, 2028, 2029]
    assert amounts[2028] is None
    assert amounts[2029] is None


def test_apply_lookup_to_standard_row_fills_missing_reits_code() -> None:
    row = {
        "REITs名称": "测试基金",
        "项目名称": "项目整体",
        "年份": 2027,
        "预测现金流金额（万元）": 100.0,
    }
    lookup = {
        "REITs代码": "180001.SZ",
        "REITs名称": "测试基金",
        "底层资产性质": "产权",
    }

    annual_update.apply_lookup_to_standard_row(row, lookup, "valuation")

    assert row["REITs代码"] == "180001.SZ"
    assert row["底层资产性质"] == "产权"


def test_code_level_valuation_is_not_written_to_detail_project_rows(tmp_path: Path) -> None:
    lookup_path = tmp_path / "评估价值.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "基金名称", "评估价值(元)", "资产性质"])
    worksheet.append(["180000.SZ", "测试基金", 123_450_000, "产权"])
    workbook.save(lookup_path)
    workbook.close()

    rows = [{"REITs代码": "180000.SZ", "项目名称": "明细项目", "年份": 2027, "预测现金流金额（万元）": 100}]

    enrich_rows_from_lookups(rows, {}, read_lookup_rows(lookup_path))

    assert "基础资产评估价值（万元）" not in rows[0]


def test_residual_lookup_is_project_specific_and_does_not_pollute_year_rows(tmp_path: Path) -> None:
    lookup_path = tmp_path / "残值参数辅助表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "项目名称", "残值年度", "残值基础数据预测", "折现率"])
    worksheet.append(["180000.SZ", "项目整体", 2031, 88, "6%"])
    worksheet.append(["180000.SZ", "明细项目", 2032, 99, "7%"])
    workbook.save(lookup_path)
    workbook.close()

    rows = [
        {"REITs代码": "180000.SZ", "项目名称": "项目整体", "年份": 2027, "预测现金流金额（万元）": 100},
        {"REITs代码": "180000.SZ", "项目名称": "明细项目", "年份": 2027, "预测现金流金额（万元）": 50},
    ]

    enrich_rows_from_lookups(rows, {}, {}, residual_rows=read_lookup_rows(lookup_path))

    assert rows[0]["折现率"] == 0.06
    assert rows[0][annual_update.RESIDUAL_PARAMETER_KEY]["残值年度"] == 2031
    assert rows[0][annual_update.RESIDUAL_PARAMETER_KEY]["残值基础数据预测"] == 88
    assert "残值年度" not in rows[0]
    assert rows[1]["折现率"] == 0.07
    assert rows[1][annual_update.RESIDUAL_PARAMETER_KEY]["残值基础数据预测"] == 99


def test_standard_input_residual_fields_are_used_as_parameters_not_year_cells() -> None:
    rows = [
        {
            "REITs代码": "180000.SZ",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
            "残值年度": 2032,
            "残值基础数据预测": 88,
            "考虑残值现金流": 99,
        }
    ]

    annual_update.stash_residual_parameters_from_rows(rows)

    assert "残值年度" not in rows[0]
    assert "残值基础数据预测" not in rows[0]
    assert rows[0][annual_update.RESIDUAL_PARAMETER_KEY]["残值年度"] == 2032
    assert rows[0][annual_update.RESIDUAL_PARAMETER_KEY]["残值基础数据预测"] == 88
    assert rows[0][annual_update.RESIDUAL_PARAMETER_KEY]["考虑残值现金流"] == 99


def test_property_residual_row_does_not_write_year_back_into_tail(tmp_path: Path) -> None:
    source = tmp_path / "产权表.xlsx"
    create_full_property_workbook(source)
    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
            "残值年度": 2032,
            "残值基础数据预测": 88,
        },
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2028,
            "预测现金流金额（万元）": 120,
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
        },
    ]

    output, _count = update_detail_workbook(source, rows, "property", tmp_path, [])

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    headers = [worksheet.cell(3, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    year_col = headers.index("年份") + 1
    residual_base_col = headers.index("残值基础数据预测") + 1
    residual_row = None
    for row_idx in range(4, worksheet.max_row + 1):
        if worksheet.cell(row_idx, residual_base_col).value == "残值":
            residual_row = row_idx
            break
    assert residual_row is not None
    assert worksheet.cell(residual_row, year_col).value in (None, "")
    workbook.close()


def test_property_residual_helper_rows_keep_numeric_format(tmp_path: Path) -> None:
    source = tmp_path / "产权表.xlsx"
    create_full_property_workbook(source)
    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
            "残值年度": 2032,
            "残值基础数据预测": 88,
        },
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试产权基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2028,
            "预测现金流金额（万元）": 120,
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
        },
    ]

    output, _count = update_detail_workbook(source, rows, "property", tmp_path, [])

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    headers = [worksheet.cell(3, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    residual_base_col = headers.index("残值基础数据预测") + 1
    residual_year_col = headers.index("残值年度") + 1
    helper_row = None
    for row_idx in range(4, worksheet.max_row + 1):
        value = worksheet.cell(row_idx, residual_base_col).value
        if isinstance(value, str) and value.startswith("="):
            helper_row = row_idx
            break
    assert helper_row is not None
    assert worksheet.cell(helper_row, residual_base_col).number_format == "0.00"
    assert worksheet.cell(helper_row, residual_year_col).number_format == "yyyy-mm-dd"
    workbook.close()


def test_lookup_enrichment_replaces_invalid_ocr_code_by_reits_name(tmp_path: Path) -> None:
    lookup_path = tmp_path / "评估价值.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["代码", "名称", "底层资产性质"])
    worksheet.append(["508099.SH", "华夏中海商业资产封闭式基础设施证券投资基金", "产权"])
    workbook.save(lookup_path)
    workbook.close()

    rows = [
        {
            "REITs代码": "I8UOU",
            "REITs名称": "华夏中海商业资产封闭式基础设施证券投资基金",
            "项目名称": "中海商业资产",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]

    enrich_rows_from_lookups(rows, {}, read_lookup_rows(lookup_path))

    assert rows[0]["REITs代码"] == "508099.SH"
    assert rows[0]["底层资产性质"] == "产权"


def test_asset_nature_inference_routes_concession_rows_without_lookup() -> None:
    rows = [
        {
            "REITs名称": "华夏越秀高速公路封闭式基础设施证券投资基金",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]

    enrich_rows_from_lookups(rows, {}, {})

    assert rows[0]["底层资产性质"] == "特许经营权"


def test_local_ocr_extracts_terminal_recovery_for_concession_rows() -> None:
    rows, warnings = annual_update.standardize_ocr_locally(
        [
            OcrItem(
                source_file=Path("平安广州交投广河高速公路封闭式基础设施证券投资基金2025年评估报告.pdf"),
                page=1,
                method="pdf-text",
                text="\n".join(
                    [
                        "运营净收益",
                        "2026年 2027年 2028年",
                        "69708.84 78021.02 88215.13",
                        "期末回收金额 2559.53",
                        "折现率 7.77%",
                        "评估报告评估价值 775030 万元",
                    ]
                ),
            )
        ]
    )

    assert warnings == []
    assert len(rows) == 3
    assert rows[0]["经营期末"] == 2559.53
    assert rows[0]["折现率"] == 0.0777
    assert rows[0]["基础资产评估价值（万元）"] == 775030


def test_metric_like_project_name_is_cleaned_to_whole_project() -> None:
    row = normalize_one_standard_row({"项目名称": "估价对象全周期运营净收益", "年份": 2027, "预测现金流金额（万元）": 100})

    assert row["项目名称"] == "项目整体"


def test_fee_rate_fields_keep_percentage_point_units() -> None:
    row = normalize_one_standard_row(
        {
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
            "固定管理费率(%)": "0.0365%",
            "托管费率(%)": "0.015%",
            "折现率": "6%",
            "预测现金流增长率": "3%",
        }
    )

    assert row["固定管理费率(%)"] == 0.0365
    assert row["托管费率(%)"] == 0.015
    assert row["折现率"] == 0.06
    assert row["预测现金流增长率"] == 0.03


def test_project_name_can_be_canonicalized_from_existing_workbook(tmp_path: Path) -> None:
    concession = tmp_path / "特许经营权表.xlsx"
    create_concession_workbook(concession)
    workbook = load_workbook(concession)
    worksheet = workbook.active
    worksheet.cell(4, 1).value = "508015.SH"
    worksheet.cell(4, 5).value = "洁源黄骅新能源资产组"
    workbook.save(concession)
    workbook.close()

    rows = [
        {
            "REITs代码": "508015.SH",
            "项目名称": "洁源黄华新能源资产组",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]

    canonicalize_project_names_from_existing_workbooks(rows, [concession])

    assert rows[0]["项目名称"] == "洁源黄骅新能源资产组"
    assert "项目名称按原表模糊修正" in rows[0]["备注"]


def test_project_name_canonicalization_prefers_detail_workbook_over_future_table(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    future = tmp_path / "未来现金流.xlsx"
    create_standard_context_workbook(
        detail,
        [
            [
                "180101.SZ",
                "招商局光明科技园",
                "产权",
                "招商局光明科技园（科技企业加速器二期）",
                20210621,
                20710606,
                0.0365,
                0.015,
                2027,
                100,
            ]
        ],
    )
    create_future_cashflow(future)
    workbook = load_workbook(future)
    worksheet = workbook.active
    worksheet.cell(2, 2).value = "180101.SZ"
    worksheet.cell(2, 11).value = "招商局光明科技园"
    workbook.save(future)
    workbook.close()

    rows = [
        {
            "REITs代码": "180101.SZ",
            "项目名称": "招商局光明科技园",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]

    canonicalize_project_names_from_existing_workbooks(rows, [detail, None, future])

    assert rows[0]["项目名称"] == "招商局光明科技园（科技企业加速器二期）"
    assert "项目名称按原表模糊修正" in rows[0]["备注"]


def test_project_name_canonicalization_keeps_whole_project_when_source_has_whole_project(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    create_standard_context_workbook(
        detail,
        [
            ["180001.SZ", "模拟产业园REIT", "产权", "项目整体", 20210531, 20600630, 0.01, 0.0015, 2027, 12000],
            ["180001.SZ", "模拟产业园REIT", "产权", "产业园一期", 20210531, 20600630, 0.01, 0.0015, 2027, 6100],
        ],
    )
    rows = [
        {
            "REITs代码": "180001.SZ",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 12000,
        }
    ]

    canonicalize_project_names_from_existing_workbooks(rows, [detail])

    assert rows[0]["项目名称"] == "项目整体"
    assert "备注" not in rows[0]


def test_existing_detail_context_preserves_standard_dates_but_not_current_fee_rates(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    create_standard_context_workbook(
        detail,
        [
            ["180000.SZ", "旧名称", "产权", "项目整体", 20210531, 20710606, 0.0365, 0.015, 2026, 1],
            ["180000.SZ", "旧名称", "产权", "项目整体", 20210531, 20710607, 0.0365, 0.015, 2027, 2],
        ],
    )
    rows = [
        {
            "REITs代码": "180000.SZ",
            "项目名称": "项目整体",
            "年份": 2027,
            "上市日期": 20210601,
            "到期日": 20710608,
            "固定管理费率(%)": 0.0365,
            "托管费率(%)": 0.00015,
            "预测现金流金额（万元）": 100,
        }
    ]
    review_items: list[dict[str, object]] = []

    apply_existing_detail_context_to_rows(rows, [detail], review_items)

    assert rows[0]["上市日期"] == 20210531
    assert rows[0]["到期日"] == 20710607
    assert rows[0]["固定管理费率(%)"] == 0.0365
    assert rows[0]["托管费率(%)"] == 0.00015
    assert "标准审核表保留" in rows[0]["备注"]
    assert any(item["类型"] == "标准表口径覆盖" and "基金到期日" in item["对象"] for item in review_items)
    assert not any(item["类型"] == "标准表口径覆盖" and "托管费率" in item["对象"] for item in review_items)


def test_strict_mode_does_not_backfill_missing_discount_rate_from_last_year(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(["代码", "名称", "项目名称", "基金上市日", "基金到期日", "折现率", "年份"])
    worksheet.append(["180000.SZ", "旧名称", "项目整体", 20210531, 20710607, 0.065, 2026])
    workbook.save(detail)
    workbook.close()
    rows = [
        {
            "REITs代码": "180000.SZ",
            "项目名称": "项目整体",
            "年份": 2027,
            "上市日期": 20240101,
            "到期日": 20881231,
            "折现率": None,
        }
    ]
    review_items: list[dict[str, object]] = []

    apply_existing_detail_context_to_rows(rows, [detail], review_items, allow_fill_missing=False)

    assert rows[0]["上市日期"] == 20210531
    assert rows[0]["到期日"] == 20710607
    assert rows[0]["折现率"] is None
    assert not any(item["类型"] == "源表口径补齐" for item in review_items)


def test_strict_mode_uses_default_announcement_date_instead_of_last_year_context(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(["代码", "名称", "项目名称", "公告日期", "年份"])
    worksheet.append(["180000.SZ", "旧名称", "项目整体", 20250326, 2026])
    workbook.save(detail)
    workbook.close()
    rows = [{"REITs代码": "180000.SZ", "项目名称": "项目整体", "年份": 2027}]
    review_items: list[dict[str, object]] = []

    apply_annual_period_defaults_to_rows(rows, [detail], 2027, review_items, allow_existing_context_fill=False)

    assert rows[0]["报告期"] == "2026年评估报告"
    assert rows[0]["评估基准日"] == date(2026, 12, 31)
    assert rows[0]["公告日期"] == date(2027, 3, 31)


def test_project_alias_mapping_table_is_discovered_and_applied(tmp_path: Path) -> None:
    alias_path = tmp_path / "项目别名映射表.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["REITs代码", "OCR项目名称", "标准项目名称"])
    worksheet.append(["180000.SZ", "OCR简称项目", "正式披露项目名称"])
    workbook.save(alias_path)
    workbook.close()
    rows = [{"REITs代码": "180000.SZ", "项目名称": "OCR简称项目", "年份": 2027}]
    review_items: list[dict[str, object]] = []

    discovered = discover_annual_files(tmp_path)
    aliases = annual_update.read_project_alias_rows(discovered.project_alias_workbook)
    annual_update.apply_project_aliases_to_rows(rows, aliases, review_items)

    assert discovered.project_alias_workbook == alias_path
    assert rows[0]["项目名称"] == "正式披露项目名称"
    assert "项目名称按别名映射修正" in rows[0]["备注"]
    assert any(item["类型"] == "项目别名映射" for item in review_items)


def test_future_cashflow_uses_future_table_names_while_detail_uses_standard_names(tmp_path: Path) -> None:
    future = tmp_path / "未来现金流.xlsx"
    create_future_cashflow(future)
    workbook = load_workbook(future)
    worksheet = workbook.active
    worksheet.delete_rows(3)
    worksheet.cell(2, 2).value = "180101.SZ"
    worksheet.cell(2, 11).value = "招商局光明科技园"
    workbook.save(future)
    workbook.close()
    property_workbook = tmp_path / "产权表.xlsx"
    create_property_workbook(property_workbook)
    workbook = load_workbook(property_workbook)
    worksheet = workbook.active
    worksheet.cell(4, 1).value = "180101.SZ"
    worksheet.cell(4, 5).value = "招商局光明科技园（科技企业加速器二期）"
    workbook.save(property_workbook)
    workbook.close()
    output_dir = tmp_path / "out"

    result = run_annual_update(
        AnnualUpdateOptions(
            workspace_path=tmp_path,
            output_dir=output_dir,
            standard_input_path=future,
            ocr_engine="pdf_text",
            output_start_year=2026,
            excel_open_check=False,
        )
    )

    future_rows = read_future_cashflow_rows(result.future_cashflow_file)
    future_names = {row["项目名称"] for row in future_rows if row["REITs代码"] == "180101.SZ"}
    assert future_names == {"招商局光明科技园"}

    workbook = load_workbook(result.property_file, data_only=True)
    worksheet = workbook.active
    assert worksheet.cell(4, 5).value == "招商局光明科技园（科技企业加速器二期）"
    workbook.close()


def test_new_project_keeps_imported_dates_and_rates_without_standard_context(tmp_path: Path) -> None:
    detail = tmp_path / "产权表.xlsx"
    create_standard_context_workbook(
        detail,
        [["180000.SZ", "旧名称", "产权", "旧项目", 20210531, 20710607, 0.0365, 0.015, 2027, 2]],
    )
    rows = [
        {
            "REITs代码": "180000.SZ",
            "项目名称": "新项目",
            "年份": 2027,
            "上市日期": 20240101,
            "到期日": 20881231,
            "固定管理费率(%)": 0.02,
            "托管费率(%)": 0.003,
            "预测现金流金额（万元）": 100,
        }
    ]
    review_items: list[dict[str, object]] = []

    apply_existing_detail_context_to_rows(rows, [detail], review_items)

    assert rows[0]["上市日期"] == 20240101
    assert rows[0]["到期日"] == 20881231
    assert rows[0]["固定管理费率(%)"] == 0.02
    assert rows[0]["托管费率(%)"] == 0.003
    assert review_items == []


def test_ai_rows_do_not_override_successfully_parsed_ocr_sources() -> None:
    source_key = "\u6765\u6e90\u6587\u4ef6"
    local_rows = [{source_key: "cashflow.png", "\u5e74\u4efd": 2027}]
    ai_rows = [
        {source_key: "cashflow.png", "\u5e74\u4efd": None},
        {source_key: "hard-case.png", "\u5e74\u4efd": 2028},
    ]

    filtered_rows, skipped_count = filter_ai_rows_for_unparsed_ocr_sources(ai_rows, local_rows)

    assert skipped_count == 1
    assert filtered_rows == [{source_key: "hard-case.png", "\u5e74\u4efd": 2028}]


def test_whole_project_can_map_to_single_existing_project(tmp_path: Path) -> None:
    concession = tmp_path / "特许经营权表.xlsx"
    create_concession_workbook(concession)
    workbook = load_workbook(concession)
    worksheet = workbook.active
    worksheet.cell(4, 1).value = "180201.SZ"
    worksheet.cell(4, 5).value = "广州至河源高速公路广州段项目"
    workbook.save(concession)
    workbook.close()

    rows = [
        {
            "REITs代码": "180201.SZ",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
        }
    ]

    canonicalize_project_names_from_existing_workbooks(rows, [concession])

    assert rows[0]["项目名称"] == "广州至河源高速公路广州段项目"
    assert "项目名称按原表唯一项目修正" in rows[0]["备注"]


def test_group_static_context_is_copied_to_ocr_only_rows() -> None:
    rows = [
        {
            "REITs代码": "508888.SH",
            "REITs名称": "模拟高速特许经营封闭式基础设施证券投资基金",
            "基础设施项目类型": "交通基础设施",
            "底层资产性质": "特许经营权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 18500,
            "折现率": 0.072,
        },
        {
            "REITs代码": "508888.SH",
            "项目名称": "项目整体",
            "年份": 2029,
            "预测现金流金额（万元）": 20150,
        },
    ]

    annual_update.fill_group_static_context(rows)

    assert rows[1]["REITs名称"] == "模拟高速特许经营封闭式基础设施证券投资基金"
    assert rows[1]["基础设施项目类型"] == "交通基础设施"
    assert rows[1]["底层资产性质"] == "特许经营权"
    assert rows[1]["折现率"] == 0.072


def test_future_cashflow_output_inherits_reference_style_with_unified_font(tmp_path: Path) -> None:
    reference = tmp_path / "202604reits未来现金流.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["名称", "代码", "公告日期", "基础设施项目类型", "底层资产性质", "股债", "上市日期", "上市年份", "到期日", "报告期", "项目名称", "自年份", "未来增长率", "经营期末", 2026, 2027])
    worksheet.append(["示例基金", "180000.SZ", date(2026, 3, 24), "园区基础设施", "产权", "股", 20210101, 2021, 20710101, "2025年评估报告", "项目整体", None, None, None, 100, 110])
    worksheet.column_dimensions["A"].width = 43.796875
    worksheet.cell(1, 1).font = Font(name="Microsoft YaHei", bold=True)
    worksheet.cell(1, 1).fill = PatternFill(fill_type=None)
    worksheet.cell(1, 3).number_format = "yyyymmdd"
    worksheet.cell(2, 1).font = Font(name="Microsoft YaHei", bold=False)
    worksheet.cell(2, 1).fill = PatternFill(fill_type=None)
    worksheet.cell(2, 3).number_format = "yyyymmdd"
    worksheet.cell(2, 16).number_format = "0"
    workbook.save(reference)
    workbook.close()

    rows = [
        {
            "REITs代码": "180001.SZ",
            "REITs名称": "输出基金",
            "公告日期": date(2027, 3, 31),
            "基础设施项目类型": "交通基础设施",
            "底层资产性质": "特许经营权",
            "项目名称": "项目整体",
            "年份": 2026,
            "预测现金流金额（万元）": 111,
        },
        {
            "REITs代码": "180001.SZ",
            "REITs名称": "输出基金",
            "公告日期": date(2027, 3, 31),
            "基础设施项目类型": "交通基础设施",
            "底层资产性质": "特许经营权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 123,
        }
    ]
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = update_future_cashflow_workbook(reference, rows, output_dir, [])

    workbook = load_workbook(output)
    worksheet = workbook.active
    try:
        assert worksheet.column_dimensions["A"].width == 43.796875
        assert worksheet.cell(1, 1).fill.fill_type is None
        assert worksheet.cell(2, 1).fill.fill_type is None
        for row_idx in range(1, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                assert worksheet.cell(row_idx, col_idx).font.name == "微软雅黑"
        assert worksheet.cell(1, 3).number_format == "yyyymmdd"
        assert worksheet.cell(2, 3).number_format == "yyyymmdd"
        assert worksheet.cell(2, 16).number_format == "0"
        assert worksheet.cell(2, 15).value == 111
        assert worksheet.cell(2, 16).value == 123
    finally:
        workbook.close()


def test_detail_reference_format_uses_one_normal_data_style_for_concession() -> None:
    reference = Workbook()
    ref_ws = reference.active
    output = Workbook()
    out_ws = output.active
    headers = ["代码", "名称", "基础资产类型", "底层资产性质", "项目名称", "公告日期", "基金上市日", "基金到期日", "年份"] + [f"列{idx}" for idx in range(10, 44)]
    headers[42] = "ccxIRR(市值)"
    for worksheet in (ref_ws, out_ws):
        worksheet.append(["基础信息"])
        worksheet.append([])
        worksheet.append(headers)
        worksheet.append([None, None, None, None, None, None, None, None, "期初"])
        worksheet.append(["180000.SZ", "基金", "交通基础设施", "特许经营权", "项目整体", None, None, None, 2027])
        worksheet.append(["180000.SZ", "基金", "交通基础设施", "特许经营权", "项目整体", None, None, None, 2028])

    ref_ws.cell(5, 43).number_format = "0.00%"
    ref_ws.cell(6, 43).number_format = "General"
    out_ws.cell(5, 43).number_format = "0.00%"
    out_ws.cell(6, 43).number_format = "0.00%"

    annual_update.copy_detail_reference_data_format(out_ws, ref_ws, "concession")

    assert out_ws.cell(5, 43).number_format == "0.00%"
    assert out_ws.cell(6, 43).number_format == "0.00%"


def test_detail_output_styles_normalize_font_and_yellow_fill() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["代码", "名称", "年份", "数据列"]
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(headers)
    worksheet.append(["180000.SZ", "基金", "期初", 1])
    worksheet.append(["180000.SZ", "基金", 2027, 2])
    worksheet.append(["180000.SZ", "基金", "期末回收", 3])
    col_map = annual_update.build_header_col_map(headers)
    thin_side = Side(style="thin", color="000000")
    thick_side = Side(style="thick", color="000000")
    normal_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    wrong_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row_idx, col_idx).font = Font(name="Microsoft YaHei", bold=(row_idx == 3))
            worksheet.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="FFFF00")
            worksheet.cell(row_idx, col_idx).border = normal_border if row_idx == 5 else wrong_border

    annual_update.normalize_detail_output_styles(worksheet, col_map, "concession", 3)

    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            assert worksheet.cell(row_idx, col_idx).font.name == "等线"
    assert worksheet.cell(3, 3).fill.fill_type is None
    assert worksheet.cell(5, 1).fill.fill_type is None
    assert worksheet.cell(6, 1).fill.fill_type is None
    assert worksheet.cell(6, 3).fill.fill_type is None
    assert worksheet.cell(6, 4).fill.fill_type is None
    for row_idx, col_idx in ((4, 1), (6, 4)):
        border = worksheet.cell(row_idx, col_idx).border
        assert border.left.style == "thin"
        assert border.right.style == "thin"
        assert border.top.style == "thin"
        assert border.bottom.style == "thin"

    annual_update.normalize_detail_output_styles(worksheet, col_map, "property", 3)
    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            assert worksheet.cell(row_idx, col_idx).font.name == "Times New Roman"
    assert worksheet.cell(5, 1).fill.fill_type is None
    assert worksheet.cell(6, 3).fill.fill_type is None


def test_property_style_does_not_force_data_borders() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["代码", "名称", "年份", "数据列"]
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(headers)
    worksheet.append(["180000.SZ", "基金", 2027, 1])
    col_map = annual_update.build_header_col_map(headers)

    annual_update.normalize_detail_output_styles(worksheet, col_map, "property", 3)

    border = worksheet.cell(4, 1).border
    assert border.left.style is None
    assert border.right.style is None
    assert border.top.style is None
    assert border.bottom.style is None


def test_detail_autofilter_range_uses_generated_table_size() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    for _ in range(5):
        worksheet.append([None for _ in range(46)])
    worksheet.auto_filter.ref = "A3:AT584"

    annual_update.normalize_detail_autofilter_range(worksheet, 3)

    assert worksheet.auto_filter.ref == "A3:AT5"


def test_internal_annual_template_paths_exist() -> None:
    assert find_internal_annual_template_path("property")
    assert find_internal_annual_template_path("concession")
    assert find_internal_annual_template_path("future_cashflow")
    assert find_internal_annual_template_path("missing") is None


def test_detail_output_keeps_explicit_reference_freeze_panes(tmp_path: Path) -> None:
    source = tmp_path / "产权表.xlsx"
    create_property_workbook(source)
    reference = tmp_path / "产权-已核_年报提取.xlsx"
    create_property_checked_reference(reference)
    workbook = load_workbook(reference)
    worksheet = workbook.active
    worksheet.freeze_panes = "A58"
    workbook.save(reference)
    workbook.close()

    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2026,
            "预测现金流金额（万元）": 100,
            "现金流折现日期": date(2026, 6, 30),
            "折现率": 0.06,
            "报告期": "2026年评估报告",
            "评估基准日": date(2025, 12, 31),
        }
    ]
    output_dir = tmp_path / "out"
    output_dir.mkdir()

    output, _count = update_detail_workbook(source, rows, "property", output_dir, [], reference)

    workbook = load_workbook(output)
    worksheet = workbook.active
    try:
        assert worksheet.freeze_panes == "A58"
    finally:
        workbook.close()


def test_property_output_reapplies_reference_header_style_after_normalization(tmp_path: Path) -> None:
    source = tmp_path / "产权表.xlsx"
    reference = tmp_path / "产权-已核_年报提取.xlsx"
    create_full_property_workbook(source)
    create_full_property_workbook(reference)
    workbook = load_workbook(reference)
    worksheet = workbook.active
    worksheet.cell(1, 37).fill = PatternFill("solid", fgColor="FFCC99")
    worksheet.cell(1, 37).font = Font(name="Times New Roman", bold=True)
    worksheet.cell(1, 1).border = Border(bottom=Side(style="thin"))
    worksheet.cell(1, 37).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    workbook.save(reference)
    workbook.close()

    rows = [
        {
            "REITs代码": "180000.SZ",
            "REITs名称": "测试基金",
            "基础设施项目类型": "园区基础设施",
            "底层资产性质": "产权",
            "项目名称": "项目整体",
            "年份": 2027,
            "预测现金流金额（万元）": 100,
            "现金流折现日期": date(2027, 6, 30),
            "基础资产评估价值（万元）": 1000,
            "折现率": 0.06,
            "报告期": "2026年评估报告",
            "评估基准日": date(2025, 12, 31),
        }
    ]

    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output, _count = update_detail_workbook(source, rows, "property", output_dir, [], reference)

    workbook = load_workbook(output)
    worksheet = workbook.active
    try:
        assert worksheet.cell(1, 37).fill.fill_type == "solid"
        assert worksheet.cell(1, 37).font.name == "Times New Roman"
        assert worksheet.cell(1, 37).font.bold is True
        assert getattr(worksheet.cell(1, 1).border.left, "style", None) is None
        assert worksheet.cell(1, 1).border.bottom.style == "thin"
        assert worksheet.cell(1, 37).border.left.style == "thin"
        assert worksheet.cell(1, 37).border.top.style == "thin"
    finally:
        workbook.close()


def test_trim_worksheet_to_reference_columns_removes_trailing_blank_columns() -> None:
    output = Workbook()
    out_ws = output.active
    reference = Workbook()
    ref_ws = reference.active
    out_ws.append([1, 2, 3])
    ref_ws.append([1, 2])

    annual_update.trim_worksheet_to_reference_columns(out_ws, ref_ws)

    assert out_ws.max_column == 2


def test_concession_interest_vat_formula_matches_reference_rate_cell() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = [
        "代码",
        "年份",
        "经调-基础资产预测现金流金额（万元）",
        "借款本金（万元）",
        "借款利息（万元）",
        "利息增值税3.26%",
    ]
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(headers)
    col_map = annual_update.build_header_col_map(headers)
    worksheet.append([None, "期初", None, None, None, None])
    worksheet.append(["180000.SZ", 2027, None, None, None, None])

    annual_update.ensure_concession_interest_vat_rate_cell(worksheet, col_map, 4)
    annual_update.apply_detail_formulas_for_row(worksheet, col_map, 5, "concession")

    assert worksheet.cell(4, 6).value == 0.0326
    assert worksheet.cell(4, 6).number_format == "0.00%"
    assert worksheet.cell(5, 6).value == "=(C5-D5-E5)*$F$4"


def test_property_interest_vat_formula_uses_inline_rate() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = [
        "代码",
        "年份",
        "经调-基础资产预测现金流金额（万元）",
        "借款本金（万元）",
        "借款利息（万元）",
        "利息增值税3.26%",
    ]
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(headers)
    worksheet.append(["180000.SZ", 2027, None, None, None, None])
    col_map = annual_update.build_header_col_map(headers)

    annual_update.apply_detail_formulas_for_row(worksheet, col_map, 4, "property")

    assert worksheet.cell(4, 6).value == "=(C4-D4-E4)*3.26%"


def test_detail_visible_format_area_extends_blank_review_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["代码", "名称", "年份", "数据列"]
    worksheet.append(["基础信息"])
    worksheet.append([])
    worksheet.append(headers)
    worksheet.append(["180000.SZ", "基金", 2027, 1])
    col_map = annual_update.build_header_col_map(headers)
    thin_side = Side(style="thin", color="000000")
    worksheet.cell(4, 1).font = Font(name="Arial")
    worksheet.cell(4, 1).border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    worksheet.cell(4, 2).font = Font(name="Arial")
    worksheet.cell(4, 2).border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    annual_update.ensure_detail_visible_format_area(worksheet, col_map, 3, min_total_rows=8)
    annual_update.normalize_detail_output_styles(worksheet, col_map, "property", 3)

    assert worksheet.max_row == 8
    assert worksheet.cell(8, 1).font.name == "Times New Roman"
    assert worksheet.cell(8, 1).border.left.style == "thin"
    assert worksheet.cell(8, 1).border.right.style == "thin"
    assert worksheet.cell(8, 1).border.top.style == "thin"
    assert worksheet.cell(8, 1).border.bottom.style == "thin"
